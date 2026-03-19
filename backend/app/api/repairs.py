from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import delete, or_, select
from sqlalchemy.orm import Session, joinedload

from app.api.access import get_repair_visibility_clause
from app.api.deps import get_current_active_user, get_current_admin, get_db
from app.core.paths import STORAGE_ROOT
from app.models.audit import AuditLog
from app.models.document import Document
from app.models.enums import CheckSeverity, RepairStatus, UserRole
from app.models.imports import ImportJob
from app.models.ocr_learning_signal import OcrLearningSignal
from app.models.repair import Repair, RepairCheck, RepairPart, RepairWork
from app.models.user import User
from app.schemas.repair import (
    RepairDeleteResponse,
    RepairCheckUpdateRequest,
    RepairDetailResponse,
    RepairReviewFieldsUpdateRequest,
    RepairServiceUpdateRequest,
    RepairUpdateRequest,
)
from app.services.document_processing import build_manual_review_check, replace_ocr_checks, resolve_service
from app.services.exporting import append_rows, safe_filename
from app.services.repair_report_analysis import build_repair_executive_report


router = APIRouter(prefix="/repairs", tags=["repairs"])

LEARNING_HEADER_FIELDS = (
    "order_number",
    "repair_date",
    "mileage",
    "service_name",
    "work_total",
    "parts_total",
    "vat_total",
    "grand_total",
)
MANUAL_REVIEW_REASON_LABELS = {
    "mileage_missing": "Не удалось определить пробег",
    "order_number_missing": "Не удалось определить номер заказ-наряда",
    "repair_date_invalid": "Дата ремонта распознана с ошибкой",
    "repair_date_missing": "Не удалось определить дату ремонта",
    "service_name_missing": "Не удалось определить сервис",
    "service_name_suspicious": "Название сервиса выглядит сомнительно",
    "service_not_found": "Сервис не найден в справочнике",
    "text_not_found": "Не удалось извлечь текст из документа",
    "vehicle_missing": "Не удалось определить технику",
    "vehicle_not_found": "Техника не найдена в базе",
}
CHECK_REPORT_SECTION_LABELS = {
    "catalogs": "Справочники",
    "labor_norms": "Нормо-часы",
    "amounts": "Суммы и структура",
    "history": "История и аномалии",
    "ocr": "OCR и ручная проверка",
}


def build_repair_snapshot(repair: Repair) -> dict:
    return {
        "order_number": repair.order_number,
        "repair_date": repair.repair_date.isoformat(),
        "mileage": repair.mileage,
        "reason": repair.reason,
        "employee_comment": repair.employee_comment,
        "service_name": repair.service.name if repair.service is not None else None,
        "work_total": float(repair.work_total),
        "parts_total": float(repair.parts_total),
        "vat_total": float(repair.vat_total),
        "grand_total": float(repair.grand_total),
        "status": repair.status.value,
        "is_preliminary": repair.is_preliminary,
        "is_partially_recognized": repair.is_partially_recognized,
        "checks": [
            {
                "id": item.id,
                "check_type": item.check_type,
                "severity": item.severity.value,
                "title": item.title,
                "is_resolved": item.is_resolved,
                "calculation_payload": item.calculation_payload,
            }
            for item in sorted(repair.checks, key=lambda item: item.id)
        ],
        "works": [
            {
                "work_code": item.work_code,
                "work_name": item.work_name,
                "quantity": item.quantity,
                "standard_hours": item.standard_hours,
                "actual_hours": item.actual_hours,
                "price": float(item.price),
                "line_total": float(item.line_total),
                "status": item.status.value,
            }
            for item in sorted(repair.works, key=lambda item: item.id)
        ],
        "parts": [
            {
                "article": item.article,
                "part_name": item.part_name,
                "quantity": item.quantity,
                "unit_name": item.unit_name,
                "price": float(item.price),
                "line_total": float(item.line_total),
                "status": item.status.value,
            }
            for item in sorted(repair.parts, key=lambda item: item.id)
        ],
    }


def build_repair_query():
    return (
        select(Repair)
        .options(
            joinedload(Repair.vehicle),
            joinedload(Repair.service),
            joinedload(Repair.works),
            joinedload(Repair.parts),
            joinedload(Repair.checks),
            joinedload(Repair.documents).joinedload(Document.versions),
        )
    )


def load_repair_for_user(db: Session, repair_id: int, current_user: User) -> Repair:
    stmt = build_repair_query().where(Repair.id == repair_id)
    if current_user.role != UserRole.ADMIN:
        stmt = stmt.where(get_repair_visibility_clause(current_user))

    repair = db.execute(stmt).unique().scalar_one_or_none()
    if repair is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Repair not found")
    return repair


def fetch_repair_history(db: Session, repair_id: int) -> list[AuditLog]:
    stmt = (
        select(AuditLog)
        .options(joinedload(AuditLog.user))
        .where(
            AuditLog.entity_type == "repair",
            AuditLog.entity_id == str(repair_id),
        )
        .order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
    )
    return db.execute(stmt).scalars().all()


def fetch_document_history(db: Session, repair: Repair) -> list[AuditLog]:
    document_ids = [str(item.id) for item in repair.documents]
    if not document_ids:
        return []

    stmt = (
        select(AuditLog)
        .options(joinedload(AuditLog.user))
        .where(
            AuditLog.entity_type == "document",
            AuditLog.entity_id.in_(document_ids),
        )
        .order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
    )
    return db.execute(stmt).scalars().all()


def serialize_document_history_entry(entry: AuditLog, documents_by_id: dict[str, Document]) -> dict:
    document = documents_by_id.get(entry.entity_id)
    new_value = entry.new_value if isinstance(entry.new_value, dict) else {}
    old_value = entry.old_value if isinstance(entry.old_value, dict) else {}
    document_id = document.id if document is not None else None
    if document_id is None and isinstance(new_value.get("document_id"), int):
        document_id = int(new_value["document_id"])

    return {
        "id": entry.id,
        "action_type": entry.action_type,
        "created_at": entry.created_at,
        "user_name": entry.user.full_name if entry.user is not None else None,
        "document_id": document_id,
        "document_filename": (
            document.original_filename
            if document is not None
            else (str(new_value.get("original_filename")) if new_value.get("original_filename") else None)
        ),
        "document_kind": (
            document.kind
            if document is not None
            else None
        ),
        "old_value": old_value or None,
        "new_value": new_value or None,
    }


def serialize_repair(
    repair: Repair,
    history_entries: list[AuditLog],
    document_history_entries: list[AuditLog],
) -> RepairDetailResponse:
    documents = sorted(repair.documents, key=lambda item: (item.created_at, item.id), reverse=True)
    documents_by_id = {str(item.id): item for item in documents}
    executive_report = build_repair_executive_report(
        repair,
        source_payload=get_report_source_payload(repair),
        manual_review_reason_labels=MANUAL_REVIEW_REASON_LABELS,
    )
    return RepairDetailResponse(
        id=repair.id,
        order_number=repair.order_number,
        repair_date=repair.repair_date,
        mileage=repair.mileage,
        reason=repair.reason,
        employee_comment=repair.employee_comment,
        work_total=float(repair.work_total),
        parts_total=float(repair.parts_total),
        vat_total=float(repair.vat_total),
        grand_total=float(repair.grand_total),
        expected_total=float(repair.expected_total) if repair.expected_total is not None else None,
        status=repair.status,
        is_preliminary=repair.is_preliminary,
        is_partially_recognized=repair.is_partially_recognized,
        is_manually_completed=repair.is_manually_completed,
        created_at=repair.created_at,
        updated_at=repair.updated_at,
        vehicle={
            "id": repair.vehicle.id,
            "external_id": repair.vehicle.external_id,
            "plate_number": repair.vehicle.plate_number,
            "brand": repair.vehicle.brand,
            "model": repair.vehicle.model,
        },
        service=(
            {
                "id": repair.service.id,
                "name": repair.service.name,
                "city": repair.service.city,
            }
            if repair.service is not None
            else None
        ),
        works=sorted(repair.works, key=lambda item: item.id),
        parts=sorted(repair.parts, key=lambda item: item.id),
        checks=sorted(repair.checks, key=lambda item: item.id),
        documents=[
            {
                "id": document.id,
                "original_filename": document.original_filename,
                "source_type": document.source_type,
                "kind": document.kind,
                "mime_type": document.mime_type,
                "status": document.status.value,
                "is_primary": document.is_primary,
                "ocr_confidence": document.ocr_confidence,
                "review_queue_priority": document.review_queue_priority,
                "notes": document.notes,
                "created_at": document.created_at,
                "updated_at": document.updated_at,
                "versions": [
                    {
                        "id": version.id,
                        "version_number": version.version_number,
                        "created_at": version.created_at,
                        "change_summary": version.change_summary,
                        "parsed_payload": version.parsed_payload,
                        "field_confidence_map": version.field_confidence_map,
                    }
                    for version in sorted(document.versions, key=lambda item: item.version_number, reverse=True)
                ],
            }
            for document in documents
        ],
        document_history=[
            serialize_document_history_entry(entry, documents_by_id)
            for entry in document_history_entries
        ],
        history=[
            {
                "id": entry.id,
                "action_type": entry.action_type,
                "created_at": entry.created_at,
                "user_name": entry.user.full_name if entry.user is not None else None,
                "old_value": entry.old_value,
                "new_value": entry.new_value,
            }
            for entry in history_entries
        ],
        executive_report=executive_report,
    )


def cleanup_storage_files(storage_keys: set[str]) -> None:
    for storage_key in storage_keys:
        if not storage_key:
            continue
        path = STORAGE_ROOT / storage_key
        if path.exists():
            path.unlink()


def replace_manual_lines(
    db: Session,
    repair: Repair,
    works_payload,
    parts_payload,
) -> None:
    if works_payload is not None:
        db.execute(delete(RepairWork).where(RepairWork.repair_id == repair.id))
        for item in works_payload:
            db.add(
                RepairWork(
                    repair_id=repair.id,
                    work_code=item.work_code,
                    work_name=item.work_name,
                    quantity=item.quantity,
                    standard_hours=item.standard_hours,
                    actual_hours=item.actual_hours,
                    price=item.price,
                    line_total=item.line_total,
                    status=item.status,
                    reference_payload=item.reference_payload,
                )
            )

    if parts_payload is not None:
        db.execute(delete(RepairPart).where(RepairPart.repair_id == repair.id))
        for item in parts_payload:
            db.add(
                RepairPart(
                    repair_id=repair.id,
                    article=item.article,
                    part_name=item.part_name,
                    quantity=item.quantity,
                    unit_name=item.unit_name,
                    price=item.price,
                    line_total=item.line_total,
                    status=item.status,
                )
            )


def get_learning_source_document(repair: Repair) -> Document | None:
    if repair.source_document_id is not None:
        matched = next((item for item in repair.documents if item.id == repair.source_document_id), None)
        if matched is not None:
            return matched
    primary = next((item for item in repair.documents if item.is_primary), None)
    if primary is not None:
        return primary
    return repair.documents[0] if repair.documents else None


def get_latest_document_version(document: Document | None):
    if document is None or not document.versions:
        return None
    return max(document.versions, key=lambda item: item.version_number)


def stringify_learning_value(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def get_report_source_document(repair: Repair) -> Document | None:
    return get_learning_source_document(repair)


def get_report_source_payload(repair: Repair) -> dict:
    source_document = get_report_source_document(repair)
    latest_version = get_latest_document_version(source_document)
    if latest_version is None or not isinstance(latest_version.parsed_payload, dict):
        return {}
    return latest_version.parsed_payload


def get_check_report_section_key(check_type: str) -> str:
    if "vehicle" in check_type or "service" in check_type:
        return "catalogs"
    if "standard_hours" in check_type:
        return "labor_norms"
    if "total" in check_type or "duplicate" in check_type or "expected_total" in check_type:
        return "amounts"
    if "repeat_repair" in check_type:
        return "history"
    return "ocr"


def build_report_status_summary(repair: Repair) -> tuple[str, str]:
    if any(document.status.value == "uploaded" for document in repair.documents):
        return "В очереди OCR", "Документ ещё находится в обработке или перепроверке."

    unresolved_checks = [item for item in repair.checks if not item.is_resolved]
    if any(item.severity in {CheckSeverity.SUSPICIOUS, CheckSeverity.ERROR} for item in unresolved_checks):
        return "Есть критичные несоответствия", "Перед подтверждением нужны ручная проверка и решение по предупреждениям."
    if unresolved_checks:
        return "Требует ручной проверки", "По заказ-наряду есть открытые предупреждения, которые нужно проверить."
    return "Готов к следующему этапу", "Открытых несоответствий по заказ-наряду не найдено."


def build_export_warning_rows(repair: Repair) -> list[tuple[object, ...]]:
    rows: list[tuple[object, ...]] = []
    report_payload = get_report_source_payload(repair)
    extracted_fields = report_payload.get("extracted_fields") if isinstance(report_payload.get("extracted_fields"), dict) else {}
    raw_reasons = report_payload.get("manual_review_reasons")
    manual_review_reasons = [str(item) for item in raw_reasons] if isinstance(raw_reasons, list) else []
    seen_reason_codes: set[str] = set()

    for check in sorted(repair.checks, key=lambda item: (item.is_resolved, item.created_at, item.id)):
        resolution = check.calculation_payload.get("resolution") if isinstance(check.calculation_payload, dict) else None
        resolved_at = resolution.get("resolved_at") if isinstance(resolution, dict) else None
        rows.append(
            (
                CHECK_REPORT_SECTION_LABELS[get_check_report_section_key(check.check_type)],
                check.severity.value,
                check.title,
                check.details or "",
                "check",
                "Да" if check.is_resolved else "Нет",
                resolved_at or "",
            )
        )
        if check.check_type.startswith("ocr_"):
            reason_code = check.check_type.removeprefix("ocr_")
            seen_reason_codes.add(reason_code)

    for reason in manual_review_reasons:
        if reason in seen_reason_codes:
            continue
        manual_check = build_manual_review_check(reason, extracted_fields=extracted_fields)
        rows.append(
            (
                CHECK_REPORT_SECTION_LABELS[get_check_report_section_key(str(manual_check["check_type"]))],
                manual_check["severity"].value,
                str(manual_check["title"]),
                str(manual_check.get("details") or ""),
                "manual_review_reason",
                "Нет",
                "",
            )
        )

    return rows


def update_service_manual_review_state(repair: Repair, service_name: str | None) -> None:
    source_document = get_learning_source_document(repair)
    latest_version = get_latest_document_version(source_document)
    if latest_version is None or not isinstance(latest_version.parsed_payload, dict):
        return

    parsed_payload = dict(latest_version.parsed_payload)
    raw_reasons = parsed_payload.get("manual_review_reasons")
    manual_review_reasons = [str(item) for item in raw_reasons] if isinstance(raw_reasons, list) else []
    manual_review_reasons = [
        item for item in manual_review_reasons if item not in {"service_not_found", "service_name_missing"}
    ]

    extracted_fields_raw = parsed_payload.get("extracted_fields")
    extracted_fields = dict(extracted_fields_raw) if isinstance(extracted_fields_raw, dict) else {}

    if service_name:
        extracted_fields["service_name"] = service_name
    else:
        extracted_fields.pop("service_name", None)
        manual_review_reasons.append("service_name_missing")

    parsed_payload["extracted_fields"] = extracted_fields
    parsed_payload["manual_review_reasons"] = manual_review_reasons
    latest_version.parsed_payload = parsed_payload


def sync_service_checks(repair: Repair, service_name: str | None, current_user: User) -> None:
    service_check_types = {"ocr_service_not_found", "ocr_service_missing"}
    service_checks = [item for item in repair.checks if item.check_type in service_check_types]

    for check in service_checks:
        payload = dict(check.calculation_payload or {})
        should_be_resolved = bool(service_name) or check.check_type == "ocr_service_not_found"
        payload["resolution"] = {
            "is_resolved": should_be_resolved,
            "comment": (
                "Сервис назначен вручную и предупреждение снято"
                if service_name
                else (
                    "Сервис очищен, актуально предупреждение об отсутствии сервиса"
                    if check.check_type == "ocr_service_not_found"
                    else "Сервис очищен, предупреждение возвращено"
                )
            ),
            "user_id": current_user.id,
            "user_name": current_user.full_name,
            "resolved_at": datetime.now(timezone.utc).isoformat(),
        }
        check.is_resolved = should_be_resolved
        check.calculation_payload = payload

    if service_name or any(item.check_type == "ocr_service_missing" for item in service_checks):
        return

    repair.checks.append(
        RepairCheck(
            repair_id=repair.id,
            check_type="ocr_service_missing",
            severity=CheckSeverity.WARNING,
            title="Не удалось определить сервис",
            details="Сервис у ремонта не назначен. Нужна ручная проверка.",
            calculation_payload={"reason": "service_name_missing"},
            is_resolved=False,
        )
    )


def build_repair_learning_snapshot(repair: Repair) -> dict[str, str | None]:
    return {
        "order_number": stringify_learning_value(repair.order_number),
        "repair_date": stringify_learning_value(repair.repair_date.isoformat()),
        "mileage": stringify_learning_value(repair.mileage),
        "service_name": stringify_learning_value(repair.service.name if repair.service is not None else None),
        "work_total": stringify_learning_value(float(repair.work_total)),
        "parts_total": stringify_learning_value(float(repair.parts_total)),
        "vat_total": stringify_learning_value(float(repair.vat_total)),
        "grand_total": stringify_learning_value(float(repair.grand_total)),
    }


def suggestion_for_learning_signal(target_field: str, signal_type: str, profile_scope: str | None) -> str:
    action = "Добавить правило извлечения" if signal_type == "missing" else "Уточнить правило извлечения"
    if profile_scope and profile_scope != "default":
        return f"{action} для поля `{target_field}` в профиле `{profile_scope}`"
    return f"{action} для поля `{target_field}` и проверить необходимость отдельного OCR-профиля"


def create_ocr_learning_signals_for_repair(
    db: Session,
    repair: Repair,
    *,
    current_admin: User,
) -> None:
    source_document = get_learning_source_document(repair)
    latest_version = get_latest_document_version(source_document)
    if latest_version is None:
        return

    parsed_payload = latest_version.parsed_payload if isinstance(latest_version.parsed_payload, dict) else {}
    extracted_fields = parsed_payload.get("extracted_fields") if isinstance(parsed_payload.get("extracted_fields"), dict) else {}
    text_excerpt = parsed_payload.get("text_excerpt") if isinstance(parsed_payload.get("text_excerpt"), str) else None
    profile_scope = parsed_payload.get("ocr_profile_scope") if isinstance(parsed_payload.get("ocr_profile_scope"), str) else None
    repair_snapshot = build_repair_learning_snapshot(repair)

    for field_name in LEARNING_HEADER_FIELDS:
        corrected_value = repair_snapshot.get(field_name)
        if not corrected_value:
            continue
        extracted_value = stringify_learning_value(extracted_fields.get(field_name))
        if extracted_value == corrected_value:
            continue

        signal_type = "missing" if not extracted_value else "mismatch"
        existing = db.scalar(
            select(OcrLearningSignal).where(
                OcrLearningSignal.repair_id == repair.id,
                OcrLearningSignal.document_version_id == latest_version.id,
                OcrLearningSignal.target_field == field_name,
                OcrLearningSignal.extracted_value == extracted_value,
                OcrLearningSignal.corrected_value == corrected_value,
            )
        )
        if existing is not None:
            continue

        db.add(
            OcrLearningSignal(
                repair_id=repair.id,
                document_id=source_document.id if source_document is not None else None,
                document_version_id=latest_version.id,
                created_by_user_id=current_admin.id,
                signal_type=signal_type,
                target_field=field_name,
                ocr_profile_scope=profile_scope,
                extracted_value=extracted_value,
                corrected_value=corrected_value,
                service_name=repair.service.name if repair.service is not None else None,
                source_type=source_document.source_type if source_document is not None else None,
                document_filename=source_document.original_filename if source_document is not None else None,
                text_excerpt=text_excerpt,
                suggestion_summary=suggestion_for_learning_signal(field_name, signal_type, profile_scope),
                status="new",
            )
        )

@router.get("/{repair_id}", response_model=RepairDetailResponse)
def get_repair(
    repair_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> RepairDetailResponse:
    repair = load_repair_for_user(db, repair_id, current_user)
    history_entries = fetch_repair_history(db, repair.id)
    document_history_entries = fetch_document_history(db, repair)
    return serialize_repair(repair, history_entries, document_history_entries)


@router.delete("/{repair_id}", response_model=RepairDeleteResponse)
def delete_repair(
    repair_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> RepairDeleteResponse:
    stmt = build_repair_query().where(Repair.id == repair_id)
    repair = db.execute(stmt).unique().scalar_one_or_none()
    if repair is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Repair not found")

    old_snapshot = build_repair_snapshot(repair)
    old_snapshot["documents"] = [
        {
            "id": document.id,
            "original_filename": document.original_filename,
            "storage_key": document.storage_key,
            "kind": document.kind.value,
            "status": document.status.value,
        }
        for document in sorted(repair.documents, key=lambda item: item.id)
    ]
    storage_keys = {
        document.storage_key
        for document in repair.documents
    } | {
        version.storage_key
        for document in repair.documents
        for version in document.versions
    }
    document_ids = [document.id for document in repair.documents]
    document_version_ids = [version.id for document in repair.documents for version in document.versions]

    db.add(
        AuditLog(
            user_id=current_admin.id,
            entity_type="repair",
            entity_id=str(repair.id),
            action_type="repair_deleted",
            old_value=old_snapshot,
            new_value={"deleted": True},
        )
    )
    db.flush()

    if document_version_ids:
        db.execute(
            delete(OcrLearningSignal).where(
                or_(
                    OcrLearningSignal.repair_id == repair.id,
                    OcrLearningSignal.document_id.in_(document_ids),
                    OcrLearningSignal.document_version_id.in_(document_version_ids),
                )
            )
        )
    else:
        db.execute(delete(OcrLearningSignal).where(OcrLearningSignal.repair_id == repair.id))

    if document_ids:
        db.execute(delete(ImportJob).where(ImportJob.document_id.in_(document_ids)))

    repair.source_document_id = None
    db.flush()

    for document in list(repair.documents):
        db.delete(document)

    db.delete(repair)
    db.commit()
    cleanup_storage_files(storage_keys)

    return RepairDeleteResponse(
        message="Заказ-наряд и связанные документы удалены",
        deleted_repair_id=repair_id,
    )


@router.get("/{repair_id}/export")
def export_repair(
    repair_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> StreamingResponse:
    repair = load_repair_for_user(db, repair_id, current_user)
    source_document = get_report_source_document(repair)
    source_payload = get_report_source_payload(repair)
    report_status, report_status_comment = build_report_status_summary(repair)
    warning_rows = build_export_warning_rows(repair)
    raw_reasons = source_payload.get("manual_review_reasons")
    manual_review_reasons = [str(item) for item in raw_reasons] if isinstance(raw_reasons, list) else []
    extracted_fields = source_payload.get("extracted_fields") if isinstance(source_payload.get("extracted_fields"), dict) else {}

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Отчет"
    append_rows(
        summary_sheet,
        [
            ("Поле", "Значение"),
            ("ID ремонта", repair.id),
            ("Номер заказ-наряда", repair.order_number or ""),
            ("Дата ремонта", repair.repair_date.isoformat()),
            ("Статус", repair.status.value),
            ("Итоговый статус отчёта", report_status),
            ("Комментарий к статусу", report_status_comment),
            ("Предварительный", "Да" if repair.is_preliminary else "Нет"),
            ("Частично распознан", "Да" if repair.is_partially_recognized else "Нет"),
            ("Госномер", repair.vehicle.plate_number or ""),
            ("VIN", repair.vehicle.vin or ""),
            ("Марка", repair.vehicle.brand or ""),
            ("Модель", repair.vehicle.model or ""),
            ("Сервис", repair.service.name if repair.service is not None else ""),
            ("Сервис по OCR", extracted_fields.get("service_name") if extracted_fields else ""),
            ("Пробег", repair.mileage),
            ("Причина ремонта", repair.reason or ""),
            ("Комментарий сотрудника", repair.employee_comment or ""),
            ("Работы, руб", float(repair.work_total)),
            ("Запчасти, руб", float(repair.parts_total)),
            ("НДС, руб", float(repair.vat_total)),
            ("Итого, руб", float(repair.grand_total)),
            ("Ожидаемая стоимость, руб", float(repair.expected_total) if repair.expected_total is not None else ""),
            ("Основной документ", source_document.original_filename if source_document is not None else ""),
            ("Статус основного документа", source_document.status.value if source_document is not None else ""),
            (
                "Причины ручной проверки OCR",
                ", ".join(MANUAL_REVIEW_REASON_LABELS.get(item, item) for item in manual_review_reasons),
            ),
            ("Открытых предупреждений", len([item for item in repair.checks if not item.is_resolved])),
            ("Создан", repair.created_at.isoformat()),
            ("Обновлен", repair.updated_at.isoformat()),
        ],
    )

    warnings_sheet = workbook.create_sheet("Несоответствия")
    append_rows(
        warnings_sheet,
        [("Раздел", "Важность", "Заголовок", "Детали", "Источник", "Решено", "Дата решения")]
        + warning_rows,
    )

    works_sheet = workbook.create_sheet("Работы")
    append_rows(
        works_sheet,
        [("Код", "Наименование", "Кол-во", "Нормо-часы", "Факт. часы", "Цена", "Сумма", "Статус")]
        + [
            (
                item.work_code or "",
                item.work_name,
                item.quantity,
                item.standard_hours if item.standard_hours is not None else "",
                item.actual_hours if item.actual_hours is not None else "",
                float(item.price),
                float(item.line_total),
                item.status.value,
            )
            for item in sorted(repair.works, key=lambda item: item.id)
        ],
    )

    parts_sheet = workbook.create_sheet("Материалы")
    append_rows(
        parts_sheet,
        [("Артикул", "Наименование", "Кол-во", "Ед.", "Цена", "Сумма", "Статус")]
        + [
            (
                item.article or "",
                item.part_name,
                item.quantity,
                item.unit_name or "",
                float(item.price),
                float(item.line_total),
                item.status.value,
            )
            for item in sorted(repair.parts, key=lambda item: item.id)
        ],
    )

    checks_sheet = workbook.create_sheet("Проверки")
    append_rows(
        checks_sheet,
        [("Тип", "Важность", "Заголовок", "Детали", "Решено", "Создано")]
        + [
            (
                item.check_type,
                item.severity.value,
                item.title,
                item.details or "",
                "Да" if item.is_resolved else "Нет",
                item.created_at.isoformat(),
            )
            for item in sorted(repair.checks, key=lambda item: item.id)
        ],
    )

    documents_sheet = workbook.create_sheet("Документы")
    append_rows(
        documents_sheet,
        [("ID", "Файл", "Вид", "Статус", "Основной", "OCR", "Создан", "Обновлен")]
        + [
            (
                item.id,
                item.original_filename,
                item.kind.value,
                item.status.value,
                "Да" if item.is_primary else "Нет",
                item.ocr_confidence if item.ocr_confidence is not None else "",
                item.created_at.isoformat(),
                item.updated_at.isoformat(),
            )
            for item in sorted(repair.documents, key=lambda item: item.id)
        ],
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = safe_filename(
        f"repair_{repair.id}_{repair.order_number or repair.vehicle.plate_number or 'card'}",
        f"repair_{repair.id}",
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


@router.patch("/{repair_id}", response_model=RepairDetailResponse)
def update_repair(
    repair_id: int,
    payload: RepairUpdateRequest,
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> RepairDetailResponse:
    stmt = build_repair_query().where(Repair.id == repair_id)
    repair = db.execute(stmt).unique().scalar_one_or_none()
    if repair is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Repair not found")
    old_snapshot = build_repair_snapshot(repair)

    update_data = payload.model_dump(exclude_unset=True)
    for field_name in (
        "order_number",
        "repair_date",
        "mileage",
        "reason",
        "employee_comment",
        "work_total",
        "parts_total",
        "vat_total",
        "grand_total",
        "status",
        "is_preliminary",
    ):
        if field_name in update_data:
            setattr(repair, field_name, update_data[field_name])

    if "service_name" in update_data:
        service_name = update_data["service_name"]
        if service_name:
            try:
                service = resolve_service(db, service_name)
            except ValueError as error:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Разрешены только сервисы из папки `Сервисы`",
                ) from error
            repair.service_id = service.id
        else:
            repair.service_id = None

    replace_manual_lines(db, repair, payload.works, payload.parts)
    replace_ocr_checks(db, repair.id, [])

    repair.is_manually_completed = True
    repair.is_partially_recognized = False

    db.commit()

    refreshed = db.execute(stmt).unique().scalar_one_or_none()
    if refreshed is None:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Repair could not be reloaded")
    create_ocr_learning_signals_for_repair(db, refreshed, current_admin=current_admin)
    new_snapshot = build_repair_snapshot(refreshed)
    action_type = "repair_archived" if new_snapshot["status"] == RepairStatus.ARCHIVED.value else "manual_update"
    db.add(
        AuditLog(
            user_id=current_admin.id,
            entity_type="repair",
            entity_id=str(refreshed.id),
            action_type=action_type,
            old_value=old_snapshot,
            new_value=new_snapshot,
        )
    )
    db.commit()
    history_entries = fetch_repair_history(db, refreshed.id)
    document_history_entries = fetch_document_history(db, refreshed)
    return serialize_repair(refreshed, history_entries, document_history_entries)


@router.patch("/{repair_id}/service", response_model=RepairDetailResponse)
def update_repair_service(
    repair_id: int,
    payload: RepairServiceUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> RepairDetailResponse:
    repair = load_repair_for_user(db, repair_id, current_user)
    old_snapshot = build_repair_snapshot(repair)

    service_name = payload.service_name.strip() if isinstance(payload.service_name, str) else None
    if service_name:
        try:
            service = resolve_service(db, service_name)
        except ValueError as error:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Указанный сервис не найден в справочнике. Выберите существующий или создайте его вручную.",
            ) from error
        repair.service_id = service.id
    else:
        repair.service_id = None

    update_service_manual_review_state(repair, service_name)
    sync_service_checks(repair, service_name, current_user)
    repair.is_manually_completed = True
    refresh_repair_status_after_check_updates(repair, current_user)

    db.flush()
    refreshed = load_repair_for_user(db, repair_id, current_user)
    new_snapshot = build_repair_snapshot(refreshed)
    db.add(
        AuditLog(
            user_id=current_user.id,
            entity_type="repair",
            entity_id=str(refreshed.id),
            action_type="service_assignment",
            old_value=old_snapshot,
            new_value=new_snapshot,
        )
    )
    db.commit()
    history_entries = fetch_repair_history(db, refreshed.id)
    document_history_entries = fetch_document_history(db, refreshed)
    return serialize_repair(refreshed, history_entries, document_history_entries)


@router.patch("/{repair_id}/review-fields", response_model=RepairDetailResponse)
def update_repair_review_fields(
    repair_id: int,
    payload: RepairReviewFieldsUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> RepairDetailResponse:
    repair = load_repair_for_user(db, repair_id, current_user)
    old_snapshot = build_repair_snapshot(repair)
    update_data = payload.model_dump(exclude_unset=True)

    if "repair_date" in update_data and update_data["repair_date"] is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Дата ремонта обязательна")
    if "mileage" in update_data and update_data["mileage"] is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Пробег обязателен")
    if "work_total" in update_data and update_data["work_total"] is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Сумма работ обязательна")
    if "parts_total" in update_data and update_data["parts_total"] is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Сумма запчастей обязательна")
    if "vat_total" in update_data and update_data["vat_total"] is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Сумма НДС обязательна")
    if "grand_total" in update_data and update_data["grand_total"] is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Итоговая сумма обязательна")

    for field_name in (
        "order_number",
        "repair_date",
        "mileage",
        "reason",
        "employee_comment",
        "work_total",
        "parts_total",
        "vat_total",
        "grand_total",
    ):
        if field_name in update_data:
            setattr(repair, field_name, update_data[field_name])

    repair.is_manually_completed = True
    db.flush()
    refreshed = load_repair_for_user(db, repair_id, current_user)
    if current_user.role == UserRole.ADMIN:
        create_ocr_learning_signals_for_repair(db, refreshed, current_admin=current_user)
    new_snapshot = build_repair_snapshot(refreshed)
    db.add(
        AuditLog(
            user_id=current_user.id,
            entity_type="repair",
            entity_id=str(refreshed.id),
            action_type="review_field_update",
            old_value=old_snapshot,
            new_value=new_snapshot,
        )
    )
    db.commit()
    history_entries = fetch_repair_history(db, refreshed.id)
    document_history_entries = fetch_document_history(db, refreshed)
    return serialize_repair(refreshed, history_entries, document_history_entries)


def update_check_resolution_payload(
    check: RepairCheck,
    is_resolved: bool,
    comment: str | None,
    current_user: User,
) -> dict:
    payload = dict(check.calculation_payload or {})
    payload["resolution"] = {
        "is_resolved": is_resolved,
        "comment": comment,
        "user_id": current_user.id,
        "user_name": current_user.full_name,
        "resolved_at": datetime.now(timezone.utc).isoformat(),
    }
    return payload


def refresh_repair_status_after_check_updates(repair: Repair, current_user: User) -> None:
    unresolved_checks = [item for item in repair.checks if not item.is_resolved]
    has_blocking_checks = any(item.severity in {CheckSeverity.SUSPICIOUS, CheckSeverity.ERROR} for item in unresolved_checks)

    if has_blocking_checks:
        repair.status = RepairStatus.SUSPICIOUS
        return

    if unresolved_checks:
        repair.status = RepairStatus.IN_REVIEW
        return

    if current_user.role == UserRole.ADMIN:
        repair.status = RepairStatus.IN_REVIEW
    else:
        repair.status = RepairStatus.EMPLOYEE_CONFIRMED


@router.patch("/{repair_id}/checks/{check_id}", response_model=RepairDetailResponse)
def update_repair_check(
    repair_id: int,
    check_id: int,
    payload: RepairCheckUpdateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> RepairDetailResponse:
    repair = load_repair_for_user(db, repair_id, current_user)
    target_check = next((item for item in repair.checks if item.id == check_id), None)
    if target_check is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Repair check not found")

    old_snapshot = build_repair_snapshot(repair)
    normalized_comment = (payload.comment or "").strip() or None

    target_check.is_resolved = payload.is_resolved
    target_check.calculation_payload = update_check_resolution_payload(
        target_check,
        is_resolved=payload.is_resolved,
        comment=normalized_comment,
        current_user=current_user,
    )
    refresh_repair_status_after_check_updates(repair, current_user)

    db.commit()

    refreshed = load_repair_for_user(db, repair_id, current_user)
    new_snapshot = build_repair_snapshot(refreshed)
    db.add(
        AuditLog(
            user_id=current_user.id,
            entity_type="repair",
            entity_id=str(refreshed.id),
            action_type="check_resolution_update",
            old_value=old_snapshot,
            new_value=new_snapshot,
        )
    )
    db.commit()

    history_entries = fetch_repair_history(db, refreshed.id)
    document_history_entries = fetch_document_history(db, refreshed)
    return serialize_repair(refreshed, history_entries, document_history_entries)
