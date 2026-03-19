from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, joinedload

from app.api.access import apply_vehicle_scope
from app.api.deps import get_current_active_user, get_current_admin, get_db
from app.models.audit import AuditLog
from app.models.enums import RepairStatus, VehicleStatus
from app.models.repair import Repair
from app.models.enums import VehicleType
from app.models.user import User
from app.models.vehicle import Vehicle, VehicleAssignmentHistory, VehicleLinkHistory
from app.schemas.vehicle import (
    VehicleDetailResponse,
    VehicleImportRequest,
    VehicleImportResponse,
    VehicleListResponse,
    VehicleRead,
    VehicleUpdateRequest,
)
from app.services.exporting import append_rows, safe_filename
from app.services.historical_repairs_import import IMPORT_REASON_PREFIX
from app.scripts.import_vehicles import (
    DEFAULT_TRAILERS_PATH,
    DEFAULT_TRUCKS_PATH,
    import_vehicles_with_session,
)


router = APIRouter(prefix="/vehicles", tags=["vehicles"])


def build_historical_summary_map(db: Session, vehicle_ids: list[int]) -> dict[int, dict[str, object]]:
    if not vehicle_ids:
        return {}

    rows = db.execute(
        select(
            Repair.vehicle_id.label("vehicle_id"),
            func.count(Repair.id).label("repairs_total"),
            func.max(Repair.repair_date).label("last_repair_date"),
        )
        .where(
            Repair.vehicle_id.in_(vehicle_ids),
            Repair.reason.is_not(None),
            Repair.reason.like(f"{IMPORT_REASON_PREFIX}%"),
        )
        .group_by(Repair.vehicle_id)
    ).all()

    return {
        int(row.vehicle_id): {
            "historical_repairs_total": int(row.repairs_total or 0),
            "historical_last_repair_date": row.last_repair_date,
        }
        for row in rows
    }


def build_history_summary(repair_history: list[Repair]) -> dict[str, object]:
    documents_total = sum(len(repair.documents) for repair in repair_history)
    confirmed_repairs = sum(1 for repair in repair_history if repair.status == RepairStatus.CONFIRMED)
    suspicious_repairs = sum(1 for repair in repair_history if repair.status == RepairStatus.SUSPICIOUS)
    last_repair = repair_history[0] if repair_history else None
    return {
        "repairs_total": len(repair_history),
        "documents_total": documents_total,
        "confirmed_repairs": confirmed_repairs,
        "suspicious_repairs": suspicious_repairs,
        "last_repair_date": last_repair.repair_date if last_repair is not None else None,
        "last_mileage": last_repair.mileage if last_repair is not None else None,
    }


def build_historical_history_summary(repair_history: list[Repair]) -> dict[str, object]:
    last_repair = repair_history[0] if repair_history else None
    first_repair = repair_history[-1] if repair_history else None
    services_total = len({repair.service_id for repair in repair_history if repair.service_id is not None})
    total_spend = round(sum(float(repair.grand_total) for repair in repair_history), 2)
    return {
        "repairs_total": len(repair_history),
        "services_total": services_total,
        "total_spend": total_spend,
        "first_repair_date": first_repair.repair_date if first_repair is not None else None,
        "last_repair_date": last_repair.repair_date if last_repair is not None else None,
        "last_mileage": last_repair.mileage if last_repair is not None else None,
    }


def build_vehicle_detail_payload(db: Session, vehicle: Vehicle) -> dict:
    today = date.today()
    links = db.scalars(
        select(VehicleLinkHistory)
        .where(
            or_(
                VehicleLinkHistory.left_vehicle_id == vehicle.id,
                VehicleLinkHistory.right_vehicle_id == vehicle.id,
            ),
            VehicleLinkHistory.starts_at <= today,
            or_(
                VehicleLinkHistory.ends_at.is_(None),
                VehicleLinkHistory.ends_at >= today,
            ),
        )
        .order_by(VehicleLinkHistory.starts_at.desc(), VehicleLinkHistory.id.desc())
    ).all()

    active_assignments = db.scalars(
        select(VehicleAssignmentHistory)
        .options(joinedload(VehicleAssignmentHistory.user))
        .where(
            VehicleAssignmentHistory.vehicle_id == vehicle.id,
            VehicleAssignmentHistory.starts_at <= today,
            or_(
                VehicleAssignmentHistory.ends_at.is_(None),
                VehicleAssignmentHistory.ends_at >= today,
            ),
        )
        .order_by(VehicleAssignmentHistory.starts_at.desc(), VehicleAssignmentHistory.id.desc())
    ).all()

    repair_history = db.scalars(
        select(Repair)
        .options(
            joinedload(Repair.service),
            joinedload(Repair.documents),
        )
        .where(Repair.vehicle_id == vehicle.id)
        .order_by(Repair.repair_date.desc(), Repair.id.desc())
    ).unique().all()

    payload = VehicleRead.model_validate(vehicle).model_dump()
    payload.update(build_historical_summary_map(db, [vehicle.id]).get(vehicle.id, {}))
    payload["active_links"] = links
    payload["active_assignments"] = [
        {
            "id": assignment.id,
            "user_id": assignment.user_id,
            "starts_at": assignment.starts_at,
            "ends_at": assignment.ends_at,
            "comment": assignment.comment,
            "user": {
                "id": assignment.user.id,
                "full_name": assignment.user.full_name,
                "email": assignment.user.email,
                "role": assignment.user.role,
            },
        }
        for assignment in active_assignments
    ]
    historical_repair_history = [
        repair
        for repair in repair_history
        if repair.reason is not None and repair.reason.startswith(IMPORT_REASON_PREFIX)
    ]
    payload["repair_history"] = [
        {
            "repair_id": repair.id,
            "order_number": repair.order_number,
            "repair_date": repair.repair_date,
            "mileage": repair.mileage,
            "status": repair.status,
            "service_name": repair.service.name if repair.service is not None else None,
            "grand_total": float(repair.grand_total),
            "documents_total": len(repair.documents),
            "created_at": repair.created_at,
            "updated_at": repair.updated_at,
        }
        for repair in repair_history
    ]
    payload["history_summary"] = build_history_summary(repair_history)
    payload["historical_repair_history"] = [
        {
            "repair_id": repair.id,
            "order_number": repair.order_number,
            "repair_date": repair.repair_date,
            "mileage": repair.mileage,
            "service_name": repair.service.name if repair.service is not None else None,
            "grand_total": float(repair.grand_total),
            "employee_comment": repair.employee_comment,
            "created_at": repair.created_at,
            "updated_at": repair.updated_at,
        }
        for repair in historical_repair_history
    ]
    payload["historical_history_summary"] = build_historical_history_summary(historical_repair_history)
    return payload


@router.get("", response_model=VehicleListResponse)
def list_vehicles(
    limit: int = Query(default=50, ge=1, le=2000),
    offset: int = Query(default=0, ge=0),
    vehicle_type: Optional[VehicleType] = Query(default=None),
    status_filter: Optional[VehicleStatus] = Query(default=None, alias="status"),
    search: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> VehicleListResponse:
    base_stmt = select(Vehicle)
    count_stmt = select(func.count(func.distinct(Vehicle.id)))

    if vehicle_type is not None:
        base_stmt = base_stmt.where(Vehicle.vehicle_type == vehicle_type)
        count_stmt = count_stmt.where(Vehicle.vehicle_type == vehicle_type)

    if status_filter is not None:
        base_stmt = base_stmt.where(Vehicle.status == status_filter)
        count_stmt = count_stmt.where(Vehicle.status == status_filter)

    if search:
        pattern = f"%{search.strip()}%"
        search_clause = or_(
            Vehicle.vin.ilike(pattern),
            Vehicle.plate_number.ilike(pattern),
            Vehicle.brand.ilike(pattern),
            Vehicle.model.ilike(pattern),
            Vehicle.external_id.ilike(pattern),
        )
        base_stmt = base_stmt.where(search_clause)
        count_stmt = count_stmt.where(search_clause)

    base_stmt = apply_vehicle_scope(base_stmt, current_user).order_by(Vehicle.created_at.desc())
    count_stmt = apply_vehicle_scope(count_stmt, current_user)

    items = db.scalars(base_stmt.offset(offset).limit(limit)).unique().all()
    total = db.scalar(count_stmt) or 0
    historical_summary_map = build_historical_summary_map(db, [item.id for item in items])

    return VehicleListResponse(
        items=[
            VehicleRead.model_validate(
                {
                    **VehicleRead.model_validate(item).model_dump(),
                    **historical_summary_map.get(item.id, {}),
                }
            )
            for item in items
        ],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.get("/{vehicle_id}", response_model=VehicleDetailResponse)
def get_vehicle(
    vehicle_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> VehicleDetailResponse:
    stmt = apply_vehicle_scope(select(Vehicle).where(Vehicle.id == vehicle_id), current_user)
    vehicle = db.scalar(stmt)
    if vehicle is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vehicle not found")

    return VehicleDetailResponse.model_validate(build_vehicle_detail_payload(db, vehicle))


@router.patch("/{vehicle_id}", response_model=VehicleDetailResponse)
def update_vehicle(
    vehicle_id: int,
    payload: VehicleUpdateRequest,
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> VehicleDetailResponse:
    vehicle = db.scalar(select(Vehicle).where(Vehicle.id == vehicle_id))
    if vehicle is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vehicle not found")

    old_snapshot = {
        "status": vehicle.status.value,
        "comment": vehicle.comment,
        "archived_at": vehicle.archived_at.isoformat() if vehicle.archived_at is not None else None,
    }

    update_data = payload.model_dump(exclude_unset=True)
    if "status" in update_data and update_data["status"] is not None:
        vehicle.status = update_data["status"]
        if vehicle.status == VehicleStatus.ARCHIVED:
            vehicle.archived_at = datetime.now(timezone.utc)
        else:
            vehicle.archived_at = None

    if "comment" in update_data:
        normalized_comment = update_data["comment"].strip() if isinstance(update_data["comment"], str) else None
        vehicle.comment = normalized_comment or None

    db.add(vehicle)
    db.flush()
    db.add(
        AuditLog(
            user_id=current_admin.id,
            entity_type="vehicle",
            entity_id=str(vehicle.id),
            action_type="vehicle_updated",
            old_value=old_snapshot,
            new_value={
                "status": vehicle.status.value,
                "comment": vehicle.comment,
                "archived_at": vehicle.archived_at.isoformat() if vehicle.archived_at is not None else None,
            },
        )
    )
    db.commit()
    db.refresh(vehicle)
    return VehicleDetailResponse.model_validate(build_vehicle_detail_payload(db, vehicle))


@router.get("/{vehicle_id}/export")
def export_vehicle(
    vehicle_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> StreamingResponse:
    stmt = apply_vehicle_scope(select(Vehicle).where(Vehicle.id == vehicle_id), current_user)
    vehicle = db.scalar(stmt)
    if vehicle is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Vehicle not found")

    payload = build_vehicle_detail_payload(db, vehicle)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Техника"
    append_rows(
        summary_sheet,
        [
            ("Поле", "Значение"),
            ("ID техники", payload["id"]),
            ("Внешний код", payload["external_id"] or ""),
            ("Тип техники", payload["vehicle_type"]),
            ("Госномер", payload["plate_number"] or ""),
            ("VIN", payload["vin"] or ""),
            ("Марка", payload["brand"] or ""),
            ("Модель", payload["model"] or ""),
            ("Год", payload["year"] or ""),
            ("Колонна", payload["column_name"] or ""),
            ("Механик", payload["mechanic_name"] or ""),
            ("Водитель", payload["current_driver_name"] or ""),
            ("Статус", payload["status"]),
            ("Комментарий", payload["comment"] or ""),
            ("Ремонтов", payload["history_summary"]["repairs_total"]),
            ("Документов", payload["history_summary"]["documents_total"]),
            ("Подтверждено ремонтов", payload["history_summary"]["confirmed_repairs"]),
            ("Подозрительных ремонтов", payload["history_summary"]["suspicious_repairs"]),
            ("Последний ремонт", payload["history_summary"]["last_repair_date"] or ""),
            ("Последний пробег", payload["history_summary"]["last_mileage"] or ""),
            ("Исторических ремонтов 2025", payload["historical_history_summary"]["repairs_total"]),
            ("Исторических сервисов", payload["historical_history_summary"]["services_total"]),
            ("Историческая сумма", payload["historical_history_summary"]["total_spend"]),
            ("Первый исторический ремонт", payload["historical_history_summary"]["first_repair_date"] or ""),
            ("Последний исторический ремонт", payload["historical_history_summary"]["last_repair_date"] or ""),
            ("Создано", payload["created_at"].isoformat()),
            ("Обновлено", payload["updated_at"].isoformat()),
        ],
    )

    assignments_sheet = workbook.create_sheet("Закрепления")
    append_rows(
        assignments_sheet,
        [("Сотрудник", "Почта", "Роль", "Начало", "Окончание", "Комментарий")]
        + [
            (
                assignment["user"]["full_name"],
                assignment["user"]["email"],
                assignment["user"]["role"],
                assignment["starts_at"].isoformat(),
                assignment["ends_at"].isoformat() if assignment["ends_at"] is not None else "",
                assignment["comment"] or "",
            )
            for assignment in payload["active_assignments"]
        ],
    )

    links_sheet = workbook.create_sheet("Связки")
    append_rows(
        links_sheet,
        [("Левая техника", "Правая техника", "Начало", "Окончание", "Комментарий")]
        + [
            (
                link.left_vehicle_id,
                link.right_vehicle_id,
                link.starts_at.isoformat(),
                link.ends_at.isoformat() if link.ends_at is not None else "",
                link.comment or "",
            )
            for link in payload["active_links"]
        ],
    )

    repairs_sheet = workbook.create_sheet("Ремонты")
    append_rows(
        repairs_sheet,
        [("ID ремонта", "Заказ-наряд", "Дата", "Пробег", "Статус", "Сервис", "Итого", "Документов", "Обновлено")]
        + [
            (
                repair["repair_id"],
                repair["order_number"] or "",
                repair["repair_date"].isoformat(),
                repair["mileage"],
                repair["status"],
                repair["service_name"] or "",
                repair["grand_total"],
                repair["documents_total"],
                repair["updated_at"].isoformat(),
            )
            for repair in payload["repair_history"]
        ],
    )

    historical_sheet = workbook.create_sheet("История 2025")
    append_rows(
        historical_sheet,
        [("ID ремонта", "Заказ-наряд", "Дата", "Пробег", "Сервис", "Итого", "Комментарий", "Обновлено")]
        + [
            (
                repair["repair_id"],
                repair["order_number"] or "",
                repair["repair_date"].isoformat(),
                repair["mileage"],
                repair["service_name"] or "",
                repair["grand_total"],
                repair["employee_comment"] or "",
                repair["updated_at"].isoformat(),
            )
            for repair in payload["historical_repair_history"]
        ],
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = safe_filename(
        f"vehicle_{payload['id']}_{payload['plate_number'] or payload['vin'] or 'card'}",
        f"vehicle_{payload['id']}",
    )
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


@router.post("/import", response_model=VehicleImportResponse)
def import_vehicle_registry(
    payload: VehicleImportRequest,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_admin),
) -> VehicleImportResponse:
    trucks_path = Path(payload.trucks_path) if payload.trucks_path else DEFAULT_TRUCKS_PATH
    trailers_path = Path(payload.trailers_path) if payload.trailers_path else DEFAULT_TRAILERS_PATH

    if not trucks_path.exists():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Trucks file not found: {trucks_path}")
    if not trailers_path.exists():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Trailers file not found: {trailers_path}",
        )

    stats = import_vehicles_with_session(db, trucks_path=trucks_path, trailers_path=trailers_path)
    return VehicleImportResponse(
        **stats.as_dict(),
        trucks_path=str(trucks_path),
        trailers_path=str(trailers_path),
    )
