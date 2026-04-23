from __future__ import annotations

from openpyxl import Workbook

from app.application.manual_review_labels import MANUAL_REVIEW_REASON_LABELS
from app.application.documents.repair_relations import ensure_repair_vehicle_relation, get_repair_source_document
from app.application.documents.support import build_manual_review_check
from app.application.documents.document_versions import get_latest_parsed_payload
from app.application.imports.document_jobs import get_document_display_import_job
from app.application.review.queue_support import has_open_suspicious_checks
from app.models.document import Document
from app.models.enums import DocumentStatus, ImportStatus, RepairStatus
from app.models.repair import Repair, RepairCheck
from app.application.reports.exporting import append_rows
from app.application.reports.repair_report_analysis import build_repair_executive_report

CHECK_REPORT_SECTION_LABELS = {
    "catalogs": "Справочники",
    "labor_norms": "Нормо-часы",
    "amounts": "Суммы и структура",
    "history": "История и аномалии",
    "ocr": "OCR и ручная проверка",
}


def get_repair_documents_for_view(
    repair: Repair,
    *,
    include_archived_documents: bool = True,
) -> list[Document]:
    documents = sorted(repair.documents, key=lambda item: (item.created_at, item.id), reverse=True)
    if include_archived_documents:
        return documents
    return [item for item in documents if item.status != DocumentStatus.ARCHIVED]


def resolve_report_document(
    repair: Repair,
    *,
    include_archived_fallback: bool = False,
    report_document: Document | None = None,
) -> Document | None:
    if report_document is not None and report_document.repair_id == repair.id:
        if report_document.status != DocumentStatus.ARCHIVED or include_archived_fallback:
            return report_document
    return get_repair_source_document(repair, include_archived_fallback=include_archived_fallback)


def get_report_source_payload(
    repair: Repair,
    *,
    include_archived_fallback: bool = False,
    report_document: Document | None = None,
) -> dict:
    selected_report_document = resolve_report_document(
        repair,
        include_archived_fallback=include_archived_fallback,
        report_document=report_document,
    )
    return get_latest_parsed_payload(selected_report_document)


def get_check_report_section_key(check_type: str) -> str:
    if "vehicle" in check_type or "service" in check_type:
        return "catalogs"
    if "standard_hours" in check_type:
        return "labor_norms"
    if "total" in check_type or "duplicate" in check_type or "expected_total" in check_type:
        return "amounts"
    if "repeat_repair" in check_type:
        return "history"
    return "ocr"


def has_blocking_repair_checks(repair: Repair) -> bool:
    return has_open_suspicious_checks(repair)


def build_report_status_summary(
    repair: Repair,
    *,
    include_archived_fallback: bool = False,
    report_document: Document | None = None,
) -> tuple[str, str]:
    selected_report_document = resolve_report_document(
        repair,
        include_archived_fallback=include_archived_fallback,
        report_document=report_document,
    )
    latest_import_job = get_document_display_import_job(selected_report_document)
    if selected_report_document is not None and (
        selected_report_document.status.value == "uploaded"
        or (
            latest_import_job is not None
            and latest_import_job.status in {ImportStatus.QUEUED, ImportStatus.RETRY, ImportStatus.PROCESSING}
        )
    ):
        return "В очереди OCR", "Документ ещё находится в обработке или перепроверке."

    executive_report = build_repair_executive_report(
        repair,
        source_payload=get_report_source_payload(
            repair,
            include_archived_fallback=include_archived_fallback,
            report_document=selected_report_document,
        ),
        manual_review_reason_labels=MANUAL_REVIEW_REASON_LABELS,
        source_document=selected_report_document,
    )
    findings = executive_report["findings"]
    if any(str(item["severity"]) == "high" for item in findings):
        return "Есть критичные несоответствия", "Перед подтверждением нужны ручная проверка и решение по предупреждениям."
    if findings:
        return "Требует ручной проверки", "По заказ-наряду есть открытые предупреждения, которые нужно проверить."
    return "Готов к следующему этапу", "Открытых несоответствий по заказ-наряду не найдено."


def build_report_workflow_summary(repair: Repair) -> tuple[str, str]:
    if repair.status == RepairStatus.ARCHIVED:
        return "Архив", "Ремонт выведен из активного потока и доступен только для просмотра, поиска и экспорта."
    if repair.status == RepairStatus.OCR_ERROR:
        return "Ошибка OCR", "Автоматическое распознавание завершилось ошибкой, поэтому ремонт требует ручной обработки."
    if repair.status == RepairStatus.SUSPICIOUS or has_blocking_repair_checks(repair):
        return "Подозрительный ремонт", "По ремонту зафиксирован риск, требующий отдельного управленческого решения."
    if repair.status == RepairStatus.DRAFT:
        return "Черновик", "Ремонт создан, но ещё не переведён в рабочий процесс проверки."
    if repair.status == RepairStatus.IN_REVIEW:
        return "Ручная проверка сотрудником", "Сотрудник сверяет документ, реквизиты и предупреждения перед подтверждением."
    if repair.status == RepairStatus.EMPLOYEE_CONFIRMED:
        return (
            "Ожидает финального подтверждения администратора",
            "Сотрудник завершил свою часть проверки. Теперь нужен финальный review администратора.",
        )
    if repair.status == RepairStatus.CONFIRMED:
        return "Подтверждён", "Ремонт прошёл employee-review и финальное подтверждение администратора."
    return repair.status.value, ""


def build_report_executive_sections(
    repair: Repair,
    *,
    include_archived_fallback: bool = False,
    report_document: Document | None = None,
) -> list[tuple[str, list[str]]]:
    selected_report_document = resolve_report_document(
        repair,
        include_archived_fallback=include_archived_fallback,
        report_document=report_document,
    )
    executive_report = build_repair_executive_report(
        repair,
        source_payload=get_report_source_payload(
            repair,
            include_archived_fallback=include_archived_fallback,
            report_document=selected_report_document,
        ),
        manual_review_reason_labels=MANUAL_REVIEW_REASON_LABELS,
        source_document=selected_report_document,
    )
    sections: list[tuple[str, list[str]]] = []
    overview_lines = [
        str(executive_report["headline"]),
        str(executive_report["summary"]),
        f"Общий риск: {executive_report['status']}",
    ]
    highlights = executive_report.get("highlights")
    if isinstance(highlights, list):
        overview_lines.extend(str(item) for item in highlights if str(item))
    sections.append(("Короткий отчёт для руководителя", overview_lines))

    raw_sections = executive_report.get("full_report_sections")
    if not isinstance(raw_sections, list):
        return sections

    for raw_section in raw_sections:
        if not isinstance(raw_section, dict):
            continue
        title = str(raw_section.get("title") or raw_section.get("key") or "Раздел")
        items = raw_section.get("items")
        if not isinstance(items, list):
            continue
        normalized_items = [str(item) for item in items if str(item)]
        if normalized_items:
            sections.append((title, normalized_items))
    return sections


def _extract_customer_summary_items(executive_sections: list[tuple[str, list[str]]]) -> list[str]:
    for title, items in executive_sections:
        if title == "0. Краткая сводка для заказчика":
            return [str(item) for item in items if str(item)]
    return []


def _build_customer_summary_sheet_rows(executive_sections: list[tuple[str, list[str]]]) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for item in _extract_customer_summary_items(executive_sections):
        if ": " in item:
            label, value = item.split(": ", 1)
            rows.append((label, value))
        else:
            rows.append(("Сводка для заказчика", item))
    return rows


def build_export_warning_rows(
    repair: Repair,
    *,
    include_archived_fallback: bool = False,
    report_document: Document | None = None,
) -> list[tuple[object, ...]]:
    selected_report_document = resolve_report_document(
        repair,
        include_archived_fallback=include_archived_fallback,
        report_document=report_document,
    )
    executive_report = build_repair_executive_report(
        repair,
        source_payload=get_report_source_payload(
            repair,
            include_archived_fallback=include_archived_fallback,
            report_document=selected_report_document,
        ),
        manual_review_reason_labels=MANUAL_REVIEW_REASON_LABELS,
        source_document=selected_report_document,
    )
    if executive_report["findings"]:
        return [
            (
                str(item["category"]),
                str(item["severity"]),
                str(item["title"]),
                str(item["summary"]),
                "executive_report",
                "Нет",
                "",
            )
            for item in executive_report["findings"]
        ]

    rows: list[tuple[object, ...]] = []
    report_payload = get_report_source_payload(
        repair,
        include_archived_fallback=include_archived_fallback,
        report_document=selected_report_document,
    )
    extracted_fields = report_payload.get("extracted_fields") if isinstance(report_payload.get("extracted_fields"), dict) else {}
    raw_reasons = report_payload.get("manual_review_reasons")
    manual_review_reasons = [str(item) for item in raw_reasons] if isinstance(raw_reasons, list) else []
    seen_reason_codes: set[str] = set()

    for check in sorted(repair.checks, key=lambda item: (item.is_resolved, item.created_at, item.id)):
        resolution = check.calculation_payload.get("resolution") if isinstance(check.calculation_payload, dict) else None
        resolved_at = resolution.get("resolved_at") if isinstance(resolution, dict) else None
        rows.append(
            (
                CHECK_REPORT_SECTION_LABELS[get_check_report_section_key(check.check_type)],
                check.severity.value,
                check.title,
                check.details or "",
                "check",
                "Да" if check.is_resolved else "Нет",
                resolved_at or "",
            )
        )
        if check.check_type.startswith("ocr_"):
            reason_code = check.check_type.removeprefix("ocr_")
            seen_reason_codes.add(reason_code)

    for reason in manual_review_reasons:
        if reason in seen_reason_codes:
            continue
        manual_check = build_manual_review_check(reason, extracted_fields=extracted_fields)
        rows.append(
            (
                CHECK_REPORT_SECTION_LABELS[get_check_report_section_key(str(manual_check["check_type"]))],
                manual_check["severity"].value,
                str(manual_check["title"]),
                str(manual_check.get("details") or ""),
                "manual_review_reason",
                "Нет",
                "",
            )
        )

    return rows


def build_repair_pdf_sections(
    repair: Repair,
    *,
    include_archived_documents: bool = True,
    report_document: Document | None = None,
) -> list[tuple[str, list[str]]]:
    ensure_repair_vehicle_relation(repair)
    source_document = get_repair_source_document(repair, include_archived_fallback=include_archived_documents)
    source_document_id = source_document.id if source_document is not None else None
    selected_report_document = resolve_report_document(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=report_document,
    )
    report_document_id = selected_report_document.id if selected_report_document is not None else None
    source_payload = get_report_source_payload(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=selected_report_document,
    )
    report_status, report_status_comment = build_report_status_summary(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=selected_report_document,
    )
    workflow_status, workflow_comment = build_report_workflow_summary(repair)
    executive_sections = build_report_executive_sections(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=selected_report_document,
    )
    raw_reasons = source_payload.get("manual_review_reasons")
    manual_review_reasons = [str(item) for item in raw_reasons] if isinstance(raw_reasons, list) else []
    extracted_fields = source_payload.get("extracted_fields") if isinstance(source_payload.get("extracted_fields"), dict) else {}
    warning_rows = build_export_warning_rows(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=selected_report_document,
    )
    documents = get_repair_documents_for_view(repair, include_archived_documents=include_archived_documents)

    warning_lines = [
        (
            f"{index}. Раздел {row[0]} | важность {row[1]} | "
            f"{row[2]} | источник {row[4]} | {'решено' if row[5] == 'Да' else 'открыто'}"
            + (f" | {row[3]}" if row[3] else "")
            + (f" | дата решения {row[6]}" if row[6] else "")
        )
        for index, row in enumerate(warning_rows, start=1)
    ]

    sections: list[tuple[str, list[str]]] = [
        (
            "Сводка",
            [
                f"ID ремонта: {repair.id}",
                f"Номер заказ-наряда: {repair.order_number or 'Не указан'}",
                f"Дата ремонта: {repair.repair_date.isoformat()}",
                f"Статус ремонта: {repair.status.value}",
                f"Этап workflow: {workflow_status}",
                f"Комментарий к workflow: {workflow_comment}",
                f"Итоговый статус отчёта: {report_status}",
                f"Комментарий к статусу: {report_status_comment}",
                f"Предварительный: {'Да' if repair.is_preliminary else 'Нет'}",
                f"Частично распознан: {'Да' if repair.is_partially_recognized else 'Нет'}",
                f"Госномер: {repair.vehicle.plate_number or 'Не указан'}",
                f"VIN: {repair.vehicle.vin or 'Не указан'}",
                f"Марка и модель: {' '.join(part for part in (repair.vehicle.brand, repair.vehicle.model) if part) or 'Не указаны'}",
                f"Сервис: {repair.service.name if repair.service is not None else 'Не назначен'}",
                f"Сервис по OCR: {extracted_fields.get('service_name') or 'Не распознан'}",
                f"Пробег: {repair.mileage}",
                f"Причина ремонта: {repair.reason or 'Не указана'}",
                f"Комментарий сотрудника: {repair.employee_comment or 'Нет'}",
                f"Работы, руб: {float(repair.work_total):.2f}",
                f"Запчасти, руб: {float(repair.parts_total):.2f}",
                f"НДС, руб: {float(repair.vat_total):.2f}",
                f"Итого, руб: {float(repair.grand_total):.2f}",
                (
                    "Ожидаемая стоимость, руб: "
                    f"{float(repair.expected_total):.2f}"
                    if repair.expected_total is not None
                    else "Ожидаемая стоимость, руб: Не рассчитана"
                ),
                f"Документ отчета: {selected_report_document.original_filename if selected_report_document is not None else 'Не выбран'}",
                f"Статус документа отчета: {selected_report_document.status.value if selected_report_document is not None else 'Не определён'}",
                f"Основной документ: {source_document.original_filename if source_document is not None else 'Не выбран'}",
                f"Статус основного документа: {source_document.status.value if source_document is not None else 'Не определён'}",
                (
                    "Причины ручной проверки OCR: "
                    + (
                        ", ".join(MANUAL_REVIEW_REASON_LABELS.get(item, item) for item in manual_review_reasons)
                        if manual_review_reasons
                        else "Нет"
                    )
                ),
                f"Открытых предупреждений: {len(warning_rows)}",
                f"Создан: {repair.created_at.isoformat()}",
                f"Обновлён: {repair.updated_at.isoformat()}",
                *(
                    ["Сводка для заказчика:"]
                    + _extract_customer_summary_items(executive_sections)
                    if _extract_customer_summary_items(executive_sections)
                    else []
                ),
            ],
        ),
        *executive_sections,
        (
            "Несоответствия",
            warning_lines
            if warning_lines
            else ["Открытых предупреждений и ручных причин проверки не найдено."],
        ),
        (
            "Работы",
            [
                (
                    f"{item.work_code or 'Без кода'} | {item.work_name} | "
                    f"кол-во {item.quantity} | нормо-часы {item.standard_hours if item.standard_hours is not None else '—'} | "
                    f"факт {item.actual_hours if item.actual_hours is not None else '—'} | "
                    f"цена {float(item.price):.2f} | сумма {float(item.line_total):.2f} | статус {item.status.value}"
                )
                for item in sorted(repair.works, key=lambda item: item.id)
            ]
            or ["Работы не заполнены."],
        ),
        (
            "Материалы",
            [
                (
                    (
                        f"{item.article or 'Без артикула'} | {item.part_name} | "
                        f"кол-во {item.quantity} {item.unit_name or ''}"
                    ).strip()
                    + f" | цена {float(item.price):.2f} | сумма {float(item.line_total):.2f} | статус {item.status.value}"
                )
                for item in sorted(repair.parts, key=lambda item: item.id)
            ]
            or ["Материалы не заполнены."],
        ),
        (
            "Проверки",
            [
                (
                    f"{item.check_type} | {item.severity.value} | "
                    f"{item.title} | {'Решено' if item.is_resolved else 'Открыто'} | "
                    f"создано {item.created_at.isoformat()}"
                    + (f" | {item.details}" if item.details else "")
                )
                for item in sorted(repair.checks, key=lambda item: item.id)
            ]
            or ["Проверки отсутствуют."],
        ),
        (
            "Документы",
            [
                (
                    f"ID {item.id} | {item.original_filename} | {item.kind.value} | {item.status.value} | "
                    f"{'основной' if source_document_id == item.id else 'дополнительный'} | "
                    + ("документ отчета | " if report_document_id == item.id else "")
                    + f"OCR {item.ocr_confidence if item.ocr_confidence is not None else '—'} | "
                    f"создан {item.created_at.isoformat()} | обновлён {item.updated_at.isoformat()}"
                )
                for item in sorted(documents, key=lambda item: item.id)
            ]
            or ["Документы отсутствуют."],
        ),
    ]

    return sections


def build_repair_export_workbook(
    repair: Repair,
    *,
    include_archived_documents: bool = True,
    report_document: Document | None = None,
) -> Workbook:
    ensure_repair_vehicle_relation(repair)
    source_document = get_repair_source_document(repair, include_archived_fallback=include_archived_documents)
    source_document_id = source_document.id if source_document is not None else None
    selected_report_document = resolve_report_document(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=report_document,
    )
    report_document_id = selected_report_document.id if selected_report_document is not None else None
    source_payload = get_report_source_payload(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=selected_report_document,
    )
    report_status, report_status_comment = build_report_status_summary(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=selected_report_document,
    )
    workflow_status, workflow_comment = build_report_workflow_summary(repair)
    executive_sections = build_report_executive_sections(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=selected_report_document,
    )
    warning_rows = build_export_warning_rows(
        repair,
        include_archived_fallback=include_archived_documents,
        report_document=selected_report_document,
    )
    documents = get_repair_documents_for_view(repair, include_archived_documents=include_archived_documents)
    raw_reasons = source_payload.get("manual_review_reasons")
    manual_review_reasons = [str(item) for item in raw_reasons] if isinstance(raw_reasons, list) else []
    extracted_fields = source_payload.get("extracted_fields") if isinstance(source_payload.get("extracted_fields"), dict) else {}

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Отчет"
    append_rows(
        summary_sheet,
        [
            ("Поле", "Значение"),
            ("ID ремонта", repair.id),
            ("Номер заказ-наряда", repair.order_number or ""),
            ("Дата ремонта", repair.repair_date.isoformat()),
            ("Статус", repair.status.value),
            ("Этап workflow", workflow_status),
            ("Комментарий к workflow", workflow_comment),
            ("Итоговый статус отчёта", report_status),
            ("Комментарий к статусу", report_status_comment),
            ("Предварительный", "Да" if repair.is_preliminary else "Нет"),
            ("Частично распознан", "Да" if repair.is_partially_recognized else "Нет"),
            ("Госномер", repair.vehicle.plate_number or ""),
            ("VIN", repair.vehicle.vin or ""),
            ("Марка", repair.vehicle.brand or ""),
            ("Модель", repair.vehicle.model or ""),
            ("Сервис", repair.service.name if repair.service is not None else ""),
            ("Сервис по OCR", extracted_fields.get("service_name") if extracted_fields else ""),
            ("Пробег", repair.mileage),
            ("Причина ремонта", repair.reason or ""),
            ("Комментарий сотрудника", repair.employee_comment or ""),
            ("Работы, руб", float(repair.work_total)),
            ("Запчасти, руб", float(repair.parts_total)),
            ("НДС, руб", float(repair.vat_total)),
            ("Итого, руб", float(repair.grand_total)),
            ("Ожидаемая стоимость, руб", float(repair.expected_total) if repair.expected_total is not None else ""),
            ("Документ отчета", selected_report_document.original_filename if selected_report_document is not None else ""),
            (
                "Статус документа отчета",
                selected_report_document.status.value if selected_report_document is not None else "",
            ),
            ("Основной документ", source_document.original_filename if source_document is not None else ""),
            ("Статус основного документа", source_document.status.value if source_document is not None else ""),
            (
                "Причины ручной проверки OCR",
                ", ".join(MANUAL_REVIEW_REASON_LABELS.get(item, item) for item in manual_review_reasons),
            ),
            ("Открытых предупреждений", len(warning_rows)),
            *_build_customer_summary_sheet_rows(executive_sections),
            ("Создан", repair.created_at.isoformat()),
            ("Обновлен", repair.updated_at.isoformat()),
        ],
    )

    executive_sheet = workbook.create_sheet("Итоговый отчет")
    append_rows(
        executive_sheet,
        [("Раздел", "Пункт")] + [(title, item) for title, items in executive_sections for item in items],
    )

    warnings_sheet = workbook.create_sheet("Несоответствия")
    append_rows(
        warnings_sheet,
        [("Раздел", "Важность", "Заголовок", "Детали", "Источник", "Решено", "Дата решения")] + warning_rows,
    )

    works_sheet = workbook.create_sheet("Работы")
    append_rows(
        works_sheet,
        [("Код", "Наименование", "Кол-во", "Нормо-часы", "Факт. часы", "Цена", "Сумма", "Статус")]
        + [
            (
                item.work_code or "",
                item.work_name,
                item.quantity,
                item.standard_hours if item.standard_hours is not None else "",
                item.actual_hours if item.actual_hours is not None else "",
                float(item.price),
                float(item.line_total),
                item.status.value,
            )
            for item in sorted(repair.works, key=lambda item: item.id)
        ],
    )

    parts_sheet = workbook.create_sheet("Материалы")
    append_rows(
        parts_sheet,
        [("Артикул", "Наименование", "Кол-во", "Ед.", "Цена", "Сумма", "Статус")]
        + [
            (
                item.article or "",
                item.part_name,
                item.quantity,
                item.unit_name or "",
                float(item.price),
                float(item.line_total),
                item.status.value,
            )
            for item in sorted(repair.parts, key=lambda item: item.id)
        ],
    )

    checks_sheet = workbook.create_sheet("Проверки")
    append_rows(
        checks_sheet,
        [("Тип", "Важность", "Заголовок", "Детали", "Решено", "Создано")]
        + [
            (
                item.check_type,
                item.severity.value,
                item.title,
                item.details or "",
                "Да" if item.is_resolved else "Нет",
                item.created_at.isoformat(),
            )
            for item in sorted(repair.checks, key=lambda item: item.id)
        ],
    )

    documents_sheet = workbook.create_sheet("Документы")
    append_rows(
        documents_sheet,
        [("ID", "Файл", "Вид", "Статус", "Основной", "Документ отчета", "OCR", "Создан", "Обновлен")]
        + [
            (
                item.id,
                item.original_filename,
                item.kind.value,
                item.status.value,
                "Да" if source_document_id == item.id else "Нет",
                "Да" if report_document_id == item.id else "Нет",
                item.ocr_confidence if item.ocr_confidence is not None else "",
                item.created_at.isoformat(),
                item.updated_at.isoformat(),
            )
            for item in sorted(documents, key=lambda item: item.id)
        ],
    )

    return workbook
