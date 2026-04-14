from __future__ import annotations

import json
import re
import shutil
import tempfile
from collections.abc import Callable
from datetime import date, datetime
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pypdf import PdfReader, PdfWriter
from pypdf.errors import PdfReadError

try:
    from PIL import Image, ImageChops
except ImportError:  # pragma: no cover - optional dependency during bootstrap
    Image = None
    ImageChops = None

from app.services.document_processing_config import (
    TESSERACT_LANGUAGE,
    TESSERACT_PAGE_SEGMENTATION_MODES,
)
from app.services.document_runtime_facade import (
    is_pillow_available,
    score_text_quality,
    select_best_tesseract_ocr_variant,
    select_best_text_variant,
)
from app.services.ocr_runtime import (
    PDFTOPPM_BINARY,
    SIPS_BINARY,
    TESSERACT_BINARY,
    VISION_OCR_SCRIPT,
)


def decode_pdf_literal(raw_text: bytes) -> bytes:
    return (
        raw_text.replace(b"\\(", b"(")
        .replace(b"\\)", b")")
        .replace(b"\\n", b"\n")
        .replace(b"\\r", b"\r")
        .replace(b"\\t", b"\t")
        .replace(b"\\\\", b"\\")
    )


def extract_pdf_stream_text(page) -> str:
    content = page.get_contents()
    if content is None:
        return ""

    contents = content if isinstance(content, list) else [content]
    literals = []
    for item in contents:
        raw = item.get_data()
        matches = re.findall(rb"\((?:\\.|[^\\)])*\)", raw)
        for match in matches:
            literals.append(decode_pdf_literal(match[1:-1]))

    decoded_variants = []
    for encoding in ("utf-8", "cp1251", "latin1"):
        try:
            decoded_variants.append("\n".join(part.decode(encoding, errors="ignore") for part in literals))
        except LookupError:
            continue

    scored_variants = sorted(
        decoded_variants,
        key=lambda item: (
            len(re.findall(r"[А-Яа-яA-Za-z0-9]", item)),
            len(item),
        ),
        reverse=True,
    )
    return scored_variants[0].strip() if scored_variants else ""


def extract_pdf_text(path: Path) -> str:
    try:
        reader = PdfReader(path.as_posix())
    except (OSError, PdfReadError, ValueError) as error:
        raise ValueError("Не удалось прочитать PDF документ") from error
    chunks = []
    for page in reader.pages:
        page_text = (page.extract_text() or "").strip()
        stream_text = extract_pdf_stream_text(page)
        if score_text_quality(stream_text) > score_text_quality(page_text):
            chunks.append(stream_text)
        else:
            chunks.append(page_text)
    return "\n".join(filter(None, chunks)).strip()


def run_vision_ocr(
    image_paths: list[Path],
    *,
    is_vision_available: Callable[[], bool],
    run_subprocess: Callable[..., object],
) -> dict[str, str]:
    if not image_paths:
        return {}
    if not is_vision_available():
        raise RuntimeError("Apple Vision OCR is not available in the current environment")

    command = ["swift", VISION_OCR_SCRIPT.as_posix(), *[path.as_posix() for path in image_paths]]
    result = run_subprocess(
        command,
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "Vision OCR command failed")

    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as error:
        raise RuntimeError("Vision OCR returned an invalid response payload") from error
    if not isinstance(payload, dict) or not isinstance(payload.get("results"), list):
        raise RuntimeError("Vision OCR returned an invalid response payload")

    normalized_results: dict[str, str] = {}
    for item in payload["results"]:
        if not isinstance(item, dict):
            raise RuntimeError("Vision OCR returned an invalid response payload")
        path_value = item.get("path")
        text_value = item.get("text")
        if not isinstance(path_value, str) or not isinstance(text_value, str):
            raise RuntimeError("Vision OCR returned an invalid response payload")
        normalized_results[path_value] = text_value
    return normalized_results


def run_tesseract_ocr_with_modes(
    image_paths: list[Path],
    *,
    page_segmentation_modes: tuple[str, ...] | list[str],
    is_tesseract_available: Callable[[], bool],
    run_subprocess: Callable[..., object],
) -> dict[str, str]:
    if not image_paths:
        return {}
    if not is_tesseract_available():
        raise RuntimeError("Tesseract OCR is not available in the current environment")

    payload: dict[str, str] = {}
    for image_path in image_paths:
        variants: list[str] = []
        last_error = "Tesseract OCR command failed"
        for psm in page_segmentation_modes:
            command = [
                TESSERACT_BINARY,
                image_path.as_posix(),
                "stdout",
                "-l",
                TESSERACT_LANGUAGE,
                "--psm",
                str(psm),
            ]
            result = run_subprocess(
                command,
                capture_output=True,
                text=True,
                check=False,
            )
            if result.returncode != 0:
                last_error = result.stderr.strip() or last_error
                continue
            variants.append(result.stdout)
        if not variants:
            raise RuntimeError(last_error)
        payload[image_path.as_posix()] = select_best_tesseract_ocr_variant(variants)
    return payload


def run_tesseract_ocr(
    image_paths: list[Path],
    *,
    run_tesseract_ocr_with_modes_fn: Callable[..., dict[str, str]],
) -> dict[str, str]:
    return run_tesseract_ocr_with_modes_fn(
        image_paths,
        page_segmentation_modes=tuple(TESSERACT_PAGE_SEGMENTATION_MODES),
    )


def run_ocr_backend(
    image_paths: list[Path],
    *,
    get_ocr_backend: Callable[[], str | None],
    run_vision_ocr_fn: Callable[[list[Path]], dict[str, str]],
    run_tesseract_ocr_fn: Callable[[list[Path]], dict[str, str]],
) -> tuple[dict[str, str], str]:
    backend = get_ocr_backend()
    if backend == "vision":
        return run_vision_ocr_fn(image_paths), backend
    if backend == "tesseract":
        return run_tesseract_ocr_fn(image_paths), backend
    raise RuntimeError("No supported OCR backend is available in the current environment")


def save_pillow_optimized_image(source_path: Path, output_path: Path) -> bool:
    if not is_pillow_available():
        return False

    try:
        with Image.open(source_path) as image:
            rgba_image = image.convert("RGBA")
            alpha_channel = rgba_image.getchannel("A")
            alpha_bbox = alpha_channel.getbbox()

            background = Image.new("RGBA", rgba_image.size, (255, 255, 255, 255))
            rgb_image = Image.alpha_composite(background, rgba_image).convert("RGB")

            grayscale_image = rgb_image.convert("L")
            diff_image = ImageChops.difference(grayscale_image, Image.new("L", grayscale_image.size, 255))
            content_bbox = diff_image.point(lambda value: 255 if value > 12 else 0).getbbox()

            if alpha_bbox is not None:
                content_bbox = alpha_bbox if content_bbox is None else (
                    min(alpha_bbox[0], content_bbox[0]),
                    min(alpha_bbox[1], content_bbox[1]),
                    max(alpha_bbox[2], content_bbox[2]),
                    max(alpha_bbox[3], content_bbox[3]),
                )

            if content_bbox is not None:
                left, top, right, bottom = content_bbox
                padding = max(24, int(max(rgb_image.size) * 0.03))
                crop_box = (
                    max(0, left - padding),
                    max(0, top - padding),
                    min(rgb_image.width, right + padding),
                    min(rgb_image.height, bottom + padding),
                )
                rgb_image = rgb_image.crop(crop_box)

            longest_side = max(rgb_image.size)
            if longest_side and longest_side < 2400:
                scale = 2400 / float(longest_side)
                resized_size = (
                    max(1, int(round(rgb_image.width * scale))),
                    max(1, int(round(rgb_image.height * scale))),
                )
                rgb_image = rgb_image.resize(resized_size, Image.Resampling.LANCZOS)

            rgb_image.save(output_path, format="JPEG", quality=95, optimize=True)
            return True
    except Exception:
        return False


def save_pdf_header_crop_for_ocr(source_path: Path, output_path: Path) -> bool:
    if not is_pillow_available():
        return False

    try:
        with Image.open(source_path) as image:
            rgb_image = image.convert("RGB")
            image_width, image_height = rgb_image.size
            if image_width <= 0 or image_height <= 0:
                return False

            top = max(0, int(image_height * 0.10))
            bottom = min(image_height, int(image_height * 0.26))
            if bottom <= top:
                return False

            cropped_image = rgb_image.crop((0, top, image_width, bottom))
            longest_side = max(cropped_image.size)
            if longest_side and longest_side < 2200:
                scale = 2200 / float(longest_side)
                resized_size = (
                    max(1, int(round(cropped_image.width * scale))),
                    max(1, int(round(cropped_image.height * scale))),
                )
                cropped_image = cropped_image.resize(resized_size, Image.Resampling.LANCZOS)

            cropped_image.save(output_path, format="JPEG", quality=95, optimize=True)
            return True
    except Exception:
        return False


def optimize_existing_image_for_ocr(
    path: Path,
    *,
    save_pillow_optimized_image_fn: Callable[[Path, Path], bool],
) -> None:
    temporary_output_path = path.with_name(f"{path.stem}_optimized.jpg")
    if save_pillow_optimized_image_fn(path, temporary_output_path):
        temporary_output_path.replace(path)


def preprocess_image_for_ocr(
    path: Path,
    *,
    save_pillow_optimized_image_fn: Callable[[Path, Path], bool],
    is_sips_available_fn: Callable[[], bool],
    run_subprocess: Callable[..., object],
    optimize_existing_image_for_ocr_fn: Callable[[Path], None],
) -> tuple[tempfile.TemporaryDirectory, Path]:
    temp_dir = tempfile.TemporaryDirectory()
    processed_path = Path(temp_dir.name) / f"{path.stem}_ocr.jpg"
    if save_pillow_optimized_image_fn(path, processed_path):
        return temp_dir, processed_path

    if is_sips_available_fn():
        command = [
            SIPS_BINARY,
            "-s",
            "format",
            "jpeg",
            "-s",
            "formatOptions",
            "best",
            "-Z",
            "2400",
            path.as_posix(),
            "--out",
            processed_path.as_posix(),
        ]
        result = run_subprocess(command, capture_output=True, text=True, check=False)
        if result.returncode != 0:
            temp_dir.cleanup()
            raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "Failed to preprocess image for OCR")
        optimize_existing_image_for_ocr_fn(processed_path)
        return temp_dir, processed_path

    passthrough_suffix = path.suffix if path.suffix else ".img"
    passthrough_path = Path(temp_dir.name) / f"{path.stem}_ocr{passthrough_suffix}"
    shutil.copy2(path, passthrough_path)
    return temp_dir, passthrough_path


def extract_image_text(
    path: Path,
    *,
    preprocess_image_for_ocr_fn: Callable[[Path], tuple[tempfile.TemporaryDirectory, Path]],
    run_ocr_backend_fn: Callable[[list[Path]], tuple[dict[str, str], str]],
) -> tuple[str, str]:
    temp_dir, processed_path = preprocess_image_for_ocr_fn(path)
    try:
        ocr_results, backend = run_ocr_backend_fn([processed_path])
        return select_best_text_variant(ocr_results.get(processed_path.as_posix(), "")), backend
    finally:
        temp_dir.cleanup()


def render_single_page_pdf_for_ocr(
    source_path: Path,
    output_path: Path,
    *,
    is_pdftoppm_available_fn: Callable[[], bool],
    is_sips_available_fn: Callable[[], bool],
    run_subprocess: Callable[..., object],
    optimize_existing_image_for_ocr_fn: Callable[[Path], None],
) -> None:
    if is_pdftoppm_available_fn():
        output_prefix = output_path.with_suffix("")
        command = [
            PDFTOPPM_BINARY,
            "-jpeg",
            "-r",
            "300",
            "-singlefile",
            source_path.as_posix(),
            output_prefix.as_posix(),
        ]
    elif is_sips_available_fn():
        command = [
            SIPS_BINARY,
            "-s",
            "format",
            "jpeg",
            "-s",
            "formatOptions",
            "best",
            "-Z",
            "2400",
            source_path.as_posix(),
            "--out",
            output_path.as_posix(),
        ]
    else:
        raise RuntimeError("Failed to render PDF page for OCR: no supported PDF renderer is available")

    result = run_subprocess(command, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "Failed to render PDF page for OCR")
    optimize_existing_image_for_ocr_fn(output_path)


def render_pdf_pages_for_ocr(
    path: Path,
    *,
    max_pages: int = 5,
    render_single_page_pdf_for_ocr_fn: Callable[[Path, Path], None],
) -> tuple[tempfile.TemporaryDirectory, list[Path]]:
    temp_dir = tempfile.TemporaryDirectory()
    image_paths: list[Path] = []
    try:
        reader = PdfReader(path.as_posix())
    except (OSError, PdfReadError, ValueError) as error:
        temp_dir.cleanup()
        raise ValueError("Не удалось прочитать PDF документ") from error
    page_count = max(1, min(len(reader.pages), max_pages))
    for page_index in range(page_count):
        single_page_pdf_path = Path(temp_dir.name) / f"ocr_page_{page_index + 1}.pdf"
        image_path = Path(temp_dir.name) / f"ocr_page_{page_index + 1}.jpg"
        writer = PdfWriter()
        writer.add_page(reader.pages[page_index])
        with single_page_pdf_path.open("wb") as output_stream:
            writer.write(output_stream)
        try:
            render_single_page_pdf_for_ocr_fn(single_page_pdf_path, image_path)
        except RuntimeError:
            temp_dir.cleanup()
            raise
        image_paths.append(image_path)

    return temp_dir, image_paths


def render_pdf_pages_for_raw_pdftoppm_ocr(
    path: Path,
    *,
    max_pages: int = 5,
    is_pdftoppm_available_fn: Callable[[], bool],
    run_subprocess: Callable[..., object],
) -> tuple[tempfile.TemporaryDirectory, list[Path]]:
    if not is_pdftoppm_available_fn():
        raise RuntimeError("Failed to render PDF page for AXB OCR fallback: pdftoppm is not available")

    temp_dir = tempfile.TemporaryDirectory()
    try:
        page_count = max(1, min(len(PdfReader(path.as_posix()).pages), max_pages))
    except (OSError, PdfReadError, ValueError) as error:
        temp_dir.cleanup()
        raise ValueError("Не удалось прочитать PDF документ") from error
    output_prefix = Path(temp_dir.name) / "ocr_raw_page"
    command = [
        PDFTOPPM_BINARY,
        "-jpeg",
        "-r",
        "300",
        "-f",
        "1",
        "-l",
        str(page_count),
        path.as_posix(),
        output_prefix.as_posix(),
    ]
    result = run_subprocess(command, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        temp_dir.cleanup()
        raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "Failed to render AXB OCR fallback pages")

    image_paths = sorted(Path(temp_dir.name).glob("ocr_raw_page-*.jpg"))
    if len(image_paths) != page_count:
        temp_dir.cleanup()
        raise RuntimeError("AXB OCR fallback did not render the expected number of pages")
    return temp_dir, image_paths


def extract_scanned_pdf_text(
    path: Path,
    *,
    render_pdf_pages_for_ocr_fn: Callable[[Path], tuple[tempfile.TemporaryDirectory, list[Path]]],
    run_ocr_backend_fn: Callable[[list[Path]], tuple[dict[str, str], str]],
    save_pdf_header_crop_for_ocr_fn: Callable[[Path, Path], bool],
) -> tuple[str, str]:
    temp_dir, image_paths = render_pdf_pages_for_ocr_fn(path)
    try:
        ocr_results, backend = run_ocr_backend_fn(image_paths)
        chunks = [select_best_text_variant(ocr_results.get(image_path.as_posix(), "")) for image_path in image_paths]

        header_crop_path = Path(temp_dir.name) / "ocr_page_1_header.jpg"
        if image_paths and save_pdf_header_crop_for_ocr_fn(image_paths[0], header_crop_path):
            header_results, _ = run_ocr_backend_fn([header_crop_path])
            header_text = select_best_text_variant(header_results.get(header_crop_path.as_posix(), ""))
            if header_text and re.search(r"(?:гос\.?\s*номер|vin|пробег|tc\s*:|автомобиль)", header_text, re.IGNORECASE):
                chunks.insert(0, header_text)

        return "\n".join(filter(None, chunks)).strip(), backend
    finally:
        temp_dir.cleanup()


def extract_axb_raw_scanned_pdf_text(
    path: Path,
    *,
    render_pdf_pages_for_raw_pdftoppm_ocr_fn: Callable[[Path], tuple[tempfile.TemporaryDirectory, list[Path]]],
    run_tesseract_ocr_with_modes_fn: Callable[..., dict[str, str]],
    save_pdf_header_crop_for_ocr_fn: Callable[[Path, Path], bool],
) -> str:
    temp_dir, image_paths = render_pdf_pages_for_raw_pdftoppm_ocr_fn(path)
    try:
        ocr_results = run_tesseract_ocr_with_modes_fn(image_paths, page_segmentation_modes=("4",))
        chunks = [select_best_text_variant(ocr_results.get(image_path.as_posix(), "")) for image_path in image_paths]

        header_crop_path = Path(temp_dir.name) / "ocr_raw_page_1_header.jpg"
        if image_paths and save_pdf_header_crop_for_ocr_fn(image_paths[0], header_crop_path):
            header_results = run_tesseract_ocr_with_modes_fn([header_crop_path], page_segmentation_modes=("4",))
            header_text = select_best_text_variant(header_results.get(header_crop_path.as_posix(), ""))
            if header_text and re.search(r"(?:гос\.?\s*номер|vin|пробег|tc\s*:|автомобиль)", header_text, re.IGNORECASE):
                chunks.insert(0, header_text)

        return "\n".join(filter(None, chunks)).strip()
    finally:
        temp_dir.cleanup()


def format_spreadsheet_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour == value.minute == value.second == value.microsecond == 0:
            return value.strftime("%d.%m.%Y")
        return value.strftime("%d.%m.%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\xa0", " ").strip()


def extract_spreadsheet_text(path: Path) -> str:
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except (OSError, BadZipFile, InvalidFileException, ValueError) as error:
        raise ValueError("Не удалось прочитать Excel документ") from error
    try:
        chunks: list[str] = []
        multiple_sheets = len(workbook.worksheets) > 1
        for sheet in workbook.worksheets:
            if multiple_sheets:
                chunks.append(f"Лист: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [format_spreadsheet_cell_value(value) for value in row]
                values = [value for value in values if value]
                if values:
                    chunks.append(" ".join(values))
        return "\n".join(chunks).strip()
    finally:
        workbook.close()


def extract_document_text(
    path: Path,
    source_type: str,
    *,
    get_ocr_backend: Callable[[], str | None],
    extract_image_text_fn: Callable[[Path], tuple[str, str]],
    extract_pdf_text_fn: Callable[[Path], str],
    extract_scanned_pdf_text_fn: Callable[[Path], tuple[str, str]],
) -> tuple[str, str, str | None]:
    if source_type == "image":
        if get_ocr_backend() is None:
            return "", "manual_review", "image_ocr_unavailable"
        text, backend = extract_image_text_fn(path)
        return text, f"image_{backend}_ocr", None if text else "image_text_not_found"

    if source_type == "xlsx":
        text = extract_spreadsheet_text(path)
        return text, "xlsx_text", None if text else "xlsx_text_not_found"

    text = select_best_text_variant(extract_pdf_text_fn(path))
    extracted_from = "pdf_text"
    failure_reason = None

    if score_text_quality(text)[0] >= 2 or score_text_quality(text)[1] >= 6:
        return text, extracted_from, None

    if get_ocr_backend() is None:
        return text, extracted_from, "pdf_ocr_unavailable" if not text else None

    try:
        scanned_text, backend = extract_scanned_pdf_text_fn(path)
    except RuntimeError:
        return text, extracted_from, "pdf_renderer_unavailable" if not text else None
    if score_text_quality(scanned_text) > score_text_quality(text):
        return scanned_text, f"pdf_{backend}_ocr", None if scanned_text else "pdf_text_not_found"

    if not text and not scanned_text:
        failure_reason = "pdf_text_not_found"
    return text or scanned_text, extracted_from if text else f"pdf_{backend}_ocr", failure_reason
