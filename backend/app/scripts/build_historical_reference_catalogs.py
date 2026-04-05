from __future__ import annotations

import argparse
import csv
import statistics
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.paths import resolve_user_path
from app.services.historical_repairs_import import (
    EXPECTED_HEADERS,
    HistoricalRepairLine,
    classify_line_kind,
    extract_order_number,
)


@dataclass
class ServiceCatalogRow:
    supplier_name: str
    lines_count: int = 0
    repairs: set[str] = field(default_factory=set)
    plates: set[str] = field(default_factory=set)
    models: set[str] = field(default_factory=set)
    vehicle_types: Counter[str] = field(default_factory=Counter)
    auto_groups: Counter[str] = field(default_factory=Counter)
    expense_items: Counter[str] = field(default_factory=Counter)
    first_date: date | None = None
    last_date: date | None = None
    total_amount: float = 0.0


@dataclass
class VehicleCatalogRow:
    plate: str
    vehicle_type: str | None
    model_name: str | None
    lines_count: int = 0
    repairs: set[str] = field(default_factory=set)
    suppliers: Counter[str] = field(default_factory=Counter)
    auto_groups: Counter[str] = field(default_factory=Counter)
    expense_items: Counter[str] = field(default_factory=Counter)
    first_date: date | None = None
    last_date: date | None = None
    max_mileage: int = 0
    total_amount: float = 0.0


@dataclass
class ModelCatalogRow:
    model_name: str
    vehicle_types: Counter[str] = field(default_factory=Counter)
    plates: set[str] = field(default_factory=set)
    suppliers: Counter[str] = field(default_factory=Counter)
    lines_count: int = 0
    repairs: set[str] = field(default_factory=set)
    first_date: date | None = None
    last_date: date | None = None
    max_mileage: int = 0
    total_amount: float = 0.0


@dataclass
class NomenclatureCatalogRow:
    key: str
    line_kind: str
    nomenclature: str
    article: str | None
    auto_group: str | None
    expense_item: str | None
    lines_count: int = 0
    repairs: set[str] = field(default_factory=set)
    suppliers: Counter[str] = field(default_factory=Counter)
    models: Counter[str] = field(default_factory=Counter)
    vehicle_types: Counter[str] = field(default_factory=Counter)
    quantities: list[float] = field(default_factory=list)
    amounts: list[float] = field(default_factory=list)
    first_date: date | None = None
    last_date: date | None = None


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    try:
        return datetime.fromisoformat(str(value).strip()).date()
    except ValueError:
        return None


def normalize_mileage(value: object) -> int:
    if value in (None, ""):
        return 0
    return max(0, int(round(float(value))))


def normalize_amount(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return round(float(value), 2)


def normalize_quantity(value: object) -> float:
    if value in (None, ""):
        return 1.0
    quantity = float(value)
    return quantity if quantity > 0 else 1.0


def format_date(value: date | None) -> str:
    return value.isoformat() if value is not None else ""


def format_counter(counter: Counter[str], limit: int = 5) -> str:
    return "; ".join(f"{name} ({count})" for name, count in counter.most_common(limit))


def format_median(values: list[float]) -> str:
    if not values:
        return ""
    return f"{statistics.median(values):.2f}"


def validate_headers(headers: tuple[object, ...]) -> None:
    normalized = tuple(str(item).strip() if item is not None else "" for item in headers)
    if normalized != EXPECTED_HEADERS:
        raise ValueError("Неожиданный формат файла. Ожидалась выгрузка `2025 для ИИ.xlsx`.")


def build_repair_key(registrator: str | None, repair_date: date | None, plate: str | None) -> str:
    order_number = extract_order_number(registrator)
    if order_number:
        return order_number
    return " | ".join([part for part in [registrator or "", format_date(repair_date), plate or ""] if part]) or "unknown"


def update_date_range(current_first: date | None, current_last: date | None, current_value: date | None) -> tuple[date | None, date | None]:
    if current_value is None:
        return current_first, current_last
    if current_first is None or current_value < current_first:
        current_first = current_value
    if current_last is None or current_value > current_last:
        current_last = current_value
    return current_first, current_last


def parse_workbook(file_obj: BinaryIO) -> tuple[dict[str, ServiceCatalogRow], dict[str, VehicleCatalogRow], dict[str, ModelCatalogRow], dict[str, NomenclatureCatalogRow], Counter[str], int]:
    file_obj.seek(0)
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except (OSError, BadZipFile, InvalidFileException, ValueError) as error:
        raise ValueError("Не удалось прочитать файл исторических справочников") from error

    try:
        worksheet = workbook[workbook.sheetnames[0]]
        iterator = worksheet.iter_rows(min_row=1, values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            raise ValueError("Файл пуст")
        validate_headers(header_row)

        services: dict[str, ServiceCatalogRow] = {}
        vehicles: dict[str, VehicleCatalogRow] = {}
        models: dict[str, ModelCatalogRow] = {}
        nomenclature_catalog: dict[str, NomenclatureCatalogRow] = {}
        expense_items = Counter[str]()
        rows_total = 0

        for row in iterator:
            rows_total += 1
            (
                raw_plate,
                vehicle_type_label,
                period_value,
                mileage_value,
                supplier_name,
                _column_name,
                registrator,
                auto_group,
                article,
                expense_item,
                vehicle_model,
                nomenclature,
                quantity,
                amount,
            ) = row

            plate = normalize_text(raw_plate)
            vehicle_type = normalize_text(vehicle_type_label)
            repair_date = normalize_date(period_value)
            mileage = normalize_mileage(mileage_value)
            supplier = normalize_text(supplier_name) or "Без поставщика"
            registrator_text = normalize_text(registrator)
            auto_group_text = normalize_text(auto_group)
            article_text = normalize_text(article)
            expense_item_text = normalize_text(expense_item)
            model_name = normalize_text(vehicle_model)
            nomenclature_name = normalize_text(nomenclature) or "Без названия"
            quantity_value = normalize_quantity(quantity)
            amount_value = normalize_amount(amount)
            repair_key = build_repair_key(registrator_text, repair_date, plate)

            line_kind = classify_line_kind(
                HistoricalRepairLine(
                    source_row=rows_total + 1,
                    nomenclature=nomenclature_name,
                    article=article_text,
                    quantity=quantity_value,
                    amount=amount_value,
                    expense_item=expense_item_text,
                    auto_group=auto_group_text,
                )
            )

            service_entry = services.setdefault(supplier, ServiceCatalogRow(supplier_name=supplier))
            service_entry.lines_count += 1
            service_entry.repairs.add(repair_key)
            if plate:
                service_entry.plates.add(plate)
            if model_name:
                service_entry.models.add(model_name)
            if vehicle_type:
                service_entry.vehicle_types[vehicle_type] += 1
            if auto_group_text:
                service_entry.auto_groups[auto_group_text] += 1
            if expense_item_text:
                service_entry.expense_items[expense_item_text] += 1
                expense_items[expense_item_text] += 1
            service_entry.total_amount += amount_value
            service_entry.first_date, service_entry.last_date = update_date_range(
                service_entry.first_date,
                service_entry.last_date,
                repair_date,
            )

            if plate:
                vehicle_entry = vehicles.setdefault(
                    plate,
                    VehicleCatalogRow(plate=plate, vehicle_type=vehicle_type, model_name=model_name),
                )
                vehicle_entry.lines_count += 1
                vehicle_entry.repairs.add(repair_key)
                vehicle_entry.suppliers[supplier] += 1
                if auto_group_text:
                    vehicle_entry.auto_groups[auto_group_text] += 1
                if expense_item_text:
                    vehicle_entry.expense_items[expense_item_text] += 1
                vehicle_entry.max_mileage = max(vehicle_entry.max_mileage, mileage)
                vehicle_entry.total_amount += amount_value
                vehicle_entry.first_date, vehicle_entry.last_date = update_date_range(
                    vehicle_entry.first_date,
                    vehicle_entry.last_date,
                    repair_date,
                )
                if vehicle_entry.vehicle_type is None and vehicle_type:
                    vehicle_entry.vehicle_type = vehicle_type
                if vehicle_entry.model_name is None and model_name:
                    vehicle_entry.model_name = model_name

            if model_name:
                model_entry = models.setdefault(model_name, ModelCatalogRow(model_name=model_name))
                model_entry.lines_count += 1
                model_entry.repairs.add(repair_key)
                if plate:
                    model_entry.plates.add(plate)
                if vehicle_type:
                    model_entry.vehicle_types[vehicle_type] += 1
                model_entry.suppliers[supplier] += 1
                model_entry.max_mileage = max(model_entry.max_mileage, mileage)
                model_entry.total_amount += amount_value
                model_entry.first_date, model_entry.last_date = update_date_range(
                    model_entry.first_date,
                    model_entry.last_date,
                    repair_date,
                )

            nomenclature_key = f"{line_kind}|{article_text or '-'}|{nomenclature_name.casefold()}"
            nomenclature_entry = nomenclature_catalog.setdefault(
                nomenclature_key,
                NomenclatureCatalogRow(
                    key=nomenclature_key,
                    line_kind=line_kind,
                    nomenclature=nomenclature_name,
                    article=article_text,
                    auto_group=auto_group_text,
                    expense_item=expense_item_text,
                ),
            )
            nomenclature_entry.lines_count += 1
            nomenclature_entry.repairs.add(repair_key)
            nomenclature_entry.suppliers[supplier] += 1
            if model_name:
                nomenclature_entry.models[model_name] += 1
            if vehicle_type:
                nomenclature_entry.vehicle_types[vehicle_type] += 1
            nomenclature_entry.quantities.append(quantity_value)
            nomenclature_entry.amounts.append(amount_value)
            nomenclature_entry.first_date, nomenclature_entry.last_date = update_date_range(
                nomenclature_entry.first_date,
                nomenclature_entry.last_date,
                repair_date,
            )

        return services, vehicles, models, nomenclature_catalog, expense_items, rows_total
    finally:
        workbook.close()


def write_services_catalog(path: Path, services: dict[str, ServiceCatalogRow]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "supplier_name",
                "lines_count",
                "repairs_count",
                "vehicles_count",
                "models_count",
                "vehicle_types_top",
                "expense_items_top",
                "auto_groups_top",
                "total_amount",
                "first_date",
                "last_date",
            ]
        )
        for item in sorted(services.values(), key=lambda current: (current.lines_count, len(current.repairs)), reverse=True):
            writer.writerow(
                [
                    item.supplier_name,
                    item.lines_count,
                    len(item.repairs),
                    len(item.plates),
                    len(item.models),
                    format_counter(item.vehicle_types),
                    format_counter(item.expense_items),
                    format_counter(item.auto_groups),
                    f"{item.total_amount:.2f}",
                    format_date(item.first_date),
                    format_date(item.last_date),
                ]
            )


def write_vehicles_catalog(path: Path, vehicles: dict[str, VehicleCatalogRow]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "plate",
                "vehicle_type",
                "model_name",
                "lines_count",
                "repairs_count",
                "suppliers_count",
                "top_suppliers",
                "top_expense_items",
                "max_mileage",
                "total_amount",
                "first_date",
                "last_date",
            ]
        )
        for item in sorted(vehicles.values(), key=lambda current: (current.lines_count, len(current.repairs)), reverse=True):
            writer.writerow(
                [
                    item.plate,
                    item.vehicle_type or "",
                    item.model_name or "",
                    item.lines_count,
                    len(item.repairs),
                    len(item.suppliers),
                    format_counter(item.suppliers),
                    format_counter(item.expense_items),
                    item.max_mileage,
                    f"{item.total_amount:.2f}",
                    format_date(item.first_date),
                    format_date(item.last_date),
                ]
            )


def write_models_catalog(path: Path, models: dict[str, ModelCatalogRow]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "model_name",
                "vehicle_types_top",
                "vehicles_count",
                "lines_count",
                "repairs_count",
                "suppliers_count",
                "top_suppliers",
                "max_mileage",
                "total_amount",
                "first_date",
                "last_date",
            ]
        )
        for item in sorted(models.values(), key=lambda current: (current.lines_count, len(current.repairs)), reverse=True):
            writer.writerow(
                [
                    item.model_name,
                    format_counter(item.vehicle_types),
                    len(item.plates),
                    item.lines_count,
                    len(item.repairs),
                    len(item.suppliers),
                    format_counter(item.suppliers),
                    item.max_mileage,
                    f"{item.total_amount:.2f}",
                    format_date(item.first_date),
                    format_date(item.last_date),
                ]
            )


def write_nomenclature_catalog(path: Path, items: list[NomenclatureCatalogRow]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "line_kind",
                "article",
                "nomenclature",
                "auto_group",
                "expense_item",
                "lines_count",
                "repairs_count",
                "suppliers_count",
                "models_count",
                "vehicle_types_top",
                "top_suppliers",
                "median_quantity",
                "median_amount",
                "min_amount",
                "max_amount",
                "first_date",
                "last_date",
            ]
        )
        for item in items:
            writer.writerow(
                [
                    item.line_kind,
                    item.article or "",
                    item.nomenclature,
                    item.auto_group or "",
                    item.expense_item or "",
                    item.lines_count,
                    len(item.repairs),
                    len(item.suppliers),
                    len(item.models),
                    format_counter(item.vehicle_types),
                    format_counter(item.suppliers),
                    format_median(item.quantities),
                    format_median(item.amounts),
                    f"{min(item.amounts):.2f}" if item.amounts else "",
                    f"{max(item.amounts):.2f}" if item.amounts else "",
                    format_date(item.first_date),
                    format_date(item.last_date),
                ]
            )


def write_expense_items_catalog(path: Path, expense_items: Counter[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["expense_item", "lines_count"])
        for name, count in expense_items.most_common():
            writer.writerow([name, count])


def write_report(
    path: Path,
    *,
    source_path: Path,
    rows_total: int,
    services: dict[str, ServiceCatalogRow],
    vehicles: dict[str, VehicleCatalogRow],
    models: dict[str, ModelCatalogRow],
    works: list[NomenclatureCatalogRow],
    parts: list[NomenclatureCatalogRow],
    expense_items: Counter[str],
    output_dir: Path,
) -> None:
    total_work_lines = sum(item.lines_count for item in works)
    total_part_lines = sum(item.lines_count for item in parts)
    lines = [
        "# Справочники из `2025 для ИИ.xlsx`",
        "",
        f"- Источник: `{source_path}`",
        f"- Строк обработано: `{rows_total}`",
        f"- Справочник сервисов: `{len(services)}` записей",
        f"- Справочник техники: `{len(vehicles)}` записей",
        f"- Справочник моделей: `{len(models)}` записей",
        f"- Справочник работ: `{len(works)}` записей (`{total_work_lines}` строк источника)",
        f"- Справочник запчастей и материалов: `{len(parts)}` записей (`{total_part_lines}` строк источника)",
        f"- Справочник статей затрат: `{len(expense_items)}` записей",
        "",
        "## Файлы",
        "",
        f"- `{output_dir / 'services_catalog.csv'}`",
        f"- `{output_dir / 'vehicles_catalog.csv'}`",
        f"- `{output_dir / 'vehicle_models_catalog.csv'}`",
        f"- `{output_dir / 'works_catalog.csv'}`",
        f"- `{output_dir / 'parts_catalog.csv'}`",
        f"- `{output_dir / 'expense_items_catalog.csv'}`",
        "",
        "## Топ поставщиков",
        "",
    ]
    for item in sorted(services.values(), key=lambda current: current.lines_count, reverse=True)[:10]:
        lines.append(f"- {item.supplier_name}: {item.lines_count} строк, {len(item.repairs)} ремонтов")
    lines.extend(
        [
            "",
            "## Топ моделей",
            "",
        ]
    )
    for item in sorted(models.values(), key=lambda current: current.lines_count, reverse=True)[:10]:
        lines.append(f"- {item.model_name}: {item.lines_count} строк, {len(item.plates)} машин")
    lines.extend(
        [
            "",
            "## Топ работ",
            "",
        ]
    )
    for item in sorted(works, key=lambda current: current.lines_count, reverse=True)[:10]:
        article_suffix = f" [{item.article}]" if item.article else ""
        lines.append(f"- {item.nomenclature}{article_suffix}: {item.lines_count} строк, {len(item.repairs)} ремонтов")
    lines.extend(
        [
            "",
            "## Топ запчастей",
            "",
        ]
    )
    for item in sorted(parts, key=lambda current: current.lines_count, reverse=True)[:10]:
        article_suffix = f" [{item.article}]" if item.article else ""
        lines.append(f"- {item.nomenclature}{article_suffix}: {item.lines_count} строк, {len(item.repairs)} ремонтов")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build_catalogs(source_path: Path, output_dir: Path) -> None:
    with source_path.open("rb") as handle:
        services, vehicles, models, nomenclature_catalog, expense_items, rows_total = parse_workbook(handle)

    works = sorted(
        (item for item in nomenclature_catalog.values() if item.line_kind == "work"),
        key=lambda current: (current.lines_count, len(current.repairs)),
        reverse=True,
    )
    parts = sorted(
        (item for item in nomenclature_catalog.values() if item.line_kind == "part"),
        key=lambda current: (current.lines_count, len(current.repairs)),
        reverse=True,
    )

    output_dir.mkdir(parents=True, exist_ok=True)
    write_services_catalog(output_dir / "services_catalog.csv", services)
    write_vehicles_catalog(output_dir / "vehicles_catalog.csv", vehicles)
    write_models_catalog(output_dir / "vehicle_models_catalog.csv", models)
    write_nomenclature_catalog(output_dir / "works_catalog.csv", works)
    write_nomenclature_catalog(output_dir / "parts_catalog.csv", parts)
    write_expense_items_catalog(output_dir / "expense_items_catalog.csv", expense_items)
    write_report(
        output_dir / "README.md",
        source_path=source_path,
        rows_total=rows_total,
        services=services,
        vehicles=vehicles,
        models=models,
        works=works,
        parts=parts,
        expense_items=expense_items,
        output_dir=output_dir,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build reference catalogs from `2025 для ИИ.xlsx`.")
    parser.add_argument("source", type=Path, help="Path to the historical Excel file.")
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path(".private/2025_ai_reference_catalogs"),
        help="Directory where derived catalogs will be written.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source_path = resolve_user_path(args.source)
    output_dir = resolve_user_path(args.output_dir)
    build_catalogs(source_path, output_dir)
    print(f"Catalogs written to: {output_dir}")


if __name__ == "__main__":
    main()
