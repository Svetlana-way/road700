from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.paths import PROJECT_ROOT
from app.db.session import SessionLocal
from app.models.enums import CatalogStatus
from app.models.labor_norm import LaborNorm
from app.services.labor_norms import (
    DEFAULT_DONGFENG_BRAND_FAMILY,
    DEFAULT_DONGFENG_CATALOG_NAME,
    DEFAULT_DONGFENG_LABOR_NORM_SCOPE,
    build_normalized_name,
    build_search_text,
    default_labor_norms_path,
    normalize_brand_family,
    normalize_labor_norm_code,
    normalize_labor_norm_scope,
    upsert_labor_norm_catalog,
)

DEFAULT_LABOR_NORMS_PATH = default_labor_norms_path(PROJECT_ROOT)
SKIPPED_SHEETS = {"Лист1", "00 Все нормы времени"}


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0

    def as_dict(self) -> dict[str, int]:
        return {
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
        }


def normalize_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    return text


def normalize_hours(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = normalize_text(value)
    if not text:
        return None
    try:
        return round(float(text.replace(",", ".")), 2)
    except ValueError:
        return None


def read_xlsx_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, object]] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIPPED_SHEETS:
            continue
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(values_only=True):
            rows.append(
                {
                    "category": sheet_name,
                    "source_sheet": sheet_name,
                    "code": row[0] if len(row) > 0 else None,
                    "name_ru": row[1] if len(row) > 1 else None,
                    "name_cn": row[2] if len(row) > 2 else None,
                    "name_en": row[3] if len(row) > 3 else None,
                    "standard_hours": row[4] if len(row) > 4 else None,
                    "name_ru_alt": row[5] if len(row) > 5 else None,
                }
            )
    return rows


def read_csv_rows(path: Path) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source)
        rows: list[dict[str, object]] = []
        for row in reader:
            rows.append(
                {
                    "category": row.get("category"),
                    "source_sheet": row.get("source_sheet") or row.get("category"),
                    "code": row.get("code"),
                    "name_ru": row.get("name_ru"),
                    "name_cn": row.get("name_cn"),
                    "name_en": row.get("name_en"),
                    "standard_hours": row.get("standard_hours"),
                    "name_ru_alt": row.get("name_ru_alt"),
                }
            )
    return rows


def upsert_labor_norm(
    db: Session,
    *,
    scope: str,
    brand_family: Optional[str],
    catalog_name: Optional[str],
    code: str,
    category: str,
    name_ru: str,
    name_ru_alt: Optional[str],
    name_cn: Optional[str],
    name_en: Optional[str],
    standard_hours: float,
    source_sheet: str,
    source_file: str,
    stats: ImportStats,
) -> None:
    existing = db.scalar(select(LaborNorm).where(LaborNorm.scope == scope, LaborNorm.code == code))
    is_new = existing is None
    labor_norm = existing or LaborNorm(scope=scope, code=code)

    labor_norm.scope = scope
    labor_norm.brand_family = brand_family
    labor_norm.catalog_name = catalog_name
    labor_norm.category = category
    labor_norm.name_ru = name_ru
    labor_norm.name_ru_alt = name_ru_alt
    labor_norm.name_cn = name_cn
    labor_norm.name_en = name_en
    labor_norm.normalized_name = build_normalized_name(name_ru, name_ru_alt, name_en)
    labor_norm.search_text = build_search_text(code, category, name_ru, name_ru_alt, name_en, name_cn)
    labor_norm.standard_hours = standard_hours
    labor_norm.source_sheet = source_sheet
    labor_norm.source_file = source_file
    labor_norm.status = CatalogStatus.CONFIRMED

    db.add(labor_norm)
    db.flush()

    if is_new:
        stats.created += 1
    else:
        stats.updated += 1


def import_labor_norms_with_session(
    db: Session,
    path: Path = DEFAULT_LABOR_NORMS_PATH,
    *,
    scope: str = DEFAULT_DONGFENG_LABOR_NORM_SCOPE,
    brand_family: Optional[str] = DEFAULT_DONGFENG_BRAND_FAMILY,
    catalog_name: Optional[str] = DEFAULT_DONGFENG_CATALOG_NAME,
) -> ImportStats:
    normalized_scope = normalize_labor_norm_scope(scope)
    if not normalized_scope:
        raise ValueError("Labor norm scope is required")

    normalized_brand_family = normalize_brand_family(brand_family)
    stats = ImportStats()
    normalized_catalog_name = catalog_name.strip() if isinstance(catalog_name, str) and catalog_name.strip() else None
    upsert_labor_norm_catalog(
        db,
        scope=normalized_scope,
        catalog_name=normalized_catalog_name,
        brand_family=normalized_brand_family,
        auto_match_enabled=False if normalized_scope != DEFAULT_DONGFENG_LABOR_NORM_SCOPE else True,
    )
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        source_rows = read_xlsx_rows(path)
    elif suffix == ".csv":
        source_rows = read_csv_rows(path)
    else:
        raise ValueError(f"Unsupported labor norms file format: {path.suffix}")

    for row in source_rows:
        code = normalize_labor_norm_code(normalize_text(row.get("code")))
        name_ru = normalize_text(row.get("name_ru"))
        name_cn = normalize_text(row.get("name_cn"))
        name_en = normalize_text(row.get("name_en"))
        standard_hours = normalize_hours(row.get("standard_hours"))
        name_ru_alt = normalize_text(row.get("name_ru_alt"))
        category = normalize_text(row.get("category")) or "Общее"
        source_sheet = normalize_text(row.get("source_sheet")) or category

        if not code or not name_ru or standard_hours is None:
            stats.skipped += 1
            continue

        upsert_labor_norm(
            db,
            scope=normalized_scope,
            brand_family=normalized_brand_family,
            catalog_name=normalized_catalog_name,
            code=code,
            category=category,
            name_ru=name_ru,
            name_ru_alt=name_ru_alt,
            name_cn=name_cn,
            name_en=name_en,
            standard_hours=standard_hours,
            source_sheet=source_sheet,
            source_file=path.name,
            stats=stats,
        )

    db.commit()
    return stats


def main() -> None:
    with SessionLocal() as db:
        stats = import_labor_norms_with_session(db)
    print(stats.as_dict())


if __name__ == "__main__":
    main()
