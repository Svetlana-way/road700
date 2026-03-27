from __future__ import annotations

from collections import Counter
import json
import logging
import re
import shutil
import statistics
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
try:
    from PIL import Image, ImageChops
except ImportError:  # pragma: no cover - optional dependency during bootstrap
    Image = None
    ImageChops = None

from pypdf import PdfReader, PdfWriter
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session, joinedload

from app.models.document import Document, DocumentVersion
from app.models.enums import (
    CatalogStatus,
    CheckSeverity,
    DocumentKind,
    DocumentStatus,
    ImportStatus,
    RepairStatus,
    VehicleStatus,
    VehicleType,
)
from app.models.imports import ImportJob
from app.models.ocr_profile_matcher import OcrProfileMatcher
from app.models.ocr_rule import OcrRule
from app.models.repair import Repair, RepairCheck, RepairPart, RepairWork
from app.models.service import Service
from app.models.vehicle import Vehicle
from app.services.import_jobs import mark_job_completed, mark_job_failed
from app.services.service_catalog import find_service_name_in_text, normalize_service_key, resolve_service_by_name
from app.services.labor_norms import (
    LaborNormApplicability,
    LaborNormEnrichmentSummary,
    assess_labor_norm_applicability,
    build_normalized_name,
    classify_known_non_catalog_operation,
    find_best_labor_norm_match,
    normalize_labor_norm_code,
)
from app.core.paths import STORAGE_ROOT


LOCAL_STORAGE_ROOT = STORAGE_ROOT
VISION_OCR_SCRIPT = Path(__file__).with_name("vision_ocr.swift")
logger = logging.getLogger(__name__)
TESSERACT_BINARY = "tesseract"
PDFTOPPM_BINARY = "pdftoppm"
SIPS_BINARY = "sips"
TESSERACT_LANGUAGE = "rus+eng"
TESSERACT_PAGE_SEGMENTATION_MODES = ("6", "4")

OCR_CONFUSABLE_CHARSETS = {
    "а": "аa4@",
    "б": "б6b",
    "в": "вb8",
    "г": "гgr",
    "д": "дdg",
    "е": "еe",
    "з": "з3",
    "и": "иu",
    "к": "кk",
    "м": "мm",
    "н": "нh",
    "о": "оo0",
    "п": "пnp",
    "р": "рpr",
    "с": "сc",
    "т": "тt",
    "у": "уy",
    "х": "хx",
}
SOURCE_PATH_KEY = "source_path"
RUSSIAN_MONTHS = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
}
FILENAME_SERVICE_PREFIXES = (
    "заказ наряды",
    "заказ-наряды",
    "заказ наряды ",
    "заказ-наряд",
)
REPEAT_REPAIR_WINDOW_DAYS = 30
EXPECTED_TOTAL_THRESHOLD_MULTIPLIER = 1.1
EXPECTED_TOTAL_SERVICE_SAMPLE_THRESHOLD = 3
EXPECTED_TOTAL_HOURLY_SAMPLE_THRESHOLD = 5
WORK_REFERENCE_MIN_SAMPLES = 3
WORK_REFERENCE_SERVICE_SAMPLE_THRESHOLD = 3
WORK_REFERENCE_VEHICLE_SAMPLE_THRESHOLD = 2
WORK_REFERENCE_WARNING_MULTIPLIER = 1.2
WORK_REFERENCE_SUSPICIOUS_MULTIPLIER = 1.35
WORK_REFERENCE_WARNING_LOWER_MULTIPLIER = 0.8
WORK_REFERENCE_SUSPICIOUS_LOWER_MULTIPLIER = 0.65
WORK_REFERENCE_MILEAGE_MARGIN_RATIO = 0.2
WORK_REFERENCE_MIN_MILEAGE_MARGIN = 10000
HISTORICAL_IMPORT_REASON_PREFIX = "historical_import:"
PLACEHOLDER_VEHICLE_EXTERNAL_ID = "__batch_import_placeholder__"
WORK_REFERENCE_OPERATIONAL_STATUSES = (
    RepairStatus.CONFIRMED,
    RepairStatus.EMPLOYEE_CONFIRMED,
)
EXPECTED_TOTAL_REPAIR_STATUSES = (
    RepairStatus.CONFIRMED,
    RepairStatus.EMPLOYEE_CONFIRMED,
    RepairStatus.SUSPICIOUS,
    RepairStatus.ARCHIVED,
)


def fuzzy_char_pattern(char: str) -> str:
    lower_char = char.lower()
    charset = OCR_CONFUSABLE_CHARSETS.get(lower_char)
    if charset is None:
        return re.escape(char)
    return f"[{re.escape(charset)}]"


def fuzzy_word_pattern(word: str) -> str:
    return "".join(fuzzy_char_pattern(char) for char in word)


def fuzzy_phrase_pattern(*words: str) -> str:
    return r"[\s-]*".join(fuzzy_word_pattern(word) for word in words)


ORDER_LABEL_PATTERN = fuzzy_phrase_pattern("заказ", "наряд")
REVERSED_ORDER_LABEL_PATTERN = fuzzy_phrase_pattern("наряд", "заказ")
ACT_LABEL_PATTERN = fuzzy_word_pattern("акт")
ACT_WORKS_LABEL_PATTERN = fuzzy_phrase_pattern("акт", "выполненных", "работ")
DOCUMENT_LABEL_PATTERN = fuzzy_word_pattern("документ")
NUMBER_MARKER_PATTERN = r"(?:№|N[Ooº°]?|#)"
DATE_CLOSED_LABEL_PATTERN = fuzzy_phrase_pattern("дата", "закрытия")
DATE_COMPLETED_LABEL_PATTERN = fuzzy_phrase_pattern("дата", "окончания", "работ")
DATE_LABEL_PATTERN = fuzzy_word_pattern("дата")
FROM_LABEL_PATTERN = fuzzy_word_pattern("от")
MILEAGE_LABEL_PATTERN = fuzzy_word_pattern("пробег")
ODOMETER_LABEL_PATTERN = fuzzy_word_pattern("одометр")
SERVICE_LABEL_PATTERN = "|".join(
    [
        fuzzy_word_pattern("поставщик"),
        fuzzy_word_pattern("исполнитель"),
        fuzzy_word_pattern("подрядчик"),
        fuzzy_word_pattern("контрагент"),
    ]
)
WORK_TOTAL_LABEL_PATTERN = "|".join(
    [
        fuzzy_phrase_pattern("работы", "итого"),
        fuzzy_phrase_pattern("стоимость", "работ"),
        fuzzy_phrase_pattern("итого", "работ"),
    ]
)
PARTS_TOTAL_LABEL_PATTERN = "|".join(
    [
        fuzzy_phrase_pattern("запчасти", "итого"),
        fuzzy_phrase_pattern("материалы", "итого"),
        fuzzy_phrase_pattern("стоимость", "запчастей"),
        fuzzy_phrase_pattern("стоимость", "материалов"),
        fuzzy_word_pattern("запчасти"),
        fuzzy_word_pattern("материалы"),
    ]
)
VAT_LABEL_PATTERN = fuzzy_word_pattern("ндс")
GRAND_TOTAL_LABEL_PATTERN = "|".join(
    [
        fuzzy_phrase_pattern("итого", "к", "оплате"),
        fuzzy_phrase_pattern("к", "оплате"),
        fuzzy_word_pattern("итого"),
        fuzzy_word_pattern("всего"),
    ]
)

ORDER_PATTERNS = [
    rf"(?:{ORDER_LABEL_PATTERN}|{REVERSED_ORDER_LABEL_PATTERN})\s*{NUMBER_MARKER_PATTERN}?\s*([A-Za-zА-Яа-я0-9/_-]{{3,}})",
    rf"(?:{ACT_WORKS_LABEL_PATTERN}|{ACT_LABEL_PATTERN})\s*{NUMBER_MARKER_PATTERN}?\s*([A-Za-zА-Яа-я0-9/_-]{{3,}})",
    rf"\b{NUMBER_MARKER_PATTERN}\s*([A-Za-zА-Яа-я0-9/_-]{{4,}})",
    r"\b([A-Z]{2,}[A-Z0-9/_-]*-\d{2,})\b",
]
DATE_PATTERNS = [
    rf"(?:{ORDER_LABEL_PATTERN}|{REVERSED_ORDER_LABEL_PATTERN}|{ACT_WORKS_LABEL_PATTERN}|{ACT_LABEL_PATTERN}|{DOCUMENT_LABEL_PATTERN})"
    rf"[^\n\r]{{0,80}}?(?:{FROM_LABEL_PATTERN})\s*[:№]?\s*(\d{{2}}[./-]\d{{2}}[./-]\d{{2,4}})",
    rf"(?:{ORDER_LABEL_PATTERN}|{REVERSED_ORDER_LABEL_PATTERN}|{ACT_WORKS_LABEL_PATTERN}|{ACT_LABEL_PATTERN}|{DOCUMENT_LABEL_PATTERN})"
    rf"[^\n\r]{{0,80}}?(?:{FROM_LABEL_PATTERN})\s*[:№]?\s*(\d{{1,2}}\s+[А-Яа-я]+\s+\d{{4}})",
    rf"(?:{DATE_CLOSED_LABEL_PATTERN}|{DATE_COMPLETED_LABEL_PATTERN})\s*[:№]?\s*(\d{{2}}[./-]\d{{2}}[./-]\d{{2,4}})",
    rf"(?:{DATE_CLOSED_LABEL_PATTERN}|{DATE_COMPLETED_LABEL_PATTERN})\s*[:№]?\s*(\d{{1,2}}\s+[А-Яа-я]+\s+\d{{4}})",
    rf"(?:{DATE_LABEL_PATTERN}|{FROM_LABEL_PATTERN})\s*[:№]?\s*(\d{{2}}[./-]\d{{2}}[./-]\d{{2,4}})",
    rf"(?:{DATE_LABEL_PATTERN}|{FROM_LABEL_PATTERN})\s*[:№]?\s*(\d{{1,2}}\s+[А-Яа-я]+\s+\d{{4}})",
    r"\b(\d{2}[./-]\d{2}[./-]\d{2,4})\b",
]
MILEAGE_PATTERNS = [
    rf"(?:{MILEAGE_LABEL_PATTERN}|{ODOMETER_LABEL_PATTERN})(?:\s*\([^)]*\))?\s*[:№]?\s*(\d[\d\s]{{1,}})",
]
PLATE_PATTERNS = [
    r"([А-ЯA-Z]\s*\d{3}\s*[А-ЯA-Z]{2}\s*\d{2,3})(?!\d)",
    r"(\d{3}\s*[А-ЯA-Z]{3}\s*\d{2,3})(?!\d)",
    r"([А-ЯA-Z]{2}\s*\d{4}\s*\d{2,3})(?!\d)",
    r"(\d{4}\s*[А-ЯA-Z]{2}\s*\d{2,3})(?!\d)",
]
VIN_PATTERNS = [
    r"(?<![A-HJ-NPR-Z0-9])([A-HJ-NPR-Z0-9]{17})(?![A-HJ-NPR-Z0-9])",
]
PLATE_LABEL_PATTERNS = [
    rf"(?:гос\.?\s*(?:номер|ном\.?\s*знак)|госномер|г/н|г\.\s*н\.)\s*[:№]?\s*(?P<value>[^\n\r]{{0,48}})",
]
VIN_LABEL_PATTERNS = [
    rf"(?:\bvin\b|vin)\s*[:№]?\s*(?P<value>[A-HJ-NPR-Z0-9]{{17}})",
    rf"(?:\bvin\b|vin)[^\n\r]{{0,60}}?(?P<value>[A-HJ-NPR-Z0-9]{{17}})",
]
CHASSIS_LABEL_PATTERNS = [
    r"(?:№\s*шасси|шасси|рама)\s*[:№]?\s*(?P<value>[A-Z0-9-]{6,20})",
]
MILEAGE_SECTION_PATTERNS = [
    rf"(?:{MILEAGE_LABEL_PATTERN}|{ODOMETER_LABEL_PATTERN})(?:\s*\([^)]*\))?\s*[:№]?\s*(?P<value>\d[\d\s]{{0,}})",
]
VEHICLE_ROW_MILEAGE_PATTERNS = [
    r"\b(\d{3}(?:\s\d{3}){0,3})\b(?=\s+(?:закрыт|открыт|rub|rur|руб|usd|eur)\b)",
    r"\b(\d{3}(?:\s\d{3}){0,3})\b(?=\s+(?:vin|вид\s+ремонта|дата\s+приема)\b)",
]
VEHICLE_SECTION_START_PATTERN = re.compile(
    r"(?:TC|ТС|Автомобиль|ТРАНСПОРТНОЕ\s+СРЕДСТВО|Модель автомобиля|Марка:)\b",
    re.IGNORECASE,
)
VEHICLE_SECTION_STOP_PATTERN = re.compile(
    r"(?:^|\n)\s*(?:ВЛАДЕЛЕЦ|Заказчик|ЗАКАЗЧИК|ПЛАТЕЛЬЩИК|Плательщик|Причина(?:\s+обращения)?|Вид ремонта|"
    r"Выполненные работы|ДОГОВОР|Контактное\s+лицо)\b",
    re.IGNORECASE,
)
SERVICE_PATTERNS = [
    rf"(?:{SERVICE_LABEL_PATTERN})\b\s*[:№]?\s*(.+?)(?=(?:\b(?:инн|кпп|адрес|тел(?:ефон)?|заказчик|плательщик|автомобиль|шасси|vin|договор|документ|заказ[- ]наряд|акт)\b)|$)",
]
SERVICE_CANDIDATE_PATTERNS = [
    re.compile(
        rf"(?:^|\n)\s*(?:{SERVICE_LABEL_PATTERN})\b\s*[:№]?\s*(?P<value>.+?)(?=(?:\b(?:инн|кпп|адрес|тел(?:ефон)?|заказчик|плательщик|автомобиль|шасси|vin|договор|документ|заказ[- ]наряд|акт)\b)|\Z)",
        re.IGNORECASE | re.MULTILINE | re.DOTALL,
    ),
    re.compile(
        r"^(?:официальный\s+дилер[^\n\r]{0,80})?(?P<value>.+?)(?=(?:\b(?:инн|кпп|адрес|тел(?:ефон)?|цех|заказ[- ]наряд|акт\s+выполненных\s+работ|документ)\b)|\Z)",
        re.IGNORECASE | re.MULTILINE | re.DOTALL,
    ),
]
REASON_LABEL_PATTERN = fuzzy_phrase_pattern("причина", "обращения")
REASON_SECTION_PATTERNS = [
    re.compile(
        rf"(?:^|\n)\s*(?:{REASON_LABEL_PATTERN})\s*[:№]?\s*(?P<value>.+?)"
        r"(?=(?:\n\s*(?:Выполненные\s+работы|Расходная\s+накладная|Акт\s+выполненных\s+работ|"
        r"Перечень\s+работ|Выполненные\s+сервисные\s+услуги|№\s+Артикул|Итого\s+работ|"
        r"Итого\s+по\s+причине\s+обращения)\b)|\Z)",
        re.IGNORECASE | re.MULTILINE | re.DOTALL,
    ),
]
TOTAL_PATTERNS = {
    "work_total": [
        r"Итого\s+работ:\s*\d+\s+\d+(?:[.,]\d+)?\s+(\d[\d\s]*(?:[.,]\d{2}))\s+\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})",
        r"Итого\s+работ:\s*(?:\d+(?:[.,]\d+)?)?[^\n\r]{0,80}?на\s+сумму:\s*(\d[\d\s]*(?:[.,]\d{2})?)",
        r"Всего\s+выполнено\s+работ\s+на\s+сумму\s*(\d[\d\s]*(?:[.,]\d{2})?)",
        rf"(?:{WORK_TOTAL_LABEL_PATTERN})\b[^\d\r\n]{{0,20}}(\d[\d\s]*(?:[.,]\d{{2}})?)(?=\s*(?:запчаст|материал|ндс|итого|сервис|сто|$))",
    ],
    "parts_total": [
        r"Итого\s+материалов:\s*\d+\s+(\d[\d\s]*(?:[.,]\d{2}))\s+\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})",
        r"Итого\s+материал(?:ов|ы):\s*(?:\d+(?:[.,]\d+)?)?[^\n\r]{0,80}?на\s+сумму:\s*(\d[\d\s]*(?:[.,]\d{2})?)",
        r"Итого\s+по\s+странице\s+материалов:\s*(?:\d+)?[^\n\r]{0,40}?на\s+сумму:\s*(\d[\d\s]*(?:[.,]\d{2})?)",
        r"Затрачено\s+материалов\s+на\s+сумму\s*(\d[\d\s]*(?:[.,]\d{2})?)",
        rf"(?:{PARTS_TOTAL_LABEL_PATTERN})\b[^\d\r\n]{{0,20}}(\d[\d\s]*(?:[.,]\d{{2}})?)(?=\s*(?:ндс|итого|сервис|сто|$))",
    ],
    "vat_total": [
        rf"(?:{VAT_LABEL_PATTERN})\b[^\d\r\n]{{0,20}}(\d[\d\s]*(?:[.,]\d{{2}})?)(?=\s*(?:итого|сервис|сто|$))",
    ],
    "grand_total": [
        r"Итого\s+по\s+акту\s+выполненных\s+работ\s*:?\s*\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})\s+(\d[\d\s]*(?:[.,]\d{2}))",
        r"Всего\s+по\s+наряд[- ]заказу:\s*\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})\s+(\d[\d\s]*(?:[.,]\d{2}))",
        r"Итого\s+по\s+заказ[- ]наряду\s*:?\s*\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})\s+(\d[\d\s]*(?:[.,]\d{2}))",
        r"Всего\s+к\s+оплате:\s*(\d[\d\s']*(?:[.,-]\d{2})?)",
        r"Итого\s+по\s+причине\s+обращения:\s*(\d[\d\s]*(?:[.,]\d{2})?)",
        rf"(?:{GRAND_TOTAL_LABEL_PATTERN})\b[^\d\r\n]{{0,20}}(\d[\d\s]*(?:[.,]\d{{2}})?)(?=\s*(?:сервис|сто|$))",
    ],
}
LINE_ITEM_PATTERN = re.compile(
    r"^(?P<name>[^\d].*?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)"
    r"(?:\s+(?P<unit>[A-Za-zА-Яа-я./-]{1,8}))?"
    r"\s+(?P<price>\d[\d\s]*(?:[.,]\d{2})?)"
    r"\s+(?P<total>\d[\d\s]*(?:[.,]\d{2})?)$",
    re.IGNORECASE,
)
PART_LINE_WITH_ARTICLE_PATTERN = re.compile(
    r"^(?P<article>[A-Za-zА-Яа-я0-9-]{3,})\s+"
    r"(?P<name>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)"
    r"(?:\s+(?P<unit>[A-Za-zА-Яа-я./-]{1,8}))?"
    r"\s+(?P<price>\d[\d\s]*(?:[.,]\d{2})?)"
    r"\s+(?P<total>\d[\d\s]*(?:[.,]\d{2})?)$",
    re.IGNORECASE,
)
WORK_SECTION_MARKERS = ("работы", "услуги", "работа:")
PART_SECTION_MARKERS = ("запчасти", "материалы", "запчасть:")
SECTION_FOOTER_MARKERS = (
    "стоимость работ",
    "работы итого",
    "итого работ",
    "запчасти итого",
    "материалы итого",
    "стоимость запчастей",
    "стоимость материалов",
    "ндс",
    "итого к оплате",
    "к оплате",
    "всего",
    "сервис",
    "сто",
    "исполнитель",
    "подрядчик",
)
ITEM_UNIT_MARKERS = {
    "шт",
    "нч",
    "ч",
    "час",
    "часа",
    "часов",
    "компл",
    "усл",
    "ед",
    "л",
    "литр",
    "кг",
    "г",
    "гр",
    "мл",
    "м",
    "к-т",
}
AXB_WORK_NAME_KEYWORDS = (
    "то ",
    "то-",
    "техническое обслуживание",
    "подсоединение",
    "отсоединение",
    "диагност",
    "нормокомплект",
    "замена",
    "снятие",
    "свароч",
    "мойка",
    "поиск неисправности",
    "ремонт",
)
AXB_PART_NAME_KEYWORDS = (
    "фильтр",
    "фитинг",
    "соединитель",
    "головка",
    "шланг",
    "смазка",
    "амортизатор",
    "палм",
)
ARTICLE_TOKEN_PATTERN = re.compile(r"^(?=.*[\d-])[A-Za-zА-Яа-я0-9/_-]{3,}$")
AXB_ARTICLE_TOKEN_PATTERN = re.compile(r"^(?=.*[\d.-])[A-Za-zА-Яа-я0-9/._-]{3,}$")
WORK_CODE_TOKEN_PATTERN = re.compile(r"^(?=.*\d)(?=.*[A-Za-zА-Яа-я])[A-Za-zА-Яа-я0-9/_-]{3,}$")
TEXT_KEYWORD_PATTERN = re.compile(
    "|".join(
        [
            fuzzy_word_pattern("заказ"),
            fuzzy_word_pattern("наряд"),
            fuzzy_word_pattern("дата"),
            fuzzy_word_pattern("госномер"),
            fuzzy_word_pattern("пробег"),
            fuzzy_word_pattern("работ"),
            fuzzy_word_pattern("запчаст"),
            fuzzy_word_pattern("итого"),
            fuzzy_word_pattern("сервис"),
            fuzzy_word_pattern("ндс"),
        ]
    ),
    re.IGNORECASE,
)
TEXT_CHAR_REPLACEMENTS = str.maketrans(
    {
        "\xa0": " ",
        "¹": "№",
        "–": "-",
        "—": "-",
        "«": '"',
        "»": '"',
    }
)
SERVICE_NAME_BLOCKLIST = (
    "стоимость",
    "работ",
    "запчаст",
    "материал",
    "ндс",
    "итого",
    "к оплате",
    "пробег",
    "госномер",
    "заказ",
    "наряд",
    "дата",
)
OCR_RULE_TARGET_FIELDS = (
    "order_number",
    "repair_date",
    "mileage",
    "plate_number",
    "vin",
    "service_name",
    "work_total",
    "parts_total",
    "vat_total",
    "grand_total",
)
OCR_RULE_VALUE_PARSERS = {
    "raw",
    "date",
    "amount",
    "digits_int",
}
OCR_TOKEN_CHAR_REPLACEMENTS = str.maketrans(
    {
        "О": "O",
        "о": "o",
        "о": "o",
        "А": "A",
        "а": "a",
        "В": "B",
        "в": "b",
        "Е": "E",
        "е": "e",
        "К": "K",
        "к": "k",
        "М": "M",
        "м": "m",
        "Н": "H",
        "н": "h",
        "Р": "P",
        "р": "p",
        "С": "C",
        "с": "c",
        "Т": "T",
        "т": "t",
        "У": "Y",
        "у": "y",
        "Х": "X",
        "х": "x",
        "І": "I",
        "і": "i",
        "—": "-",
        "–": "-",
        "−": "-",
        "‑": "-",
    }
)
DEFAULT_OCR_PROFILE_MATCHER_DEFINITIONS: list[dict[str, object]] = [
    {
        "profile_scope": "axb",
        "title": "AXB Truck Service scanned order form",
        "source_type": "pdf",
        "filename_pattern": None,
        "text_pattern": r"АХВ\s*Трак\s*Сервис",
        "service_name_pattern": None,
        "priority": 10,
    },
    {
        "profile_scope": "antares",
        "title": "Антарес order form",
        "source_type": "pdf",
        "filename_pattern": None,
        "text_pattern": r"Антарес",
        "service_name_pattern": None,
        "priority": 20,
    },
    {
        "profile_scope": "gruzovye_rezervy",
        "title": "Грузовые резервы order form",
        "source_type": "pdf",
        "filename_pattern": None,
        "text_pattern": r"ГРУЗОВЫЕ\s+РЕЗЕРВЫ|ГПТ?\d{6,}",
        "service_name_pattern": None,
        "priority": 30,
    },
    {
        "profile_scope": "ets_act",
        "title": "Енисей Трак Сервис work act",
        "source_type": "pdf",
        "filename_pattern": None,
        "text_pattern": r"Енисей\s+Трак\s+Сервис[\s\S]{0,400}Акт\s+выполненных\s+работ",
        "service_name_pattern": None,
        "priority": 40,
    },
    {
        "profile_scope": "ets_invoice",
        "title": "Енисей Трак Сервис invoice",
        "source_type": "pdf",
        "filename_pattern": None,
        "text_pattern": r"Енисей\s+Трак\s+Сервис[\s\S]{0,400}(?:СЧЕТ|Счет[- ]фактура)",
        "service_name_pattern": None,
        "priority": 50,
    },
    {
        "profile_scope": "leader_trak",
        "title": "ЛидерТрак order form",
        "source_type": "pdf",
        "filename_pattern": None,
        "text_pattern": r"ЛидерТрак|НАРЯД-ЗАКАЗ",
        "service_name_pattern": None,
        "priority": 60,
    },
    {
        "profile_scope": "logistics",
        "title": "Логистика order form",
        "source_type": "pdf",
        "filename_pattern": None,
        "text_pattern": r"ЛОГИСТИКА\",\s*ИНН|ЛОГИСТИКА\",\s*КПП",
        "service_name_pattern": None,
        "priority": 70,
    },
    {
        "profile_scope": "klever_trak",
        "title": "Клевер Трак XLSX order form",
        "source_type": "xlsx",
        "filename_pattern": None,
        "text_pattern": r"КЛЕВЕР\s+ТРАК|Заказ-наряд\s*№\s*КТ\d{6,}",
        "service_name_pattern": None,
        "priority": 75,
    },
    {
        "profile_scope": "sibtrakscan",
        "title": "СибТракСкан order form",
        "source_type": "pdf",
        "filename_pattern": None,
        "text_pattern": r"СИБТРАКСКАН|ЗСТ\d{6,}",
        "service_name_pattern": None,
        "priority": 80,
    },
]
PROFILE_SPECIFIC_OCR_RULE_DEFINITIONS: list[dict[str, object]] = [
    {"profile_scope": "axb", "target_field": "service_name", "pattern": r"(АХВ\s*Трак\s*Сервис)", "value_parser": "raw", "confidence": 0.98, "priority": 1},
    {"profile_scope": "ets_act", "target_field": "service_name", "pattern": r"(Енисей\s*Трак\s*Сервис)", "value_parser": "raw", "confidence": 0.98, "priority": 1},
    {"profile_scope": "ets_invoice", "target_field": "service_name", "pattern": r"(Енисей\s*Трак\s*Сервис)", "value_parser": "raw", "confidence": 0.98, "priority": 1},
    {"profile_scope": "leader_trak", "target_field": "service_name", "pattern": r"(ЛидерТрак)", "value_parser": "raw", "confidence": 0.98, "priority": 1},
    {"profile_scope": "leader_trak", "target_field": "order_number", "pattern": r"План\s+ремонта\s*№\s*([A-Za-zА-Яа-я0-9/_-]{3,})", "value_parser": "raw", "confidence": 0.97, "priority": 1},
    {"profile_scope": "leader_trak", "target_field": "order_number", "pattern": r"Предварительный\s+счет\s+на\s+оплату\s*№\s*([A-Za-zА-Яа-я0-9/_-]{3,})", "value_parser": "raw", "confidence": 0.92, "priority": 2},
    {"profile_scope": "klever_trak", "target_field": "service_name", "pattern": r"(КЛЕВЕР\s*ТРАК)", "value_parser": "raw", "confidence": 0.98, "priority": 1},
    {"profile_scope": "sibtrakscan", "target_field": "service_name", "pattern": r"(СИБТРАКСКАН)", "value_parser": "raw", "confidence": 0.98, "priority": 1},
    {"profile_scope": "gruzovye_rezervy", "target_field": "service_name", "pattern": r"(ГРУЗОВЫЕ\s+РЕЗЕРВЫ)", "value_parser": "raw", "confidence": 0.98, "priority": 1},
    {"profile_scope": "sibtrakscan", "target_field": "repair_date", "pattern": rf"(?:{DATE_CLOSED_LABEL_PATTERN})\s*[:№]?\s*(\d{{2}}[./-]\d{{2}}[./-]\d{{2,4}})", "value_parser": "date", "confidence": 0.96, "priority": 1},
    {"profile_scope": "sibtrakscan", "target_field": "repair_date", "pattern": rf"(?:{DATE_CLOSED_LABEL_PATTERN})\s*[:№]?\s*(\d{{1,2}}\s+[А-Яа-я]+\s+\d{{4}})", "value_parser": "date", "confidence": 0.96, "priority": 2},
    {"profile_scope": "sibtrakscan", "target_field": "repair_date", "pattern": rf"(?:{DATE_COMPLETED_LABEL_PATTERN})\s*[:№]?\s*(\d{{2}}[./-]\d{{2}}[./-]\d{{2,4}})", "value_parser": "date", "confidence": 0.92, "priority": 3},
    {"profile_scope": "sibtrakscan", "target_field": "repair_date", "pattern": rf"(?:{DATE_COMPLETED_LABEL_PATTERN})\s*[:№]?\s*(\d{{1,2}}\s+[А-Яа-я]+\s+\d{{4}})", "value_parser": "date", "confidence": 0.92, "priority": 4},
    {"profile_scope": "gruzovye_rezervy", "target_field": "repair_date", "pattern": r"(?:счет(?:\s+на\s+оплату)?)\s*[^\n\r]{0,120}?\bот\b\s*[:№]?\s*(\d{1,2}\s+[А-Яа-я]+\s+\d{4})", "value_parser": "date", "confidence": 0.97, "priority": 1},
    {"profile_scope": "gruzovye_rezervy", "target_field": "repair_date", "pattern": r"(?:счет(?:\s+на\s+оплату)?)\s*[^\n\r]{0,120}?\bот\b\s*[:№]?\s*(\d{2}[./-]\d{2}[./-]\d{2,4})", "value_parser": "date", "confidence": 0.97, "priority": 2},
    {"profile_scope": "gruzovye_rezervy", "target_field": "repair_date", "pattern": rf"(?:{ORDER_LABEL_PATTERN}|{REVERSED_ORDER_LABEL_PATTERN})[^\n\r]{{0,120}}?(?:{FROM_LABEL_PATTERN})\s*[:№]?\s*(\d{{1,2}}\s+[А-Яа-я]+\s+\d{{4}})", "value_parser": "date", "confidence": 0.96, "priority": 3},
    {"profile_scope": "gruzovye_rezervy", "target_field": "repair_date", "pattern": rf"(?:{ORDER_LABEL_PATTERN}|{REVERSED_ORDER_LABEL_PATTERN})[^\n\r]{{0,120}}?(?:{FROM_LABEL_PATTERN})\s*[:№]?\s*(\d{{2}}[./-]\d{{2}}[./-]\d{{2,4}})", "value_parser": "date", "confidence": 0.96, "priority": 4},
]
UNIT_ALIASES = {
    "шт": "шт",
    "шг": "шт",
    "шt": "шт",
    "шт.": "шт",
    "ед": "ед",
    "ед.": "ед",
    "компл": "компл",
    "компл.": "компл",
    "к-т": "компл",
    "кт": "компл",
    "усл": "усл",
    "усл.": "усл",
    "л": "л",
    "л.": "л",
    "литр": "л",
    "литров": "л",
    "кг": "кг",
    "кг.": "кг",
    "г": "г",
    "гр": "г",
    "гр.": "г",
    "мл": "мл",
    "мл.": "мл",
    "м": "м",
    "м.": "м",
    "ч": "ч",
    "ч.": "ч",
    "час": "ч",
    "часа": "ч",
    "часов": "ч",
    "нч": "нч",
    "н.ч": "нч",
    "н/ч": "нч",
    "hч": "нч",
    "h.ч": "нч",
    "h/ч": "нч",
}


@dataclass
class ProcessingResult:
    document: Document
    job: ImportJob
    message: str


@dataclass(frozen=True)
class OcrProfileSelection:
    profile_scope: str
    source: str
    reason: str


DEFAULT_OCR_RULE_DEFINITIONS: list[dict[str, object]] = [
    {"profile_scope": "default", "target_field": "order_number", "pattern": pattern, "value_parser": "raw", "confidence": 0.74, "priority": (index + 1) * 10}
    for index, pattern in enumerate(ORDER_PATTERNS)
] + [
    {"profile_scope": "default", "target_field": "repair_date", "pattern": pattern, "value_parser": "date", "confidence": 0.7, "priority": (index + 1) * 10}
    for index, pattern in enumerate(DATE_PATTERNS)
] + [
    {"profile_scope": "default", "target_field": "mileage", "pattern": pattern, "value_parser": "digits_int", "confidence": 0.82, "priority": (index + 1) * 10}
    for index, pattern in enumerate(MILEAGE_PATTERNS)
] + [
    {"profile_scope": "default", "target_field": "plate_number", "pattern": pattern, "value_parser": "raw", "confidence": 0.77, "priority": (index + 1) * 10}
    for index, pattern in enumerate(PLATE_PATTERNS)
] + [
    {"profile_scope": "default", "target_field": "vin", "pattern": pattern, "value_parser": "raw", "confidence": 0.88, "priority": (index + 1) * 10}
    for index, pattern in enumerate(VIN_PATTERNS)
] + [
    {"profile_scope": "default", "target_field": "service_name", "pattern": pattern, "value_parser": "raw", "confidence": 0.58, "priority": (index + 1) * 10}
    for index, pattern in enumerate(SERVICE_PATTERNS)
] + [
    {"profile_scope": "default", "target_field": target_field, "pattern": pattern, "value_parser": "amount", "confidence": 0.8 if target_field == "grand_total" else 0.72, "priority": (index + 1) * 10}
    for target_field, patterns in TOTAL_PATTERNS.items()
    for index, pattern in enumerate(patterns)
] + PROFILE_SPECIFIC_OCR_RULE_DEFINITIONS


def get_storage_path(storage_key: str) -> Path:
    return LOCAL_STORAGE_ROOT / storage_key


def normalize_text(text: str) -> str:
    return re.sub(r"[ \t]+", " ", text).strip()


def normalize_multiline_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def score_text_quality(text: str) -> tuple[int, int, int]:
    cyrillic_count = len(re.findall(r"[А-Яа-я]", text))
    alnum_count = len(re.findall(r"[А-Яа-яA-Za-z0-9]", text))
    keyword_hits = len(TEXT_KEYWORD_PATTERN.findall(text))
    return (keyword_hits, cyrillic_count, alnum_count)


def clean_text_lines(text: str) -> str:
    text = text.translate(TEXT_CHAR_REPLACEMENTS).replace("\r\n", "\n").replace("\r", "\n")
    return "\n".join(normalize_line(line) for line in text.splitlines() if normalize_line(line))


def generate_text_variants(text: str) -> list[str]:
    base = clean_text_lines(text)
    variants: list[str] = [base]
    seen = {base}

    for source_encoding, target_encoding in (
        ("latin1", "cp1251"),
        ("cp1252", "cp1251"),
        ("latin1", "utf-8"),
        ("cp1252", "utf-8"),
    ):
        try:
            repaired = base.encode(source_encoding, errors="ignore").decode(target_encoding, errors="ignore")
        except (LookupError, UnicodeError):
            continue
        cleaned = clean_text_lines(repaired)
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            variants.append(cleaned)

    return variants


def select_best_text_variant(text: str) -> str:
    variants = generate_text_variants(text)
    if not variants:
        return ""

    best_variant = variants[0]
    best_score = score_text_quality(best_variant)
    for candidate in variants[1:]:
        candidate_score = score_text_quality(candidate)
        if candidate_score > best_score:
            best_variant = candidate
            best_score = candidate_score
    return best_variant


def score_tesseract_ocr_variant(text: str) -> tuple[int, int, int, int, int]:
    normalized_text = select_best_text_variant(text)
    keyword_hits, cyrillic_count, alnum_count = score_text_quality(normalized_text)
    amount_hits = len(re.findall(r"\d[\d\s]*(?:[.,]\d{2})", normalized_text))
    table_hits = len(
        re.findall(
            r"(?:ремонт|итого|всего|артикул|наименование|кол[-. ]?во|норма|гос\.?\s*номер|vin|пробег)",
            normalized_text,
            re.IGNORECASE,
        )
    )
    return (table_hits, amount_hits, keyword_hits, cyrillic_count, alnum_count)


def select_best_tesseract_ocr_variant(candidates: list[str]) -> str:
    non_empty_candidates = [candidate for candidate in candidates if candidate and candidate.strip()]
    if not non_empty_candidates:
        return ""

    best_variant = select_best_text_variant(non_empty_candidates[0])
    best_score = score_tesseract_ocr_variant(non_empty_candidates[0])
    for candidate in non_empty_candidates[1:]:
        candidate_variant = select_best_text_variant(candidate)
        candidate_score = score_tesseract_ocr_variant(candidate)
        if candidate_score > best_score:
            best_variant = candidate_variant
            best_score = candidate_score
    return best_variant


def extract_reason_from_text(text: str) -> Optional[str]:
    for pattern in REASON_SECTION_PATTERNS:
        match = pattern.search(text)
        if match is None:
            continue
        value = normalize_multiline_text(match.group("value")).strip(' "')
        if value:
            return value[:2000]
    return None


def extract_recommendations_from_text(text: str) -> Optional[str]:
    fragment = extract_fragment_after_marker(
        text,
        r"Рекомендации\s*:",
        stop_patterns=(
            r"НЕ\s+ДЕЛАЕМ\s*:",
            r"После\s+подписания",
            r"Сервисные\s+услуги\s+сдал",
            r"С\s+условиями\s+Программы",
        ),
        max_chars=2000,
    )
    if not fragment:
        return None

    lines = [normalize_multiline_text(line) for line in fragment.splitlines()]
    cleaned_lines = [line.strip(" -*") for line in lines if line.strip(" -*")]
    if not cleaned_lines:
        return None
    return " ".join(cleaned_lines)[:2000]


def detect_document_flags(text: str) -> list[str]:
    normalized = normalize_multiline_text(text).lower()
    flags: list[str] = []
    if "exchange program" in normalized:
        flags.append("exchange_program_present")
    return flags


def extract_not_done_items_from_text(text: str) -> list[str]:
    fragment = extract_fragment_after_marker(
        text,
        r"НЕ\s+ДЕЛАЕМ\s*:",
        stop_patterns=(
            r"После\s+подписания",
            r"Сервисные\s+услуги\s+сдал",
            r"Страница\s+\d+\s+из\s+\d+",
        ),
        max_chars=2000,
    )
    if not fragment:
        return []

    items: list[str] = []
    for raw_line in fragment.splitlines():
        line = normalize_multiline_text(raw_line).strip()
        if not line:
            continue
        line = re.sub(r"^\d+[.)]\s*", "", line).strip()
        if not line:
            continue
        items.append(line[:300])
    return items[:6]


def is_vision_ocr_available() -> bool:
    return shutil.which("swift") is not None and VISION_OCR_SCRIPT.exists()


def is_tesseract_ocr_available() -> bool:
    return shutil.which(TESSERACT_BINARY) is not None


def is_pdftoppm_available() -> bool:
    return shutil.which(PDFTOPPM_BINARY) is not None


def is_sips_available() -> bool:
    return shutil.which(SIPS_BINARY) is not None


def get_available_ocr_backend() -> str | None:
    if is_vision_ocr_available():
        return "vision"
    if is_tesseract_ocr_available():
        return "tesseract"
    return None


def get_available_pdf_renderer() -> str | None:
    if is_pdftoppm_available():
        return "pdftoppm"
    if is_sips_available():
        return "sips"
    return None


def get_ocr_runtime_status() -> dict[str, object]:
    ocr_backend = get_available_ocr_backend()
    pdf_renderer = get_available_pdf_renderer()
    return {
        "ocr_backend": ocr_backend,
        "pdf_renderer": pdf_renderer,
        "image_ocr_available": ocr_backend is not None,
        "pdf_scan_ocr_available": ocr_backend is not None and pdf_renderer is not None,
        "vision_available": is_vision_ocr_available(),
        "tesseract_available": is_tesseract_ocr_available(),
        "pdftoppm_available": is_pdftoppm_available(),
        "sips_available": is_sips_available(),
    }


def get_ocr_runtime_issues(*, require_pdf_scan_ocr: bool = True) -> list[str]:
    status = get_ocr_runtime_status()
    issues: list[str] = []
    if not bool(status["image_ocr_available"]):
        issues.append("OCR backend for images is not available")
    if require_pdf_scan_ocr and not bool(status["pdf_scan_ocr_available"]):
        if status["ocr_backend"] is None:
            issues.append("OCR backend for scanned PDFs is not available")
        if status["pdf_renderer"] is None:
            issues.append("PDF renderer for OCR is not available")
    return issues


def format_ocr_runtime_status_lines(*, require_pdf_scan_ocr: bool = True) -> list[str]:
    status = get_ocr_runtime_status()
    issues = get_ocr_runtime_issues(require_pdf_scan_ocr=require_pdf_scan_ocr)
    lines = [
        f"OCR backend: {status['ocr_backend'] or 'missing'}",
        f"PDF renderer: {status['pdf_renderer'] or 'missing'}",
        f"Image OCR available: {'yes' if status['image_ocr_available'] else 'no'}",
        f"Scanned PDF OCR available: {'yes' if status['pdf_scan_ocr_available'] else 'no'}",
    ]
    if issues:
        lines.append(f"Issues: {'; '.join(issues)}")
    else:
        lines.append("Issues: none")
    return lines


def ensure_ocr_runtime(*, require_pdf_scan_ocr: bool = True) -> None:
    issues = get_ocr_runtime_issues(require_pdf_scan_ocr=require_pdf_scan_ocr)
    if issues:
        raise RuntimeError("; ".join(issues))


def is_pillow_available() -> bool:
    return Image is not None and ImageChops is not None


def parse_amount(value: str) -> Optional[float]:
    cleaned = value.replace(" ", "").replace("\xa0", "").replace("'", "").replace("|", "").replace(",", ".")
    cleaned = re.sub(r"(?<=\d)-(?=\d{2}$)", ".", cleaned)
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return None


def parse_date_value(value: str) -> Optional[date]:
    normalized_value = normalize_text(value).strip().lower()
    for fmt in (
        "%d.%m.%Y",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d.%m.%y",
        "%d-%m-%y",
        "%d/%m/%y",
        "%Y.%m.%d",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(normalized_value, fmt).date()
        except ValueError:
            continue
    textual_match = re.search(r"(\d{1,2})\s+([а-я]+)\s+(\d{4})", normalized_value)
    if textual_match is not None:
        day = int(textual_match.group(1))
        month = RUSSIAN_MONTHS.get(textual_match.group(2))
        year = int(textual_match.group(3))
        if month is not None:
            try:
                return date(year, month, day)
            except ValueError:
                return None
    return None


def normalize_identifier_token(value: str | None) -> Optional[str]:
    if not value:
        return None
    normalized = re.sub(r"[^A-Za-zА-Яа-я0-9]+", "", value).upper()
    return normalized or None


def is_plausible_order_number(value: str | None) -> bool:
    if not value:
        return False
    normalized = normalize_text(value)
    if len(normalized) < 3:
        return False
    return any(char.isdigit() for char in normalized)


def extract_document_source_path(document: Document) -> Optional[str]:
    candidates: list[tuple[int, str]] = []
    for version in document.versions:
        payload = version.parsed_payload if isinstance(version.parsed_payload, dict) else {}
        source_path = payload.get(SOURCE_PATH_KEY)
        if isinstance(source_path, str) and source_path.strip():
            candidates.append((version.version_number, source_path.strip()))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def derive_service_name_from_source_path(source_path: str | None) -> Optional[str]:
    if not source_path:
        return None
    parent_name = Path(source_path).parent.name.strip()
    if not parent_name:
        return None
    normalized = parent_name.replace("_", " ").strip()
    lowered = normalized.lower()
    for prefix in FILENAME_SERVICE_PREFIXES:
        if lowered.startswith(prefix):
            normalized = normalized[len(prefix):].strip(" -_")
            break
    return normalized.strip() or None


def build_document_hint_text(document: Document) -> str:
    parts = [document.original_filename or ""]
    source_path = extract_document_source_path(document)
    if source_path:
        parts.append(source_path)
        service_hint = derive_service_name_from_source_path(source_path)
        if service_hint:
            parts.append(service_hint)
    return " | ".join(part for part in parts if part)


def remove_manual_review_reason(reasons: list[str], value: str) -> None:
    while value in reasons:
        reasons.remove(value)


def add_manual_review_reason(reasons: list[str], value: str) -> None:
    if value not in reasons:
        reasons.append(value)


def apply_document_metadata_fallbacks(
    document: Document,
    *,
    extracted_fields: dict[str, object],
    confidence_map: dict[str, float],
    manual_review_reasons: list[str],
    normalization_notes: list[str],
) -> None:
    hint_text = build_document_hint_text(document)
    if not hint_text:
        return

    if "plate_number" not in extracted_fields:
        plate_number = first_match(PLATE_PATTERNS, hint_text)
        normalized_plate = normalize_identifier_token(plate_number)
        if normalized_plate:
            extracted_fields["plate_number"] = normalized_plate
            confidence_map["plate_number"] = max(confidence_map.get("plate_number", 0.0), 0.55)
            normalization_notes.append("Госномер дополнен по имени файла или пути источника.")

    if "vin" not in extracted_fields:
        vin = first_match(VIN_PATTERNS, hint_text)
        normalized_vin = normalize_identifier_token(vin)
        if normalized_vin:
            extracted_fields["vin"] = normalized_vin
            confidence_map["vin"] = max(confidence_map.get("vin", 0.0), 0.6)
            normalization_notes.append("VIN дополнен по имени файла или пути источника.")

    if "order_number" not in extracted_fields:
        order_number = first_match(ORDER_PATTERNS, hint_text)
        if order_number:
            extracted_fields["order_number"] = order_number
            confidence_map["order_number"] = max(confidence_map.get("order_number", 0.0), 0.5)
            remove_manual_review_reason(manual_review_reasons, "order_number_missing")
            normalization_notes.append("Номер документа дополнен по имени файла или пути источника.")

    if "repair_date" not in extracted_fields:
        raw_date = first_match(DATE_PATTERNS, hint_text)
        parsed_date = parse_date_value(raw_date) if raw_date else None
        if parsed_date is not None:
            extracted_fields["repair_date"] = parsed_date.isoformat()
            confidence_map["repair_date"] = max(confidence_map.get("repair_date", 0.0), 0.48)
            remove_manual_review_reason(manual_review_reasons, "repair_date_missing")
            remove_manual_review_reason(manual_review_reasons, "repair_date_invalid")
            normalization_notes.append("Дата ремонта дополнена по имени файла или пути источника.")

    if "service_name" not in extracted_fields:
        service_name = derive_service_name_from_source_path(extract_document_source_path(document))
        if service_name and not is_service_name_suspicious(service_name):
            extracted_fields["service_name"] = service_name[:120]
            confidence_map["service_name"] = max(confidence_map.get("service_name", 0.0), 0.45)
            normalization_notes.append("Сервис дополнен по папке источника документа.")


def enrich_vehicle_fields_from_repair(
    repair: Repair,
    *,
    extracted_fields: dict[str, object],
    confidence_map: dict[str, float],
    normalization_notes: list[str],
) -> None:
    vehicle = repair.vehicle
    if vehicle is None or vehicle.external_id == PLACEHOLDER_VEHICLE_EXTERNAL_ID:
        return

    if "plate_number" not in extracted_fields and vehicle.plate_number:
        normalized_plate = normalize_identifier_token(vehicle.plate_number)
        if normalized_plate:
            extracted_fields["plate_number"] = normalized_plate
            confidence_map["plate_number"] = max(confidence_map.get("plate_number", 0.0), 0.35)
            normalization_notes.append("Госномер дополнен по карточке техники.")

    if "vin" not in extracted_fields and vehicle.vin:
        normalized_vin = normalize_identifier_token(vehicle.vin)
        if normalized_vin:
            extracted_fields["vin"] = normalized_vin
            confidence_map["vin"] = max(confidence_map.get("vin", 0.0), 0.35)
            normalization_notes.append("VIN дополнен по карточке техники.")


def find_vehicle_by_identifiers(
    db: Session,
    *,
    plate_number: str | None,
    vin: str | None,
    chassis_number: str | None = None,
) -> Vehicle | None:
    normalized_plate = normalize_compare_token(plate_number)
    normalized_vin = normalize_compare_token(vin)
    normalized_chassis = normalize_compare_token(chassis_number)
    if not normalized_plate and not normalized_vin and not normalized_chassis:
        return None

    vehicles = db.scalars(select(Vehicle)).all()
    exact_matches: dict[int, Vehicle] = {}
    partial_plate_matches: dict[int, Vehicle] = {}

    for vehicle in vehicles:
        if vehicle.external_id == PLACEHOLDER_VEHICLE_EXTERNAL_ID:
            continue
        vehicle_plate = normalize_compare_token(vehicle.plate_number)
        vehicle_vin = normalize_compare_token(vehicle.vin)
        if normalized_vin and vehicle_vin == normalized_vin:
            exact_matches[vehicle.id] = vehicle
            continue
        if normalized_chassis and vehicle_vin and vehicle_vin.endswith(normalized_chassis):
            exact_matches[vehicle.id] = vehicle
            continue
        if normalized_plate and vehicle_plate == normalized_plate:
            exact_matches[vehicle.id] = vehicle
            continue
        if (
            normalized_plate
            and vehicle_plate
            and len(normalized_plate) >= 6
            and (
                vehicle_plate.startswith(normalized_plate)
                or normalized_plate.startswith(vehicle_plate)
            )
        ):
            partial_plate_matches[vehicle.id] = vehicle

    if len(exact_matches) == 1:
        return next(iter(exact_matches.values()))
    if len(exact_matches) > 1:
        return None
    if len(partial_plate_matches) == 1:
        return next(iter(partial_plate_matches.values()))
    return None


def enrich_vehicle_fields_from_registry(
    db: Session,
    *,
    extracted_fields: dict[str, object],
    confidence_map: dict[str, float],
    normalization_notes: list[str],
) -> None:
    vehicle = find_vehicle_by_identifiers(
        db,
        plate_number=str(extracted_fields.get("plate_number")) if extracted_fields.get("plate_number") else None,
        vin=str(extracted_fields.get("vin")) if extracted_fields.get("vin") else None,
        chassis_number=str(extracted_fields.get("chassis_number")) if extracted_fields.get("chassis_number") else None,
    )
    if vehicle is None:
        return

    if "plate_number" not in extracted_fields and vehicle.plate_number:
        normalized_plate = normalize_identifier_token(vehicle.plate_number)
        if normalized_plate:
            extracted_fields["plate_number"] = normalized_plate
            confidence_map["plate_number"] = max(confidence_map.get("plate_number", 0.0), 0.4)
            normalization_notes.append("Госномер дополнен по совпадению с реестром техники.")

    if "vin" not in extracted_fields and vehicle.vin:
        normalized_vin = normalize_identifier_token(vehicle.vin)
        if normalized_vin:
            extracted_fields["vin"] = normalized_vin
            confidence_map["vin"] = max(confidence_map.get("vin", 0.0), 0.4)
            normalization_notes.append("VIN дополнен по совпадению с реестром техники.")


def auto_link_repair_vehicle_from_registry(
    db: Session,
    repair: Repair,
    *,
    extracted_fields: dict[str, object],
    normalization_notes: list[str],
) -> None:
    vehicle = repair.vehicle
    if vehicle is None or vehicle.external_id != PLACEHOLDER_VEHICLE_EXTERNAL_ID:
        return

    matched_vehicle = find_vehicle_by_identifiers(
        db,
        plate_number=str(extracted_fields.get("plate_number")) if extracted_fields.get("plate_number") else None,
        vin=str(extracted_fields.get("vin")) if extracted_fields.get("vin") else None,
        chassis_number=str(extracted_fields.get("chassis_number")) if extracted_fields.get("chassis_number") else None,
    )
    if matched_vehicle is None or matched_vehicle.id == repair.vehicle_id:
        return

    repair.vehicle_id = matched_vehicle.id
    repair.vehicle = matched_vehicle
    normalization_notes.append(
        f"Ремонт автоматически перепривязан к технике {matched_vehicle.plate_number or matched_vehicle.id} по совпадению с реестром."
    )


def infer_vehicle_type_from_document_text(text: str) -> VehicleType:
    vehicle_section = extract_vehicle_section_text(text)
    if re.search(
        r"\b(?:п/п|полуприцеп|прицеп|schmitz|cargobull|koluman|orthaus|krone|kogel|wielton|тонар|рефрижератор)\b",
        vehicle_section,
        re.IGNORECASE,
    ):
        return VehicleType.TRAILER
    return VehicleType.TRUCK


def extract_vehicle_year_from_document_text(text: str) -> Optional[int]:
    vehicle_section = extract_vehicle_section_text(text)
    match = re.search(r"год\s+вып\.?\s*(?P<value>20\d{2}|19\d{2})", vehicle_section, re.IGNORECASE)
    if match is None:
        return None
    try:
        return int(match.group("value"))
    except ValueError:
        return None


def extract_vehicle_brand_model_from_document_text(text: str) -> tuple[Optional[str], Optional[str]]:
    vehicle_section = extract_vehicle_section_text(text)
    descriptor_match = re.search(
        r"(?:ТС|Автомобиль)\s*:\s*(?P<value>.*?)(?=\b(?:гос\.?\s*номер|vin|пробег|год\s+вып)\b|$)",
        vehicle_section,
        re.IGNORECASE | re.DOTALL,
    )
    if descriptor_match is None:
        return None, None

    descriptor = normalize_line(descriptor_match.group("value")).strip(" ,.;:-")
    if not descriptor:
        return None, None

    descriptor = re.sub(r"\b(?:п/п|полуприцеп|прицеп)\b\s*", "", descriptor, flags=re.IGNORECASE).strip(" ,.;:-")
    if not descriptor:
        return None, None

    tokens = descriptor.split()
    if not tokens:
        return None, None

    brand = tokens[0].upper()
    model = " ".join(tokens[1:]).strip() or None
    return brand[:120], model[:255] if model else None


def auto_create_repair_vehicle_from_document(
    repair: Repair,
    document: Document,
    *,
    extracted_fields: dict[str, object],
    text: str,
    normalization_notes: list[str],
) -> bool:
    vehicle = repair.vehicle
    if vehicle is None or vehicle.external_id != PLACEHOLDER_VEHICLE_EXTERNAL_ID:
        return False

    plate_number = str(extracted_fields.get("plate_number")) if extracted_fields.get("plate_number") else None
    vin = str(extracted_fields.get("vin")) if extracted_fields.get("vin") else None
    chassis_number = str(extracted_fields.get("chassis_number")) if extracted_fields.get("chassis_number") else None

    normalized_plate = find_plate_candidate(plate_number) or normalize_identifier_token(plate_number)
    normalized_vin = find_vin_candidate(vin) or normalize_identifier_token(vin)
    normalized_chassis = find_chassis_candidate(chassis_number) or normalize_identifier_token(chassis_number)

    # Safe OCR auto-create requires a strong hardware identifier.
    if not normalized_vin and not normalized_chassis:
        return False

    vehicle_type = infer_vehicle_type_from_document_text(text)
    brand, model = extract_vehicle_brand_model_from_document_text(text)
    year = extract_vehicle_year_from_document_text(text)

    created_vehicle = Vehicle(
        external_id=None,
        vehicle_type=vehicle_type,
        vin=normalized_vin,
        plate_number=normalized_plate,
        brand=brand,
        model=model,
        year=year,
        status=VehicleStatus.ACTIVE,
        comment="Карточка техники автоматически создана из OCR документа.",
        source_payload={
            "created_from": "document_ocr_auto_create",
            "created_from_document_id": document.id,
            "created_from_repair_id": repair.id,
            "plate_number": normalized_plate,
            "vin": normalized_vin,
            "chassis_number": normalized_chassis,
            "brand": brand,
            "model": model,
            "year": year,
        },
    )
    repair.vehicle = created_vehicle
    normalization_notes.append(
        "Для техники не найдено совпадение в реестре, создана новая карточка по OCR документа."
    )
    return True


def parse_decimal_value(value: str) -> Optional[float]:
    normalized = value.replace(" ", "").replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def normalize_ocr_rule_code(value: str | None) -> Optional[str]:
    if value is None:
        return None
    normalized = normalize_text(str(value)).lower()
    normalized = re.sub(r"[^a-z0-9_:-]+", "_", normalized)
    normalized = normalized.strip("_")
    return normalized or None


def ensure_default_ocr_rules(db: Session) -> None:
    existing_signatures = {
        (
            str(rule.profile_scope),
            str(rule.target_field),
            str(rule.pattern),
        )
        for rule in db.scalars(select(OcrRule)).all()
    }
    for item in DEFAULT_OCR_RULE_DEFINITIONS:
        signature = (
            str(item["profile_scope"]),
            str(item["target_field"]),
            str(item["pattern"]),
        )
        if signature in existing_signatures:
            continue
        db.add(
            OcrRule(
                profile_scope=str(item["profile_scope"]),
                target_field=str(item["target_field"]),
                pattern=str(item["pattern"]),
                value_parser=str(item["value_parser"]),
                confidence=float(item["confidence"]),
                priority=int(item["priority"]),
                is_active=True,
            )
        )
        existing_signatures.add(signature)
    db.flush()


def ensure_default_ocr_profile_matchers(db: Session) -> None:
    existing_count = db.scalar(select(func.count(OcrProfileMatcher.id))) or 0
    if existing_count > 0:
        return
    for item in DEFAULT_OCR_PROFILE_MATCHER_DEFINITIONS:
        db.add(
            OcrProfileMatcher(
                profile_scope=str(item["profile_scope"]),
                title=str(item["title"]),
                source_type=str(item["source_type"]) if item.get("source_type") else None,
                filename_pattern=str(item["filename_pattern"]) if item.get("filename_pattern") else None,
                text_pattern=str(item["text_pattern"]) if item.get("text_pattern") else None,
                service_name_pattern=str(item["service_name_pattern"]) if item.get("service_name_pattern") else None,
                priority=int(item["priority"]),
                is_active=True,
            )
        )
    db.flush()


def load_active_ocr_rules(db: Session, *, profile_scope: str | None = None) -> list[OcrRule]:
    ensure_default_ocr_rules(db)
    stmt = (
        select(OcrRule)
        .where(OcrRule.is_active.is_(True))
        .order_by(OcrRule.profile_scope.asc(), OcrRule.target_field.asc(), OcrRule.priority.asc(), OcrRule.id.asc())
    )
    normalized_profile_scope = normalize_ocr_rule_code(profile_scope) if profile_scope else None
    if normalized_profile_scope:
        stmt = stmt.where(OcrRule.profile_scope.in_(("default", normalized_profile_scope)))
    return db.scalars(stmt).all()


def load_active_ocr_profile_matchers(db: Session) -> list[OcrProfileMatcher]:
    ensure_default_ocr_profile_matchers(db)
    stmt = (
        select(OcrProfileMatcher)
        .where(OcrProfileMatcher.is_active.is_(True))
        .order_by(OcrProfileMatcher.priority.asc(), OcrProfileMatcher.id.asc())
    )
    return db.scalars(stmt).all()


def infer_builtin_profile_scope_from_text(text: str, *, source_type: str = "pdf") -> Optional[str]:
    matched: list[dict[str, object]] = []
    for item in DEFAULT_OCR_PROFILE_MATCHER_DEFINITIONS:
        matcher_source_type = str(item.get("source_type") or "")
        if matcher_source_type and matcher_source_type != source_type:
            continue
        text_pattern = str(item.get("text_pattern") or "")
        if not text_pattern:
            continue
        try:
            if re.search(text_pattern, text, re.IGNORECASE | re.MULTILINE | re.DOTALL) is None:
                continue
        except re.error:
            continue
        matched.append(item)

    if not matched:
        return None

    matched.sort(key=lambda item: int(item["priority"]))
    best = matched[0]
    runner_up = matched[1] if len(matched) > 1 else None
    if runner_up is not None and int(runner_up["priority"]) == int(best["priority"]) and runner_up["profile_scope"] != best["profile_scope"]:
        return None
    return str(best["profile_scope"])


def extract_profile_history_scope(document: Document) -> Optional[str]:
    repair = document.repair
    if repair is None:
        return None
    candidate_versions = []
    for sibling in repair.documents:
        if sibling.id == document.id:
            continue
        for version in sibling.versions:
            payload = version.parsed_payload if isinstance(version.parsed_payload, dict) else {}
            profile_scope = payload.get("ocr_profile_scope")
            if isinstance(profile_scope, str) and profile_scope.strip():
                candidate_versions.append((version.created_at, profile_scope.strip()))
    if not candidate_versions:
        return None
    candidate_versions.sort(key=lambda item: item[0], reverse=True)
    return candidate_versions[0][1]


def profile_matcher_applies(
    matcher: OcrProfileMatcher,
    *,
    document: Document,
    text: str,
) -> bool:
    if matcher.source_type and matcher.source_type != document.source_type:
        return False

    filename = document.original_filename or ""
    if matcher.filename_pattern:
        try:
            if re.search(matcher.filename_pattern, filename, re.IGNORECASE | re.MULTILINE) is None:
                return False
        except re.error:
            return False

    if matcher.text_pattern:
        try:
            if re.search(matcher.text_pattern, text, re.IGNORECASE | re.MULTILINE | re.DOTALL) is None:
                return False
        except re.error:
            return False

    if matcher.service_name_pattern:
        service_name = document.repair.service.name if document.repair and document.repair.service else ""
        try:
            if re.search(matcher.service_name_pattern, service_name, re.IGNORECASE | re.MULTILINE | re.DOTALL) is None:
                return False
        except re.error:
            return False

    return True


def select_ocr_profile_scope(db: Session, document: Document, text: str) -> OcrProfileSelection:
    history_scope = extract_profile_history_scope(document)
    matchers = load_active_ocr_profile_matchers(db)
    matched = [item for item in matchers if profile_matcher_applies(item, document=document, text=text)]
    if matched:
        matched.sort(key=lambda item: (item.priority, item.id))
        best = matched[0]
        runner_up = matched[1] if len(matched) > 1 else None
        if (
            runner_up is not None
            and runner_up.priority == best.priority
            and runner_up.profile_scope != best.profile_scope
        ):
            if history_scope:
                return OcrProfileSelection(
                    profile_scope=history_scope,
                    source="history_fallback",
                    reason="Есть несколько одинаково подходящих matcher-правил, выбран последний профиль ремонта",
                )
            return OcrProfileSelection(
                profile_scope="default",
                source="ambiguous_fallback",
                reason="Есть несколько одинаково подходящих matcher-правил, выбран default-профиль",
            )
        return OcrProfileSelection(
            profile_scope=best.profile_scope,
            source="matcher",
            reason=best.title,
        )

    if history_scope:
        return OcrProfileSelection(
            profile_scope=history_scope,
            source="history",
            reason="Использован последний OCR-профиль из истории ремонта",
        )

    return OcrProfileSelection(
        profile_scope="default",
        source="default",
        reason="Подходящий профиль не найден, использован default",
    )


def build_ocr_rule_sort_key(rule: OcrRule, *, profile_scope: str | None = None) -> tuple[int, int, int, int]:
    normalized_profile_scope = normalize_ocr_rule_code(profile_scope)
    normalized_rule_scope = normalize_ocr_rule_code(rule.profile_scope)
    if normalized_profile_scope and normalized_rule_scope == normalized_profile_scope:
        scope_rank = 0
    elif normalized_rule_scope == "default":
        scope_rank = 1
    else:
        scope_rank = 2
    return (scope_rank, int(rule.priority), -len(str(rule.pattern or "")), int(rule.id or 0))


def group_ocr_rules_by_field(rules: list[OcrRule], *, profile_scope: str | None = None) -> dict[str, list[OcrRule]]:
    grouped: dict[str, list[OcrRule]] = {}
    for rule in rules:
        grouped.setdefault(rule.target_field, []).append(rule)
    for field_rules in grouped.values():
        field_rules.sort(key=lambda item: build_ocr_rule_sort_key(item, profile_scope=profile_scope))
    return grouped


def match_custom_ocr_rule(text: str, rules: list[OcrRule]) -> tuple[Optional[str], Optional[float], Optional[OcrRule]]:
    for rule in rules:
        try:
            match = re.search(rule.pattern, text, re.IGNORECASE | re.MULTILINE)
        except re.error:
            continue
        if not match:
            continue
        captured = match.group(1) if match.groups() else match.group(0)
        return normalize_text(captured), float(rule.confidence), rule
    return None, None, None


def parse_ocr_rule_value(raw_value: str, value_parser: str) -> Optional[object]:
    if value_parser == "date":
        parsed_date = parse_date_value(raw_value)
        return parsed_date.isoformat() if parsed_date else None
    if value_parser == "amount":
        return parse_amount(raw_value)
    if value_parser == "digits_int":
        digits_only = re.sub(r"\D", "", raw_value)
        return int(digits_only) if digits_only else None
    return raw_value


def extract_header_field(
    text: str,
    *,
    target_field: str,
    fallback_patterns: list[str],
    fallback_parser: str,
    fallback_confidence: float,
    rule_map: dict[str, list[OcrRule]],
) -> tuple[Optional[object], Optional[float], bool]:
    rules = rule_map.get(target_field, [])
    custom_match, custom_confidence, matched_rule = match_custom_ocr_rule(text, rules)
    if custom_match is not None:
        parser_name = matched_rule.value_parser if matched_rule is not None else fallback_parser
        parsed_value = parse_ocr_rule_value(custom_match, parser_name)
        if parsed_value is not None:
            return parsed_value, custom_confidence, False
        return None, None, True

    fallback_match = first_match(fallback_patterns, text)
    if fallback_match is None:
        return None, None, False
    parsed_value = parse_ocr_rule_value(fallback_match, fallback_parser)
    if parsed_value is None:
        return None, None, True
    return parsed_value, fallback_confidence, False


def first_match(patterns: list[str], text: str) -> Optional[str]:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            return normalize_text(match.group(1))
    return None


def normalize_line(line: str) -> str:
    return normalize_text(line.replace("\xa0", " "))


def normalize_service_candidate(value: str | None) -> Optional[str]:
    if not value:
        return None
    normalized_value = normalize_text(str(value).replace("\n", " ").replace("\r", " "))
    normalized_value = re.sub(r"\s+", " ", normalized_value).strip(" -:;,")
    return normalized_value or None


def extract_header_text(text: str, limit: int = 2500) -> str:
    return text[:limit]


def extract_vehicle_section_text(text: str, limit: int = 1800) -> str:
    head = text[:3000]
    match = VEHICLE_SECTION_START_PATTERN.search(head)
    if match is None:
        return extract_header_text(text, limit=limit)

    fragment = head[match.start(): match.start() + limit]
    stop_match = VEHICLE_SECTION_STOP_PATTERN.search(fragment[1:])
    if stop_match is not None:
        fragment = fragment[: stop_match.start() + 1]
    return fragment


def normalize_compare_token(value: str | None) -> Optional[str]:
    if not value:
        return None
    normalized = normalize_identifier_token(normalize_ocr_code_token(value))
    return normalized or None


def find_pattern_value(patterns: list[str], text: str) -> Optional[str]:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match is None:
            continue
        captured = match.group("value") if "value" in match.groupdict() else match.group(1)
        normalized = normalize_text(captured)
        if normalized:
            return normalized
    return None


def find_plate_candidate(value: str | None) -> Optional[str]:
    if not value:
        return None

    search_variants = [normalize_text(value)]
    translated_variant = normalize_text(normalize_ocr_code_token(value))
    if translated_variant and translated_variant not in search_variants:
        search_variants.append(translated_variant)

    for candidate_text in search_variants:
        for pattern in PLATE_PATTERNS:
            match = re.search(pattern, candidate_text, re.IGNORECASE | re.MULTILINE)
            if match is None:
                continue
            normalized = normalize_identifier_token(match.group(1))
            if normalized:
                return normalized
    return None


def find_vin_candidate(value: str | None) -> Optional[str]:
    if not value:
        return None
    normalized_value = normalize_ocr_code_token(normalize_text(value)).upper()
    for pattern in VIN_PATTERNS:
        match = re.search(pattern, normalized_value, re.IGNORECASE | re.MULTILINE)
        if match is None:
            continue
        normalized = normalize_identifier_token(match.group(1))
        if normalized:
            return normalized
    return None


def find_chassis_candidate(value: str | None) -> Optional[str]:
    if not value:
        return None
    normalized_value = normalize_ocr_code_token(normalize_text(value)).upper()
    for pattern in CHASSIS_LABEL_PATTERNS:
        match = re.search(pattern, normalized_value, re.IGNORECASE | re.MULTILINE)
        if match is None:
            continue
        normalized = normalize_identifier_token(match.group("value"))
        if normalized and 6 <= len(normalized) <= 17:
            return normalized
    standalone_match = re.search(r"(?<![A-Z0-9])([A-Z]\d{6,8})(?![A-Z0-9])", normalized_value)
    if standalone_match is not None:
        normalized = normalize_identifier_token(standalone_match.group(1))
        if normalized:
            return normalized
    return None


def parse_mileage_candidate(value: str | None) -> Optional[int]:
    if not value:
        return None
    digits_only = re.sub(r"\D", "", value)
    if not digits_only:
        return None
    try:
        mileage = int(digits_only)
    except ValueError:
        return None
    if mileage < 1:
        return None
    return mileage


def has_explicit_missing_mileage(text: str) -> bool:
    return bool(
        re.search(
            rf"(?:{MILEAGE_LABEL_PATTERN}|{ODOMETER_LABEL_PATTERN})(?:\s*\([^)]*\))?\s*[:№]?\s*[-—]+(?:\s|$)",
            text,
            re.IGNORECASE | re.MULTILINE,
        )
    )


def has_logistics_blank_mileage_field(text: str) -> bool:
    section_text = extract_vehicle_section_text(text)
    return bool(
        re.search(
            rf"(?:{MILEAGE_LABEL_PATTERN}|{ODOMETER_LABEL_PATTERN})(?:\s*\([^)]*\))?\s*[:№]?\s*(?:\r?\n\s*)?(?=Цена\s+автомототранспортного\s+средства)",
            section_text,
            re.IGNORECASE | re.MULTILINE,
        )
    )


def is_logistics_trailer_vehicle_context(text: str) -> bool:
    section_text = extract_vehicle_section_text(text)
    return bool(
        re.search(
            r"\b(?:п/п|полуприцеп|прицеп|koluman|orthaus|schmitz)\b",
            section_text,
            re.IGNORECASE,
        )
    )


def is_gruzovye_rezervy_invoice_only_document(text: str) -> bool:
    header_text = extract_header_text(text, limit=4000)
    if not re.search(r"(?:^|\n)\s*счет(?:\s+на\s+оплату)?\b", header_text, re.IGNORECASE):
        return False
    if re.search(r"\b(?:заказ[- ]наряд|наряд[- ]заказ|акт(?:\s+выполненных\s+работ)?)\b", text, re.IGNORECASE):
        return False
    if re.search(r"(?:^|\n)\s*виды\s+работ:\s*$", text, re.IGNORECASE | re.MULTILINE):
        return False
    return True


def is_leader_trak_invoice_only_document(text: str) -> bool:
    header_text = extract_header_text(text, limit=5000)
    if not (
        re.search(r"внимание!\s*оплата\s+данного\s+счета\s+означает", header_text, re.IGNORECASE)
        or re.search(r"(?:^|\n)\s*счет(?:\s+на\s+оплату)?\b", header_text, re.IGNORECASE)
    ):
        return False
    if re.search(r"\bнаряд[- ]заказ\b", text, re.IGNORECASE):
        return False
    if re.search(
        r"выполненные\s+сервисные\s+услуги\s+и\s+использованные\s+материалы",
        text,
        re.IGNORECASE,
    ):
        return False
    return True


def extract_vehicle_identifiers_from_section(text: str) -> tuple[Optional[str], Optional[str], Optional[int]]:
    section_text = extract_vehicle_section_text(text)
    plate_number = find_plate_candidate(find_pattern_value(PLATE_LABEL_PATTERNS, section_text))
    if plate_number is None:
        plate_number = find_plate_candidate(section_text)

    vin = find_vin_candidate(find_pattern_value(VIN_LABEL_PATTERNS, section_text))
    if vin is None:
        vin = find_vin_candidate(section_text)

    mileage: Optional[int] = None
    if not has_explicit_missing_mileage(section_text):
        for pattern in MILEAGE_SECTION_PATTERNS:
            match = re.search(pattern, section_text, re.IGNORECASE | re.MULTILINE)
            if match is None:
                continue
            mileage = parse_mileage_candidate(match.group("value"))
            if mileage is not None:
                break

        if mileage is None:
            vehicle_lines = [normalize_line(line) for line in section_text.splitlines() if normalize_line(line)]
            for index, line in enumerate(vehicle_lines):
                window = line
                if re.search(r"пробег|одометр", line, re.IGNORECASE) and index + 1 < len(vehicle_lines):
                    window = f"{line} {vehicle_lines[index + 1]}"
                mileage_window = re.split(r"(?:пробег|одометр)", window, maxsplit=1, flags=re.IGNORECASE)
                search_window = mileage_window[1] if len(mileage_window) > 1 else window
                for pattern in VEHICLE_ROW_MILEAGE_PATTERNS:
                    match = re.search(pattern, search_window, re.IGNORECASE | re.MULTILINE)
                    if match is None:
                        continue
                    mileage = parse_mileage_candidate(match.group(1))
                    if mileage is not None:
                        break
                if mileage is not None:
                    break

    return plate_number, vin, mileage


def extract_service_candidate_from_text(text: str) -> Optional[str]:
    text_head = text[:2000]
    for pattern in SERVICE_CANDIDATE_PATTERNS:
        match = pattern.search(text_head)
        if match is None:
            continue
        candidate = normalize_service_candidate(match.group("value"))
        if not candidate:
            continue
        candidate = re.split(
            r"\b(?:инн|кпп|адрес|тел(?:ефон)?|заказчик|плательщик|автомобиль|шасси|vin|договор|документ|заказ[- ]наряд|акт)\b",
            candidate,
            maxsplit=1,
            flags=re.IGNORECASE,
        )[0].strip(" ,.;:-")
        if candidate:
            return candidate[:200]
    return None


def extract_ets_act_late_fragment(text: str) -> str:
    pattern = re.compile(r"Акт\s+выполненных\s+работ", re.IGNORECASE)
    matches = list(pattern.finditer(text))
    for match in reversed(matches):
        fragment = text[match.start() : match.start() + 5000]
        if re.search(r"Модель\s+автомобиля|Пробег\s*\(м/ч\)|VIN|Гос\.\s*номер", fragment, re.IGNORECASE):
            return fragment
    return ""


def is_ets_act_work_code_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    return bool(
        re.fullmatch(r"[A-Za-zА-Яа-я0-9.-]{5,}", normalized_line)
        and "-" in normalized_line
        and any(char.isdigit() for char in normalized_line)
    )


def is_ets_act_sparse_work_noise_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    lower_line = normalized_line.lower()
    if not normalized_line:
        return True
    if lower_line in {"n", "артикул", "наименование", "кол. оп.", "цена", "н/ч", "норма"}:
        return True
    if lower_line.startswith("итого работ"):
        return True
    if re.fullmatch(r"\d+", normalized_line):
        return True
    if parse_amount(normalized_line) is not None:
        return True
    if re.search(r"\bруб(?:\.|ля|лей)?\b", lower_line):
        return True
    return False


def parse_ets_act_sparse_summary_amount(line: str) -> Optional[float]:
    normalized_line = normalize_line(line)
    if not normalized_line:
        return None
    if re.fullmatch(r"\d{1,2}", normalized_line):
        return None
    if normalized_line in {"T", "НДС"}:
        return None
    return parse_amount(normalized_line)


def extract_ets_act_sparse_summary_blocks(fragment: str) -> list[dict[str, list[float]]]:
    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    blocks: list[dict[str, list[float]]] = []
    index = 0

    while index < len(lines):
        if lines[index].lower() != "сумма":
            index += 1
            continue

        index += 1
        net_values: list[float] = []
        while index < len(lines) and not lines[index].lower().startswith("сумма ндс"):
            amount_value = parse_ets_act_sparse_summary_amount(lines[index])
            if amount_value is not None:
                net_values.append(amount_value)
            index += 1

        if index >= len(lines):
            break

        index += 1
        vat_values: list[float] = []
        while index < len(lines) and not lines[index].lower().startswith("сумма с учетом"):
            amount_value = parse_ets_act_sparse_summary_amount(lines[index])
            if amount_value is not None:
                vat_values.append(amount_value)
            index += 1

        if index >= len(lines):
            break

        index += 1
        total_values: list[float] = []
        while index < len(lines) and lines[index].lower() != "сумма" and not lines[index].lower().startswith("итого по акту"):
            amount_value = parse_ets_act_sparse_summary_amount(lines[index])
            if amount_value is not None:
                total_values.append(amount_value)
            index += 1

        if net_values and vat_values and total_values:
            blocks.append(
                {
                    "net_values": net_values,
                    "vat_values": vat_values,
                    "total_values": total_values,
                }
            )

    return blocks


def extract_ets_act_sparse_work_layout(fragment: str) -> list[tuple[str, str]]:
    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    start_index = next(
        (index for index, line in enumerate(lines) if line.lower().startswith("выполненные работы по акту выполненных работ")),
        None,
    )
    end_index = next((index for index, line in enumerate(lines) if line.lower().startswith("итого работ")), None)
    if start_index is None or end_index is None or end_index <= start_index:
        return []

    section_lines = lines[start_index + 1 : end_index]
    first_code_index = next((index for index, line in enumerate(section_lines) if is_ets_act_work_code_line(line)), None)
    if first_code_index is None:
        return []
    section_lines = section_lines[first_code_index:]

    layout: list[tuple[str, str]] = []
    index = 0
    while index < len(section_lines):
        if not is_ets_act_work_code_line(section_lines[index]):
            index += 1
            continue

        run_start = index
        while index < len(section_lines) and is_ets_act_work_code_line(section_lines[index]):
            index += 1
        code_run = section_lines[run_start:index]

        name_lines: list[str] = []
        while index < len(section_lines) and not is_ets_act_work_code_line(section_lines[index]):
            candidate_line = section_lines[index]
            if not is_ets_act_sparse_work_noise_line(candidate_line):
                name_lines.append(candidate_line)
            index += 1

        if len(code_run) > 1:
            for code_index, code_value in enumerate(code_run):
                if code_index < len(code_run) - 1 and code_index < len(name_lines):
                    name_value = name_lines[code_index]
                else:
                    tail_lines = name_lines[len(code_run) - 1 :] if len(name_lines) >= len(code_run) - 1 else []
                    name_value = normalize_line(" ".join(tail_lines))
                if name_value:
                    layout.append((normalize_article_value(code_value) or code_value, name_value))
            continue

        name_value = normalize_line(" ".join(name_lines))
        if name_value:
            layout.append((normalize_article_value(code_run[0]) or code_run[0], name_value))

    return layout


def extract_ets_act_sparse_work_norms(fragment: str) -> list[float]:
    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    start_index = next((index for index, line in enumerate(lines) if line.lower().startswith("итого работ")), None)
    end_index = next((index for index, line in enumerate(lines) if line.lower().startswith("расходная накладная")), None)
    if start_index is None or end_index is None or end_index <= start_index:
        return []

    norms: list[float] = []
    for line in lines[start_index + 1 : end_index]:
        if not re.fullmatch(r"\d+(?:[.,]\d+)?", line) or not re.search(r"[.,]", line):
            continue
        norm_value = parse_decimal_value(line)
        if norm_value is not None:
            norms.append(norm_value)
    return norms


def extract_ets_act_sparse_scanned_work_items(text: str) -> list[dict[str, object]]:
    fragment = extract_ets_act_late_fragment(text)
    if not fragment:
        return []

    layout = extract_ets_act_sparse_work_layout(fragment)
    summary_blocks = extract_ets_act_sparse_summary_blocks(fragment)
    if not layout or not summary_blocks:
        return []

    work_block = next(
        (
            block
            for block in summary_blocks
            if len(block["net_values"]) >= len(layout)
            and len(block["vat_values"]) >= len(layout)
            and len(block["total_values"]) >= len(layout)
        ),
        None,
    )
    if work_block is None:
        return []

    net_values = work_block["net_values"][:-1] if len(work_block["net_values"]) == len(layout) + 1 else work_block["net_values"]
    total_values = work_block["total_values"][:-1] if len(work_block["total_values"]) == len(layout) + 1 else work_block["total_values"]
    norms = extract_ets_act_sparse_work_norms(fragment)
    if len(norms) == len(layout) + 1:
        norms = norms[:-1]

    item_count = min(len(layout), len(net_values), len(total_values), len(norms) if norms else len(layout))
    if item_count <= 0:
        return []

    works: list[dict[str, object]] = []
    for index in range(item_count):
        work_code, work_name = layout[index]
        standard_hours = norms[index] if index < len(norms) else None
        net_total = net_values[index]
        price = round(net_total / standard_hours, 2) if standard_hours and standard_hours > 0 else net_total
        works.append(
            {
                "work_code": normalize_article_value(work_code),
                "work_name": work_name[:500],
                "quantity": 1.0,
                "unit_name": None,
                "price": price,
                "line_total": net_total,
                "standard_hours": standard_hours,
            }
        )

    return works


def extract_ets_act_sparse_scanned_totals(text: str) -> dict[str, float]:
    fragment = extract_ets_act_late_fragment(text)
    if not fragment:
        return {}

    summary_blocks = extract_ets_act_sparse_summary_blocks(fragment)
    if len(summary_blocks) < 3:
        return {}

    item_blocks = [block for block in summary_blocks if len(block["net_values"]) >= 2 and len(block["vat_values"]) >= 2 and len(block["total_values"]) >= 2]
    overall_block = summary_blocks[-1]
    if len(item_blocks) < 2 or not overall_block["net_values"] or not overall_block["vat_values"] or not overall_block["total_values"]:
        return {}

    work_total = item_blocks[0]["net_values"][-1]
    parts_total = item_blocks[1]["net_values"][-1]
    overall_net_total = overall_block["net_values"][-1]
    vat_total = overall_block["vat_values"][-1]
    grand_total = overall_block["total_values"][-1]

    if not amounts_match(round(work_total + parts_total, 2), overall_net_total, tolerance=0.5):
        return {}
    if not amounts_match(round(overall_net_total + vat_total, 2), grand_total, tolerance=0.5):
        return {}

    return {
        "work_total": work_total,
        "parts_total": parts_total,
        "vat_total": vat_total,
        "grand_total": grand_total,
    }


def extract_ets_act_scanned_header_fields(text: str) -> dict[str, object]:
    fragment = extract_ets_act_late_fragment(text)
    if not fragment:
        return {}

    extracted: dict[str, object] = {}

    order_match = re.search(
        r"Акт\s+выполненных\s+работ\s*(?:N|№|No|Nº)?\s*(?P<value>[A-Za-zА-Яа-я0-9]{4,})",
        fragment,
        re.IGNORECASE,
    )
    if order_match:
        extracted["order_number"] = normalize_identifier_token(order_match.group("value"))

    plate_match = re.search(r"Гос\.\s*номер\s*(?P<value>[^\n\r]{0,24})", fragment, re.IGNORECASE)
    if plate_match:
        raw_plate_value = plate_match.group("value")
        plate_number = find_plate_candidate(raw_plate_value) or find_plate_candidate(normalize_identifier_token(raw_plate_value))
        if plate_number:
            extracted["plate_number"] = plate_number

    vin_match = re.search(r"VIN\s*(?P<value>[^\n\r]{0,24})", fragment, re.IGNORECASE)
    if vin_match:
        vin = find_vin_candidate(vin_match.group("value")) or find_vin_candidate(fragment)
        if vin:
            extracted["vin"] = vin

    mileage_match = re.search(
        rf"(?:{MILEAGE_LABEL_PATTERN}|{ODOMETER_LABEL_PATTERN})(?:\s*\([^)]*\))?\s*[:№]?\s*(?P<value>\d[\d\s]{{2,}})",
        fragment,
        re.IGNORECASE,
    )
    if mileage_match:
        mileage = parse_mileage_candidate(mileage_match.group("value"))
        if mileage is not None:
            extracted["mileage"] = mileage

    return extracted


def build_section_body_pattern(markers: tuple[str, ...], stop_markers: tuple[str, ...]) -> re.Pattern[str]:
    marker_pattern = "|".join(re.escape(marker.rstrip(":")) + ":?" for marker in markers)
    stop_pattern = "|".join(re.escape(marker) for marker in stop_markers)
    return re.compile(
        rf"(?:^|\b)(?:{marker_pattern})\b\s*(?P<body>.+?)(?=(?:\b(?:{stop_pattern})\b)|$)",
        re.IGNORECASE | re.DOTALL,
    )


def tokenize_inline_section(section_text: str) -> list[str]:
    return [token for token in re.split(r"\s+", normalize_line(section_text)) if token]


def normalize_token_for_unit(token: str) -> str:
    return token.lower().strip(".,;:!?)(")


def normalize_ocr_code_token(value: str) -> str:
    return normalize_text(value).translate(OCR_TOKEN_CHAR_REPLACEMENTS)


def normalize_unit_name(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    normalized_value = normalize_token_for_unit(value)
    compact_value = normalized_value.replace(" ", "")
    translated_value = normalize_ocr_code_token(compact_value).lower()
    translated_value = translated_value.replace(".", ".").replace("/", "/")
    return (
        UNIT_ALIASES.get(compact_value)
        or UNIT_ALIASES.get(translated_value)
        or compact_value
        or None
    )


def normalize_article_value(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    normalized_value = normalize_ocr_code_token(value).upper()
    normalized_value = re.sub(r"[^A-Z0-9/_-]+", "", normalized_value)
    return normalized_value or None


def split_work_code_and_name(value: str) -> tuple[Optional[str], str]:
    normalized_value = normalize_text(value)
    parts = normalized_value.split(maxsplit=1)
    if len(parts) == 2 and (WORK_CODE_TOKEN_PATTERN.fullmatch(parts[0]) or ARTICLE_TOKEN_PATTERN.fullmatch(parts[0])):
        return normalize_article_value(parts[0]), parts[1]
    return None, normalized_value


def is_meaningful_work_name(value: object) -> bool:
    normalized_value = normalize_line(str(value or ""))
    lower_value = normalized_value.lower()
    if not normalized_value:
        return False
    if re.fullmatch(r"\d+(?:[.,]\d+)?\s*(?:н/?ч|нч|шт|г|гр|мл|литр|л|кг|м)\b", lower_value):
        return False

    alpha_tokens = re.findall(r"[A-Za-zА-Яа-я]+", lower_value)
    if any(token in {"итого", "всего", "ндс", "rub", "руб"} for token in alpha_tokens):
        return False
    if any(marker in lower_value for marker in ("итого rub", "всего rub", "сумма ндс", "в т.ч. ндс")):
        return False
    return True


def is_meaningful_part_name(value: object) -> bool:
    normalized_value = normalize_line(str(value or ""))
    lower_value = normalized_value.lower()
    if not normalized_value:
        return False
    if not re.search(r"[A-Za-zА-Яа-я]", normalized_value):
        return False
    if any(marker in lower_value for marker in ("итого", "всего", "ндс", "rub", "руб")):
        return False
    if re.fullmatch(r"\d+(?:[.,]\d+)?\s*(?:шт|г|гр|мл|литр|л|кг|м)\b", lower_value):
        return False
    return True


def has_meaningful_leader_trak_items(extracted_items: dict[str, list[dict[str, object]]]) -> bool:
    works = extracted_items.get("works") or []
    parts = extracted_items.get("parts") or []
    return any(is_meaningful_work_name(item.get("work_name")) for item in works) or any(
        is_meaningful_part_name(item.get("part_name")) for item in parts
    )


def sanitize_extracted_items(extracted_items: dict[str, list[dict[str, object]]]) -> tuple[dict[str, list[dict[str, object]]], int]:
    works = extracted_items.get("works") or []
    parts = extracted_items.get("parts") or []
    sanitized_works = [item for item in works if is_meaningful_work_name(item.get("work_name"))]
    removed_count = len(works) - len(sanitized_works)
    return {"works": sanitized_works, "parts": list(parts)}, removed_count


def is_quantity_token(token: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:[.,]\d+)?", token))


def is_amount_token(token: str) -> bool:
    return parse_amount(token) is not None


def parse_inline_item_sequence(section_text: str, target: str) -> list[dict[str, object]]:
    tokens = tokenize_inline_section(section_text)
    items: list[dict[str, object]] = []
    start_index = 0

    while start_index < len(tokens):
        matched = False
        for quantity_index in range(start_index + 1, len(tokens)):
            if not is_quantity_token(tokens[quantity_index]):
                continue

            amount_start_index = quantity_index + 1
            unit_token: Optional[str] = None
            normalized_candidate_unit = (
                normalize_unit_name(tokens[amount_start_index]) if amount_start_index < len(tokens) else None
            )
            if normalized_candidate_unit in ITEM_UNIT_MARKERS:
                unit_token = normalized_candidate_unit
                amount_start_index += 1

            if amount_start_index + 1 >= len(tokens):
                continue
            if not is_amount_token(tokens[amount_start_index]) or not is_amount_token(tokens[amount_start_index + 1]):
                continue

            prefix_tokens = tokens[start_index:quantity_index]
            if not prefix_tokens:
                continue

            payload_text = " ".join(prefix_tokens)
            if any(marker in payload_text.lower() for marker in SECTION_FOOTER_MARKERS):
                return items

            quantity = parse_decimal_value(tokens[quantity_index])
            price = parse_amount(tokens[amount_start_index])
            total = parse_amount(tokens[amount_start_index + 1])
            if quantity is None or price is None or total is None:
                continue

            payload: Optional[dict[str, object]]
            if target == "works":
                work_code, work_name = split_work_code_and_name(payload_text)
                payload = {
                    "work_code": work_code,
                    "work_name": work_name[:500],
                    "quantity": quantity,
                    "unit_name": normalize_unit_name(unit_token),
                    "price": price,
                    "line_total": total,
                }
            else:
                article = None
                name_tokens = prefix_tokens
                if len(prefix_tokens) > 1 and ARTICLE_TOKEN_PATTERN.fullmatch(prefix_tokens[0]):
                    article = normalize_article_value(prefix_tokens[0])
                    name_tokens = prefix_tokens[1:]
                part_name = " ".join(name_tokens).strip()
                if not part_name:
                    continue
                payload = {
                    "article": article,
                    "part_name": part_name[:500],
                    "quantity": quantity,
                    "unit_name": normalize_unit_name(unit_token),
                    "price": price,
                    "line_total": total,
                }

            items.append(payload)
            start_index = amount_start_index + 2
            matched = True
            break

        if not matched:
            break

    return items


def extract_inline_section_items(text: str) -> dict[str, list[dict[str, object]]]:
    works: list[dict[str, object]] = []
    parts: list[dict[str, object]] = []
    normalized_text = normalize_line(text.replace("\n", " "))

    work_pattern = build_section_body_pattern(
        WORK_SECTION_MARKERS,
        PART_SECTION_MARKERS + SECTION_FOOTER_MARKERS,
    )
    part_pattern = build_section_body_pattern(
        PART_SECTION_MARKERS,
        SECTION_FOOTER_MARKERS,
    )

    work_match = work_pattern.search(normalized_text)
    if work_match:
        works = parse_inline_item_sequence(work_match.group("body"), "works")

    part_match = part_pattern.search(normalized_text)
    if part_match:
        parts = parse_inline_item_sequence(part_match.group("body"), "parts")

    return {"works": works, "parts": parts}


def extract_line_items(text: str) -> dict[str, list[dict[str, object]]]:
    works: list[dict[str, object]] = []
    parts: list[dict[str, object]] = []
    current_section: Optional[str] = None

    for raw_line in text.splitlines():
        line = normalize_line(raw_line)
        if not line:
            continue

        lower_line = line.lower()
        if any(marker in lower_line for marker in WORK_SECTION_MARKERS):
            current_section = "works"
            if lower_line.startswith("работа:"):
                payload = parse_work_line(line.split(":", 1)[1].strip())
                if payload:
                    works.append(payload)
            continue

        if any(marker in lower_line for marker in PART_SECTION_MARKERS):
            current_section = "parts"
            if lower_line.startswith("запчасть:"):
                payload = parse_part_line(line.split(":", 1)[1].strip())
                if payload:
                    parts.append(payload)
            continue

        if current_section == "works":
            payload = parse_work_line(line)
            if payload:
                works.append(payload)
                continue

        if current_section == "parts":
            payload = parse_part_line(line)
            if payload:
                parts.append(payload)
                continue

        if lower_line.startswith("работа:"):
            payload = parse_work_line(line.split(":", 1)[1].strip())
            if payload:
                works.append(payload)
            continue

        if lower_line.startswith("запчасть:"):
            payload = parse_part_line(line.split(":", 1)[1].strip())
            if payload:
                parts.append(payload)

    if works or parts:
        return {"works": works, "parts": parts}

    inline_items = extract_inline_section_items(text)
    return inline_items


def parse_work_line(line: str) -> Optional[dict[str, object]]:
    match = LINE_ITEM_PATTERN.match(line)
    if not match:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    price = parse_amount(match.group("price"))
    total = parse_amount(match.group("total"))
    work_code, name = split_work_code_and_name(normalize_text(match.group("name")))
    if quantity is None or price is None or total is None or not name:
        return None

    return {
        "work_code": work_code,
        "work_name": name[:500],
        "quantity": quantity,
        "unit_name": normalize_unit_name(match.group("unit")),
        "price": price,
        "line_total": total,
    }


def parse_part_line(line: str) -> Optional[dict[str, object]]:
    match = PART_LINE_WITH_ARTICLE_PATTERN.match(line)
    article = None
    if match is None:
        match = LINE_ITEM_PATTERN.match(line)
    else:
        article = normalize_text(match.group("article"))

    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    price = parse_amount(match.group("price"))
    total = parse_amount(match.group("total"))
    name = normalize_text(match.group("name"))
    if quantity is None or price is None or total is None or not name:
        return None

    return {
        "article": normalize_article_value(article),
        "part_name": name[:500],
        "quantity": quantity,
        "unit_name": normalize_unit_name(match.groupdict().get("unit")),
        "price": price,
        "line_total": total,
    }


def is_service_name_suspicious(value: str) -> bool:
    normalized_value = normalize_line(value).lower()
    if not normalized_value:
        return True
    if any(marker in normalized_value for marker in SERVICE_NAME_BLOCKLIST):
        return True
    if len(re.findall(r"\d", normalized_value)) >= 6:
        return True
    if re.search(r"\d[\d\s]*(?:[.,]\d{2})", normalized_value):
        return True
    return False


def summarize_line_totals(extracted_items: dict[str, list[dict[str, object]]]) -> tuple[Optional[float], Optional[float]]:
    works = extracted_items.get("works") or []
    parts = extracted_items.get("parts") or []
    works_total = round(sum(float(item["line_total"]) for item in works), 2) if works else None
    parts_total = round(sum(float(item["line_total"]) for item in parts), 2) if parts else None
    return works_total, parts_total


def amounts_match(left: Optional[float], right: Optional[float], tolerance: float = 0.0) -> bool:
    if left is None or right is None:
        return False
    return abs(left - right) <= tolerance


def reconcile_header_totals_with_line_items(
    extracted_fields: dict[str, object],
    extracted_items: dict[str, list[dict[str, object]]],
    confidence_map: dict[str, float],
) -> list[str]:
    notes: list[str] = []
    works_total_from_lines, parts_total_from_lines = summarize_line_totals(extracted_items)

    if works_total_from_lines is not None and "work_total" not in extracted_fields:
        extracted_fields["work_total"] = works_total_from_lines
        confidence_map["work_total"] = 0.68
        notes.append("work_total_restored_from_lines")

    if parts_total_from_lines is not None and "parts_total" not in extracted_fields:
        extracted_fields["parts_total"] = parts_total_from_lines
        confidence_map["parts_total"] = 0.68
        notes.append("parts_total_restored_from_lines")

    grand_total = float(extracted_fields["grand_total"]) if "grand_total" in extracted_fields else None
    vat_total = float(extracted_fields.get("vat_total", 0) or 0) if "vat_total" in extracted_fields else 0.0
    if grand_total is None:
        return notes

    if works_total_from_lines is not None and parts_total_from_lines is not None:
        inferred_grand_total = round(works_total_from_lines + parts_total_from_lines + vat_total, 2)
        if amounts_match(inferred_grand_total, grand_total):
            current_work_total = float(extracted_fields["work_total"]) if "work_total" in extracted_fields else None
            current_parts_total = float(extracted_fields["parts_total"]) if "parts_total" in extracted_fields else None

            if not amounts_match(current_work_total, works_total_from_lines):
                extracted_fields["work_total"] = works_total_from_lines
                confidence_map["work_total"] = max(confidence_map.get("work_total", 0), 0.68)
                notes.append("work_total_aligned_with_lines")

            if not amounts_match(current_parts_total, parts_total_from_lines):
                extracted_fields["parts_total"] = parts_total_from_lines
                confidence_map["parts_total"] = max(confidence_map.get("parts_total", 0), 0.68)
                notes.append("parts_total_aligned_with_lines")

    return notes


def extract_fragment_after_marker(
    text: str,
    marker_pattern: str,
    *,
    stop_patterns: tuple[str, ...] = (),
    max_chars: int = 2500,
) -> Optional[str]:
    marker_match = re.search(marker_pattern, text, re.IGNORECASE | re.MULTILINE)
    if marker_match is None:
        return None

    fragment = text[marker_match.end(): marker_match.end() + max_chars]
    stop_offsets = []
    for pattern in stop_patterns:
        stop_match = re.search(pattern, fragment, re.IGNORECASE | re.MULTILINE)
        if stop_match is not None:
            stop_offsets.append(stop_match.start())
    if stop_offsets:
        fragment = fragment[: min(stop_offsets)]
    return fragment


def extract_largest_amount_around_marker(
    text: str,
    marker_pattern: str,
    *,
    before_chars: int = 0,
    after_chars: int = 2500,
    stop_patterns: tuple[str, ...] = (),
) -> Optional[float]:
    marker_match = re.search(marker_pattern, text, re.IGNORECASE | re.MULTILINE)
    if marker_match is None:
        return None

    prefix = text[max(0, marker_match.start() - before_chars) : marker_match.start()]
    suffix = text[marker_match.end() : marker_match.end() + after_chars]
    stop_offsets = []
    for pattern in stop_patterns:
        stop_match = re.search(pattern, suffix, re.IGNORECASE | re.MULTILINE)
        if stop_match is not None:
            stop_offsets.append(stop_match.start())
    if stop_offsets:
        suffix = suffix[: min(stop_offsets)]

    fragment = f"{prefix}\n{suffix}".strip()
    return extract_largest_amount_from_fragment(fragment)


def extract_largest_amount_from_fragment(fragment: str | None) -> Optional[float]:
    if not fragment:
        return None
    candidates = [
        parse_amount(match.group(0))
        for match in re.finditer(r"\d[\d\s']*(?:[.,-]\d{2})", fragment)
    ]
    normalized_candidates = [value for value in candidates if value is not None and value > 0]
    if not normalized_candidates:
        return None
    return max(normalized_candidates)


def extract_amount_candidates_from_fragment(fragment: str | None) -> list[float]:
    if not fragment:
        return []
    values = [
        parse_amount(match.group(0))
        for match in re.finditer(r"\d[\d\s']*(?:[.,-]\d{2})", fragment)
    ]
    return [value for value in values if value is not None and value > 0]


def is_axb_invoice_header_line(line: str) -> bool:
    normalized_line = normalize_line(line).lower().rstrip(":")
    compact_line = normalized_line.replace(" ", "")
    return normalized_line in {
        "артикул",
        "кол-во",
        "ед",
        "ед.",
        "цена",
        "скидка",
        "итого",
        "rub",
        "rub:",
        "в валюте",
        "в валюте:",
        "в т.ч. ндс",
        "в т ч ндс",
    } or "rub" in compact_line


def is_axb_invoice_stop_line(line: str) -> bool:
    normalized_line = normalize_line(line).lower()
    return normalized_line.startswith("всего наименований") or normalized_line in {"руководитель", "бухгалтер"}


def is_axb_invoice_total_marker(line: str) -> bool:
    compact_line = re.sub(r"[^A-Za-zА-Яа-яЁё]", "", normalize_line(line)).lower()
    return compact_line in {"всего", "всег"} or bool(re.fullmatch(r"[bв][cс][eе][rг](?:[oо])?", compact_line))


def extract_axb_invoice_fragment(text: str) -> str:
    invoice_fragment = extract_fragment_after_marker(text, r"Счет\s+на\s+оплату", max_chars=7000)
    if invoice_fragment:
        return invoice_fragment

    lines = [normalize_line(line) for line in text.splitlines()]
    for index, line in enumerate(lines):
        lowered_line = line.lower()
        if not lowered_line or "товар" in lowered_line or "артикул" in lowered_line:
            continue
        if "на основании" not in lowered_line and "покупатель" not in lowered_line:
            continue

        search_window = "\n".join(lines[index : index + 48]).lower()
        if "товар" not in search_window or "артикул" not in search_window:
            continue
        if "заказ-наряд" not in search_window and "автомобиль" not in search_window:
            continue

        start_index = max(0, index - 6)
        while start_index > 0:
            previous_line = lines[start_index - 1].lower()
            if not previous_line:
                break
            if any(marker in previous_line for marker in ("поставщик", "получатель", "банк получателя", "образец заполнения")):
                start_index -= 1
                continue
            if previous_line.startswith("сч.") or previous_line.startswith("cч.") or re.fullmatch(r"\d{6,}", previous_line):
                start_index -= 1
                continue
            break

        return "\n".join(filter(None, lines[start_index : min(len(lines), start_index + 160)])).strip()

    return ""


def parse_axb_quantity_candidate(line: str) -> Optional[tuple[float, Optional[str]]]:
    normalized_line = normalize_line(line).replace("|", "").strip()
    match = re.fullmatch(r"(?P<qty>\d+(?:[.,]\d+)?)\s*(?P<unit>[A-Za-zА-Яа-я4Ff-]{0,4})", normalized_line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    if quantity is None or quantity <= 0 or quantity > 20:
        return None

    raw_unit = (match.group("unit") or "").strip()
    if not raw_unit and abs(quantity - round(quantity)) <= 0.001:
        return None
    if not raw_unit and quantity > 0.5:
        return None
    if not raw_unit and re.search(r"[.,]00$", normalized_line):
        return None

    return quantity, raw_unit or None


def normalize_axb_explicit_unit(raw_unit: Optional[str]) -> Optional[str]:
    if not raw_unit:
        return None
    translated_unit = raw_unit.lower()
    if translated_unit in {"4", "f"}:
        return "ч"
    if translated_unit == "-":
        return None
    return normalize_unit_name(raw_unit)


ETS_ACT_WORK_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<code>[A-Za-zА-Яа-я0-9.-]+)\s+"
    r"(?P<name>.+?)"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<norm>\d+(?:[.,]\d+)?)\s+"
    r"(?P<net>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<total>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)
ETS_ACT_PART_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<article>[A-Za-zА-Яа-я0-9.-]+)\s+"
    r"(?P<name>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<unit>[A-Za-zА-Яа-я./-]+)\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<net>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<total>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)
SIBTRAKSCAN_ROW_START_PATTERN = re.compile(r"^\d+\s+[A-Za-zА-Яа-я0-9/-]{2,}(?:\s|$)", re.IGNORECASE)
SIBTRAKSCAN_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<code>[A-Za-zА-Яа-я0-9/-]+)\s+"
    r"(?P<name>.+?)\s+"
    r"(?P<unit>н/ч|шт|л|кг|м|компл|ед)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<net>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<total>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)
LEADER_TRAK_ROW_START_PATTERN = re.compile(r"^\d+\s+\S+", re.IGNORECASE)
LEADER_TRAK_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<unit>н/ч|шт|г|гр|мл|литр|л|кг|м)\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<discount>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<net>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<gross>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)
LEADER_TRAK_INVOICE_ROW_PATTERN = re.compile(
    r"(?P<row>\d{1,3})\s*"
    r"(?P<body>.+?)\s*"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<unit>н/ч|шт|г|гр|мл|литр|л|кг|м)\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2})?)\s+"
    r"(?P<discount>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<total>\d[\d\s]*(?:[.,]\d{2}))"
    r"(?=(?:\s*\d{1,3}\s*[A-Za-zА-Яа-яЁё№(]|Итого\s+RUB:|$))",
    re.IGNORECASE | re.DOTALL,
)
LOGISTICS_ROW_START_PATTERN = re.compile(r"^\d+\s+\S+", re.IGNORECASE)
LOGISTICS_WORK_TAIL_PATTERN = re.compile(
    r"^\d+(?:[.,]\d+)?\s+\d[\d\s]*(?:[.,]\d{2})\s+\d+(?:[.,]\d+)?\s+\d[\d\s]*(?:[.,]\d{0,2})?\s+\d+%\s+\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})$",
    re.IGNORECASE,
)
LOGISTICS_PART_TAIL_PATTERN = re.compile(
    r"^\d+(?:[.,]\d+)?\s+[A-Za-zА-Яа-я./-]{1,8}\s+\d[\d\s]*(?:[.,]\d{2})\s+\d+%\s+\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})$",
    re.IGNORECASE,
)
LOGISTICS_WORK_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<norm>\d+(?:[.,]\d+)?)\s+"
    r"(?P<rate>\d[\d\s]*(?:[.,]\d{0,2})?)\s+"
    r"\d+%\s+"
    r"(?P<gross>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)
LOGISTICS_PART_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<unit>[A-Za-zА-Яа-я./-]{1,8})\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"\d+%\s+"
    r"(?P<gross>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)
KLEVER_TRAK_ROW_START_PATTERN = re.compile(r"^\d+\s+\S+", re.IGNORECASE)
KLEVER_TRAK_WORK_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{1,2})?)\s+"
    r"(?P<norm>\d+(?:[.,]\d+)?)\s+"
    r"(?P<unit>н/?ч)\s+"
    r"(?P<gross>\d[\d\s]*(?:[.,]\d{1,2})?)\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{1,2})?)$",
    re.IGNORECASE,
)
KLEVER_TRAK_PART_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<unit>[A-Za-zА-Яа-я./-]{1,8})\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{1,2})?)\s+"
    r"(?P<gross>\d[\d\s]*(?:[.,]\d{1,2})?)\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{1,2})?)$",
    re.IGNORECASE,
)
KLEVER_TRAK_VEHICLE_ROW_PATTERN = re.compile(
    r"(?:^|\n)\s*\d{2}\.\d{2}\.\d{4}\s+.+?\s+(?P<mileage>\d[\d\s]{2,})\s+"
    r"(?:нет|да)\s+(?P<vin>[A-HJ-NPR-Z0-9]{17})\s+(?P<plate>[А-ЯA-Z0-9 ]{6,16})(?=\s*(?:\n|$))",
    re.IGNORECASE,
)
ANTARES_ROW_START_PATTERN = re.compile(r"^\d+\s+\S+", re.IGNORECASE)
ANTARES_PART_TAIL_PATTERN = re.compile(
    r"^\d+(?:[.,]\d+)?\s+(?:шт\.?|кг|л|м|мл|г|гр|ед\.?|компл\.?|к-т)\b",
    re.IGNORECASE,
)
ANTARES_WORK_TAIL_PATTERN = re.compile(
    r"^\d+(?:[.,]\d+)?\s+\d[\d\s]*(?:[.,]\d{2})\b",
    re.IGNORECASE,
)
ANTARES_AMOUNT_CONTINUATION_PATTERN = re.compile(
    r"^\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})$",
    re.IGNORECASE,
)
ANTARES_PART_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<unit>[A-Za-zА-Яа-я./-]{1,8})\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?:(?P<rate>\d+%)\s+)?"
    r"(?P<discount>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<net>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)
ANTARES_WORK_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<norm>\d+(?:[.,]\d+)?)\s+"
    r"(?P<middle>.+?)\s+"
    r"(?:(?P<rate>\d+%)\s+)?"
    r"(?P<net>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<vat>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)
GRUZOVYE_REZERVY_ROW_START_PATTERN = re.compile(r"^\d+\s+\S+", re.IGNORECASE)
GRUZOVYE_REZERVY_ROW_TAIL_PATTERN = re.compile(
    r"^\d+(?:[.,]\d+)?\s+[A-Za-zА-Яа-я./-]{1,8}\s+\d[\d\s]*(?:[.,]\d{2})\s+\d[\d\s]*(?:[.,]\d{2})$",
    re.IGNORECASE,
)
GRUZOVYE_REZERVY_ROW_PATTERN = re.compile(
    r"^\d+\s+"
    r"(?P<body>.+?)\s+"
    r"(?P<qty>\d+(?:[.,]\d+)?)\s+"
    r"(?P<unit>[A-Za-zА-Яа-я./-]{1,8})\s+"
    r"(?P<price>\d[\d\s]*(?:[.,]\d{2}))\s+"
    r"(?P<net>\d[\d\s]*(?:[.,]\d{2}))$",
    re.IGNORECASE,
)


def extract_ets_act_section_rows(
    text: str,
    *,
    marker_pattern: str,
    stop_patterns: tuple[str, ...],
    max_chars: int = 8000,
) -> list[str]:
    fragment = extract_fragment_after_marker(
        text,
        marker_pattern,
        stop_patterns=stop_patterns,
        max_chars=max_chars,
    )
    if not fragment:
        return []

    section_text = normalize_text(fragment.replace("\n", " ").replace("\r", " "))
    section_text = re.sub(r".*?1 2 3 4 5 6 7 8 9\s*", "", section_text, count=1)
    section_text = re.sub(
        r"(?<=\d[.,]\d{2})\s+(?=\d{1,2}\s+[A-Za-zА-Яа-я0-9.-]{1,}\s)",
        "\n",
        section_text,
    )
    section_text = re.sub(
        r"(?<=\d[.,]\d{2})(?=\d{1,2}\s+[A-Za-zА-Яа-я0-9.-]{1,}\s)",
        "\n",
        section_text,
    )
    return [normalize_line(line) for line in section_text.splitlines() if normalize_line(line)]


def parse_ets_act_work_row(line: str) -> Optional[dict[str, object]]:
    match = ETS_ACT_WORK_ROW_PATTERN.match(line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    price = parse_amount(match.group("price"))
    standard_hours = parse_decimal_value(match.group("norm"))
    net_total = parse_amount(match.group("net"))
    if quantity is None or price is None or standard_hours is None or net_total is None:
        return None

    return {
        "work_code": normalize_article_value(match.group("code")),
        "work_name": normalize_line(match.group("name"))[:500],
        "quantity": quantity,
        "unit_name": None,
        "price": price,
        "line_total": net_total,
        "standard_hours": standard_hours,
    }


def parse_ets_act_part_row(line: str) -> Optional[dict[str, object]]:
    match = ETS_ACT_PART_ROW_PATTERN.match(line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    price = parse_amount(match.group("price"))
    net_total = parse_amount(match.group("net"))
    if quantity is None or price is None or net_total is None:
        return None

    return {
        "article": normalize_article_value(match.group("article")),
        "part_name": normalize_line(match.group("name"))[:500],
        "quantity": quantity,
        "unit_name": normalize_unit_name(match.group("unit")),
        "price": price,
        "line_total": net_total,
    }


def extract_ets_act_items(text: str) -> dict[str, list[dict[str, object]]]:
    work_rows = extract_ets_act_section_rows(
        text,
        marker_pattern=r"Выполненные\s+работы\s+по\s+акту\s+выполненных\s+работ",
        stop_patterns=(r"Итого\s+работ:",),
        max_chars=5000,
    )
    part_rows = extract_ets_act_section_rows(
        text,
        marker_pattern=r"Расходная\s+накладная\s+к\s+акту\s+выполненных\s+работ",
        stop_patterns=(r"Итого\s+материал",),
        max_chars=9000,
    )

    works = [payload for payload in (parse_ets_act_work_row(line) for line in work_rows) if payload]
    parts = [payload for payload in (parse_ets_act_part_row(line) for line in part_rows) if payload]
    if not works:
        works = extract_ets_act_sparse_scanned_work_items(text)
    return {"works": works, "parts": parts}


def is_sibtrakscan_noise_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    lower_line = normalized_line.lower()
    if not normalized_line:
        return True
    if lower_line.startswith("стр."):
        return True
    if lower_line.startswith("страница "):
        return True
    if lower_line.startswith("передан через диадок"):
        return True
    if re.fullmatch(r"[0-9a-f-]{16,}", lower_line):
        return True
    return False


def extract_sibtrakscan_row_buffers(text: str) -> list[str]:
    fragment = extract_fragment_after_marker(
        text,
        r"ЗАДАНИЕ\s*:",
        stop_patterns=(r"Итого\s+нормочасов", r"Всего\s+по\s+работам", r"Итого\s+по\s+заказ[- ]наряду", r"К\s+оплате"),
        max_chars=20000,
    )
    if not fragment:
        return []

    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    buffers: list[str] = []
    current_buffer: list[str] = []

    for line in lines:
        lower_line = line.lower()
        if is_sibtrakscan_noise_line(line):
            continue
        if lower_line.startswith("задание:") or lower_line.startswith("итого по заданию"):
            if current_buffer:
                buffers.append(normalize_line(" ".join(current_buffer)))
                current_buffer = []
            continue
        if SIBTRAKSCAN_ROW_START_PATTERN.match(line):
            if current_buffer:
                buffers.append(normalize_line(" ".join(current_buffer)))
            current_buffer = [line]
            continue
        if current_buffer:
            current_buffer.append(line)

    if current_buffer:
        buffers.append(normalize_line(" ".join(current_buffer)))
    return buffers


def parse_sibtrakscan_row(line: str) -> Optional[tuple[str, dict[str, object]]]:
    match = SIBTRAKSCAN_ROW_PATTERN.match(line)
    if match is None:
        return None

    unit_name = normalize_unit_name(match.group("unit"))
    quantity = parse_decimal_value(match.group("qty"))
    price = parse_amount(match.group("price"))
    net_total = parse_amount(match.group("net"))
    if unit_name is None or quantity is None or price is None or net_total is None:
        return None

    code_value = normalize_article_value(match.group("code"))
    name_value = normalize_line(match.group("name"))[:500]
    if unit_name == "нч":
        return (
            "works",
            {
                "work_code": code_value,
                "work_name": name_value,
                "quantity": quantity,
                "unit_name": unit_name,
                "price": price,
                "line_total": net_total,
                "standard_hours": quantity,
            },
        )

    return (
        "parts",
        {
            "article": code_value,
            "part_name": name_value,
            "quantity": quantity,
            "unit_name": unit_name,
            "price": price,
            "line_total": net_total,
        },
    )


def extract_sibtrakscan_items(text: str) -> dict[str, list[dict[str, object]]]:
    works: list[dict[str, object]] = []
    parts: list[dict[str, object]] = []
    for buffer in extract_sibtrakscan_row_buffers(text):
        parsed_row = parse_sibtrakscan_row(buffer)
        if parsed_row is None:
            continue
        item_kind, payload = parsed_row
        if item_kind == "works":
            works.append(payload)
        else:
            parts.append(payload)
    return {"works": works, "parts": parts}


def is_leader_trak_noise_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    lower_line = normalized_line.lower()
    if not normalized_line:
        return True
    if lower_line.startswith("страница "):
        return True
    if lower_line.startswith("стр. "):
        return True
    if lower_line.startswith("передан через диадок"):
        return True
    if lower_line.startswith("сервисные услуги сдал"):
        return True
    if lower_line.startswith("сервисные услуги принял"):
        return True
    if lower_line.startswith("после подписания"):
        return True
    if re.fullmatch(r"[0-9a-f-]{16,}", lower_line):
        return True
    if normalized_line in {"№ Номер", "операции или", "запчасти", "Наименование работ,", "запчастей и материалов"}:
        return True
    if lower_line in {
        "кол-во ед.",
        "измер.",
        "цена за",
        "единицу",
        "сумма",
        "скидки",
        "стоимость",
        "без налога",
        "сумма налога",
        "стоимость с налогом *",
    }:
        return True
    return False


def is_leader_trak_tail_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    return bool(
        re.match(
            r"^\d+(?:[.,]\d+)?\s+(?:н/ч|шт|г|гр|мл|литр|л|кг|м)\b",
            normalized_line,
            re.IGNORECASE,
        )
    )


def split_leader_trak_body(value: str) -> tuple[Optional[str], str]:
    normalized_value = normalize_line(value)
    normalized_value = re.sub(r"(?<=[A-Za-zА-Я0-9/-])(?=[А-Я][а-я])", " ", normalized_value, count=1)
    normalized_value = re.sub(r"(?<=[A-Za-zА-Я0-9/-])(?=[A-Z][a-z])", " ", normalized_value, count=1)
    parts = normalized_value.split(maxsplit=1)
    if len(parts) == 2 and (ARTICLE_TOKEN_PATTERN.fullmatch(parts[0]) or WORK_CODE_TOKEN_PATTERN.fullmatch(parts[0])):
        return normalize_article_value(parts[0]), parts[1]
    return None, normalized_value


def extract_leader_trak_row_buffers(text: str) -> list[str]:
    fragment = extract_fragment_after_marker(
        text,
        r"Выполненные\s+сервисные\s+услуги\s+и\s+использованные\s+материалы",
        stop_patterns=(r"Всего\s+по\s+наряд[- ]заказу", r"После\s+подписания", r"Рекомендации:"),
        max_chars=25000,
    )
    if not fragment:
        return []

    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    buffers: list[str] = []
    current_buffer: list[str] = []

    for line in lines:
        lower_line = line.lower()
        if is_leader_trak_noise_line(line):
            continue
        if lower_line.startswith("всего по странице:"):
            if current_buffer:
                buffers.append(normalize_line(" ".join(current_buffer)))
                current_buffer = []
            continue
        if lower_line.startswith("всего по наряд-заказу"):
            if current_buffer:
                buffers.append(normalize_line(" ".join(current_buffer)))
            break
        if current_buffer and is_leader_trak_tail_line(line):
            current_buffer.append(line)
            continue
        if LEADER_TRAK_ROW_START_PATTERN.match(line):
            if current_buffer:
                buffers.append(normalize_line(" ".join(current_buffer)))
            current_buffer = [line]
            continue
        if current_buffer:
            current_buffer.append(line)

    if current_buffer:
        buffers.append(normalize_line(" ".join(current_buffer)))
    return buffers


def extract_leader_trak_invoice_fragment(text: str) -> str:
    fragment = extract_fragment_after_marker(
        text,
        r"№\s*Товар\s+Артикул\s+Кол-во\s+Ед\.\s+Цена\s+Скидка\s+НДС\s+Всего",
        stop_patterns=(r"Итого\s+RUB:", r"Всего\s+наименований", r"Сумма\s+прописью"),
        max_chars=25000,
    )
    if fragment:
        return fragment
    return ""


def extract_leader_trak_order_number(text: str) -> Optional[str]:
    for pattern in (
        r"План\s+ремонта\s*№\s*([A-Za-zА-Яа-я0-9/_-]{3,})",
        r"Предварительный\s+счет\s+на\s+оплату\s*№\s*([A-Za-zА-Яа-я0-9/_-]{3,})",
    ):
        match = re.search(pattern, text, re.IGNORECASE)
        if match is None:
            continue
        order_number = normalize_line(match.group(1))
        if is_plausible_order_number(order_number):
            return order_number
    return None


def parse_leader_trak_row(line: str) -> Optional[tuple[str, dict[str, object]]]:
    match = LEADER_TRAK_ROW_PATTERN.match(line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    unit_name = normalize_unit_name(match.group("unit"))
    price = parse_amount(match.group("price"))
    net_total = parse_amount(match.group("net"))
    if quantity is None or unit_name is None or price is None or net_total is None:
        return None

    code_value, name_value = split_leader_trak_body(match.group("body"))
    if not name_value:
        return None

    if unit_name == "нч":
        return (
            "works",
            {
                "work_code": normalize_article_value(code_value),
                "work_name": name_value[:500],
                "quantity": quantity,
                "unit_name": unit_name,
                "price": price,
                "line_total": net_total,
                "standard_hours": quantity,
            },
        )

    return (
        "parts",
        {
            "article": normalize_article_value(code_value),
            "part_name": name_value[:500],
            "quantity": quantity,
            "unit_name": unit_name,
            "price": price,
            "line_total": net_total,
        },
    )


def parse_leader_trak_invoice_row(fragment: str) -> Optional[tuple[str, dict[str, object]]]:
    match = LEADER_TRAK_INVOICE_ROW_PATTERN.match(fragment)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    unit_name = normalize_unit_name(match.group("unit"))
    price = parse_amount(match.group("price"))
    vat_total = parse_amount(match.group("vat"))
    gross_total = parse_amount(match.group("total"))
    if quantity is None or unit_name is None or price is None or vat_total is None or gross_total is None:
        return None

    net_total = round(gross_total - vat_total, 2)
    if net_total <= 0:
        return None

    code_value, name_value = split_leader_trak_body(match.group("body"))
    if not name_value:
        return None

    if unit_name == "нч":
        return (
            "works",
            {
                "work_code": normalize_article_value(code_value),
                "work_name": name_value[:500],
                "quantity": quantity,
                "unit_name": unit_name,
                "price": price,
                "line_total": net_total,
                "standard_hours": quantity,
            },
        )

    return (
        "parts",
        {
            "article": normalize_article_value(code_value),
            "part_name": name_value[:500],
            "quantity": quantity,
            "unit_name": unit_name,
            "price": price,
            "line_total": net_total,
        },
    )


def extract_leader_trak_invoice_items(text: str) -> dict[str, list[dict[str, object]]]:
    fragment = normalize_line(extract_leader_trak_invoice_fragment(text))
    if not fragment:
        return {"works": [], "parts": []}

    works: list[dict[str, object]] = []
    parts: list[dict[str, object]] = []
    for match in LEADER_TRAK_INVOICE_ROW_PATTERN.finditer(fragment):
        parsed_row = parse_leader_trak_invoice_row(match.group(0))
        if parsed_row is None:
            continue
        item_kind, payload = parsed_row
        if item_kind == "works":
            works.append(payload)
        else:
            parts.append(payload)

    if len(works) + len(parts) < 2:
        return {"works": [], "parts": []}
    return {"works": works, "parts": parts}


def extract_leader_trak_items(text: str) -> dict[str, list[dict[str, object]]]:
    works: list[dict[str, object]] = []
    parts: list[dict[str, object]] = []
    for buffer in extract_leader_trak_row_buffers(text):
        parsed_row = parse_leader_trak_row(buffer)
        if parsed_row is None:
            continue
        item_kind, payload = parsed_row
        if item_kind == "works":
            works.append(payload)
        else:
            parts.append(payload)
    service_items = {"works": works, "parts": parts}
    invoice_items = extract_leader_trak_invoice_items(text)
    if is_leader_trak_invoice_only_document(text) and (invoice_items["works"] or invoice_items["parts"]):
        return invoice_items
    return service_items


def is_logistics_noise_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    lower_line = normalized_line.lower()
    if not normalized_line:
        return True
    if lower_line.startswith("выполненные работы по заказ-наряду"):
        return True
    if lower_line.startswith("расходная накладная к заказ-наряду"):
        return True
    if lower_line.startswith("итого работ:"):
        return True
    if lower_line.startswith("итого материалов:"):
        return True
    if lower_line.startswith("итого по странице материалов:"):
        return True
    if lower_line.startswith("итого по причине обращения:"):
        return True
    if lower_line.startswith("итого по заказ-наряду"):
        return True
    if lower_line.startswith("всего по причине обращения:"):
        return True
    if lower_line.startswith("всего по заказ-наряду:"):
        return True
    if lower_line.startswith("рекомендации:"):
        return True
    if normalized_line in {
        "№ Артикул Наименование Кол. оп. Цена н/ч Норма н/ч Ставка НДС Всего в т.ч. НДС",
        "№ Артикул Наименование Кол-во Ед.изм. Цена Ставка НДС Всего в т.ч. НДС",
        "1 2 3 4 5 6 7 8 9",
        "1 2 3 4 5 6 7 8 9 10",
    }:
        return True
    if re.fullmatch(r"[0-9a-f-]{16,}", lower_line):
        return True
    if re.fullmatch(r"страница:\s*\d+", lower_line):
        return True
    return False


def split_logistics_part_body(value: str) -> tuple[Optional[str], str]:
    normalized_value = normalize_line(value)
    tokens = normalized_value.split()
    if len(tokens) < 2:
        return None, normalized_value

    article_tokens: list[str] = []
    for token in tokens:
        if re.search(r"[А-Яа-я]", token):
            break
        article_tokens.append(token)

    if not article_tokens:
        return None, normalized_value

    part_name = normalize_line(" ".join(tokens[len(article_tokens) :]))
    article = normalize_article_value("".join(article_tokens))
    if not part_name:
        return None, normalized_value
    return article, part_name


def extract_logistics_row_buffers(
    text: str,
    *,
    marker_pattern: str,
    stop_patterns: tuple[str, ...],
    tail_pattern: re.Pattern[str],
    max_chars: int = 30000,
) -> list[str]:
    fragment = extract_fragment_after_marker(
        text,
        marker_pattern,
        stop_patterns=stop_patterns,
        max_chars=max_chars,
    )
    if not fragment:
        return []

    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    buffers: list[str] = []
    current_buffer: list[str] = []

    for line in lines:
        if is_logistics_noise_line(line):
            continue
        if current_buffer and tail_pattern.match(line):
            current_buffer.append(line)
            buffers.append(normalize_line(" ".join(current_buffer)))
            current_buffer = []
            continue
        if LOGISTICS_ROW_START_PATTERN.match(line):
            if current_buffer:
                buffers.append(normalize_line(" ".join(current_buffer)))
            current_buffer = [line]
            continue
        if current_buffer:
            current_buffer.append(line)

    if current_buffer:
        buffers.append(normalize_line(" ".join(current_buffer)))
    return buffers


def parse_logistics_work_row(line: str) -> Optional[dict[str, object]]:
    match = LOGISTICS_WORK_ROW_PATTERN.match(line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    price = parse_amount(match.group("price"))
    standard_hours = parse_decimal_value(match.group("norm"))
    line_total = parse_amount(match.group("gross"))
    work_name = normalize_line(match.group("body"))
    if quantity is None or price is None or standard_hours is None or line_total is None or not work_name:
        return None

    return {
        "work_code": None,
        "work_name": work_name[:500],
        "quantity": quantity,
        "unit_name": None,
        "price": price,
        "line_total": line_total,
        "standard_hours": standard_hours,
    }


def parse_logistics_part_row(line: str) -> Optional[dict[str, object]]:
    match = LOGISTICS_PART_ROW_PATTERN.match(line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    price = parse_amount(match.group("price"))
    line_total = parse_amount(match.group("gross"))
    if quantity is None or price is None or line_total is None:
        return None

    article, part_name = split_logistics_part_body(match.group("body"))
    if not part_name:
        return None

    return {
        "article": article,
        "part_name": part_name[:500],
        "quantity": quantity,
        "unit_name": normalize_unit_name(match.group("unit")),
        "price": price,
        "line_total": line_total,
    }


def extract_logistics_items(text: str) -> dict[str, list[dict[str, object]]]:
    work_rows = extract_logistics_row_buffers(
        text,
        marker_pattern=r"Выполненные\s+работы\s+по\s+заказ[- ]наряду",
        stop_patterns=(r"Итого\s+работ:",),
        tail_pattern=LOGISTICS_WORK_TAIL_PATTERN,
    )
    part_rows = extract_logistics_row_buffers(
        text,
        marker_pattern=r"Расходная\s+накладная\s+к\s+заказ[- ]наряду",
        stop_patterns=(r"Итого\s+материал", r"Итого\s+по\s+причине\s+обращения"),
        tail_pattern=LOGISTICS_PART_TAIL_PATTERN,
        max_chars=40000,
    )

    works = [payload for payload in (parse_logistics_work_row(line) for line in work_rows) if payload]
    parts = [payload for payload in (parse_logistics_part_row(line) for line in part_rows) if payload]
    return {"works": works, "parts": parts}


def is_klever_trak_noise_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    lower_line = normalized_line.lower()
    if not normalized_line:
        return True
    if lower_line.startswith("выполненные работы по заказ-наряду"):
        return True
    if lower_line.startswith("расходная накладная к заказ-наряду"):
        return True
    if lower_line.startswith("итого работ:"):
        return True
    if lower_line.startswith("итого материалов:"):
        return True
    if lower_line.startswith("итого по заказ-наряду"):
        return True
    if lower_line.startswith("всего по заказ-наряду:"):
        return True
    if lower_line.startswith("рекомендации:"):
        return True
    if lower_line.startswith("страница:"):
        return True
    if normalized_line in {
        "№ Артикул Наименование Кол. оп. Цена н/ч Норма н/ч Всего в т.ч. НДС",
        "№ Артикул Наименование Кол-во Ед.изм. Цена Всего в т.ч. НДС",
        "1 2 3 4 5 6 7 8 9",
        "1 2 3 4 5 6 7 8",
    }:
        return True
    return False


def split_klever_trak_work_body(value: str) -> tuple[Optional[str], str]:
    normalized_value = normalize_line(value)
    tokens = normalized_value.split()
    if len(tokens) >= 3 and ARTICLE_TOKEN_PATTERN.fullmatch(tokens[0]) and re.fullmatch(r"\d{4,}", tokens[1]):
        article = normalize_article_value(tokens[0] + tokens[1])
        name = normalize_line(" ".join(tokens[2:]))
        if name:
            return article, name
    return split_work_code_and_name(normalized_value)


def split_klever_trak_part_body(value: str) -> tuple[Optional[str], str]:
    normalized_value = normalize_line(value)
    tokens = normalized_value.split()
    if len(tokens) >= 3 and ARTICLE_TOKEN_PATTERN.fullmatch(tokens[0]) and re.fullmatch(r"\d{4,}", tokens[1]):
        article = normalize_article_value(tokens[0] + tokens[1])
        name = normalize_line(" ".join(tokens[2:]))
        if name:
            return article, name
    if len(tokens) >= 2 and ARTICLE_TOKEN_PATTERN.fullmatch(tokens[0]) and any(char.isdigit() for char in tokens[0]):
        article = normalize_article_value(tokens[0])
        name = normalize_line(" ".join(tokens[1:]))
        if name:
            return article, name
    return None, normalized_value


def extract_klever_trak_row_buffers(
    text: str,
    *,
    marker_pattern: str,
    stop_patterns: tuple[str, ...],
    max_chars: int = 30000,
) -> list[str]:
    fragment = extract_fragment_after_marker(
        text,
        marker_pattern,
        stop_patterns=stop_patterns,
        max_chars=max_chars,
    )
    if not fragment:
        return []

    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    return [
        line
        for line in lines
        if KLEVER_TRAK_ROW_START_PATTERN.match(line) and not is_klever_trak_noise_line(line)
    ]


def parse_klever_trak_work_row(line: str) -> Optional[dict[str, object]]:
    normalized_line = normalize_line(line)
    tokens = normalized_line.split()
    if len(tokens) < 8 or not tokens[0].isdigit():
        return None

    quantity = parse_decimal_value(tokens[-6])
    price = parse_amount(tokens[-5])
    standard_hours = parse_decimal_value(tokens[-4])
    unit_name = normalize_unit_name(tokens[-3])
    line_total = parse_amount(tokens[-2])
    vat_total = parse_amount(tokens[-1])
    if quantity is None or price is None or standard_hours is None or line_total is None or vat_total is None:
        return None

    work_code, work_name = split_klever_trak_work_body(" ".join(tokens[1:-6]))
    if not work_name:
        return None

    return {
        "work_code": normalize_article_value(work_code),
        "work_name": work_name[:500],
        "quantity": quantity,
        "unit_name": unit_name,
        "price": price,
        "line_total": line_total,
        "standard_hours": standard_hours,
    }


def parse_klever_trak_part_row(line: str) -> Optional[dict[str, object]]:
    normalized_line = normalize_line(line)
    tokens = normalized_line.split()
    if len(tokens) < 7 or not tokens[0].isdigit():
        return None

    quantity = parse_decimal_value(tokens[-5])
    unit_name = normalize_unit_name(tokens[-4])
    price = parse_amount(tokens[-3])
    line_total = parse_amount(tokens[-2])
    vat_total = parse_amount(tokens[-1])
    if quantity is None or unit_name is None or price is None or line_total is None or vat_total is None:
        return None

    article, part_name = split_klever_trak_part_body(" ".join(tokens[1:-5]))
    if not part_name:
        return None

    return {
        "article": article,
        "part_name": part_name[:500],
        "quantity": quantity,
        "unit_name": unit_name,
        "price": price,
        "line_total": line_total,
    }


def extract_klever_trak_items(text: str) -> dict[str, list[dict[str, object]]]:
    work_rows = extract_klever_trak_row_buffers(
        text,
        marker_pattern=r"Выполненные\s+работы\s+по\s+заказ[- ]наряду",
        stop_patterns=(r"Итого\s+работ:",),
    )
    part_rows = extract_klever_trak_row_buffers(
        text,
        marker_pattern=r"Расходная\s+накладная\s+к\s+заказ[- ]наряду",
        stop_patterns=(r"Итого\s+материал", r"Итого\s+по\s+заказ[- ]наряду"),
        max_chars=40000,
    )
    works = [payload for payload in (parse_klever_trak_work_row(line) for line in work_rows) if payload]
    parts = [payload for payload in (parse_klever_trak_part_row(line) for line in part_rows) if payload]
    return {"works": works, "parts": parts}


def is_antares_noise_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    lower_line = normalized_line.lower()
    if not normalized_line:
        return True
    if lower_line.startswith("передан через диадок"):
        return True
    if lower_line.startswith("страница "):
        return True
    if lower_line.startswith("заказ-наряд №"):
        return True
    if lower_line.startswith("выполненные работы по заказ-наряду"):
        return True
    if lower_line.startswith("расходная накладная к заказ-наряду"):
        return True
    if lower_line.startswith("№ артикул"):
        return True
    if lower_line in {"1 2 3 4 5 6 7 8 9", "1 2 3 4 5 6 7 8 9 10"}:
        return True
    if lower_line.startswith("принят:"):
        return True
    if lower_line.startswith("вид ремонта:"):
        return True
    if lower_line.startswith("диспетчер:"):
        return True
    if lower_line.startswith("мастер:"):
        return True
    if lower_line.startswith("срок исполнения:"):
        return True
    if lower_line == "закрыт":
        return True
    if re.fullmatch(r"[0-9a-f-]{16,}", lower_line):
        return True
    return False


def repair_antares_numeric_splits(value: str) -> str:
    normalized_value = normalize_line(value)
    previous_value = None
    while previous_value != normalized_value:
        previous_value = normalized_value
        normalized_value = re.sub(
            r"(\d[\d\s]*,\d)\s+(\d)(?=\s+\d[\d\s]*(?:[.,]\d{2})\b)",
            r"\1\2",
            normalized_value,
        )
    return normalized_value


def extract_antares_row_buffers(
    text: str,
    *,
    marker_pattern: str,
    stop_patterns: tuple[str, ...],
    tail_pattern: re.Pattern[str],
    final_total_prefix: str,
    max_chars: int = 25000,
) -> list[str]:
    fragment = extract_fragment_after_marker(
        text,
        marker_pattern,
        stop_patterns=stop_patterns,
        max_chars=max_chars,
    )
    if not fragment:
        return []

    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    buffers: list[str] = []
    current_buffer: list[str] = []

    for line in lines:
        lower_line = line.lower()
        if is_antares_noise_line(line):
            continue
        if lower_line.startswith(final_total_prefix):
            if current_buffer:
                buffers.append(repair_antares_numeric_splits(" ".join(current_buffer)))
                current_buffer = []
            break
        if lower_line.startswith("итого по странице"):
            if current_buffer:
                buffers.append(repair_antares_numeric_splits(" ".join(current_buffer)))
                current_buffer = []
            continue
        if current_buffer and (tail_pattern.match(line) or ANTARES_AMOUNT_CONTINUATION_PATTERN.match(line)):
            current_buffer.append(line)
            continue
        if ANTARES_ROW_START_PATTERN.match(line):
            if current_buffer:
                buffers.append(repair_antares_numeric_splits(" ".join(current_buffer)))
            current_buffer = [line]
            continue
        if current_buffer:
            current_buffer.append(line)

    if current_buffer:
        buffers.append(repair_antares_numeric_splits(" ".join(current_buffer)))
    return buffers


def split_antares_part_body(value: str) -> tuple[Optional[str], str]:
    normalized_value = normalize_line(value)
    tokens = normalized_value.split()
    if not tokens:
        return None, ""

    first_token = tokens[0]
    if not (ARTICLE_TOKEN_PATTERN.fullmatch(first_token) or WORK_CODE_TOKEN_PATTERN.fullmatch(first_token)):
        return None, normalized_value

    article_tokens = [first_token]
    index = 1
    while index < len(tokens):
        token = tokens[index]
        if token.startswith("("):
            break
        if token.isdigit() and len(token) <= 3:
            article_tokens.append(token)
            index += 1
            continue
        if re.fullmatch(r"[A-Z0-9_/-]{1,4}", token):
            article_tokens.append(token)
            index += 1
            continue
        break

    article = normalize_article_value("".join(article_tokens))
    part_name = normalize_line(" ".join(tokens[index:]))
    return article, part_name


def parse_antares_work_row(line: str) -> Optional[dict[str, object]]:
    match = ANTARES_WORK_ROW_PATTERN.match(line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    price = parse_amount(match.group("price"))
    standard_hours = parse_decimal_value(match.group("norm"))
    net_total = parse_amount(match.group("net"))
    if quantity is None or price is None or standard_hours is None or net_total is None:
        return None

    work_code, work_name = split_work_code_and_name(match.group("body"))
    if not work_name:
        return None

    return {
        "work_code": normalize_article_value(work_code),
        "work_name": normalize_line(work_name)[:500],
        "quantity": quantity,
        "unit_name": None,
        "price": price,
        "line_total": net_total,
        "standard_hours": standard_hours,
    }


def parse_antares_part_row(line: str) -> Optional[dict[str, object]]:
    match = ANTARES_PART_ROW_PATTERN.match(line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    unit_name = normalize_unit_name(match.group("unit"))
    price = parse_amount(match.group("price"))
    net_total = parse_amount(match.group("net"))
    if quantity is None or unit_name is None or price is None or net_total is None:
        return None

    article, part_name = split_antares_part_body(match.group("body"))
    if not part_name:
        return None

    return {
        "article": article,
        "part_name": part_name[:500],
        "quantity": quantity,
        "unit_name": unit_name,
        "price": price,
        "line_total": net_total,
    }


def extract_antares_items(text: str) -> dict[str, list[dict[str, object]]]:
    work_rows = extract_antares_row_buffers(
        text,
        marker_pattern=r"Выполненные\s+работы\s+по\s+заказ[- ]наряду",
        stop_patterns=(r"Итого\s+работ:",),
        tail_pattern=ANTARES_WORK_TAIL_PATTERN,
        final_total_prefix="итого работ:",
    )
    part_rows = extract_antares_row_buffers(
        text,
        marker_pattern=r"Расходная\s+накладная\s+к\s+заказ[- ]наряду",
        stop_patterns=(r"Итого\s+материал",),
        tail_pattern=ANTARES_PART_TAIL_PATTERN,
        final_total_prefix="итого материалов:",
    )

    works = [payload for payload in (parse_antares_work_row(line) for line in work_rows) if payload]
    parts = [payload for payload in (parse_antares_part_row(line) for line in part_rows) if payload]
    return {"works": works, "parts": parts}


def is_gruzovye_rezervy_noise_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    lower_line = normalized_line.lower()
    if not normalized_line:
        return True
    if lower_line.startswith("передан через диадок"):
        return True
    if lower_line.startswith("страница "):
        return True
    if lower_line.startswith("стр. "):
        return True
    if lower_line in {"виды работ:", "материалы:"}:
        return True
    if normalized_line == "№ Наименование работ, услуг Кол-во Ед. Цена Сумма":
        return True
    if re.fullmatch(r"[0-9a-f-]{16,}", lower_line):
        return True
    return False


def is_gruzovye_rezervy_total_line(line: str) -> bool:
    return bool(re.fullmatch(r"\d[\d\s]*(?:[.,]\d{2})\s*руб\.?", normalize_line(line), re.IGNORECASE))


def extract_gruzovye_rezervy_row_buffers(
    text: str,
    *,
    marker_pattern: str,
    stop_patterns: tuple[str, ...],
    max_chars: int = 25000,
) -> list[str]:
    fragment = extract_fragment_after_marker(
        text,
        marker_pattern,
        stop_patterns=stop_patterns,
        max_chars=max_chars,
    )
    if not fragment:
        return []

    lines = [normalize_line(line) for line in fragment.splitlines() if normalize_line(line)]
    buffers: list[str] = []
    current_buffer: list[str] = []

    for line in lines:
        if is_gruzovye_rezervy_noise_line(line):
            continue
        if is_gruzovye_rezervy_total_line(line):
            if current_buffer:
                buffers.append(normalize_line(" ".join(current_buffer)))
                current_buffer = []
            continue
        if current_buffer and GRUZOVYE_REZERVY_ROW_TAIL_PATTERN.match(line):
            current_buffer.append(line)
            continue
        if GRUZOVYE_REZERVY_ROW_START_PATTERN.match(line):
            if current_buffer:
                buffers.append(normalize_line(" ".join(current_buffer)))
            current_buffer = [line]
            continue
        if current_buffer:
            current_buffer.append(line)

    if current_buffer:
        buffers.append(normalize_line(" ".join(current_buffer)))
    return buffers


def split_gruzovye_rezervy_part_body(value: str) -> tuple[Optional[str], str]:
    normalized_value = normalize_line(value)
    parts = normalized_value.split(maxsplit=1)
    if len(parts) == 2 and ARTICLE_TOKEN_PATTERN.fullmatch(parts[0]) and any(char.isdigit() for char in parts[0]):
        return normalize_article_value(parts[0]), parts[1]
    return None, normalized_value


def parse_gruzovye_rezervy_row(line: str, *, item_kind: str) -> Optional[dict[str, object]]:
    match = GRUZOVYE_REZERVY_ROW_PATTERN.match(line)
    if match is None:
        return None

    quantity = parse_decimal_value(match.group("qty"))
    unit_name = normalize_unit_name(match.group("unit"))
    price = parse_amount(match.group("price"))
    net_total = parse_amount(match.group("net"))
    if quantity is None or unit_name is None or price is None or net_total is None:
        return None

    body_value = normalize_line(match.group("body"))
    if item_kind == "works":
        return {
            "work_code": None,
            "work_name": body_value[:500],
            "quantity": quantity,
            "unit_name": unit_name,
            "price": price,
            "line_total": net_total,
        }

    article, part_name = split_gruzovye_rezervy_part_body(body_value)
    if not part_name:
        return None
    return {
        "article": article,
        "part_name": part_name[:500],
        "quantity": quantity,
        "unit_name": unit_name,
        "price": price,
        "line_total": net_total,
    }


def extract_gruzovye_rezervy_items(text: str) -> dict[str, list[dict[str, object]]]:
    work_rows = extract_gruzovye_rezervy_row_buffers(
        text,
        marker_pattern=r"Виды\s+Работ:",
        stop_patterns=(r"Материалы:",),
    )
    part_rows = extract_gruzovye_rezervy_row_buffers(
        text,
        marker_pattern=r"Материалы:",
        stop_patterns=(r"Итого:",),
    )

    works = [payload for payload in (parse_gruzovye_rezervy_row(line, item_kind="works") for line in work_rows) if payload]
    parts = [payload for payload in (parse_gruzovye_rezervy_row(line, item_kind="parts") for line in part_rows) if payload]
    return {"works": works, "parts": parts}


def is_axb_article_candidate(line: str) -> bool:
    normalized_line = normalize_line(line)
    compact_line = normalized_line.replace(" ", "")
    if not compact_line:
        return False
    if is_axb_invoice_header_line(normalized_line) or is_axb_invoice_stop_line(normalized_line):
        return False
    if parse_axb_quantity_candidate(normalized_line) is not None:
        return False
    if compact_line.isdigit() and len(compact_line) >= 3:
        return True
    if compact_line.isdigit() and len(compact_line) <= 2:
        return True
    if WORK_CODE_TOKEN_PATTERN.fullmatch(compact_line) or ARTICLE_TOKEN_PATTERN.fullmatch(compact_line) or AXB_ARTICLE_TOKEN_PATTERN.fullmatch(compact_line):
        return True
    if parse_amount(normalized_line) is not None:
        return False
    if normalized_line.isdigit() and len(normalized_line) == 1:
        return False
    return False


def is_axb_material_article_candidate(line: str) -> bool:
    normalized_line = normalize_line(line)
    compact_line = normalized_line.replace(" ", "").rstrip("*|")
    if not compact_line:
        return False
    if " " in normalized_line and not compact_line.isdigit():
        return False
    if compact_line.isdigit() and len(compact_line) <= 2:
        return False
    if WORK_CODE_TOKEN_PATTERN.fullmatch(compact_line) or ARTICLE_TOKEN_PATTERN.fullmatch(compact_line) or AXB_ARTICLE_TOKEN_PATTERN.fullmatch(compact_line):
        return True
    return compact_line.isdigit() and len(compact_line) >= 3


def extract_axb_invoice_totals(lines: list[str]) -> tuple[list[float], int]:
    total_marker_index = -1
    for index, line in enumerate(lines):
        if is_axb_invoice_total_marker(line):
            total_marker_index = index

    if total_marker_index < 0:
        return [], 0

    totals = [
        amount
        for amount in (parse_amount(line) for line in lines[total_marker_index + 1 :])
        if amount is not None and amount > 0
    ]
    if not totals:
        return [], 0

    expected_count = len(totals)
    if len(totals) >= 2 and totals[-1] > max(totals[:-1]):
        expected_count -= 1
    return totals[:expected_count], expected_count


def axb_name_merge_score(left: str, right: str) -> int:
    left_normalized = normalize_line(left)
    right_normalized = normalize_line(right)
    score = 0

    if re.match(r"^[a-zа-я(]", right_normalized):
        score += 8
    if re.match(r"^\d", right_normalized):
        score += 7
    if len(right_normalized) <= 12:
        score += 3
    if re.match(r"^[A-ZА-Я0-9/-]{2,}$", right_normalized):
        score += 3
    if re.search(r"\b(?:с|в|на|под|для|по|от|до|над|при)\s*$", left_normalized.lower()):
        score += 5
    if left_normalized.endswith(("-", "/", ",")):
        score += 4
    if len(left_normalized) <= 24:
        score += 2
    if left_normalized.endswith((".", "!", "?", ")", '"')):
        score -= 6
    return score


def collapse_axb_name_lines(name_lines: list[str], expected_count: int) -> list[str]:
    groups = [normalize_line(line) for line in name_lines if normalize_line(line)]
    if expected_count <= 0 or not groups:
        return groups

    while len(groups) > expected_count:
        best_index = 0
        best_score = None
        for index in range(len(groups) - 1):
            score = axb_name_merge_score(groups[index], groups[index + 1])
            if best_score is None or score > best_score:
                best_index = index
                best_score = score
        groups[best_index] = normalize_line(f"{groups[best_index]} {groups[best_index + 1]}")
        del groups[best_index + 1]

    return groups


def infer_axb_item_kind(name: str, *, unit_name: Optional[str], quantity: Optional[float]) -> str:
    normalized_name = normalize_line(name).lower()
    if any(keyword in normalized_name for keyword in AXB_WORK_NAME_KEYWORDS):
        return "works"
    if any(keyword in normalized_name for keyword in AXB_PART_NAME_KEYWORDS):
        return "parts"
    if unit_name == "шт":
        return "parts"
    if quantity is not None and abs(quantity - round(quantity)) > 0.001:
        return "works"
    return "parts"


def infer_axb_price(line_total: float, quantity: float) -> Optional[float]:
    if quantity <= 0:
        return None
    price = round(line_total / quantity / 0.95, 2)
    return price if price > 0 else None


def infer_axb_quantity(line_total: float, price: float) -> Optional[float]:
    if price <= 0:
        return None
    quantity = round(line_total / price / 0.95, 2)
    return quantity if quantity > 0 else None


def score_axb_price_sequence(
    prices: list[float],
    quantities: list[float],
    total_slice: list[float],
) -> float:
    if len(prices) != len(quantities) or len(prices) != len(total_slice):
        return float("-inf")

    score = 0.0
    for price, quantity, line_total in zip(prices, quantities, total_slice):
        if price <= 0 or quantity <= 0:
            return float("-inf")
        if price * quantity < line_total * 0.98:
            return float("-inf")
        ratio = line_total / (price * quantity)
        if not 0.75 <= ratio <= 1.05:
            score -= abs(ratio - 0.95) * 40
        else:
            score += 8 - abs(ratio - 0.95) * 40
        if price >= 100:
            score += 1.5
        else:
            score -= 4
    return score


def select_axb_batch_prices(
    quantities: list[float],
    amount_batch: list[float],
    total_slice: list[float],
) -> list[float]:
    batch_size = len(quantities)
    if batch_size == 0 or not amount_batch:
        return []

    candidate_sequences: list[list[float]] = []
    if len(amount_batch) >= batch_size:
        candidate_sequences.append(amount_batch[:batch_size])
    if len(amount_batch) >= batch_size * 3:
        candidate_sequences.append([amount_batch[index * 3] for index in range(batch_size)])

    best_sequence: list[float] = []
    best_score = float("-inf")
    for sequence in candidate_sequences:
        score = score_axb_price_sequence(sequence, quantities, total_slice[: len(sequence)])
        if score > best_score:
            best_score = score
            best_sequence = sequence

    return best_sequence


def split_axb_leading_item_code(line: str) -> tuple[Optional[str], str]:
    normalized_line = normalize_line(line)
    if not normalized_line:
        return None, ""

    match = re.fullmatch(
        r"(?:(?P<row>\d{1,2})\s+)?(?P<code>[A-Za-zА-Яа-я0-9/._*-]{3,})(?:\s+(?P<name>.+))?",
        normalized_line,
    )
    if match is None:
        return None, normalized_line

    code = (match.group("code") or "").rstrip("*|")
    inline_name = normalize_line(match.group("name") or "")
    compact_code = code.replace(" ", "")
    is_numeric_article = compact_code.isdigit() and len(compact_code) >= 5
    if not compact_code or (parse_amount(compact_code) is not None and not is_numeric_article) or not any(char.isdigit() for char in compact_code):
        return None, normalized_line
    if inline_name.startswith(("(", "[", "{")):
        return None, normalized_line

    alpha_suffix = re.sub(r"[^A-Za-zА-Яа-я]+", "", compact_code).lower()
    if compact_code[0].isdigit() and alpha_suffix in {"мл", "ml", "л", "l", "кг", "kg", "г", "гр"}:
        return None, normalized_line
    if not (
        WORK_CODE_TOKEN_PATTERN.fullmatch(compact_code)
        or ARTICLE_TOKEN_PATTERN.fullmatch(compact_code)
        or AXB_ARTICLE_TOKEN_PATTERN.fullmatch(compact_code)
    ):
        return None, normalized_line

    return normalize_article_value(code) or normalize_text(code), inline_name


def split_axb_leading_work_code(line: str) -> tuple[Optional[str], str]:
    normalized_line = normalize_line(line)
    if not normalized_line:
        return None, ""

    match = re.fullmatch(
        r"(?P<code>[A-Za-zА-Яа-я0-9/_-]{3,})(?:\s+(?P<name>.+))?",
        normalized_line,
    )
    if match is None:
        return None, normalized_line

    code = match.group("code") or ""
    compact_code = code.replace(" ", "")
    if not compact_code or not any(char.isdigit() for char in compact_code):
        return None, normalized_line
    if re.fullmatch(r"(?:то|to)[-_]?\d+", compact_code, re.IGNORECASE):
        return None, normalized_line
    if not (WORK_CODE_TOKEN_PATTERN.fullmatch(compact_code) or ARTICLE_TOKEN_PATTERN.fullmatch(compact_code)):
        return None, normalized_line

    return normalize_article_value(code) or normalize_text(code), normalize_line(match.group("name") or "")


def is_axb_material_name_header_line(line: str) -> bool:
    normalized_line = normalize_line(line).lower()
    return (
        "наимен" in normalized_line
        and ("артик" in normalized_line or "товар" in normalized_line)
        and "кол" in normalized_line
    )


def is_axb_compact_work_row_start(line: str) -> bool:
    normalized_line = normalize_line(line)
    if not normalized_line or is_axb_invoice_total_marker(normalized_line):
        return False
    if "ремонт" not in normalized_line.lower():
        return False
    return len(extract_amount_candidates_from_fragment(normalized_line)) >= 3


def is_axb_compact_work_continuation_line(line: str) -> bool:
    normalized_line = normalize_line(line)
    if not normalized_line or is_axb_compact_work_row_start(normalized_line):
        return False
    if is_axb_invoice_total_marker(normalized_line):
        return False
    lowered_line = normalized_line.lower()
    if lowered_line.startswith(("выполненные работы", "итого работ", "расходная накладная")):
        return False
    if parse_amount(normalized_line) is not None:
        return False
    if parse_axb_quantity_candidate(normalized_line) is not None:
        return False
    return bool(re.search(r"[A-Za-zА-Яа-я]", normalized_line))


def extract_axb_compact_work_row(line: str) -> Optional[dict[str, object]]:
    normalized_line = normalize_line(re.sub(r"[\[\]{}_]+", " ", line))
    if not is_axb_compact_work_row_start(normalized_line):
        return None

    amount_pattern = r"(?:\d{1,3}(?:\s\d{3})+|\d{1,6})(?:[.,-]\d{2})"
    amount_matches = [
        (match.start(), match.end(), value)
        for match in re.finditer(amount_pattern, normalized_line)
        if (value := parse_amount(match.group(0))) is not None and value > 0
    ]
    if len(amount_matches) < 3:
        return None

    line_total = float(amount_matches[-2][2])
    prefix_end = amount_matches[-3][0]
    prefix_text = normalize_line(normalized_line[:prefix_end].rstrip(" -:;,|"))
    if not prefix_text:
        return None

    prefix_amounts = [
        (match.start(), match.end(), value)
        for match in re.finditer(amount_pattern, prefix_text)
        if (value := parse_amount(match.group(0))) is not None and value > 0
    ]
    price_payload = next((payload for payload in reversed(prefix_amounts) if payload[2] >= 100 and payload[2] < 10000), None)
    standard_hours = next((float(payload[2]) for payload in reversed(prefix_amounts) if payload[2] < 20), None)

    name_prefix = prefix_text[: price_payload[0]] if price_payload is not None else prefix_text
    name_prefix = re.sub(r"\b\d+(?:[.,]\d+)?[!;)}|]*\s*$", "", name_prefix).strip(" -:;,|[]()")
    name_prefix = normalize_line(name_prefix)
    if not name_prefix:
        return None

    row_stripped_prefix = re.sub(r"^\d{1,2}\s+", "", name_prefix)
    work_code, inline_name = split_axb_leading_work_code(row_stripped_prefix)
    work_name = inline_name if work_code else row_stripped_prefix
    work_name = re.sub(r"^(?:\d+\s+)+", "", work_name)
    work_name = re.sub(r"^[^A-Za-zА-Яа-я0-9]+", "", work_name)
    work_name = normalize_line(work_name).strip(" -:;,|[]()")
    if not work_name or not re.search(r"[A-Za-zА-Яа-я]", work_name):
        return None

    price = float(price_payload[2]) if price_payload is not None else None
    if standard_hours is not None and price is None:
        price = infer_axb_price(line_total, standard_hours)
    if standard_hours is None and price is not None:
        standard_hours = infer_axb_quantity(line_total, price)

    return {
        "work_code": work_code,
        "name_lines": [work_name],
        "line_total": line_total,
        "price": float(price) if price is not None else line_total,
        "standard_hours": float(standard_hours) if standard_hours is not None else None,
    }


def extract_axb_compact_work_items(text: str) -> list[dict[str, object]]:
    work_body_match = re.search(
        r"Выполненные\s+работы\s+(?:по|no|No|ho|но)?\s*заказ[- ]наряду(?P<body>.*?)(?:Итого\s+работ:)",
        text,
        re.IGNORECASE | re.MULTILINE | re.DOTALL,
    )
    if work_body_match is None:
        return []

    work_body_lines = [normalize_line(line) for line in work_body_match.group("body").splitlines() if normalize_line(line)]
    compact_rows: list[dict[str, object]] = []
    current_row: Optional[dict[str, object]] = None

    for line in work_body_lines:
        if is_axb_compact_work_row_start(line):
            row_payload = extract_axb_compact_work_row(line)
            if row_payload is None:
                continue
            if current_row is not None and current_row.get("name_lines"):
                compact_rows.append(current_row)
            current_row = row_payload
            continue
        if current_row is not None and is_axb_compact_work_continuation_line(line):
            current_row.setdefault("name_lines", []).append(line)

    if current_row is not None and current_row.get("name_lines"):
        compact_rows.append(current_row)

    works: list[dict[str, object]] = []
    for row_payload in compact_rows:
        work_name = clean_axb_work_name(" ".join(str(line) for line in row_payload.get("name_lines", [])))[:500]
        if not work_name:
            continue
        standard_hours = row_payload.get("standard_hours")
        works.append(
            {
                "work_code": row_payload.get("work_code"),
                "work_name": work_name,
                "quantity": float(standard_hours) if isinstance(standard_hours, (int, float)) else 1.0,
                "standard_hours": float(standard_hours) if isinstance(standard_hours, (int, float)) else None,
                "unit_name": "ч" if isinstance(standard_hours, (int, float)) else "усл",
                "price": float(row_payload.get("price") or row_payload.get("line_total") or 0.0),
                "line_total": float(row_payload["line_total"]),
            }
        )

    return works


def extract_axb_work_line_totals_from_summary(
    summary_lines: list[str],
    *,
    expected_work_total: Optional[float] = None,
) -> list[float]:
    candidate_amounts: list[float] = []
    for line in summary_lines:
        normalized_line = normalize_line(line)
        lowered_line = normalized_line.lower()
        if not normalized_line:
            continue
        if any(marker in lowered_line for marker in ("расходная накладная", "итого материалов", "итого по заказ")):
            break
        if re.fullmatch(r"\d+", normalized_line):
            continue
        candidate_amounts.extend(extract_amount_candidates_from_fragment(normalized_line))

    if not candidate_amounts:
        return []

    vat_counter: dict[int, int] = {}
    for amount in candidate_amounts:
        amount_key = int(round(amount * 100))
        vat_counter[amount_key] = vat_counter.get(amount_key, 0) + 1

    line_totals: list[float] = []
    for amount in candidate_amounts:
        if expected_work_total is not None and amounts_match(amount, expected_work_total, tolerance=0.01):
            continue
        if amount <= 100:
            continue
        vat_candidate = round(amount / 6, 2)
        vat_key = int(round(vat_candidate * 100))
        if vat_counter.get(vat_key, 0) <= 0:
            continue
        vat_counter[vat_key] -= 1
        line_totals.append(float(amount))

    return line_totals


def extract_axb_expected_work_line_totals_from_summary(
    summary_lines: list[str],
    *,
    expected_work_total: float,
    name_line_hint_count: int,
) -> list[float]:
    normalized_lines: list[str] = []
    candidate_amounts: list[float] = []
    for line in summary_lines:
        normalized_line = normalize_line(line)
        lowered_line = normalized_line.lower()
        if not normalized_line:
            continue
        if any(marker in lowered_line for marker in ("расходная накладная", "итого материалов", "итого по заказ")):
            break
        normalized_lines.append(normalized_line)
        if re.fullmatch(r"\d+", normalized_line):
            continue
        candidate_amounts.extend(extract_amount_candidates_from_fragment(normalized_line))

    if not candidate_amounts:
        return []

    gross_section_amounts: list[float] = []
    gross_section_started = False
    for line in normalized_lines:
        lowered_line = line.lower()
        if "всего" in lowered_line:
            gross_section_started = True
            continue
        if gross_section_started and "ндс" in lowered_line:
            break
        if not gross_section_started:
            continue
        if re.fullmatch(r"\d+", line):
            continue
        gross_section_amounts.extend(extract_amount_candidates_from_fragment(line))

    filtered_candidates = [
        float(amount)
        for amount in (gross_section_amounts or candidate_amounts)
        if amount > 100 and amount <= float(expected_work_total) + 3.0
    ]
    if not filtered_candidates:
        return []

    target_cents = int(round(float(expected_work_total) * 100))
    tolerance_cents = 300
    subset_candidates = [
        amount
        for amount in filtered_candidates
        if not amounts_match(amount, float(expected_work_total), tolerance=0.2)
    ]
    subset_states: dict[int, list[int]] = {0: []}
    for index, amount in enumerate(subset_candidates):
        amount_cents = int(round(amount * 100))
        if amount_cents <= 0 or amount_cents > target_cents + tolerance_cents:
            continue
        for current_sum, current_indexes in list(subset_states.items()):
            new_sum = current_sum + amount_cents
            if new_sum > target_cents + tolerance_cents:
                continue
            new_indexes = current_indexes + [index]
            existing_indexes = subset_states.get(new_sum)
            if existing_indexes is None or len(new_indexes) > len(existing_indexes):
                subset_states[new_sum] = new_indexes

    best_sum = None
    best_indexes: list[int] | None = None
    for reachable_sum, indexes in subset_states.items():
        if not indexes:
            continue
        if abs(reachable_sum - target_cents) > tolerance_cents:
            continue
        if best_sum is None:
            best_sum = reachable_sum
            best_indexes = indexes
            continue
        current_delta = abs(reachable_sum - target_cents)
        best_delta = abs(best_sum - target_cents)
        if current_delta < best_delta or (current_delta == best_delta and len(indexes) > len(best_indexes or [])):
            best_sum = reachable_sum
            best_indexes = indexes
    if best_indexes:
        return [subset_candidates[index] for index in best_indexes]

    if name_line_hint_count <= 4:
        single_total = next(
            (amount for amount in filtered_candidates if amounts_match(amount, float(expected_work_total), tolerance=0.2)),
            None,
        )
        if single_total is not None:
            return [float(single_total)]

    return []


def clean_axb_work_name(name: str) -> str:
    cleaned = normalize_line(name)
    cleaned = re.sub(r"^скидка\s+", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(
        r"[З3з]\s*\d[\d\s]*(?:[.,]\d{2})\s+\d+(?:[.,]\d+)?\s*Рем\w*",
        " ",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\b[З3з]\s*\d[\d\s]*(?:[.,]\d{2})\b", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b\d+(?:[.,]\d+)?\s*Рем\w*\b", " ", cleaned, flags=re.IGNORECASE)
    return normalize_line(cleaned).strip(" -:;,|[]")


def score_axb_material_entries(entries: list[dict[str, Optional[str]]]) -> int:
    score = 0
    for entry in entries:
        part_name = normalize_line(str(entry.get("part_name") or ""))
        if not part_name:
            score -= 40
            continue
        keyword_hits, cyrillic_count, alnum_count = score_text_quality(part_name)
        score += keyword_hits * 18 + min(cyrillic_count, 80) + min(alnum_count, 80)
        if entry.get("article"):
            score += 24
        if re.search(r"(?:итого|всего|насумму|на\s+сумму)", part_name, re.IGNORECASE):
            score -= 140
        score -= part_name.count("|") * 4
        if len(part_name) > 240:
            score -= 50
    score += len(entries) * 12
    return score


def has_polluted_axb_material_entry_names(entries: list[dict[str, Optional[str]]]) -> bool:
    for entry in entries:
        part_name = normalize_line(str(entry.get("part_name") or ""))
        if not part_name:
            continue
        if len(extract_amount_candidates_from_fragment(part_name)) >= 2:
            return True
    return False


def clean_axb_inline_material_name(name: str) -> str:
    cleaned = normalize_line(re.sub(r"[\[\]|_]+", " ", name)).strip(" -:;,")
    cleaned = re.sub(r"\b(?:шт|шт\.|метр|мер|едиизм|едизм|tm)\b.*$", "", cleaned, flags=re.IGNORECASE).strip(" -:;,")
    cleaned = re.sub(r"(?:\s+\d+(?:[.,]\d+)?)+\s*$", "", cleaned).strip(" -:;,")
    return normalize_line(cleaned)


def extract_axb_inline_material_entries(section_text: str) -> list[dict[str, object]]:
    lines = [normalize_line(line) for line in section_text.splitlines() if normalize_line(line)]
    entries: list[dict[str, object]] = []

    for line in lines:
        lowered_line = line.lower()
        if lowered_line.startswith("расходная накладная"):
            continue
        if re.search(r"итого\s+материал|итого\s+по\s+странице|всего\s+по\s+причине|итого\s+по\s+заказ[- ]наряду", lowered_line):
            continue
        if is_axb_material_name_header_line(line):
            continue

        amount_values = extract_amount_candidates_from_fragment(line)
        if len(amount_values) < 2:
            continue

        article_match = next(
            (
                match
                for match in re.finditer(r"[A-Za-zА-Яа-я0-9/._-]{2,}", line)
                if is_axb_material_article_candidate(match.group(0).strip("._-"))
            ),
            None,
        )
        if article_match is None:
            continue

        tail = line[article_match.end():]
        amount_match = re.search(r"\d[\d\s']*(?:[.,-]\d{2})", tail)
        if amount_match is not None:
            tail = tail[: amount_match.start()]
        part_name = clean_axb_inline_material_name(tail)
        if not part_name or not re.search(r"[A-Za-zА-Яа-я]", part_name):
            continue

        article_value = normalize_article_value(article_match.group(0).strip("._-")) or normalize_text(article_match.group(0))
        entries.append(
            {
                "article": article_value,
                "part_name": part_name[:500],
                "line_total": float(amount_values[-2]),
            }
        )

    return entries


def extract_axb_work_items(text: str, *, expected_work_total: Optional[float] = None) -> list[dict[str, object]]:
    work_body_match = re.search(
        r"Выполненные\s+работы\s+(?:по|no|No|ho|но)?\s*заказ[- ]наряду(?P<body>.*?)(?:Итого\s+работ:)",
        text,
        re.IGNORECASE | re.MULTILINE | re.DOTALL,
    )
    if work_body_match is None:
        return extract_axb_compact_work_items(text)

    work_body_lines = [normalize_line(line) for line in work_body_match.group("body").splitlines() if normalize_line(line)]
    filtered_name_lines: list[tuple[int, str]] = []
    for index, line in enumerate(work_body_lines):
        lowered_line = line.lower()
        if "выполненные работы по заказ-наряду" in lowered_line or lowered_line.startswith("обращения "):
            continue
        if "к причине" in lowered_line:
            continue
        if any(marker in lowered_line for marker in ("кол. оп.", "цена н/ч", "норма н/ч")):
            continue
        if any(marker in lowered_line for marker in ("скидк", "всего", "ндс")):
            continue
        if lowered_line in {"n", "№", "артикул", "наименование", "кол. оп.", "цена н/ч", "норма", "н/ч", "скидка", "ремонт"}:
            continue
        if lowered_line.startswith(("nº ", "№ ", "цена н/ч")):
            continue
        if re.fullmatch(r"\d+", line):
            continue
        if parse_axb_quantity_candidate(line) is not None:
            continue
        if re.fullmatch(r"\d+(?:[.,]\d+)?\s*ремонт", lowered_line):
            continue
        if parse_axb_standard_hours_candidate(line) is not None:
            continue
        if parse_amount(line) is not None:
            continue
        if extract_axb_rate_candidate(line) is not None:
            continue
        compact_row = re.sub(r"^\d{1,2}\s+", "", line)
        _inline_code, inline_name = split_axb_leading_work_code(compact_row)
        normalized_candidate_line = normalize_line(inline_name or compact_row)
        if not normalized_candidate_line:
            continue
        if " " not in normalized_candidate_line and is_axb_article_candidate(normalized_candidate_line):
            continue
        filtered_name_lines.append((index, normalized_candidate_line))
    work_totals_fragment = extract_fragment_after_marker(
        text,
        r"Итого\s+работ:",
        stop_patterns=(r"Расходная\s+накладная", r"Итого\s+материал", r"Итого\s+по\s+заказ[- ]наряду"),
        max_chars=1400,
    )
    if not work_totals_fragment:
        return extract_axb_compact_work_items(text)

    totals_lines = [normalize_line(line) for line in work_totals_fragment.splitlines() if normalize_line(line)]
    explicit_expected_work_total = expected_work_total
    total_marker_index = -1
    for index, line in enumerate(totals_lines):
        if is_axb_invoice_total_marker(line):
            total_marker_index = index
    if total_marker_index < 0 and explicit_expected_work_total is not None:
        total_marker_index = 0
    if total_marker_index < 0:
        return extract_axb_compact_work_items(text)

    expected_work_total = expected_work_total or extract_largest_amount_from_fragment(work_totals_fragment)
    if explicit_expected_work_total is not None:
        line_totals = extract_axb_expected_work_line_totals_from_summary(
            totals_lines[total_marker_index + 1 :],
            expected_work_total=float(expected_work_total),
            name_line_hint_count=len(filtered_name_lines),
        )
    else:
        line_totals = extract_axb_work_line_totals_from_summary(
            totals_lines[total_marker_index + 1 :],
            expected_work_total=expected_work_total,
        )
    if not line_totals:
        if explicit_expected_work_total is not None and 0 < len(filtered_name_lines) <= 4:
            aggregated_work_name = clean_axb_work_name(" ".join(line for _index, line in filtered_name_lines))[:500]
            if aggregated_work_name:
                return [
                    {
                        "work_code": None,
                        "work_name": aggregated_work_name,
                        "quantity": 1.0,
                        "standard_hours": None,
                        "unit_name": "усл",
                        "price": float(explicit_expected_work_total),
                        "line_total": float(explicit_expected_work_total),
                    }
                ]
        return extract_axb_compact_work_items(text)

    grouped_name_lines = [[line] for _index, line in filtered_name_lines]
    grouped_positions = [[index] for index, _line in filtered_name_lines]
    while len(grouped_name_lines) > len(line_totals):
        best_index = 0
        best_score = None
        for index in range(len(grouped_name_lines) - 1):
            score = axb_name_merge_score(grouped_name_lines[index][-1], grouped_name_lines[index + 1][0])
            if best_score is None or score > best_score:
                best_index = index
                best_score = score
        grouped_name_lines[best_index].extend(grouped_name_lines[best_index + 1])
        grouped_positions[best_index].extend(grouped_positions[best_index + 1])
        del grouped_name_lines[best_index + 1]
        del grouped_positions[best_index + 1]

    if len(grouped_name_lines) != len(line_totals):
        return extract_axb_compact_work_items(text)

    works: list[dict[str, object]] = []
    previous_group_end = -1
    for group_index, (name_group, position_group, line_total) in enumerate(zip(grouped_name_lines, grouped_positions, line_totals)):
        work_name = clean_axb_work_name(" ".join(name_group))[:500]
        if not work_name:
            previous_group_end = position_group[-1]
            continue

        current_group_start = position_group[0]
        current_group_end = position_group[-1]
        next_group_start = grouped_positions[group_index + 1][0] if group_index + 1 < len(grouped_positions) else len(work_body_lines)

        work_code: Optional[str] = None
        for line in work_body_lines[previous_group_end + 1 : current_group_start]:
            if " " in line:
                continue
            if re.fullmatch(r"\d{1,2}", line):
                continue
            if not re.fullmatch(r"[A-Za-zА-Яа-я0-9/_-]{3,}", line):
                continue
            if is_axb_article_candidate(line):
                normalized_code = normalize_article_value(line) or normalize_text(line)
                if normalized_code and any(char.isdigit() for char in normalized_code):
                    work_code = normalized_code

        standard_hours: Optional[float] = None
        price: Optional[float] = None
        forward_lines = work_body_lines[current_group_start + 1 : next_group_start]
        for index, line in enumerate(forward_lines):
            candidate_price = extract_axb_rate_candidate(line)
            if candidate_price is not None and price is None:
                price = candidate_price

            candidate_hours = parse_axb_standard_hours_candidate(line)
            if candidate_hours is None:
                continue

            tail_window = " ".join(forward_lines[index : index + 3]).lower()
            if "ремонт" in tail_window or price is not None:
                standard_hours = candidate_hours

        if standard_hours is not None and price is None:
            price = infer_axb_price(line_total, standard_hours)
        if standard_hours is None and price is not None:
            standard_hours = infer_axb_quantity(line_total, price)

        works.append(
            {
                "work_code": work_code,
                "work_name": work_name,
                "quantity": float(standard_hours) if standard_hours is not None else 1.0,
                "standard_hours": float(standard_hours) if standard_hours is not None else None,
                "unit_name": "ч" if standard_hours is not None else "усл",
                "price": float(price) if price is not None else float(line_total),
                "line_total": float(line_total),
            }
        )
        previous_group_end = current_group_end

    compact_works = extract_axb_compact_work_items(text)
    if compact_works and (
        len(compact_works) > len(works)
        or sum(item["line_total"] for item in compact_works) > sum(item["line_total"] for item in works) + 3.0
    ):
        return compact_works

    return works


def _extract_axb_material_section_entries_from_lines(
    lines: list[str],
    *,
    name_header_index: int,
) -> list[dict[str, Optional[str]]]:
    try:
        body_lines = lines[name_header_index + 1 :]
    except IndexError:
        return []
    lookback_codes = [
        normalize_article_value(line) or normalize_text(line)
        for line in lines[max(0, name_header_index - 8) : name_header_index]
        if is_axb_material_article_candidate(line)
    ]
    body_end_index = next(
        (
            index
            for index, line in enumerate(body_lines)
            if is_axb_invoice_total_marker(line)
            or re.search(
                r"итого\s+по\s+странице\s+материалов|итого\s+материалов|всего\s+по\s+причине|итого\s+по\s+заказ[- ]наряду",
                line,
                re.IGNORECASE,
            )
        ),
        len(body_lines),
    )

    entries: list[dict[str, object]] = []
    current_entry: Optional[dict[str, object]] = None
    for line in body_lines[:body_end_index]:
        lowered_line = line.lower()
        if lowered_line in {
            "артикул",
            "наименование",
            "кол-во",
            "ед. изм. цена",
            "1 скидка",
            "no",
        }:
            continue
        if any(marker in lowered_line for marker in ("кол-во", "ед. изм", "ед.изм", "едизм", "скидка")):
            continue
        if re.fullmatch(r"\d{1,2}", line):
            continue
        if parse_axb_quantity_candidate(line) is not None:
            continue
        if parse_amount(line) is not None and not is_axb_material_article_candidate(line):
            continue
        if re.match(r"^(?:шт|шт\.|л/дмз|л/дм3|л|кг|г|м)\b", lowered_line):
            continue

        code, inline_name = split_axb_leading_item_code(line)
        if code:
            current_entry = {"article": code, "name_lines": []}
            if inline_name:
                current_entry["name_lines"].append(inline_name)
            entries.append(current_entry)
            continue

        if current_entry is None:
            seeded_code = lookback_codes.pop(0) if lookback_codes else None
            current_entry = {"article": seeded_code, "name_lines": []}
            entries.append(current_entry)

        current_entry.setdefault("name_lines", []).append(line)

    normalized_entries: list[dict[str, Optional[str]]] = []
    for entry in entries:
        part_name = normalize_line(" ".join(entry.get("name_lines", [])))[:500]
        if not part_name:
            continue
        normalized_entries.append(
            {
                "article": normalize_article_value(str(entry.get("article"))) if entry.get("article") else None,
                "part_name": part_name,
            }
        )
    return normalized_entries


def extract_axb_batched_material_section_entries_from_lines(
    lines: list[str],
    *,
    name_header_index: int,
) -> list[dict[str, Optional[str]]]:
    try:
        body_lines = lines[name_header_index + 1 :]
    except IndexError:
        return []

    body_end_index = next(
        (
            index
            for index, line in enumerate(body_lines)
            if is_axb_invoice_total_marker(line)
            or re.search(
                r"итого\s+по\s+странице\s+материалов|итого\s+материалов|всего\s+по\s+причине|итого\s+по\s+заказ[- ]наряду",
                line,
                re.IGNORECASE,
            )
        ),
        len(body_lines),
    )

    codes: list[str] = [
        normalize_article_value(line) or normalize_text(line)
        for line in lines[max(0, name_header_index - 8) : name_header_index]
        if is_axb_material_article_candidate(line)
    ]
    name_lines: list[str] = []
    for line in body_lines[:body_end_index]:
        lowered_line = line.lower()
        if lowered_line in {
            "артикул",
            "наименование",
            "кол-во",
            "ед. изм. цена",
            "1 скидка",
            "no",
        }:
            continue
        if any(marker in lowered_line for marker in ("кол-во", "ед. изм", "ед.изм", "едизм", "скидка")):
            continue

        code, inline_name = split_axb_leading_item_code(line)
        if code:
            codes.append(code)
            if inline_name:
                name_lines.append(inline_name)
            continue

        if re.fullmatch(r"\d{1,2}", line):
            continue
        if parse_axb_quantity_candidate(line) is not None:
            continue
        if parse_amount(line) is not None and not is_axb_material_article_candidate(line):
            continue
        if re.match(r"^(?:шт|шт\.|л/дмз|л/дм3|л|кг|г|м)\b", lowered_line):
            continue
        if not re.search(r"[A-Za-zА-Яа-я]", line):
            continue

        name_lines.append(line)

    if not codes or len(name_lines) < len(codes):
        return []

    merged_names = collapse_axb_name_lines(name_lines, len(codes))
    if len(merged_names) != len(codes):
        return []

    entries: list[dict[str, Optional[str]]] = []
    for code, part_name in zip(codes, merged_names):
        normalized_name = normalize_line(part_name)[:500]
        if not normalized_name:
            continue
        entries.append(
            {
                "article": normalize_article_value(code) or normalize_text(code),
                "part_name": normalized_name,
            }
        )
    return entries


def extract_axb_material_section_entries(section_text: str) -> list[dict[str, Optional[str]]]:
    lines = [normalize_line(line) for line in section_text.splitlines() if normalize_line(line)]
    if not lines:
        return []

    header_indexes = [index for index, line in enumerate(lines) if line.lower() == "наименование" or is_axb_material_name_header_line(line)]
    best_entries: list[dict[str, Optional[str]]] = []
    best_score = float("-inf")
    for header_index in header_indexes:
        for entries in (
            _extract_axb_material_section_entries_from_lines(lines, name_header_index=header_index),
            extract_axb_batched_material_section_entries_from_lines(lines, name_header_index=header_index),
        ):
            entry_score = score_axb_material_entries(entries)
            if entries and entry_score > best_score:
                best_entries = entries
                best_score = entry_score
    return best_entries


def extract_axb_material_section_amounts(section_text: str) -> list[float]:
    lines = [normalize_line(line) for line in section_text.splitlines() if normalize_line(line)]
    if not lines:
        return []

    header_anchor_index = -1
    for index, line in enumerate(lines):
        lowered_line = line.lower()
        if is_axb_material_name_header_line(line) or any(marker in lowered_line for marker in ("кол-во", "ед. изм", "ед.изм", "цена", "скидка")):
            header_anchor_index = index
            break

    footer_index = next(
        (
            index
            for index, line in enumerate(lines[header_anchor_index + 1 :], start=header_anchor_index + 1)
            if line.lower().startswith("заказчик подтверждает")
            or line.lower().startswith("заказ-наряд и сч")
            or line.lower().startswith("универсальный передаточный")
        ),
        len(lines),
    )

    total_marker_indexes = [
        index
        for index, line in enumerate(lines[header_anchor_index + 1 : footer_index], start=header_anchor_index + 1)
        if is_axb_invoice_total_marker(line)
    ]
    if not total_marker_indexes:
        return []

    best_amounts: list[float] = []
    best_score: tuple[int, int, float] | None = None
    for total_marker_index in total_marker_indexes:
        amounts: list[float] = []
        trailing_window = lines[total_marker_index + 1 : min(footer_index, total_marker_index + 25)]
        for line in trailing_window:
            amounts.extend(extract_amount_candidates_from_fragment(line))
        while amounts and float(amounts[0]).is_integer() and amounts[0] <= 20:
            amounts = amounts[1:]
        if not amounts:
            continue

        duplicate_count = sum(1 for count in Counter(round(float(amount), 2) for amount in amounts).values() if count >= 2)
        candidate_score = (len(amounts), duplicate_count, max(amounts))
        if best_score is None or candidate_score > best_score:
            best_amounts = amounts
            best_score = candidate_score

    return best_amounts


def choose_axb_material_line_totals(
    raw_amounts: list[float],
    *,
    expected_count: int,
    section_total: Optional[float],
) -> list[float]:
    if expected_count <= 0 or not raw_amounts:
        return []

    if expected_count == 1:
        normalized_amounts = [round(float(amount), 2) for amount in raw_amounts if float(amount) > 0]
        duplicated_amounts = [
            amount
            for amount, count in Counter(normalized_amounts).items()
            if count >= 2 and (section_total is None or amount <= float(section_total) + 0.01)
        ]
        if duplicated_amounts:
            return [max(duplicated_amounts)]

    if len(raw_amounts) == expected_count:
        return raw_amounts
    if len(raw_amounts) == expected_count + 1 and raw_amounts[-1] > max(raw_amounts[:-1]):
        return raw_amounts[:-1]

    if section_total is None:
        return []

    prefix: list[float] = []
    for amount in raw_amounts:
        if amounts_match(amount, section_total, tolerance=0.2):
            break
        prefix.append(amount)

    if len(prefix) >= expected_count:
        direct_totals = prefix[:expected_count]
        if amounts_match(round(sum(direct_totals), 2), section_total, tolerance=3.0):
            return direct_totals

    if len(prefix) < expected_count * 2:
        return []

    candidate_totals = prefix[: expected_count * 2 : 2]
    if len(candidate_totals) != expected_count:
        return []
    if not amounts_match(round(sum(candidate_totals), 2), section_total, tolerance=3.0):
        return []
    return candidate_totals


def select_axb_material_section_total(
    raw_amounts: list[float],
    *,
    expected_count: int,
    expected_parts_total: Optional[float],
) -> Optional[float]:
    if not raw_amounts:
        return None

    candidate_totals: list[float] = []
    seen_candidates: set[float] = set()
    for amount in raw_amounts:
        normalized_amount = round(float(amount), 2)
        if normalized_amount <= 0 or normalized_amount in seen_candidates:
            continue
        seen_candidates.add(normalized_amount)
        candidate_totals.append(normalized_amount)

    def candidate_sort_key(amount: float) -> tuple[int, float]:
        is_expected_total = int(
            expected_parts_total is not None and amounts_match(amount, float(expected_parts_total), tolerance=0.2)
        )
        return (is_expected_total, amount)

    for candidate_total in sorted(candidate_totals, key=candidate_sort_key, reverse=True):
        if choose_axb_material_line_totals(
            raw_amounts,
            expected_count=expected_count,
            section_total=candidate_total,
        ):
            return candidate_total

    if expected_parts_total is not None and any(
        amounts_match(amount, float(expected_parts_total), tolerance=0.2) for amount in raw_amounts
    ):
        return float(expected_parts_total)

    smaller_candidates = [
        amount
        for amount in candidate_totals
        if expected_parts_total is None or amount < float(expected_parts_total) - 0.01
    ]
    return max(smaller_candidates) if smaller_candidates else None


def parse_axb_standard_hours_candidate(line: str) -> Optional[float]:
    normalized_line = normalize_line(line).lower()
    if not normalized_line:
        return None

    match = re.fullmatch(r"(?P<hours>\d+(?:[.,]\d+)?)\s*(?:[a-zа-я]+)?", normalized_line)
    if match is None:
        return None
    if "," not in match.group("hours") and "." not in match.group("hours"):
        return None

    hours = parse_decimal_value(match.group("hours"))
    if hours is None or hours <= 0 or hours > 12:
        return None
    return float(hours)


def extract_axb_rate_candidate(line: str) -> Optional[float]:
    normalized_line = normalize_line(line)
    if not re.search(r"[З3]\s*\d", normalized_line):
        return None
    compact_match = re.fullmatch(r"[З3]\s*(\d[\d\s]*(?:[.,]\d{2}))", normalized_line)
    if compact_match is not None:
        return parse_amount(f"3 {compact_match.group(1)}")
    amounts = extract_amount_candidates_from_fragment(normalized_line)
    if not amounts:
        return None
    return max(amounts)


def extract_axb_material_parts(text: str, *, expected_parts_total: Optional[float] = None) -> list[dict[str, object]]:
    section_pattern = re.compile(
        r"Расходная\s+накладная\s+к\s+заказ[- ]наряду",
        re.IGNORECASE,
    )
    matches = list(section_pattern.finditer(text))
    if not matches:
        return []

    parts: list[dict[str, object]] = []
    for index, match in enumerate(matches):
        section_start = match.start()
        section_end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        section_text = text[section_start:section_end]
        section_entries = extract_axb_material_section_entries(section_text)
        inline_entries = extract_axb_inline_material_entries(section_text)
        inline_totals = [
            float(entry["line_total"])
            for entry in inline_entries
            if isinstance(entry.get("line_total"), (int, float))
        ]
        inline_score = score_axb_material_entries(
            [{"article": entry.get("article"), "part_name": entry.get("part_name")} for entry in inline_entries]
        )
        section_score = score_axb_material_entries(section_entries)
        inline_totals_match_expected = (
            expected_parts_total is not None
            and len(inline_totals) == len(inline_entries)
            and len(inline_entries) >= 2
            and amounts_match(round(sum(inline_totals), 2), float(expected_parts_total), tolerance=3.0)
        )
        use_inline_entries = bool(inline_entries) and (
            inline_score > section_score
            or inline_totals_match_expected
            or (
                len(inline_entries) > len(section_entries)
                and len(inline_totals) == len(inline_entries)
                and has_polluted_axb_material_entry_names(section_entries)
            )
        )
        if use_inline_entries:
            section_entries = [
                {"article": entry.get("article"), "part_name": entry.get("part_name")}
                for entry in inline_entries
            ]
        if not section_entries:
            continue

        raw_amounts = extract_axb_material_section_amounts(section_text)
        section_total = select_axb_material_section_total(
            raw_amounts,
            expected_count=len(section_entries),
            expected_parts_total=expected_parts_total,
        )
        if section_total is None and expected_parts_total is not None:
            section_total = float(expected_parts_total)
        line_totals = choose_axb_material_line_totals(
            raw_amounts,
            expected_count=len(section_entries),
            section_total=section_total,
        )
        if use_inline_entries:
            if len(inline_totals) == len(section_entries):
                line_totals = inline_totals
                if section_total is not None and not amounts_match(round(sum(line_totals), 2), section_total, tolerance=3.0):
                    line_totals = []

        if len(line_totals) == len(section_entries):
            for item_payload, line_total in zip(section_entries, line_totals):
                parts.append(
                    {
                        "article": item_payload["article"],
                        "part_name": str(item_payload["part_name"])[:500],
                        "quantity": 1.0,
                        "unit_name": "шт",
                        "price": float(line_total),
                        "line_total": float(line_total),
                    }
                )
            continue

        if section_total is None:
            continue

        parts.append(
            {
                "article": section_entries[0]["article"] if len(section_entries) == 1 else None,
                "part_name": "; ".join(str(item["part_name"]) for item in section_entries if item.get("part_name"))[:500],
                "quantity": 1.0,
                "unit_name": "шт",
                "price": float(section_total),
                "line_total": float(section_total),
            }
        )

    if parts:
        return parts

    lines = [normalize_line(line) for line in text.splitlines() if normalize_line(line)]
    name_header_indexes = [
        index for index, line in enumerate(lines) if line.lower() == "наименование" or is_axb_material_name_header_line(line)
    ]
    if len(name_header_indexes) < 2:
        return parts

    fallback_name_header_index = name_header_indexes[-1]
    fallback_entries = _extract_axb_material_section_entries_from_lines(lines, name_header_index=fallback_name_header_index)
    if not fallback_entries:
        return parts

    fallback_slice = "\n".join(lines[max(0, fallback_name_header_index - 8) :])
    fallback_amounts = extract_axb_material_section_amounts(fallback_slice)
    fallback_total = None
    if expected_parts_total is not None and any(
        amounts_match(amount, float(expected_parts_total), tolerance=0.2) for amount in fallback_amounts
    ):
        fallback_total = float(expected_parts_total)
    elif fallback_amounts:
        fallback_total = max(fallback_amounts)
    if fallback_total is None:
        return parts

    parts.append(
        {
            "article": fallback_entries[0]["article"] if len(fallback_entries) == 1 else None,
            "part_name": "; ".join(str(item["part_name"]) for item in fallback_entries if item.get("part_name"))[:500],
            "quantity": 1.0,
            "unit_name": "шт",
            "price": float(fallback_total),
            "line_total": float(fallback_total),
        }
    )
    return parts


def extract_axb_invoice_items(text: str) -> dict[str, list[dict[str, object]]]:
    invoice_fragment = extract_axb_invoice_fragment(text)
    if not invoice_fragment:
        return {"works": [], "parts": []}

    lines = [normalize_line(line) for line in invoice_fragment.splitlines() if normalize_line(line)]
    totals, expected_count = extract_axb_invoice_totals(lines)
    if expected_count == 0:
        return {"works": [], "parts": []}

    try:
        names_start = next(index for index, line in enumerate(lines) if line.lower() == "товар")
        article_header_index = next(
            index for index, line in enumerate(lines[names_start + 1 :], start=names_start + 1) if line.lower() == "артикул"
        )
        quantity_header_index = next(
            index for index, line in enumerate(lines[article_header_index + 1 :], start=article_header_index + 1) if line.lower().startswith("кол-во")
        )
    except StopIteration:
        return {"works": [], "parts": []}

    pending_codes: list[str] = []
    trailing_name_lines: list[str] = []
    seen_non_article_after_header = False
    for line in lines[article_header_index + 1 : quantity_header_index]:
        if not seen_non_article_after_header and is_axb_article_candidate(line):
            normalized_code = normalize_article_value(line) or normalize_text(line)
            if normalized_code:
                pending_codes.append(normalized_code)
            continue
        seen_non_article_after_header = True
        if is_axb_article_candidate(line):
            continue
        trailing_name_lines.append(line)

    names = collapse_axb_name_lines(
        lines[names_start + 1 : article_header_index] + trailing_name_lines,
        expected_count,
    )
    if len(names) != expected_count:
        return {"works": [], "parts": []}

    body_lines = lines[quantity_header_index + 1 :]
    total_marker_index = next((index for index, line in enumerate(body_lines) if line.lower() == "всего"), len(body_lines))
    body_lines = body_lines[:total_marker_index]
    invoice_total_index = next(
        (index for index, line in enumerate(body_lines) if normalize_line(line).lower().startswith("итого rub")),
        len(body_lines),
    )
    body_lines = body_lines[:invoice_total_index]

    normalized_body_lines: list[str] = []
    index = 0
    while index < len(body_lines):
        current_line = body_lines[index]
        if (
            current_line.endswith("-")
            and index + 1 < len(body_lines)
            and not is_axb_invoice_header_line(body_lines[index + 1])
            and not is_axb_invoice_stop_line(body_lines[index + 1])
            and not parse_axb_quantity_candidate(body_lines[index + 1])
            and parse_amount(body_lines[index + 1]) is None
        ):
            normalized_body_lines.append(normalize_line(f"{current_line}{body_lines[index + 1]}"))
            index += 2
            continue
        normalized_body_lines.append(current_line)
        index += 1

    rows: list[dict[str, object]] = []
    body_index = 0
    trailing_amounts: list[float] = []

    while body_index < len(normalized_body_lines) and len(rows) < expected_count:
        current_line = normalized_body_lines[body_index]
        if is_axb_invoice_header_line(current_line) or is_axb_invoice_stop_line(current_line):
            body_index += 1
            continue
        if is_axb_article_candidate(current_line):
            normalized_code = normalize_article_value(current_line) or normalize_text(current_line)
            if normalized_code:
                pending_codes.append(normalized_code)
            body_index += 1
            continue

        quantity_candidate = parse_axb_quantity_candidate(current_line)
        if quantity_candidate is None:
            body_index += 1
            continue

        qty_batch: list[tuple[float, Optional[str]]] = []
        while body_index < len(normalized_body_lines):
            quantity_candidate = parse_axb_quantity_candidate(normalized_body_lines[body_index])
            if quantity_candidate is None:
                break
            qty_batch.append(quantity_candidate)
            body_index += 1

        amount_batch: list[float] = []
        non_amount_streak = 0
        while body_index < len(normalized_body_lines):
            current_line = normalized_body_lines[body_index]
            if is_axb_invoice_header_line(current_line) or is_axb_invoice_stop_line(current_line):
                break
            if is_axb_article_candidate(current_line) or parse_axb_quantity_candidate(current_line) is not None:
                break

            amount_value = parse_amount(current_line)
            body_index += 1
            if amount_value is None:
                if amount_batch:
                    non_amount_streak += 1
                    if non_amount_streak >= 2:
                        break
                continue
            amount_batch.append(amount_value)
            non_amount_streak = 0

        current_row_index = len(rows)
        batch_quantities = [quantity for quantity, _raw_unit in qty_batch]
        batch_prices = select_axb_batch_prices(
            batch_quantities,
            amount_batch,
            totals[current_row_index : current_row_index + len(qty_batch)],
        )
        trailing_amounts = amount_batch[len(batch_prices) :]

        for batch_index, (quantity, raw_unit) in enumerate(qty_batch):
            if len(rows) >= expected_count:
                break
            row_payload: dict[str, object] = {
                "quantity": quantity,
                "raw_unit": raw_unit,
            }
            if batch_index < len(batch_prices):
                row_payload["price"] = batch_prices[batch_index]
            if pending_codes:
                row_payload["code"] = pending_codes.pop(0)
            rows.append(row_payload)

    missing_rows = expected_count - len(rows)
    if missing_rows > 0 and trailing_amounts:
        trailing_prices = [trailing_amounts[index] for index in range(len(trailing_amounts) - 3, -1, -3)]
        trailing_prices.reverse()
        for price in trailing_prices[-missing_rows:]:
            rows.append({"price": price})
            if len(rows) >= expected_count:
                break

    works: list[dict[str, object]] = []
    parts: list[dict[str, object]] = []

    for index, (name, line_total) in enumerate(zip(names, totals)):
        row_payload = rows[index] if index < len(rows) else {}
        quantity = float(row_payload["quantity"]) if row_payload.get("quantity") is not None else None
        price = float(row_payload["price"]) if row_payload.get("price") is not None else None
        explicit_unit = normalize_axb_explicit_unit(str(row_payload.get("raw_unit")) if row_payload.get("raw_unit") else None)

        if quantity is None and price is not None:
            quantity = infer_axb_quantity(line_total, price)
        if price is None and quantity is not None:
            price = infer_axb_price(line_total, quantity)
        if quantity is None or price is None:
            continue

        item_kind = infer_axb_item_kind(name, unit_name=explicit_unit, quantity=quantity)
        unit_name = explicit_unit
        if unit_name is None:
            if item_kind == "works":
                unit_name = "ч"
            elif abs(quantity - round(quantity)) <= 0.001:
                unit_name = "шт"

        code_value = str(row_payload.get("code")) if row_payload.get("code") else None
        if item_kind == "works":
            works.append(
                {
                    "work_code": normalize_article_value(code_value),
                    "work_name": name[:500],
                    "quantity": quantity,
                    "unit_name": unit_name,
                    "price": price,
                    "line_total": float(line_total),
                }
            )
            continue

        parts.append(
            {
                "article": normalize_article_value(code_value),
                "part_name": name[:500],
                "quantity": quantity,
                "unit_name": unit_name,
                "price": price,
                "line_total": float(line_total),
            }
        )

    return {"works": works, "parts": parts}


def apply_profile_specific_item_fallbacks(
    text: str,
    *,
    profile_scope: str | None,
    extracted_items: dict[str, list[dict[str, object]]],
    extracted_fields: dict[str, object],
    normalization_notes: list[str],
) -> dict[str, list[dict[str, object]]]:
    normalized_profile_scope = normalize_ocr_rule_code(profile_scope)
    if normalized_profile_scope not in {"axb", "antares", "ets_act", "sibtrakscan", "leader_trak", "gruzovye_rezervy", "logistics", "klever_trak"}:
        return extracted_items

    header_work_total = float(extracted_fields["work_total"]) if isinstance(extracted_fields.get("work_total"), (int, float)) else None
    header_parts_total = float(extracted_fields["parts_total"]) if isinstance(extracted_fields.get("parts_total"), (int, float)) else None
    header_grand_total = float(extracted_fields["grand_total"]) if isinstance(extracted_fields.get("grand_total"), (int, float)) else None

    if normalized_profile_scope == "antares":
        fallback_items = extract_antares_items(text)
    elif normalized_profile_scope == "ets_act":
        fallback_items = extract_ets_act_items(text)
    elif normalized_profile_scope == "logistics":
        fallback_items = extract_logistics_items(text)
    elif normalized_profile_scope == "klever_trak":
        fallback_items = extract_klever_trak_items(text)
    elif normalized_profile_scope == "sibtrakscan":
        fallback_items = extract_sibtrakscan_items(text)
    elif normalized_profile_scope == "leader_trak":
        fallback_items = extract_leader_trak_items(text)
    elif normalized_profile_scope == "gruzovye_rezervy":
        fallback_items = extract_gruzovye_rezervy_items(text)
    else:
        fallback_items = extract_axb_invoice_items(text)
        profile_work_items = extract_axb_work_items(text)
        tuned_profile_work_items = (
            extract_axb_work_items(text, expected_work_total=header_work_total)
            if header_work_total is not None
            else []
        )
        material_parts = extract_axb_material_parts(text, expected_parts_total=header_parts_total)
        raw_profile_work_total = round(sum(float(item.get("line_total") or 0) for item in profile_work_items), 2) if profile_work_items else None
        tuned_profile_work_total = (
            round(sum(float(item.get("line_total") or 0) for item in tuned_profile_work_items), 2)
            if tuned_profile_work_items
            else None
        )
        if tuned_profile_work_items and (
            not profile_work_items
            or (
                header_work_total is not None
                and amounts_match(tuned_profile_work_total, header_work_total, tolerance=3.0)
                and not amounts_match(raw_profile_work_total, header_work_total, tolerance=3.0)
            )
        ):
            profile_work_items = tuned_profile_work_items
        fallback_work_total, _ = summarize_line_totals(fallback_items)
        profile_work_total = round(sum(float(item.get("line_total") or 0) for item in profile_work_items), 2) if profile_work_items else None
        if profile_work_items and (
            len(profile_work_items) >= len(fallback_items["works"])
            or (
                header_work_total is not None
                and amounts_match(profile_work_total, header_work_total, tolerance=3.0)
                and not amounts_match(fallback_work_total, header_work_total, tolerance=3.0)
            )
        ):
            fallback_items["works"] = profile_work_items
        if material_parts and len(material_parts) >= len(fallback_items["parts"]):
            fallback_items["parts"] = material_parts

    fallback_items, fallback_removed_count = sanitize_extracted_items(fallback_items)
    if fallback_removed_count:
        normalization_notes.append(f"noise_work_items_removed_from_profile_fallback:{fallback_removed_count}")

    current_count = len(extracted_items.get("works") or []) + len(extracted_items.get("parts") or [])
    fallback_fills_missing_section = (
        bool(fallback_items["works"]) and not bool(extracted_items.get("works"))
    ) or (
        bool(fallback_items["parts"]) and not bool(extracted_items.get("parts"))
    )
    fallback_drops_existing_section = (
        bool(extracted_items.get("works")) and not bool(fallback_items["works"])
    ) or (
        bool(extracted_items.get("parts")) and not bool(fallback_items["parts"])
    )
    candidate_items = fallback_items
    if normalized_profile_scope == "ets_act" and fallback_fills_missing_section and fallback_drops_existing_section:
        candidate_items = {
            "works": list(fallback_items["works"] or extracted_items.get("works") or []),
            "parts": list(fallback_items["parts"] or extracted_items.get("parts") or []),
        }

    fallback_count = len(candidate_items["works"]) + len(candidate_items["parts"])
    if fallback_count == 0:
        return extracted_items

    fallback_work_total, fallback_parts_total = summarize_line_totals(candidate_items)
    fallback_grand_total = None
    if fallback_work_total is not None and fallback_parts_total is not None:
        fallback_grand_total = round(fallback_work_total + fallback_parts_total, 2)

    current_work_total, current_parts_total = summarize_line_totals(extracted_items)
    current_grand_total = None
    if current_work_total is not None and current_parts_total is not None:
        current_grand_total = round(current_work_total + current_parts_total, 2)

    fallback_matches_header = header_grand_total is not None and amounts_match(fallback_grand_total, header_grand_total, tolerance=3.0)
    current_matches_header = header_grand_total is not None and amounts_match(current_grand_total, header_grand_total, tolerance=3.0)
    fallback_match_score = int(amounts_match(fallback_work_total, header_work_total, tolerance=3.0)) + int(
        amounts_match(fallback_parts_total, header_parts_total, tolerance=3.0)
    )
    current_match_score = int(amounts_match(current_work_total, header_work_total, tolerance=3.0)) + int(
        amounts_match(current_parts_total, header_parts_total, tolerance=3.0)
    )

    if (
        (fallback_fills_missing_section and not fallback_drops_existing_section)
        or
        fallback_count > current_count
        or fallback_match_score > current_match_score
        or (fallback_matches_header and not current_matches_header)
    ):
        if normalized_profile_scope == "antares":
            normalization_notes.append("antares_items_restored_from_tabular_sections")
        elif normalized_profile_scope == "ets_act":
            normalization_notes.append("ets_act_items_restored_from_tabular_sections")
        elif normalized_profile_scope == "sibtrakscan":
            normalization_notes.append("sibtrakscan_items_restored_from_task_sections")
        elif normalized_profile_scope == "leader_trak":
            if is_leader_trak_invoice_only_document(text):
                normalization_notes.append("leader_trak_items_restored_from_invoice_table")
            else:
                normalization_notes.append("leader_trak_items_restored_from_service_table")
        elif normalized_profile_scope == "logistics":
            normalization_notes.append("logistics_items_restored_from_tabular_sections")
        elif normalized_profile_scope == "klever_trak":
            normalization_notes.append("klever_trak_items_restored_from_spreadsheet_rows")
        elif normalized_profile_scope == "gruzovye_rezervy":
            normalization_notes.append("gruzovye_rezervy_items_restored_from_sections")
        else:
            normalization_notes.append("axb_invoice_items_restored_from_payment_invoice")
        return candidate_items

    if normalized_profile_scope == "axb" and fallback_items["parts"] and not extracted_items.get("parts"):
        normalization_notes.append("axb_parts_restored_from_material_sections")
        return {
            "works": extracted_items.get("works", []),
            "parts": fallback_items["parts"],
        }

    return extracted_items


def apply_profile_specific_total_fallbacks(
    text: str,
    *,
    profile_scope: str | None,
    extracted_fields: dict[str, object],
    confidence_map: dict[str, float],
    normalization_notes: list[str],
) -> None:
    normalized_profile_scope = normalize_ocr_rule_code(profile_scope)
    if normalized_profile_scope == "klever_trak":
        work_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Итого\s+работ:",
                stop_patterns=(r"Расходная\s+накладная", r"Итого\s+материал", r"Итого\s+по\s+заказ[- ]наряду"),
                max_chars=180,
            )
        )
        if len(work_amounts) >= 2:
            extracted_fields["work_total"] = work_amounts[0]
            confidence_map["work_total"] = max(confidence_map.get("work_total", 0.0), 0.92)
            normalization_notes.append("work_total_restored_from_klever_trak_summary")

        parts_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Итого\s+материал(?:ов|ы):",
                stop_patterns=(r"Итого\s+по\s+заказ[- ]наряду", r"Всего\s+по\s+заказ[- ]наряду"),
                max_chars=180,
            )
        )
        if len(parts_amounts) >= 2:
            extracted_fields["parts_total"] = parts_amounts[0]
            confidence_map["parts_total"] = max(confidence_map.get("parts_total", 0.0), 0.92)
            normalization_notes.append("parts_total_restored_from_klever_trak_summary")

        overall_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Итого\s+по\s+заказ[- ]наряду",
                stop_patterns=(r"Всего\s+по\s+заказ[- ]наряду", r"Мастер-приемщик", r"Рекомендации"),
                max_chars=180,
            )
        )
        if len(overall_amounts) >= 2:
            extracted_fields["grand_total"] = overall_amounts[0]
            extracted_fields["vat_total"] = overall_amounts[1]
            confidence_map["grand_total"] = max(confidence_map.get("grand_total", 0.0), 0.94)
            confidence_map["vat_total"] = max(confidence_map.get("vat_total", 0.0), 0.9)
            normalization_notes.append("grand_total_restored_from_klever_trak_summary")
        return

    if normalized_profile_scope == "antares":
        work_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Итого\s+работ:",
                stop_patterns=(r"Расходная\s+накладная", r"Принят:", r"Вид\s+ремонта", r"Итого\s+по\s+причине\s+обращения"),
                max_chars=160,
            )
        )
        if work_amounts:
            extracted_fields["work_total"] = work_amounts[0]
            confidence_map["work_total"] = max(confidence_map.get("work_total", 0.0), 0.9)
            normalization_notes.append("work_total_restored_from_antares_summary")

        parts_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Итого\s+материал(?:ов|ы):",
                stop_patterns=(r"Итого\s+по\s+причине\s+обращения", r"Итого\s+по\s+заказ[- ]наряду", r"Руб"),
                max_chars=220,
            )
        )
        if len(parts_amounts) >= 2:
            extracted_fields["parts_total"] = parts_amounts[-2]
            confidence_map["parts_total"] = max(confidence_map.get("parts_total", 0.0), 0.9)
            normalization_notes.append("parts_total_restored_from_antares_summary")

        overall_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Итого\s+по\s+причине\s+обращения",
                stop_patterns=(r"Всего\s+по\s+причине\s+обращения", r"Итого\s+по\s+заказ[- ]наряду", r"Гарантии", r"Мастер"),
                max_chars=180,
            )
        )
        if len(overall_amounts) < 2:
            overall_amounts = extract_amount_candidates_from_fragment(
                extract_fragment_after_marker(
                    text,
                    r"Итого\s+по\s+заказ[- ]наряду",
                    stop_patterns=(r"Всего\s+по\s+заказ[- ]наряду", r"Гарантии", r"Мастер"),
                    max_chars=180,
                )
            )
        if len(overall_amounts) >= 2:
            extracted_fields["grand_total"] = overall_amounts[0]
            extracted_fields["vat_total"] = overall_amounts[1]
            confidence_map["grand_total"] = max(confidence_map.get("grand_total", 0.0), 0.92)
            confidence_map["vat_total"] = max(confidence_map.get("vat_total", 0.0), 0.88)
            normalization_notes.append("grand_total_restored_from_antares_summary")

            if isinstance(extracted_fields.get("work_total"), (int, float)):
                derived_parts_total = round(overall_amounts[0] - float(extracted_fields["work_total"]), 2)
                if derived_parts_total > 0:
                    extracted_fields["parts_total"] = derived_parts_total
                    confidence_map["parts_total"] = max(confidence_map.get("parts_total", 0.0), 0.92)
                    normalization_notes.append("parts_total_derived_from_antares_overall_total")
        return

    if normalized_profile_scope == "ets_act":
        work_fragment = extract_fragment_after_marker(
            text,
            r"Итого\s+работ:",
            stop_patterns=(r"Расходная\s+накладная", r"Итого\s+материал", r"Руб", r"руб"),
            max_chars=250,
        )
        work_amounts = extract_amount_candidates_from_fragment(work_fragment)
        work_total_candidate = None
        if len(work_amounts) >= 8 and work_amounts[-1] > max(work_amounts[:-1]):
            work_total_candidate = work_amounts[-1]
        elif len(work_amounts) >= 3:
            work_total_candidate = work_amounts[-3]
        if work_total_candidate is not None and (
            not isinstance(extracted_fields.get("work_total"), (int, float))
            or float(extracted_fields["work_total"]) < float(work_total_candidate) * 0.5
        ):
            extracted_fields["work_total"] = work_total_candidate
            confidence_map["work_total"] = max(confidence_map.get("work_total", 0.0), 0.9)
            normalization_notes.append("work_total_restored_from_ets_act_summary")

        overall_fragment = extract_fragment_after_marker(
            text,
            r"Итого\s+по\s+акту\s+выполненных\s+работ",
            stop_patterns=(r"Всего\s+по\s+акту", r"Причина\s+обращения", r"Рекомендации"),
            max_chars=250,
        )
        overall_amounts = extract_amount_candidates_from_fragment(overall_fragment)
        if len(overall_amounts) >= 3:
            overall_net_total, overall_vat_total, overall_grand_total = overall_amounts[-3:]
            extracted_fields["vat_total"] = overall_vat_total
            extracted_fields["grand_total"] = overall_grand_total
            confidence_map["vat_total"] = max(confidence_map.get("vat_total", 0.0), 0.9)
            confidence_map["grand_total"] = max(confidence_map.get("grand_total", 0.0), 0.9)
            normalization_notes.append("grand_total_restored_from_ets_act_summary")

            if isinstance(extracted_fields.get("work_total"), (int, float)):
                derived_parts_total = round(overall_net_total - float(extracted_fields["work_total"]), 2)
                if derived_parts_total > 0:
                    extracted_fields["parts_total"] = derived_parts_total
                    confidence_map["parts_total"] = max(confidence_map.get("parts_total", 0.0), 0.88)
                    normalization_notes.append("parts_total_derived_from_ets_act_summary")

        invoice_total_fragment = extract_fragment_after_marker(
            text,
            r"Всего\s+к\s+оплате:",
            stop_patterns=(r"Всего\s+наименований", r"Предпр", r"Счет-фактура"),
            max_chars=180,
        )
        invoice_total_amounts = extract_amount_candidates_from_fragment(invoice_total_fragment)
        if len(invoice_total_amounts) >= 3:
            if len(invoice_total_amounts) >= 4 and amounts_match(invoice_total_amounts[0], invoice_total_amounts[1], tolerance=0.01):
                invoice_net_total = invoice_total_amounts[0]
                invoice_vat_total = invoice_total_amounts[2]
                invoice_grand_total = invoice_total_amounts[3]
            else:
                invoice_net_total = invoice_total_amounts[-3]
                invoice_vat_total = invoice_total_amounts[-2]
                invoice_grand_total = invoice_total_amounts[-1]

            if (
                not isinstance(extracted_fields.get("vat_total"), (int, float))
                or float(extracted_fields["vat_total"]) < invoice_vat_total * 0.5
            ):
                extracted_fields["vat_total"] = invoice_vat_total
                confidence_map["vat_total"] = max(confidence_map.get("vat_total", 0.0), 0.88)
                normalization_notes.append("vat_total_restored_from_ets_invoice_summary")

            if (
                not isinstance(extracted_fields.get("grand_total"), (int, float))
                or float(extracted_fields["grand_total"]) < invoice_grand_total * 0.5
            ):
                extracted_fields["grand_total"] = invoice_grand_total
                confidence_map["grand_total"] = max(confidence_map.get("grand_total", 0.0), 0.9)
                normalization_notes.append("grand_total_restored_from_ets_invoice_summary")

            if isinstance(extracted_fields.get("work_total"), (int, float)):
                derived_parts_total = round(invoice_net_total - float(extracted_fields["work_total"]), 2)
                if derived_parts_total > 0 and (
                    not isinstance(extracted_fields.get("parts_total"), (int, float))
                    or float(extracted_fields["parts_total"]) < derived_parts_total * 0.5
                ):
                    extracted_fields["parts_total"] = derived_parts_total
                    confidence_map["parts_total"] = max(confidence_map.get("parts_total", 0.0), 0.88)
                    normalization_notes.append("parts_total_derived_from_ets_invoice_summary")

        sparse_totals = extract_ets_act_sparse_scanned_totals(text)
        if sparse_totals:
            for field_name, confidence, note in (
                ("work_total", 0.9, "work_total_restored_from_ets_act_sparse_scan"),
                ("parts_total", 0.88, "parts_total_restored_from_ets_act_sparse_scan"),
                ("vat_total", 0.88, "vat_total_restored_from_ets_act_sparse_scan"),
                ("grand_total", 0.9, "grand_total_restored_from_ets_act_sparse_scan"),
            ):
                sparse_value = sparse_totals.get(field_name)
                if sparse_value is None:
                    continue
                current_value = extracted_fields.get(field_name)
                should_replace = not isinstance(current_value, (int, float))
                if isinstance(current_value, (int, float)) and float(current_value) < float(sparse_value) * 0.5:
                    should_replace = True
                if should_replace:
                    extracted_fields[field_name] = sparse_value
                    confidence_map[field_name] = max(confidence_map.get(field_name, 0.0), confidence)
                    normalization_notes.append(note)
        return

    if normalized_profile_scope == "sibtrakscan":
        work_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Всего\s+по\s+работам",
                stop_patterns=(r"Всего\s+по\s+материалам", r"Всего:", r"К\s+оплате"),
                max_chars=120,
            )
        )
        if work_amounts:
            extracted_fields["work_total"] = work_amounts[0]
            confidence_map["work_total"] = max(confidence_map.get("work_total", 0.0), 0.9)
            normalization_notes.append("work_total_restored_from_sibtrakscan_summary")

        parts_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Всего\s+по\s+материалам",
                stop_patterns=(r"Всего:", r"в\s+т\.ч\.\s+НДС", r"Итого\s+по\s+заказ[- ]наряду"),
                max_chars=120,
            )
        )
        if parts_amounts:
            extracted_fields["parts_total"] = parts_amounts[0]
            confidence_map["parts_total"] = max(confidence_map.get("parts_total", 0.0), 0.9)
            normalization_notes.append("parts_total_restored_from_sibtrakscan_summary")

        grand_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"К\s+оплате",
                stop_patterns=(r"Сто\s", r"Дата\s+оплаты", r"ГАРАНТИИ", r"РЕКОМЕНДАЦИИ"),
                max_chars=120,
            )
        )
        if not grand_amounts:
            grand_amounts = extract_amount_candidates_from_fragment(
                extract_fragment_after_marker(
                    text,
                    r"Итого\s+по\s+заказ[- ]наряду",
                    stop_patterns=(r"К\s+оплате", r"Сто\s", r"Дата\s+оплаты", r"ГАРАНТИИ", r"РЕКОМЕНДАЦИИ"),
                    max_chars=180,
                )
            )
        if grand_amounts:
            extracted_fields["grand_total"] = grand_amounts[-1]
            confidence_map["grand_total"] = max(confidence_map.get("grand_total", 0.0), 0.92)
            normalization_notes.append("grand_total_restored_from_sibtrakscan_summary")

        vat_amounts = extract_amount_candidates_from_fragment(
            extract_fragment_after_marker(
                text,
                r"в\s+т\.ч\.\s+НДС",
                stop_patterns=(r"Итого\s+по\s+заказ[- ]наряду", r"К\s+оплате", r"Сто\s"),
                max_chars=120,
            )
        )
        if vat_amounts:
            extracted_fields["vat_total"] = vat_amounts[0]
            confidence_map["vat_total"] = max(confidence_map.get("vat_total", 0.0), 0.9)
            normalization_notes.append("vat_total_restored_from_sibtrakscan_summary")
        return

    if normalized_profile_scope == "leader_trak":
        summary_fragment = extract_fragment_after_marker(
            text,
            r"Всего\s+по\s+наряд[- ]заказу",
            stop_patterns=(r"Всего:", r"Сумма\s+прописью", r"Сервисные\s+услуги", r"После\s+подписания"),
            max_chars=180,
        )
        summary_amounts = extract_amount_candidates_from_fragment(summary_fragment)
        if len(summary_amounts) >= 3:
            net_total, vat_total, grand_total = summary_amounts[-3:]
            extracted_fields["vat_total"] = vat_total
            extracted_fields["grand_total"] = grand_total
            confidence_map["vat_total"] = max(confidence_map.get("vat_total", 0.0), 0.9)
            confidence_map["grand_total"] = max(confidence_map.get("grand_total", 0.0), 0.92)
            normalization_notes.append("leader_trak_totals_restored_from_summary")

        invoice_summary_fragment = extract_fragment_after_marker(
            text,
            r"Итого\s+RUB:",
            stop_patterns=(r"Всего\s+наименований", r"Сумма\s+прописью"),
            max_chars=160,
        )
        invoice_summary_amounts = extract_amount_candidates_from_fragment(invoice_summary_fragment)
        if len(invoice_summary_amounts) >= 3:
            extracted_fields["vat_total"] = invoice_summary_amounts[-2]
            extracted_fields["grand_total"] = invoice_summary_amounts[-1]
            confidence_map["vat_total"] = max(confidence_map.get("vat_total", 0.0), 0.9)
            confidence_map["grand_total"] = max(confidence_map.get("grand_total", 0.0), 0.92)
            normalization_notes.append("leader_trak_totals_restored_from_invoice_summary")
        return

    if normalized_profile_scope != "axb":
        return

    axb_material_parts = extract_axb_material_parts(text)
    axb_material_parts_total = (
        round(sum(float(item.get("line_total") or 0) for item in axb_material_parts), 2)
        if axb_material_parts
        else None
    )

    profile_total_candidates = {
        "work_total": extract_largest_amount_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Итого\s+работ:",
                stop_patterns=(r"Расходная\s+накладная", r"Итого\s+материал", r"Итого\s+по\s+заказ[- ]наряду", r"Всего\s+по\s+заказ[- ]наряду"),
                max_chars=800,
            )
        ),
        "parts_total": extract_largest_amount_from_fragment(
            extract_fragment_after_marker(
                text,
                r"Итого\s+материал(?:ов|ы):",
                stop_patterns=(r"Итого\s+по\s+причине\s+обращения", r"Всего\s+по\s+причине\s+обращения", r"Итого\s+по\s+заказ[- ]наряду", r"Всего\s+по\s+заказ[- ]наряду"),
                max_chars=1800,
            )
        ),
        "grand_total": extract_largest_amount_around_marker(
            text,
            r"(?:Итого|Всего)\s+по\s+заказ[- ]наряду",
            before_chars=180,
            after_chars=180,
            stop_patterns=(r"Заказчик\s+подтверждает", r"Заказ-наряд\s+и\s+Сч[её]т", r"Универсальный\s+передаточный"),
        ),
    }

    if axb_material_parts_total is not None and (
        profile_total_candidates["parts_total"] is None
        or float(profile_total_candidates["parts_total"]) < round(axb_material_parts_total * 0.7, 2)
    ):
        profile_total_candidates["parts_total"] = axb_material_parts_total

    for field_name, candidate_amount in profile_total_candidates.items():
        if candidate_amount is None:
            continue

        current_value_raw = extracted_fields.get(field_name)
        current_value = float(current_value_raw) if isinstance(current_value_raw, (int, float)) else None
        if current_value is not None and current_value >= candidate_amount:
            continue

        extracted_fields[field_name] = candidate_amount
        confidence_map[field_name] = max(confidence_map.get(field_name, 0.0), 0.86 if field_name == "grand_total" else 0.82)
        normalization_notes.append(f"{field_name}_restored_from_axb_profile_totals")

    work_total = extracted_fields.get("work_total")
    parts_total = extracted_fields.get("parts_total")
    grand_total = extracted_fields.get("grand_total")
    if (
        axb_material_parts_total is None
        and isinstance(work_total, (int, float))
        and isinstance(grand_total, (int, float))
    ):
        derived_parts_total = round(float(grand_total) - float(work_total), 2)
        if derived_parts_total > 0 and (
            not isinstance(parts_total, (int, float)) or float(parts_total) < round(derived_parts_total * 0.7, 2)
        ):
            extracted_fields["parts_total"] = derived_parts_total
            confidence_map["parts_total"] = max(confidence_map.get("parts_total", 0.0), 0.84)
            normalization_notes.append("parts_total_derived_from_axb_grand_total")


def enrich_work_payloads_with_labor_norms(
    db: Session,
    works_payload: list[dict[str, object]],
    applicability: LaborNormApplicability,
) -> tuple[list[str], LaborNormEnrichmentSummary]:
    notes: list[str] = []
    matched_count = 0
    unmatched_count = 0
    applicable_item_count = 0
    for item in works_payload:
        work_name = str(item.get("work_name") or "").strip()
        if not work_name:
            continue

        work_code = normalize_labor_norm_code(str(item.get("work_code"))) if item.get("work_code") else None
        if work_code:
            item["work_code"] = work_code

        reference_payload = item.get("reference_payload")
        if not isinstance(reference_payload, dict):
            reference_payload = {}
        reference_payload["normalized_work_name"] = build_normalized_name(work_name)
        reference_payload["labor_norm_applicable"] = applicability.eligible
        reference_payload["labor_norm_scope"] = applicability.scope
        reference_payload["labor_norm_applicability_reason_code"] = applicability.reason_code
        reference_payload["labor_norm_applicability_reason"] = applicability.reason
        reference_payload["labor_norm_item_applicable"] = applicability.eligible
        reference_payload["labor_norm_item_reason_code"] = applicability.reason_code
        reference_payload["labor_norm_item_reason"] = applicability.reason
        if applicability.catalog_name:
            reference_payload["labor_norm_catalog_name"] = applicability.catalog_name
        if applicability.brand_family:
            reference_payload["labor_norm_brand_family"] = applicability.brand_family
        if item.get("standard_hours") is not None:
            try:
                reference_payload["document_standard_hours"] = float(item["standard_hours"])
            except (TypeError, ValueError):
                reference_payload.pop("document_standard_hours", None)

        is_non_catalog_service, non_catalog_reason = classify_known_non_catalog_operation(
            work_code=work_code,
            work_name=work_name,
        )
        if is_non_catalog_service:
            reference_payload["labor_norm_item_applicable"] = False
            reference_payload["labor_norm_item_reason_code"] = "outside_catalog_service"
            reference_payload["labor_norm_item_reason"] = non_catalog_reason
            reference_payload["labor_norm_reference_status"] = "outside_catalog_service"
            item["reference_payload"] = reference_payload
            notes.append("labor_norm_skip:outside_catalog_service")
            continue

        if not applicability.eligible:
            reference_payload["labor_norm_reference_status"] = "catalog_not_applicable"
            item["reference_payload"] = reference_payload
            continue

        applicable_item_count += 1
        match = find_best_labor_norm_match(
            db,
            work_code=work_code,
            work_name=work_name,
            scope=applicability.scope,
        )
        if match is None:
            reference_payload["labor_norm_reference_status"] = "catalog_gap"
            item["reference_payload"] = reference_payload
            unmatched_count += 1
            continue

        if not item.get("work_code"):
            item["work_code"] = match.norm.code
        reference_payload.update(
            {
                "labor_norm_id": match.norm.id,
                "labor_norm_code": match.norm.code,
                "labor_norm_scope": match.norm.scope,
                "labor_norm_catalog_name": match.norm.catalog_name,
                "labor_norm_brand_family": match.norm.brand_family,
                "labor_norm_name": match.norm.name_ru,
                "labor_norm_category": match.norm.category,
                "labor_norm_standard_hours": float(match.norm.standard_hours),
                "labor_norm_match_score": match.score,
                "labor_norm_matched_by": match.matched_by,
                "labor_norm_reference_status": "matched",
            }
        )
        item["reference_payload"] = reference_payload
        notes.append(f"labor_norm_match:{match.norm.code}")
        matched_count += 1

    if works_payload and not applicability.eligible:
        notes.append(f"labor_norm_skipped:{applicability.reason_code}")
    elif applicable_item_count > 0 and matched_count == 0:
        notes.append("labor_norm_match_missing")

    return notes, LaborNormEnrichmentSummary(
        matched_count=matched_count,
        unmatched_count=unmatched_count,
    )


def build_standard_hours_checks(
    works_payload: list[dict[str, object]],
) -> list[dict[str, object]]:
    checks: list[dict[str, object]] = []
    for item in works_payload:
        reference_payload = item.get("reference_payload")
        if not isinstance(reference_payload, dict):
            reference_payload = {}
        normalized_unit_name = normalize_unit_name(str(item.get("unit_name")) if item.get("unit_name") else None)
        actual_hours = item.get("actual_hours")
        if actual_hours is None and normalized_unit_name in {"нч", "ч"} and item.get("quantity") is not None:
            actual_hours = float(item["quantity"])
        document_standard_hours = reference_payload.get("document_standard_hours")
        if document_standard_hours is None:
            document_standard_hours = item.get("standard_hours")
        catalog_standard_hours = reference_payload.get("labor_norm_standard_hours")

        actual_value: Optional[float] = None
        if actual_hours is not None:
            try:
                actual_value = float(actual_hours)
            except (TypeError, ValueError):
                actual_value = None

        document_standard_value: Optional[float] = None
        if document_standard_hours is not None:
            try:
                document_standard_value = float(document_standard_hours)
            except (TypeError, ValueError):
                document_standard_value = None

        catalog_standard_value: Optional[float] = None
        if catalog_standard_hours is not None:
            try:
                catalog_standard_value = float(catalog_standard_hours)
            except (TypeError, ValueError):
                catalog_standard_value = None

        comparison_standard_value = document_standard_value if document_standard_value is not None else catalog_standard_value
        if (
            actual_value is not None
            and comparison_standard_value is not None
            and comparison_standard_value > 0
            and actual_value > round(comparison_standard_value * 1.1, 2)
        ):
            checks.append(
                {
                    "check_type": "ocr_standard_hours_exceeded",
                    "severity": CheckSeverity.SUSPICIOUS,
                    "title": "Фактические часы превышают норматив",
                    "details": (
                        f"{item.get('work_name', 'Работа')} · факт {actual_value:.2f} ч, "
                        f"норма {comparison_standard_value:.2f} ч"
                    ),
                    "payload": {
                        "work_code": item.get("work_code"),
                        "work_name": item.get("work_name"),
                        "actual_hours": actual_value,
                        "standard_hours": comparison_standard_value,
                        "document_standard_hours": document_standard_value,
                        "catalog_standard_hours": catalog_standard_value,
                        "reference_payload": reference_payload,
                    },
                }
            )

        if (
            document_standard_value is not None
            and catalog_standard_value is not None
            and document_standard_value - catalog_standard_value > 0.01
        ):
            checks.append(
                {
                    "check_type": "ocr_document_standard_hours_exceeded",
                    "severity": CheckSeverity.SUSPICIOUS,
                    "title": "Норма в заказ-наряде выше нормы справочника",
                    "details": (
                        f"{item.get('work_name', 'Работа')} · в документе {document_standard_value:.2f} ч, "
                        f"в справочнике {catalog_standard_value:.2f} ч"
                    ),
                    "payload": {
                        "work_code": item.get("work_code"),
                        "work_name": item.get("work_name"),
                        "document_standard_hours": document_standard_value,
                        "catalog_standard_hours": catalog_standard_value,
                        "reference_payload": reference_payload,
                    },
                }
            )
    return checks


def build_repeat_repair_checks(
    db: Session,
    repair: Repair,
    works_payload: list[dict[str, object]],
) -> list[dict[str, object]]:
    if not works_payload or repair.vehicle_id is None or repair.repair_date is None:
        return []

    checks: list[dict[str, object]] = []
    seen_keys: set[tuple[Optional[str], str]] = set()
    window_start = repair.repair_date - timedelta(days=REPEAT_REPAIR_WINDOW_DAYS)

    previous_repairs = db.execute(
        select(Repair, RepairWork)
        .join(RepairWork, RepairWork.repair_id == Repair.id)
        .where(
            Repair.vehicle_id == repair.vehicle_id,
            Repair.id != repair.id,
            Repair.repair_date >= window_start,
            Repair.repair_date <= repair.repair_date,
        )
        .order_by(Repair.repair_date.desc(), Repair.id.desc(), RepairWork.id.desc())
    ).all()

    if not previous_repairs:
        return checks

    indexed_previous: dict[tuple[Optional[str], str], tuple[Repair, RepairWork]] = {}
    for previous_repair, previous_work in previous_repairs:
        normalized_name = build_normalized_name(previous_work.work_name or "")
        match_key = (previous_work.work_code or None, normalized_name)
        if match_key not in indexed_previous:
            indexed_previous[match_key] = (previous_repair, previous_work)

    for item in works_payload:
        work_name = str(item.get("work_name") or "").strip()
        if not work_name:
            continue
        work_code = str(item.get("work_code")).strip() if item.get("work_code") else None
        normalized_name = build_normalized_name(work_name)
        match_key = (work_code, normalized_name)
        fallback_key = (None, normalized_name)

        previous_match = indexed_previous.get(match_key) or indexed_previous.get(fallback_key)
        if previous_match is None or match_key in seen_keys:
            continue

        previous_repair, previous_work = previous_match
        seen_keys.add(match_key)
        days_delta = (repair.repair_date - previous_repair.repair_date).days

        checks.append(
            {
                "check_type": "ocr_repeat_repair_detected",
                "severity": CheckSeverity.SUSPICIOUS,
                "title": "Повторный ремонт по той же работе",
                "details": (
                    f"{work_name} · уже было {previous_repair.repair_date.isoformat()} "
                    f"по заказ-наряду {previous_repair.order_number or 'без номера'} "
                    f"({days_delta} дн. назад)"
                ),
                "payload": {
                    "work_code": work_code,
                    "work_name": work_name,
                    "previous_repair_id": previous_repair.id,
                    "previous_order_number": previous_repair.order_number,
                    "previous_repair_date": previous_repair.repair_date.isoformat(),
                    "previous_service_id": previous_repair.service_id,
                    "previous_work_id": previous_work.id,
                    "previous_work_code": previous_work.work_code,
                    "previous_work_name": previous_work.work_name,
                    "days_since_previous": days_delta,
                    "window_days": REPEAT_REPAIR_WINDOW_DAYS,
                },
            }
        )

    return checks


def build_duplicate_line_checks(
    works_payload: list[dict[str, object]],
    parts_payload: list[dict[str, object]],
) -> list[dict[str, object]]:
    checks: list[dict[str, object]] = []

    def build_group_key(
        item: dict[str, object],
        *,
        code_keys: tuple[str, ...],
        name_keys: tuple[str, ...],
    ) -> tuple[str | None, str, float | None, float | None, float | None]:
        code_value: str | None = None
        for key in code_keys:
            raw_code = item.get(key)
            if raw_code:
                code_value = str(raw_code).strip() or None
                if code_value:
                    break
        name_value = ""
        for key in name_keys:
            raw_name = item.get(key)
            if raw_name:
                name_value = build_normalized_name(str(raw_name))
                if name_value:
                    break
        quantity = float(item["quantity"]) if item.get("quantity") is not None else None
        price = float(item["price"]) if item.get("price") is not None else None
        line_total = float(item["line_total"]) if item.get("line_total") is not None else None
        return code_value, name_value, quantity, price, line_total

    work_groups: dict[tuple[str | None, str, float | None, float | None, float | None], list[dict[str, object]]] = {}
    for item in works_payload:
        group_key = build_group_key(item, code_keys=("work_code",), name_keys=("work_name",))
        if not group_key[1]:
            continue
        work_groups.setdefault(group_key, []).append(item)

    for (_code, _name, quantity, price, line_total), items in work_groups.items():
        if len(items) < 2:
            continue
        sample = items[0]
        checks.append(
            {
                "check_type": "ocr_duplicate_work_lines",
                "severity": CheckSeverity.SUSPICIOUS,
                "title": "Дубли строк работ в заказ-наряде",
                "details": (
                    f"{sample.get('work_name', 'Работа')} · совпадающих строк {len(items)}"
                    f"{f' · кол-во {quantity:.2f}' if quantity is not None else ''}"
                    f"{f' · цена {price:.2f}' if price is not None else ''}"
                    f"{f' · сумма {line_total:.2f}' if line_total is not None else ''}"
                ),
                "payload": {
                    "work_code": sample.get("work_code"),
                    "work_name": sample.get("work_name"),
                    "duplicate_count": len(items),
                    "quantity": quantity,
                    "price": price,
                    "line_total": line_total,
                },
            }
        )

    part_groups: dict[tuple[str | None, str, float | None, float | None, float | None], list[dict[str, object]]] = {}
    for item in parts_payload:
        group_key = build_group_key(item, code_keys=("article",), name_keys=("part_name", "name"))
        if not group_key[1]:
            continue
        part_groups.setdefault(group_key, []).append(item)

    for (_code, _name, quantity, price, line_total), items in part_groups.items():
        if len(items) < 2:
            continue
        sample = items[0]
        checks.append(
            {
                "check_type": "ocr_duplicate_part_lines",
                "severity": CheckSeverity.SUSPICIOUS,
                "title": "Дубли строк запчастей в заказ-наряде",
                "details": (
                    f"{sample.get('part_name', sample.get('name', 'Запчасть'))} · совпадающих строк {len(items)}"
                    f"{f' · кол-во {quantity:.2f}' if quantity is not None else ''}"
                    f"{f' · цена {price:.2f}' if price is not None else ''}"
                    f"{f' · сумма {line_total:.2f}' if line_total is not None else ''}"
                ),
                "payload": {
                    "article": sample.get("article"),
                    "part_name": sample.get("part_name") or sample.get("name"),
                    "duplicate_count": len(items),
                    "quantity": quantity,
                    "price": price,
                    "line_total": line_total,
                },
            }
        )

    return checks


def resolve_work_reference_hours(item: dict[str, object]) -> Optional[float]:
    reference_payload = item.get("reference_payload")
    if not isinstance(reference_payload, dict):
        reference_payload = {}

    candidate_values = [
        reference_payload.get("document_standard_hours"),
        item.get("standard_hours"),
        reference_payload.get("labor_norm_standard_hours"),
        item.get("actual_hours"),
    ]
    normalized_unit_name = normalize_unit_name(str(item.get("unit_name")) if item.get("unit_name") else None)
    if normalized_unit_name in {"нч", "ч"} and item.get("quantity") is not None:
        candidate_values.append(item.get("quantity"))

    for value in candidate_values:
        if value is None:
            continue
        try:
            normalized_value = float(value)
        except (TypeError, ValueError):
            continue
        if normalized_value > 0:
            return normalized_value
    return None


def build_expected_total_checks(
    db: Session,
    repair: Repair,
    works_payload: list[dict[str, object]],
) -> tuple[Optional[float], list[dict[str, object]]]:
    if not works_payload or repair.vehicle is None:
        return None, []

    historical_rows = db.execute(
        select(
            Repair.id,
            Repair.repair_date,
            Repair.service_id,
            Repair.work_total,
            Repair.parts_total,
            Repair.vat_total,
            Repair.grand_total,
            RepairWork.work_code,
            RepairWork.work_name,
            RepairWork.line_total,
            RepairWork.standard_hours,
            RepairWork.actual_hours,
        )
        .join(RepairWork, RepairWork.repair_id == Repair.id)
        .join(Vehicle, Vehicle.id == Repair.vehicle_id)
        .where(
            Repair.id != repair.id,
            Repair.status.in_(EXPECTED_TOTAL_REPAIR_STATUSES),
            Repair.grand_total > 0,
            Vehicle.vehicle_type == repair.vehicle.vehicle_type,
        )
    ).all()

    if not historical_rows:
        return None, []

    current_service_id = repair.service_id
    expected_line_totals: list[float] = []
    line_breakdown: list[dict[str, object]] = []

    general_hourly_rates: list[float] = []
    service_hourly_rates: list[float] = []
    for row in historical_rows:
        reference_hours = row.standard_hours if row.standard_hours and row.standard_hours > 0 else row.actual_hours
        if reference_hours is None or reference_hours <= 0:
            continue
        hourly_rate = float(row.line_total) / float(reference_hours)
        if hourly_rate <= 0:
            continue
        general_hourly_rates.append(hourly_rate)
        if current_service_id is not None and row.service_id == current_service_id:
            service_hourly_rates.append(hourly_rate)

    for item in works_payload:
        work_name = str(item.get("work_name") or "").strip()
        if not work_name:
            continue
        work_code = str(item.get("work_code")).strip() if item.get("work_code") else None
        normalized_name = build_normalized_name(work_name)

        line_service_matches: list[float] = []
        line_general_matches: list[float] = []
        for row in historical_rows:
            same_code = bool(work_code and row.work_code and str(row.work_code).strip() == work_code)
            same_name = build_normalized_name(str(row.work_name or "")) == normalized_name
            if not same_code and not same_name:
                continue
            line_general_matches.append(float(row.line_total))
            if current_service_id is not None and row.service_id == current_service_id:
                line_service_matches.append(float(row.line_total))

        selected_matches = (
            line_service_matches
            if len(line_service_matches) >= EXPECTED_TOTAL_SERVICE_SAMPLE_THRESHOLD
            else line_general_matches
        )
        if selected_matches:
            expected_line_total = round(float(statistics.median(selected_matches)), 2)
            expected_line_totals.append(expected_line_total)
            line_breakdown.append(
                {
                    "work_code": work_code,
                    "work_name": work_name,
                    "source": "historical_work_median",
                    "samples": len(selected_matches),
                    "expected_line_total": expected_line_total,
                }
            )
            continue

        reference_hours = resolve_work_reference_hours(item)
        if reference_hours is None:
            continue

        selected_hourly_rates = (
            service_hourly_rates
            if len(service_hourly_rates) >= EXPECTED_TOTAL_HOURLY_SAMPLE_THRESHOLD
            else general_hourly_rates
        )
        if not selected_hourly_rates:
            continue

        expected_line_total = round(float(statistics.median(selected_hourly_rates)) * reference_hours, 2)
        expected_line_totals.append(expected_line_total)
        line_breakdown.append(
            {
                "work_code": work_code,
                "work_name": work_name,
                "source": "historical_hourly_rate",
                "samples": len(selected_hourly_rates),
                "reference_hours": reference_hours,
                "expected_line_total": expected_line_total,
            }
        )

    if not expected_line_totals:
        return None, []

    expected_work_total = round(sum(expected_line_totals), 2)
    expected_total = round(expected_work_total + float(repair.parts_total or 0) + float(repair.vat_total or 0), 2)
    actual_total = round(float(repair.grand_total or 0), 2)

    checks: list[dict[str, object]] = []
    if expected_total > 0 and actual_total > round(expected_total * EXPECTED_TOTAL_THRESHOLD_MULTIPLIER, 2):
        checks.append(
            {
                "check_type": "ocr_expected_total_exceeded",
                "severity": CheckSeverity.SUSPICIOUS,
                "title": "Стоимость ремонта выше ожидаемой",
                "details": (
                    f"Итого {actual_total:.2f} руб. при ожидаемой стоимости {expected_total:.2f} руб. "
                    f"по истории аналогичных работ"
                ),
                "payload": {
                    "actual_total": actual_total,
                    "expected_total": expected_total,
                    "expected_work_total": expected_work_total,
                    "actual_work_total": round(float(repair.work_total or 0), 2),
                    "actual_parts_total": round(float(repair.parts_total or 0), 2),
                    "actual_vat_total": round(float(repair.vat_total or 0), 2),
                    "threshold_multiplier": EXPECTED_TOTAL_THRESHOLD_MULTIPLIER,
                    "line_breakdown": line_breakdown,
                },
            }
        )

    return expected_total, checks


def describe_work_reference_source(source: str) -> str:
    if source == "same_vehicle":
        return "по этой же технике"
    if source == "same_service":
        return "по этому же сервису"
    return "по типу техники"


def build_dynamic_work_reference_checks(
    db: Session,
    repair: Repair,
    works_payload: list[dict[str, object]],
) -> list[dict[str, object]]:
    if not works_payload or repair.vehicle is None:
        return []

    historical_rows = db.execute(
        select(
            Repair.id.label("repair_id"),
            Repair.vehicle_id,
            Repair.service_id,
            Repair.repair_date,
            Repair.mileage,
            Repair.status,
            Repair.reason,
            RepairWork.id.label("work_id"),
            RepairWork.work_code,
            RepairWork.work_name,
            RepairWork.quantity,
            RepairWork.price,
            RepairWork.line_total,
        )
        .join(RepairWork, RepairWork.repair_id == Repair.id)
        .join(Vehicle, Vehicle.id == Repair.vehicle_id)
        .where(
            Repair.id != repair.id,
            RepairWork.work_name.is_not(None),
            Vehicle.vehicle_type == repair.vehicle.vehicle_type,
            (
                Repair.reason.like(f"{HISTORICAL_IMPORT_REASON_PREFIX}%")
                | Repair.status.in_(WORK_REFERENCE_OPERATIONAL_STATUSES)
            ),
        )
    ).all()

    if not historical_rows:
        return []

    indexed_rows: dict[tuple[Optional[str], str], list[object]] = {}
    for row in historical_rows:
        work_name = str(row.work_name or "").strip()
        if not work_name:
            continue
        work_code = str(row.work_code).strip() if row.work_code else None
        normalized_name = build_normalized_name(work_name)
        key = (work_code, normalized_name)
        fallback_key = (None, normalized_name)
        indexed_rows.setdefault(key, []).append(row)
        if fallback_key != key:
            indexed_rows.setdefault(fallback_key, []).append(row)

    checks: list[dict[str, object]] = []
    seen_missing_keys: set[tuple[Optional[str], str]] = set()
    current_service_id = repair.service_id

    for item in works_payload:
        work_name = str(item.get("work_name") or "").strip()
        if not work_name:
            continue
        work_code = str(item.get("work_code")).strip() if item.get("work_code") else None
        normalized_name = build_normalized_name(work_name)
        match_key = (work_code, normalized_name)
        matches = indexed_rows.get(match_key) or indexed_rows.get((None, normalized_name)) or []

        reference_payload = item.get("reference_payload")
        if not isinstance(reference_payload, dict):
            reference_payload = {}
            item["reference_payload"] = reference_payload

        labor_norm_item_reason_code = str(reference_payload.get("labor_norm_item_reason_code") or "").strip()
        if labor_norm_item_reason_code == "outside_catalog_service":
            reference_payload["dynamic_work_reference"] = {
                "comparison_source": "not_applicable",
                "reason_code": labor_norm_item_reason_code,
                "reason": reference_payload.get("labor_norm_item_reason"),
                "sample_lines": 0,
                "historical_sample_lines": 0,
                "operational_sample_lines": 0,
            }
            continue

        if not matches:
            if match_key not in seen_missing_keys:
                checks.append(
                    {
                        "check_type": "ocr_work_reference_missing",
                        "severity": CheckSeverity.WARNING,
                        "title": "Работа не найдена в динамическом справочнике",
                        "details": f"{work_name} · в базе пока нет подтвержденной истории для сверки",
                        "payload": {
                            "work_code": work_code,
                            "work_name": work_name,
                            "comparison_source": "none",
                            "vehicle_type": repair.vehicle.vehicle_type.value,
                            "repair_mileage": repair.mileage,
                        },
                    }
                )
                seen_missing_keys.add(match_key)
            reference_payload["dynamic_work_reference"] = {
                "comparison_source": "none",
                "sample_lines": 0,
                "historical_sample_lines": 0,
                "operational_sample_lines": 0,
            }
            continue

        vehicle_matches = [row for row in matches if row.vehicle_id == repair.vehicle_id]
        service_matches = [row for row in matches if current_service_id is not None and row.service_id == current_service_id]
        if len(vehicle_matches) >= WORK_REFERENCE_VEHICLE_SAMPLE_THRESHOLD:
            selected_matches = vehicle_matches
            comparison_source = "same_vehicle"
        elif len(service_matches) >= WORK_REFERENCE_SERVICE_SAMPLE_THRESHOLD:
            selected_matches = service_matches
            comparison_source = "same_service"
        else:
            selected_matches = matches
            comparison_source = "vehicle_type"

        line_totals = [float(row.line_total) for row in selected_matches if row.line_total is not None]
        prices = [float(row.price) for row in selected_matches if row.price is not None]
        mileages = [int(row.mileage) for row in selected_matches if row.mileage is not None and int(row.mileage) > 0]
        historical_sample_lines = sum(
            1
            for row in matches
            if row.reason is not None and str(row.reason).startswith(HISTORICAL_IMPORT_REASON_PREFIX)
        )
        operational_sample_lines = len(matches) - historical_sample_lines

        reference_payload["dynamic_work_reference"] = {
            "comparison_source": comparison_source,
            "comparison_source_label": describe_work_reference_source(comparison_source),
            "sample_lines": len(selected_matches),
            "all_sample_lines": len(matches),
            "historical_sample_lines": historical_sample_lines,
            "operational_sample_lines": operational_sample_lines,
            "median_line_total": round(float(statistics.median(line_totals)), 2) if line_totals else None,
            "median_price": round(float(statistics.median(prices)), 2) if prices else None,
            "median_mileage": int(round(float(statistics.median(mileages)))) if mileages else None,
            "min_mileage": min(mileages) if mileages else None,
            "max_mileage": max(mileages) if mileages else None,
        }

        if len(selected_matches) < WORK_REFERENCE_MIN_SAMPLES:
            continue

        current_price = float(item["price"]) if item.get("price") is not None else None
        median_price = float(statistics.median(prices)) if prices else None
        if current_price is not None and median_price is not None and median_price > 0:
            price_ratio = round(current_price / median_price, 4)
            if (
                price_ratio >= WORK_REFERENCE_SUSPICIOUS_MULTIPLIER
                or price_ratio <= WORK_REFERENCE_SUSPICIOUS_LOWER_MULTIPLIER
            ):
                severity = CheckSeverity.SUSPICIOUS
            elif (
                price_ratio >= WORK_REFERENCE_WARNING_MULTIPLIER
                or price_ratio <= WORK_REFERENCE_WARNING_LOWER_MULTIPLIER
            ):
                severity = CheckSeverity.WARNING
            else:
                severity = None

            if severity is not None:
                checks.append(
                    {
                        "check_type": "ocr_work_reference_price_deviation",
                        "severity": severity,
                        "title": "Цена работы отклоняется от динамического справочника",
                        "details": (
                            f"{work_name} · цена {current_price:.2f} руб., медиана {median_price:.2f} руб. "
                            f"{describe_work_reference_source(comparison_source)}"
                        ),
                        "payload": {
                            "work_code": work_code,
                            "work_name": work_name,
                            "current_price": current_price,
                            "median_price": round(median_price, 2),
                            "price_ratio": price_ratio,
                            "comparison_source": comparison_source,
                            "comparison_source_label": describe_work_reference_source(comparison_source),
                            "sample_lines": len(selected_matches),
                            "all_sample_lines": len(matches),
                            "historical_sample_lines": historical_sample_lines,
                            "operational_sample_lines": operational_sample_lines,
                        },
                    }
                )

        if repair.mileage > 0 and len(mileages) >= WORK_REFERENCE_MIN_SAMPLES:
            min_mileage = min(mileages)
            max_mileage = max(mileages)
            median_mileage = int(round(float(statistics.median(mileages))))
            mileage_margin = max(
                WORK_REFERENCE_MIN_MILEAGE_MARGIN,
                int(round((max_mileage - min_mileage) * WORK_REFERENCE_MILEAGE_MARGIN_RATIO)),
            )
            if repair.mileage < (min_mileage - mileage_margin) or repair.mileage > (max_mileage + mileage_margin):
                checks.append(
                    {
                        "check_type": "ocr_work_reference_mileage_outlier",
                        "severity": CheckSeverity.WARNING,
                        "title": "Работа нетипична для текущего пробега",
                        "details": (
                            f"{work_name} · пробег {repair.mileage} км, "
                            f"наблюдаемый диапазон {min_mileage}-{max_mileage} км "
                            f"{describe_work_reference_source(comparison_source)}"
                        ),
                        "payload": {
                            "work_code": work_code,
                            "work_name": work_name,
                            "repair_mileage": repair.mileage,
                            "median_mileage": median_mileage,
                            "min_mileage": min_mileage,
                            "max_mileage": max_mileage,
                            "mileage_margin": mileage_margin,
                            "comparison_source": comparison_source,
                            "comparison_source_label": describe_work_reference_source(comparison_source),
                            "sample_lines": len(selected_matches),
                        },
                    }
                )

    return checks


def decode_pdf_literal(raw_text: bytes) -> bytes:
    return (
        raw_text.replace(b"\\(", b"(")
        .replace(b"\\)", b")")
        .replace(b"\\n", b"\n")
        .replace(b"\\r", b"\r")
        .replace(b"\\t", b"\t")
        .replace(b"\\\\", b"\\")
    )


def extract_pdf_stream_text(page) -> str:
    content = page.get_contents()
    if content is None:
        return ""

    contents = content if isinstance(content, list) else [content]
    literals = []
    for item in contents:
        raw = item.get_data()
        matches = re.findall(rb"\((?:\\.|[^\\)])*\)", raw)
        for match in matches:
            literals.append(decode_pdf_literal(match[1:-1]))

    decoded_variants = []
    for encoding in ("utf-8", "cp1251", "latin1"):
        try:
            decoded_variants.append("\n".join(part.decode(encoding, errors="ignore") for part in literals))
        except LookupError:
            continue

    scored_variants = sorted(
        decoded_variants,
        key=lambda item: (
            len(re.findall(r"[А-Яа-яA-Za-z0-9]", item)),
            len(item),
        ),
        reverse=True,
    )
    return scored_variants[0].strip() if scored_variants else ""


def extract_pdf_text(path: Path) -> str:
    reader = PdfReader(path.as_posix())
    chunks = []
    for page in reader.pages:
        page_text = (page.extract_text() or "").strip()
        stream_text = extract_pdf_stream_text(page)
        if score_text_quality(stream_text) > score_text_quality(page_text):
            chunks.append(stream_text)
        else:
            chunks.append(page_text)
    return "\n".join(filter(None, chunks)).strip()


def run_vision_ocr(image_paths: list[Path]) -> dict[str, str]:
    if not image_paths:
        return {}
    if not is_vision_ocr_available():
        raise RuntimeError("Apple Vision OCR is not available in the current environment")

    command = ["swift", VISION_OCR_SCRIPT.as_posix(), *[path.as_posix() for path in image_paths]]
    result = subprocess.run(
        command,
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "Vision OCR command failed")

    payload = json.loads(result.stdout)
    return {item["path"]: item["text"] for item in payload.get("results", [])}


def run_tesseract_ocr_with_modes(
    image_paths: list[Path],
    *,
    page_segmentation_modes: tuple[str, ...] | list[str],
) -> dict[str, str]:
    if not image_paths:
        return {}
    if not is_tesseract_ocr_available():
        raise RuntimeError("Tesseract OCR is not available in the current environment")

    payload: dict[str, str] = {}
    for image_path in image_paths:
        variants: list[str] = []
        last_error = "Tesseract OCR command failed"
        for psm in page_segmentation_modes:
            command = [
                TESSERACT_BINARY,
                image_path.as_posix(),
                "stdout",
                "-l",
                TESSERACT_LANGUAGE,
                "--psm",
                str(psm),
            ]
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                check=False,
            )
            if result.returncode != 0:
                last_error = result.stderr.strip() or last_error
                continue
            variants.append(result.stdout)
        if not variants:
            raise RuntimeError(last_error)
        payload[image_path.as_posix()] = select_best_tesseract_ocr_variant(variants)
    return payload


def run_tesseract_ocr(image_paths: list[Path]) -> dict[str, str]:
    return run_tesseract_ocr_with_modes(
        image_paths,
        page_segmentation_modes=tuple(TESSERACT_PAGE_SEGMENTATION_MODES),
    )


def run_ocr_backend(image_paths: list[Path]) -> tuple[dict[str, str], str]:
    backend = get_available_ocr_backend()
    if backend == "vision":
        return run_vision_ocr(image_paths), backend
    if backend == "tesseract":
        return run_tesseract_ocr(image_paths), backend
    raise RuntimeError("No supported OCR backend is available in the current environment")


def save_pillow_optimized_image(source_path: Path, output_path: Path) -> bool:
    if not is_pillow_available():
        return False

    try:
        with Image.open(source_path) as image:
            rgba_image = image.convert("RGBA")
            alpha_channel = rgba_image.getchannel("A")
            alpha_bbox = alpha_channel.getbbox()

            background = Image.new("RGBA", rgba_image.size, (255, 255, 255, 255))
            rgb_image = Image.alpha_composite(background, rgba_image).convert("RGB")

            grayscale_image = rgb_image.convert("L")
            diff_image = ImageChops.difference(grayscale_image, Image.new("L", grayscale_image.size, 255))
            content_bbox = diff_image.point(lambda value: 255 if value > 12 else 0).getbbox()

            if alpha_bbox is not None:
                content_bbox = alpha_bbox if content_bbox is None else (
                    min(alpha_bbox[0], content_bbox[0]),
                    min(alpha_bbox[1], content_bbox[1]),
                    max(alpha_bbox[2], content_bbox[2]),
                    max(alpha_bbox[3], content_bbox[3]),
                )

            if content_bbox is not None:
                left, top, right, bottom = content_bbox
                padding = max(24, int(max(rgb_image.size) * 0.03))
                crop_box = (
                    max(0, left - padding),
                    max(0, top - padding),
                    min(rgb_image.width, right + padding),
                    min(rgb_image.height, bottom + padding),
                )
                rgb_image = rgb_image.crop(crop_box)

            longest_side = max(rgb_image.size)
            if longest_side and longest_side < 2400:
                scale = 2400 / float(longest_side)
                resized_size = (
                    max(1, int(round(rgb_image.width * scale))),
                    max(1, int(round(rgb_image.height * scale))),
                )
                rgb_image = rgb_image.resize(resized_size, Image.Resampling.LANCZOS)

            rgb_image.save(output_path, format="JPEG", quality=95, optimize=True)
            return True
    except Exception:
        return False


def save_pdf_header_crop_for_ocr(source_path: Path, output_path: Path) -> bool:
    if not is_pillow_available():
        return False

    try:
        with Image.open(source_path) as image:
            rgb_image = image.convert("RGB")
            image_width, image_height = rgb_image.size
            if image_width <= 0 or image_height <= 0:
                return False

            top = max(0, int(image_height * 0.10))
            bottom = min(image_height, int(image_height * 0.26))
            if bottom <= top:
                return False

            cropped_image = rgb_image.crop((0, top, image_width, bottom))
            longest_side = max(cropped_image.size)
            if longest_side and longest_side < 2200:
                scale = 2200 / float(longest_side)
                resized_size = (
                    max(1, int(round(cropped_image.width * scale))),
                    max(1, int(round(cropped_image.height * scale))),
                )
                cropped_image = cropped_image.resize(resized_size, Image.Resampling.LANCZOS)

            cropped_image.save(output_path, format="JPEG", quality=95, optimize=True)
            return True
    except Exception:
        return False


def optimize_existing_image_for_ocr(path: Path) -> None:
    temporary_output_path = path.with_name(f"{path.stem}_optimized.jpg")
    if save_pillow_optimized_image(path, temporary_output_path):
        temporary_output_path.replace(path)


def preprocess_image_for_ocr(path: Path) -> tuple[tempfile.TemporaryDirectory, Path]:
    temp_dir = tempfile.TemporaryDirectory()
    processed_path = Path(temp_dir.name) / f"{path.stem}_ocr.jpg"
    if save_pillow_optimized_image(path, processed_path):
        return temp_dir, processed_path

    if is_sips_available():
        command = [
            SIPS_BINARY,
            "-s",
            "format",
            "jpeg",
            "-s",
            "formatOptions",
            "best",
            "-Z",
            "2400",
            path.as_posix(),
            "--out",
            processed_path.as_posix(),
        ]
        result = subprocess.run(command, capture_output=True, text=True, check=False)
        if result.returncode != 0:
            temp_dir.cleanup()
            raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "Failed to preprocess image for OCR")
        optimize_existing_image_for_ocr(processed_path)
        return temp_dir, processed_path

    passthrough_suffix = path.suffix if path.suffix else ".img"
    passthrough_path = Path(temp_dir.name) / f"{path.stem}_ocr{passthrough_suffix}"
    shutil.copy2(path, passthrough_path)
    return temp_dir, passthrough_path


def extract_image_text(path: Path) -> tuple[str, str]:
    temp_dir, processed_path = preprocess_image_for_ocr(path)
    try:
        ocr_results, backend = run_ocr_backend([processed_path])
        return select_best_text_variant(ocr_results.get(processed_path.as_posix(), "")), backend
    finally:
        temp_dir.cleanup()


def render_single_page_pdf_for_ocr(source_path: Path, output_path: Path) -> None:
    if is_pdftoppm_available():
        output_prefix = output_path.with_suffix("")
        command = [
            PDFTOPPM_BINARY,
            "-jpeg",
            "-r",
            "300",
            "-singlefile",
            source_path.as_posix(),
            output_prefix.as_posix(),
        ]
    elif is_sips_available():
        command = [
            SIPS_BINARY,
            "-s",
            "format",
            "jpeg",
            "-s",
            "formatOptions",
            "best",
            "-Z",
            "2400",
            source_path.as_posix(),
            "--out",
            output_path.as_posix(),
        ]
    else:
        raise RuntimeError("Failed to render PDF page for OCR: no supported PDF renderer is available")

    result = subprocess.run(command, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "Failed to render PDF page for OCR")
    optimize_existing_image_for_ocr(output_path)


def render_pdf_pages_for_ocr(path: Path, max_pages: int = 5) -> tuple[tempfile.TemporaryDirectory, list[Path]]:
    temp_dir = tempfile.TemporaryDirectory()
    image_paths: list[Path] = []
    reader = PdfReader(path.as_posix())
    page_count = max(1, min(len(reader.pages), max_pages))
    for page_index in range(page_count):
        single_page_pdf_path = Path(temp_dir.name) / f"ocr_page_{page_index + 1}.pdf"
        image_path = Path(temp_dir.name) / f"ocr_page_{page_index + 1}.jpg"
        writer = PdfWriter()
        writer.add_page(reader.pages[page_index])
        with single_page_pdf_path.open("wb") as output_stream:
            writer.write(output_stream)
        try:
            render_single_page_pdf_for_ocr(single_page_pdf_path, image_path)
        except RuntimeError:
            temp_dir.cleanup()
            raise
        image_paths.append(image_path)

    return temp_dir, image_paths


def render_pdf_pages_for_raw_pdftoppm_ocr(path: Path, max_pages: int = 5) -> tuple[tempfile.TemporaryDirectory, list[Path]]:
    if not is_pdftoppm_available():
        raise RuntimeError("Failed to render PDF page for AXB OCR fallback: pdftoppm is not available")

    temp_dir = tempfile.TemporaryDirectory()
    page_count = max(1, min(len(PdfReader(path.as_posix()).pages), max_pages))
    output_prefix = Path(temp_dir.name) / "ocr_raw_page"
    command = [
        PDFTOPPM_BINARY,
        "-jpeg",
        "-r",
        "300",
        "-f",
        "1",
        "-l",
        str(page_count),
        path.as_posix(),
        output_prefix.as_posix(),
    ]
    result = subprocess.run(command, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        temp_dir.cleanup()
        raise RuntimeError(result.stderr.strip() or result.stdout.strip() or "Failed to render AXB OCR fallback pages")

    image_paths = sorted(Path(temp_dir.name).glob("ocr_raw_page-*.jpg"))
    if len(image_paths) != page_count:
        temp_dir.cleanup()
        raise RuntimeError("AXB OCR fallback did not render the expected number of pages")
    return temp_dir, image_paths


def extract_scanned_pdf_text(path: Path) -> tuple[str, str]:
    temp_dir, image_paths = render_pdf_pages_for_ocr(path)
    try:
        ocr_results, backend = run_ocr_backend(image_paths)
        chunks = [select_best_text_variant(ocr_results.get(image_path.as_posix(), "")) for image_path in image_paths]

        header_crop_path = Path(temp_dir.name) / "ocr_page_1_header.jpg"
        if image_paths and save_pdf_header_crop_for_ocr(image_paths[0], header_crop_path):
            header_results, _ = run_ocr_backend([header_crop_path])
            header_text = select_best_text_variant(header_results.get(header_crop_path.as_posix(), ""))
            if header_text and re.search(r"(?:гос\.?\s*номер|vin|пробег|tc\s*:|автомобиль)", header_text, re.IGNORECASE):
                chunks.insert(0, header_text)

        return "\n".join(filter(None, chunks)).strip(), backend
    finally:
        temp_dir.cleanup()


def extract_axb_raw_scanned_pdf_text(path: Path) -> str:
    temp_dir, image_paths = render_pdf_pages_for_raw_pdftoppm_ocr(path)
    try:
        ocr_results = run_tesseract_ocr_with_modes(image_paths, page_segmentation_modes=("4",))
        chunks = [select_best_text_variant(ocr_results.get(image_path.as_posix(), "")) for image_path in image_paths]

        header_crop_path = Path(temp_dir.name) / "ocr_raw_page_1_header.jpg"
        if image_paths and save_pdf_header_crop_for_ocr(image_paths[0], header_crop_path):
            header_results = run_tesseract_ocr_with_modes([header_crop_path], page_segmentation_modes=("4",))
            header_text = select_best_text_variant(header_results.get(header_crop_path.as_posix(), ""))
            if header_text and re.search(r"(?:гос\.?\s*номер|vin|пробег|tc\s*:|автомобиль)", header_text, re.IGNORECASE):
                chunks.insert(0, header_text)

        return "\n".join(filter(None, chunks)).strip()
    finally:
        temp_dir.cleanup()


def format_spreadsheet_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour == value.minute == value.second == value.microsecond == 0:
            return value.strftime("%d.%m.%Y")
        return value.strftime("%d.%m.%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\xa0", " ").strip()


def extract_spreadsheet_text(path: Path) -> str:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        chunks: list[str] = []
        multiple_sheets = len(workbook.worksheets) > 1
        for sheet in workbook.worksheets:
            if multiple_sheets:
                chunks.append(f"Лист: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [format_spreadsheet_cell_value(value) for value in row]
                values = [value for value in values if value]
                if values:
                    chunks.append(" ".join(values))
        return "\n".join(chunks).strip()
    finally:
        workbook.close()


def extract_document_text(path: Path, source_type: str) -> tuple[str, str, Optional[str]]:
    if source_type == "image":
        if get_available_ocr_backend() is None:
            return "", "manual_review", "image_ocr_unavailable"
        text, backend = extract_image_text(path)
        return text, f"image_{backend}_ocr", None if text else "image_text_not_found"

    if source_type == "xlsx":
        text = extract_spreadsheet_text(path)
        return text, "xlsx_text", None if text else "xlsx_text_not_found"

    text = select_best_text_variant(extract_pdf_text(path))
    extracted_from = "pdf_text"
    failure_reason = None

    if score_text_quality(text)[0] >= 2 or score_text_quality(text)[1] >= 6:
        return text, extracted_from, None

    if get_available_ocr_backend() is None:
        return text, extracted_from, "pdf_ocr_unavailable" if not text else None

    try:
        scanned_text, backend = extract_scanned_pdf_text(path)
    except RuntimeError:
        return text, extracted_from, "pdf_renderer_unavailable" if not text else None
    if score_text_quality(scanned_text) > score_text_quality(text):
        return scanned_text, f"pdf_{backend}_ocr", None if scanned_text else "pdf_text_not_found"

    if not text and not scanned_text:
        failure_reason = "pdf_text_not_found"
    return text or scanned_text, extracted_from if text else f"pdf_{backend}_ocr", failure_reason


def get_line_items_total(items: list[dict[str, object]]) -> float:
    return round(sum(float(item.get("line_total", 0) or 0) for item in items), 2)


def should_retry_axb_raw_tesseract(parsed: dict[str, object]) -> bool:
    extracted_fields = parsed.get("extracted_fields", {})
    extracted_items = parsed.get("extracted_items", {})
    works = extracted_items.get("works", []) if isinstance(extracted_items, dict) else []
    parts = extracted_items.get("parts", []) if isinstance(extracted_items, dict) else []

    if not works and not parts:
        return True

    work_total = extracted_fields.get("work_total") if isinstance(extracted_fields, dict) else None
    parts_total = extracted_fields.get("parts_total") if isinstance(extracted_fields, dict) else None
    works_sum = get_line_items_total(works)
    parts_sum = get_line_items_total(parts)

    if work_total not in {None, ""} and works and not amounts_match(works_sum, float(work_total), tolerance=0.2):
        return True
    if parts_total not in {None, ""} and parts and not amounts_match(parts_sum, float(parts_total), tolerance=0.2):
        return True
    if parts_total not in {None, ""} and float(parts_total) > 0 and len(parts) <= 1:
        return True
    if work_total not in {None, ""} and float(work_total) > 0 and not works:
        return True
    return False


def score_axb_parsed_document(parsed: dict[str, object]) -> int:
    extracted_fields = parsed.get("extracted_fields", {})
    extracted_items = parsed.get("extracted_items", {})
    confidence_map = parsed.get("confidence_map", {})
    manual_review_reasons = parsed.get("manual_review_reasons", [])
    works = extracted_items.get("works", []) if isinstance(extracted_items, dict) else []
    parts = extracted_items.get("parts", []) if isinstance(extracted_items, dict) else []
    work_total = extracted_fields.get("work_total") if isinstance(extracted_fields, dict) else None
    parts_total = extracted_fields.get("parts_total") if isinstance(extracted_fields, dict) else None
    works_sum = get_line_items_total(works)
    parts_sum = get_line_items_total(parts)

    score = len(works) * 18 + len(parts) * 30 + len(confidence_map) * 5
    if works:
        score += 100
    if parts:
        score += 120

    if work_total not in {None, ""} and works:
        score += 260 if amounts_match(works_sum, float(work_total), tolerance=0.2) else -180
    if parts_total not in {None, ""} and parts:
        score += 280 if amounts_match(parts_sum, float(parts_total), tolerance=0.2) else -220

    if parts_total not in {None, ""} and float(parts_total) > 0 and len(parts) <= 1:
        score -= 260
    if work_total not in {None, ""} and float(work_total) > 0 and not works:
        score -= 240

    score -= len(manual_review_reasons) * 12
    return score


def maybe_apply_axb_raw_tesseract_fallback(
    path: Path,
    *,
    text: str,
    extracted_from: str,
    profile_scope: str | None,
    parsed: dict[str, object],
    db: Session | None = None,
) -> tuple[str, str, dict[str, object]]:
    if normalize_ocr_rule_code(profile_scope) != "axb":
        return text, extracted_from, parsed
    if extracted_from != "pdf_tesseract_ocr":
        return text, extracted_from, parsed
    if not should_retry_axb_raw_tesseract(parsed):
        return text, extracted_from, parsed

    try:
        fallback_text = extract_axb_raw_scanned_pdf_text(path)
    except RuntimeError:
        logger.info("axb_raw_tesseract_fallback_unavailable", extra={"source_path": path.as_posix()})
        return text, extracted_from, parsed

    if not fallback_text:
        return text, extracted_from, parsed

    fallback_parsed = parse_document_text(fallback_text, db=db, profile_scope=profile_scope)
    if score_axb_parsed_document(fallback_parsed) <= score_axb_parsed_document(parsed):
        return text, extracted_from, parsed

    fallback_notes = list(fallback_parsed.get("normalization_notes", []))
    fallback_notes.append("Для AXB применён резервный OCR-проход без постобработки изображений.")
    fallback_parsed["normalization_notes"] = fallback_notes
    logger.info("axb_raw_tesseract_fallback_applied", extra={"source_path": path.as_posix()})
    return fallback_text, "pdf_tesseract_ocr_axb_raw_fallback", fallback_parsed


def parse_document_text(text: str, db: Session | None = None, *, profile_scope: str | None = None) -> dict[str, object]:
    text = select_best_text_variant(text)
    normalized_profile_scope = normalize_ocr_rule_code(profile_scope)
    if not normalized_profile_scope:
        inferred_profile_scope = infer_builtin_profile_scope_from_text(text)
        normalized_profile_scope = normalize_ocr_rule_code(inferred_profile_scope)
    header_text = extract_header_text(text)
    vehicle_section_text = extract_vehicle_section_text(text)
    field_search_texts: list[str] = []
    for candidate in (vehicle_section_text, header_text):
        normalized_candidate = normalize_text(candidate)
        if normalized_candidate and normalized_candidate not in field_search_texts:
            field_search_texts.append(normalized_candidate)
    extracted_fields = {}
    confidence_map = {}
    manual_review_reasons = []
    normalization_notes = []
    normalization_notes.extend(detect_document_flags(text))
    extracted_items = extract_line_items(text)
    extracted_items, removed_noise_work_count = sanitize_extracted_items(extracted_items)
    if removed_noise_work_count:
        normalization_notes.append(f"noise_work_items_removed:{removed_noise_work_count}")
    rule_map = (
        group_ocr_rules_by_field(
            load_active_ocr_rules(db, profile_scope=normalized_profile_scope),
            profile_scope=normalized_profile_scope,
        )
        if db is not None
        else {}
    )
    section_plate_number, section_vin, section_mileage = extract_vehicle_identifiers_from_section(text)
    section_chassis_number = find_chassis_candidate(vehicle_section_text) or find_chassis_candidate(header_text)

    order_number, order_number_confidence, _ = extract_header_field(
        header_text,
        target_field="order_number",
        fallback_patterns=ORDER_PATTERNS,
        fallback_parser="raw",
        fallback_confidence=0.74,
        rule_map=rule_map,
    )
    if isinstance(order_number, str) and is_plausible_order_number(order_number):
        extracted_fields["order_number"] = order_number
        confidence_map["order_number"] = float(order_number_confidence or 0.74)
    else:
        manual_review_reasons.append("order_number_missing")

    repair_date, repair_date_confidence, repair_date_invalid = extract_header_field(
        header_text,
        target_field="repair_date",
        fallback_patterns=DATE_PATTERNS,
        fallback_parser="date",
        fallback_confidence=0.7,
        rule_map=rule_map,
    )
    if isinstance(repair_date, str) and repair_date:
        extracted_fields["repair_date"] = repair_date
        confidence_map["repair_date"] = float(repair_date_confidence or 0.7)
    elif repair_date_invalid:
        manual_review_reasons.append("repair_date_invalid")
    else:
        manual_review_reasons.append("repair_date_missing")

    if section_mileage is not None:
        extracted_fields["mileage"] = section_mileage
        confidence_map["mileage"] = 0.9
    else:
        mileage_found = False
        explicit_missing_mileage = any(has_explicit_missing_mileage(field_text) for field_text in field_search_texts)
        logistics_blank_mileage = normalized_profile_scope == "logistics" and has_logistics_blank_mileage_field(text)
        suppress_missing_mileage_review = explicit_missing_mileage or (
            normalized_profile_scope == "logistics" and is_logistics_trailer_vehicle_context(text)
        ) or logistics_blank_mileage
        if not explicit_missing_mileage:
            for field_text in field_search_texts:
                mileage, mileage_confidence, _ = extract_header_field(
                    field_text,
                    target_field="mileage",
                    fallback_patterns=MILEAGE_PATTERNS,
                    fallback_parser="digits_int",
                    fallback_confidence=0.82,
                    rule_map=rule_map,
                )
                if isinstance(mileage, int):
                    extracted_fields["mileage"] = mileage
                    confidence_map["mileage"] = float(mileage_confidence or 0.82)
                    mileage_found = True
                    break
        if not mileage_found and not suppress_missing_mileage_review:
            manual_review_reasons.append("mileage_missing")

    if section_plate_number:
        extracted_fields["plate_number"] = section_plate_number
        confidence_map["plate_number"] = 0.9
    else:
        for field_text in field_search_texts:
            plate_number, plate_number_confidence, _ = extract_header_field(
                field_text,
                target_field="plate_number",
                fallback_patterns=PLATE_PATTERNS,
                fallback_parser="raw",
                fallback_confidence=0.77,
                rule_map=rule_map,
            )
            if isinstance(plate_number, str) and plate_number:
                normalized_plate = find_plate_candidate(plate_number) or normalize_identifier_token(plate_number)
                if normalized_plate:
                    extracted_fields["plate_number"] = normalized_plate
                    confidence_map["plate_number"] = float(plate_number_confidence or 0.77)
                    break

    if section_vin:
        extracted_fields["vin"] = section_vin
        confidence_map["vin"] = 0.92
    else:
        for field_text in field_search_texts:
            vin, vin_confidence, _ = extract_header_field(
                field_text,
                target_field="vin",
                fallback_patterns=VIN_PATTERNS,
                fallback_parser="raw",
                fallback_confidence=0.88,
                rule_map=rule_map,
            )
            if isinstance(vin, str) and vin:
                normalized_vin = find_vin_candidate(vin) or normalize_identifier_token(vin)
                if normalized_vin:
                    extracted_fields["vin"] = normalized_vin
                    confidence_map["vin"] = float(vin_confidence or 0.88)
                    break

    if section_chassis_number:
        extracted_fields["chassis_number"] = section_chassis_number
        confidence_map["chassis_number"] = 0.78

    resolved_service_match = find_service_name_in_text(text, db=db) if db is not None else find_service_name_in_text(text)
    service_name, service_name_confidence, _ = extract_header_field(
        text,
        target_field="service_name",
        fallback_patterns=SERVICE_PATTERNS,
        fallback_parser="raw",
        fallback_confidence=0.58,
        rule_map=rule_map,
    )
    service_candidate = normalize_service_candidate(service_name) if isinstance(service_name, str) else None
    labeled_service_candidate = extract_service_candidate_from_text(text)
    if labeled_service_candidate:
        labeled_key = normalize_service_key(labeled_service_candidate)
        current_key = normalize_service_key(service_candidate) if service_candidate else None
        if (
            service_candidate is None
            or not current_key
            or (labeled_key and current_key and labeled_key.startswith(current_key))
            or len(labeled_service_candidate) > len(service_candidate) + 4
        ):
            service_candidate = labeled_service_candidate

    if resolved_service_match is not None:
        extracted_fields["service_name"] = resolved_service_match[0]
        confidence_map["service_name"] = 0.92
        normalization_notes.append(f"Сервис распознан по тексту документа: {resolved_service_match[1]}")
    elif service_candidate:
        if is_service_name_suspicious(service_candidate):
            manual_review_reasons.append("service_name_suspicious")
        else:
            extracted_fields["service_name"] = service_candidate[:120]
            confidence_map["service_name"] = float(service_name_confidence or 0.58)

    reason = extract_reason_from_text(text)
    if reason:
        extracted_fields["reason"] = reason
        confidence_map["reason"] = 0.84
    employee_comment = extract_recommendations_from_text(text)
    if employee_comment:
        extracted_fields["employee_comment"] = employee_comment
        confidence_map["employee_comment"] = 0.8
    service_not_done = extract_not_done_items_from_text(text)

    if normalized_profile_scope == "leader_trak" and "order_number" not in extracted_fields:
        leader_trak_order_number = extract_leader_trak_order_number(header_text or text)
        if leader_trak_order_number:
            extracted_fields["order_number"] = leader_trak_order_number
            confidence_map["order_number"] = max(confidence_map.get("order_number", 0.0), 0.9)
            remove_manual_review_reason(manual_review_reasons, "order_number_missing")

    if normalized_profile_scope == "ets_act":
        ets_scanned_fields = extract_ets_act_scanned_header_fields(text)
        for field_name, confidence in (
            ("order_number", 0.9),
            ("plate_number", 0.88),
            ("vin", 0.9),
            ("mileage", 0.88),
        ):
            field_value = ets_scanned_fields.get(field_name)
            if field_value is None or field_name in extracted_fields:
                continue
            extracted_fields[field_name] = field_value
            confidence_map[field_name] = max(confidence_map.get(field_name, 0.0), confidence)
            if field_name == "order_number":
                remove_manual_review_reason(manual_review_reasons, "order_number_missing")
            elif field_name == "mileage":
                remove_manual_review_reason(manual_review_reasons, "mileage_missing")

    if normalized_profile_scope == "klever_trak":
        klever_match = KLEVER_TRAK_VEHICLE_ROW_PATTERN.search(text)
        if klever_match is not None:
            mileage = parse_mileage_candidate(klever_match.group("mileage"))
            if mileage is not None:
                extracted_fields["mileage"] = mileage
                confidence_map["mileage"] = max(confidence_map.get("mileage", 0.0), 0.9)
                remove_manual_review_reason(manual_review_reasons, "mileage_missing")

    for field_name, patterns in TOTAL_PATTERNS.items():
        amount, amount_confidence, _ = extract_header_field(
            text,
            target_field=field_name,
            fallback_patterns=patterns,
            fallback_parser="amount",
            fallback_confidence=0.8 if field_name == "grand_total" else 0.72,
            rule_map=rule_map,
        )
        if not isinstance(amount, (int, float)):
            continue
        extracted_fields[field_name] = float(amount)
        confidence_map[field_name] = float(amount_confidence or (0.8 if field_name == "grand_total" else 0.72))

    apply_profile_specific_total_fallbacks(
        text,
        profile_scope=normalized_profile_scope,
        extracted_fields=extracted_fields,
        confidence_map=confidence_map,
        normalization_notes=normalization_notes,
    )

    extracted_items = apply_profile_specific_item_fallbacks(
        text,
        profile_scope=normalized_profile_scope,
        extracted_items=extracted_items,
        extracted_fields=extracted_fields,
        normalization_notes=normalization_notes,
    )
    extracted_items, removed_post_fallback_noise_work_count = sanitize_extracted_items(extracted_items)
    if removed_post_fallback_noise_work_count:
        normalization_notes.append(f"noise_work_items_removed_after_fallback:{removed_post_fallback_noise_work_count}")

    if normalized_profile_scope == "leader_trak" and is_leader_trak_invoice_only_document(text):
        if (extracted_items.get("works") or extracted_items.get("parts")) and not has_meaningful_leader_trak_items(extracted_items):
            extracted_items = {"works": [], "parts": []}
            normalization_notes.append("leader_trak_invoice_only_items_suppressed")

    normalization_notes.extend(
        reconcile_header_totals_with_line_items(
            extracted_fields=extracted_fields,
            extracted_items=extracted_items,
            confidence_map=confidence_map,
        )
    )

    if normalized_profile_scope == "gruzovye_rezervy" and is_gruzovye_rezervy_invoice_only_document(text):
        removed_reasons = []
        if "order_number_missing" in manual_review_reasons:
            remove_manual_review_reason(manual_review_reasons, "order_number_missing")
            removed_reasons.append("order_number_missing")
        if "mileage_missing" in manual_review_reasons:
            remove_manual_review_reason(manual_review_reasons, "mileage_missing")
            removed_reasons.append("mileage_missing")
        if removed_reasons:
            normalization_notes.append(
                "gruzovye_rezervy_invoice_only_review_suppressed:" + ",".join(removed_reasons)
            )

    if normalized_profile_scope == "leader_trak" and is_leader_trak_invoice_only_document(text):
        removed_reasons = []
        if "mileage_missing" in manual_review_reasons:
            remove_manual_review_reason(manual_review_reasons, "mileage_missing")
            removed_reasons.append("mileage_missing")
        if removed_reasons:
            normalization_notes.append(
                "leader_trak_invoice_only_review_suppressed:" + ",".join(removed_reasons)
            )

    if db is not None:
        enrich_vehicle_fields_from_registry(
            db,
            extracted_fields=extracted_fields,
            confidence_map=confidence_map,
            normalization_notes=normalization_notes,
        )

    return {
        "extracted_fields": extracted_fields,
        "extracted_items": extracted_items,
        "confidence_map": confidence_map,
        "manual_review_reasons": manual_review_reasons,
        "normalization_notes": normalization_notes,
        "service_not_done": service_not_done,
    }


def load_document_for_processing(db: Session, document_id: int) -> Optional[Document]:
    stmt = (
        select(Document)
        .options(
            joinedload(Document.repair).joinedload(Repair.vehicle),
            joinedload(Document.versions),
        )
        .where(Document.id == document_id)
    )
    return db.execute(stmt).unique().scalar_one_or_none()


def average_confidence(confidence_map: dict[str, float]) -> Optional[float]:
    if not confidence_map:
        return None
    return round(sum(confidence_map.values()) / len(confidence_map), 2)


def resolve_service(db: Session, service_name: str) -> Service:
    service = resolve_service_by_name(db, service_name)
    if service is None:
        raise ValueError(f"Unknown service: {service_name}")
    return service


def build_manual_review_check(
    reason: str,
    *,
    extracted_fields: dict[str, object],
) -> dict[str, object]:
    service_name = str(extracted_fields.get("service_name")).strip() if extracted_fields.get("service_name") else None
    plate_number = str(extracted_fields.get("plate_number")).strip() if extracted_fields.get("plate_number") else None
    vin = str(extracted_fields.get("vin")).strip() if extracted_fields.get("vin") else None

    if reason == "vehicle_not_found":
        identifiers = []
        if plate_number:
            identifiers.append(f"госномер {plate_number}")
        if vin:
            identifiers.append(f"VIN {vin}")
        details = "Техника из документа не найдена в базе техники."
        if identifiers:
            details = f"{details} Распознаны: {', '.join(identifiers)}."
        return {
            "check_type": "ocr_vehicle_not_found",
            "severity": CheckSeverity.WARNING,
            "title": "Техника не найдена в базе",
            "details": details,
            "payload": {
                "reason": reason,
                "plate_number": plate_number,
                "vin": vin,
            },
        }

    if reason == "vehicle_missing":
        return {
            "check_type": "ocr_vehicle_missing",
            "severity": CheckSeverity.WARNING,
            "title": "Не удалось определить технику",
            "details": "В документе не удалось надёжно определить технику. Нужна ручная привязка.",
            "payload": {"reason": reason},
        }

    if reason == "service_not_found":
        details = "Сервис из документа не найден в справочнике сервисов."
        if service_name:
            details = f"{details} Распознанное значение: {service_name}."
        return {
            "check_type": "ocr_service_not_found",
            "severity": CheckSeverity.WARNING,
            "title": "Сервис не найден в справочнике",
            "details": details,
            "payload": {
                "reason": reason,
                "service_name": service_name,
            },
        }

    if reason == "service_name_missing":
        return {
            "check_type": "ocr_service_missing",
            "severity": CheckSeverity.WARNING,
            "title": "Не удалось определить сервис",
            "details": "В документе не удалось надёжно определить сервис. Нужна ручная проверка.",
            "payload": {"reason": reason},
        }

    if reason == "text_not_found":
        return {
            "check_type": "ocr_text_not_found",
            "severity": CheckSeverity.WARNING,
            "title": "Не удалось извлечь текст из документа",
            "details": "Документ сохранён, но автоматическое распознавание не извлекло текст для проверки.",
            "payload": {"reason": reason},
        }

    if reason == "image_ocr_unavailable":
        return {
            "check_type": "ocr_image_backend_unavailable",
            "severity": CheckSeverity.WARNING,
            "title": "OCR для изображений недоступен",
            "details": "В текущем окружении не найден поддерживаемый OCR backend для изображений.",
            "payload": {"reason": reason},
        }

    if reason == "pdf_ocr_unavailable":
        return {
            "check_type": "ocr_pdf_backend_unavailable",
            "severity": CheckSeverity.WARNING,
            "title": "OCR для PDF-сканов недоступен",
            "details": "В текущем окружении не найден поддерживаемый OCR backend для PDF-сканов.",
            "payload": {"reason": reason},
        }

    if reason == "pdf_renderer_unavailable":
        return {
            "check_type": "ocr_pdf_renderer_unavailable",
            "severity": CheckSeverity.WARNING,
            "title": "Рендер PDF для OCR недоступен",
            "details": "В текущем окружении не найден поддерживаемый renderer PDF-страниц для OCR.",
            "payload": {"reason": reason},
        }

    return {
        "check_type": f"ocr_{reason}",
        "severity": CheckSeverity.WARNING,
        "title": "Нужна ручная проверка OCR",
        "details": reason,
        "payload": {"reason": reason},
    }


def replace_ocr_checks(db: Session, repair_id: int, checks: list[dict[str, object]]) -> None:
    db.execute(delete(RepairCheck).where(RepairCheck.repair_id == repair_id, RepairCheck.check_type.like("ocr_%")))
    for item in checks:
        db.add(
            RepairCheck(
                repair_id=repair_id,
                check_type=str(item["check_type"]),
                severity=item["severity"],
                title=str(item["title"]),
                details=item.get("details"),
                calculation_payload=item.get("payload"),
            )
        )


def replace_repair_lines(
    db: Session,
    repair: Repair,
    works_payload: list[dict[str, object]],
    parts_payload: list[dict[str, object]],
) -> None:
    db.execute(delete(RepairWork).where(RepairWork.repair_id == repair.id))
    db.execute(delete(RepairPart).where(RepairPart.repair_id == repair.id))

    for item in works_payload:
        normalized_unit_name = normalize_unit_name(str(item.get("unit_name")) if item.get("unit_name") else None)
        reference_payload = item.get("reference_payload")
        if not isinstance(reference_payload, dict):
            reference_payload = {}
        reference_payload.update(
            {
                "source": "ocr",
                "unit_name": normalized_unit_name,
                "normalized": True,
            }
        )
        db.add(
            RepairWork(
                repair_id=repair.id,
                work_code=str(item["work_code"]) if item.get("work_code") else None,
                work_name=str(item["work_name"]),
                quantity=float(item["quantity"]),
                actual_hours=float(item["quantity"]) if normalized_unit_name in {"нч", "ч"} else None,
                standard_hours=float(item["standard_hours"]) if item.get("standard_hours") is not None else None,
                price=float(item["price"]),
                line_total=float(item["line_total"]),
                status=CatalogStatus.PRELIMINARY,
                reference_payload=reference_payload,
            )
        )

    for item in parts_payload:
        normalized_unit_name = normalize_unit_name(str(item["unit_name"]) if item.get("unit_name") else None)
        db.add(
            RepairPart(
                repair_id=repair.id,
                article=normalize_article_value(str(item["article"])) if item.get("article") else None,
                part_name=str(item["part_name"]),
                quantity=float(item["quantity"]),
                unit_name=normalized_unit_name,
                price=float(item["price"]),
                line_total=float(item["line_total"]),
                status=CatalogStatus.PRELIMINARY,
            )
        )


def process_document(db: Session, document_id: int, *, job_id: int | None = None) -> ProcessingResult:
    initial_document = load_document_for_processing(db, document_id)
    if initial_document is None or initial_document.repair is None:
        raise ValueError("Document not found or repair relation is incomplete")

    storage_path = get_storage_path(initial_document.storage_key)
    if job_id is None:
        job = ImportJob(
            document_id=initial_document.id,
            import_type="document_ocr",
            source_filename=initial_document.original_filename,
            status=ImportStatus.PROCESSING,
            summary={"document_id": initial_document.id, "stage": "started"},
            attempts=1,
            started_at=datetime.now().astimezone(),
            finished_at=None,
        )
        db.add(job)
        db.flush()
        job_id = job.id
        db.commit()
    else:
        job = db.get(ImportJob, job_id)
        if job is None:
            raise ValueError(f"Import job {job_id} not found")

    try:
        document = load_document_for_processing(db, document_id)
        job = db.get(ImportJob, job_id)
        if document is None or document.repair is None or job is None:
            raise ValueError("Document processing context could not be reloaded")

        logger.info("document_processing_started", extra={"document_id": document.id, "job_id": job.id})

        if not storage_path.exists():
            raise FileNotFoundError(f"Source document file not found: {storage_path}")

        if document.kind in {DocumentKind.ATTACHMENT, DocumentKind.CONFIRMATION}:
            document.status = DocumentStatus.CONFIRMED
            document.review_queue_priority = 0
            document.ocr_confidence = None

            version_number = max([version.version_number for version in document.versions], default=0) + 1
            parsed_payload = {
                "processor": "document_storage_only_v1",
                "document_kind": document.kind.value,
                "ocr_skipped": True,
            }
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=version_number,
                    storage_key=document.storage_key,
                    parsed_payload=parsed_payload,
                    field_confidence_map={},
                    change_summary="Stored without OCR",
                )
            )
            mark_job_completed(
                db,
                job,
                status=ImportStatus.COMPLETED,
                summary={
                    "document_id": document.id,
                    "document_kind": document.kind.value,
                    "document_status": document.status.value,
                    "ocr_skipped": True,
                },
            )
            db.commit()

            message = "Document stored without OCR"
        else:
            logger.info("document_processing_extract_text", extra={"document_id": document.id, "job_id": job.id})
            text, extracted_from, extraction_failure_reason = extract_document_text(storage_path, document.source_type)
            profile_selection = select_ocr_profile_scope(db, document, text) if text else OcrProfileSelection(
                profile_scope="default",
                source="default",
                reason="Текст не извлечён, использован default",
            )
            logger.info(
                "document_processing_parse",
                extra={"document_id": document.id, "job_id": job.id, "profile_scope": profile_selection.profile_scope},
            )
            parsed = parse_document_text(text, db=db, profile_scope=profile_selection.profile_scope) if text else {
                "extracted_fields": {},
                "extracted_items": {"works": [], "parts": []},
                "confidence_map": {},
                "manual_review_reasons": [extraction_failure_reason or "text_not_found"],
                "normalization_notes": [],
            }
            if text:
                text, extracted_from, parsed = maybe_apply_axb_raw_tesseract_fallback(
                    storage_path,
                    text=text,
                    extracted_from=extracted_from,
                    profile_scope=profile_selection.profile_scope,
                    parsed=parsed,
                    db=db,
                )

            extracted_fields = parsed["extracted_fields"]
            extracted_items = parsed["extracted_items"]
            confidence_map = parsed["confidence_map"]
            manual_review_reasons = parsed["manual_review_reasons"]
            normalization_notes = parsed.get("normalization_notes", [])
            apply_document_metadata_fallbacks(
                document,
                extracted_fields=extracted_fields,
                confidence_map=confidence_map,
                manual_review_reasons=manual_review_reasons,
                normalization_notes=normalization_notes,
            )
            repair = document.repair
            auto_link_repair_vehicle_from_registry(
                db,
                repair,
                extracted_fields=extracted_fields,
                normalization_notes=normalization_notes,
            )
            if repair.vehicle is not None and repair.vehicle.external_id == PLACEHOLDER_VEHICLE_EXTERNAL_ID:
                created_vehicle = auto_create_repair_vehicle_from_document(
                    repair,
                    document,
                    extracted_fields=extracted_fields,
                    text=text,
                    normalization_notes=normalization_notes,
                )
                if created_vehicle:
                    db.add(repair.vehicle)
                    db.flush()
                    repair.vehicle_id = repair.vehicle.id
            enrich_vehicle_fields_from_repair(
                repair,
                extracted_fields=extracted_fields,
                confidence_map=confidence_map,
                normalization_notes=normalization_notes,
            )
            enrich_vehicle_fields_from_registry(
                db,
                extracted_fields=extracted_fields,
                confidence_map=confidence_map,
                normalization_notes=normalization_notes,
            )
            if repair.vehicle is not None and repair.vehicle.external_id == PLACEHOLDER_VEHICLE_EXTERNAL_ID:
                if extracted_fields.get("plate_number") or extracted_fields.get("vin"):
                    add_manual_review_reason(manual_review_reasons, "vehicle_not_found")
                    normalization_notes.append("Техника из документа не найдена в базе и требует ручной привязки.")
                else:
                    add_manual_review_reason(manual_review_reasons, "vehicle_missing")
            labor_norm_applicability = assess_labor_norm_applicability(db, repair.vehicle)
            labor_norm_notes, labor_norm_summary = enrich_work_payloads_with_labor_norms(
                db,
                extracted_items["works"],
                labor_norm_applicability,
            )
            normalization_notes.extend(labor_norm_notes)
            checks = []

            if "order_number" in extracted_fields:
                repair.order_number = str(extracted_fields["order_number"])
            if "repair_date" in extracted_fields:
                parsed_repair_date = parse_date_value(str(extracted_fields["repair_date"]).replace("-", "."))
                if parsed_repair_date is not None:
                    repair.repair_date = parsed_repair_date
            if "mileage" in extracted_fields:
                repair.mileage = int(extracted_fields["mileage"])
            if "reason" in extracted_fields:
                repair.reason = str(extracted_fields["reason"])
            if "employee_comment" in extracted_fields:
                repair.employee_comment = str(extracted_fields["employee_comment"])
            repair.work_total = float(extracted_fields["work_total"]) if "work_total" in extracted_fields else 0.0
            repair.parts_total = float(extracted_fields["parts_total"]) if "parts_total" in extracted_fields else 0.0
            repair.vat_total = float(extracted_fields["vat_total"]) if "vat_total" in extracted_fields else 0.0
            repair.grand_total = float(extracted_fields["grand_total"]) if "grand_total" in extracted_fields else 0.0
            if "service_name" in extracted_fields:
                service = resolve_service_by_name(db, str(extracted_fields["service_name"]))
                if service is not None:
                    extracted_fields["service_name"] = service.name
                    repair.service_id = service.id
                    remove_manual_review_reason(manual_review_reasons, "service_not_found")
                else:
                    add_manual_review_reason(manual_review_reasons, "service_not_found")
                    normalization_notes.append(
                        f"Сервис из документа не найден в справочнике: {extracted_fields['service_name']}"
                    )
                remove_manual_review_reason(manual_review_reasons, "service_name_missing")
            else:
                add_manual_review_reason(manual_review_reasons, "service_name_missing")

            logger.info("document_processing_match_and_checks", extra={"document_id": document.id, "job_id": job.id})
            checks.extend(build_dynamic_work_reference_checks(db, repair, extracted_items["works"]))
            replace_repair_lines(
                db,
                repair,
                works_payload=extracted_items["works"],
                parts_payload=extracted_items["parts"],
            )

            if "plate_number" in extracted_fields and repair.vehicle.plate_number:
                extracted_plate_compare = normalize_compare_token(str(extracted_fields["plate_number"]))
                vehicle_plate_compare = normalize_compare_token(repair.vehicle.plate_number)
                if extracted_plate_compare and vehicle_plate_compare and extracted_plate_compare != vehicle_plate_compare:
                    checks.append(
                        {
                            "check_type": "ocr_vehicle_plate_mismatch",
                            "severity": CheckSeverity.WARNING,
                            "title": "Госномер в документе не совпадает с карточкой техники",
                            "details": (
                                f"В документе найден {extracted_fields['plate_number']}, "
                                f"в системе {repair.vehicle.plate_number}"
                            ),
                            "payload": {
                                "document_plate_number": extracted_fields["plate_number"],
                                "vehicle_plate_number": repair.vehicle.plate_number,
                            },
                        }
                    )

            works_sum = round(sum(float(item["line_total"]) for item in extracted_items["works"]), 2)
            parts_sum = round(sum(float(item["line_total"]) for item in extracted_items["parts"]), 2)
            checks.extend(build_standard_hours_checks(extracted_items["works"]))
            checks.extend(build_repeat_repair_checks(db, repair, extracted_items["works"]))
            checks.extend(build_duplicate_line_checks(extracted_items["works"], extracted_items["parts"]))
            expected_total, expected_total_checks = build_expected_total_checks(db, repair, extracted_items["works"])
            repair.expected_total = expected_total
            checks.extend(expected_total_checks)
            if extracted_items["works"] and "work_total" in extracted_fields:
                if not amounts_match(works_sum, float(extracted_fields["work_total"])):
                    checks.append(
                        {
                            "check_type": "ocr_work_lines_total_mismatch",
                            "severity": CheckSeverity.SUSPICIOUS,
                            "title": "Сумма строк работ не совпадает с итогом работ",
                            "details": "Нужна ручная проверка работ в заказ-наряде",
                            "payload": {
                                "lines_total": works_sum,
                                "header_total": float(extracted_fields["work_total"]),
                            },
                        }
                    )

            if extracted_items["parts"] and "parts_total" in extracted_fields:
                if not amounts_match(parts_sum, float(extracted_fields["parts_total"])):
                    checks.append(
                        {
                            "check_type": "ocr_part_lines_total_mismatch",
                            "severity": CheckSeverity.SUSPICIOUS,
                            "title": "Сумма строк запчастей не совпадает с итогом материалов",
                            "details": "Нужна ручная проверка состава материалов",
                            "payload": {
                                "lines_total": parts_sum,
                                "header_total": float(extracted_fields["parts_total"]),
                            },
                        }
                    )

            if "grand_total" in extracted_fields:
                work_total = float(extracted_fields.get("work_total", 0) or 0)
                parts_total = float(extracted_fields.get("parts_total", 0) or 0)
                vat_total = float(extracted_fields.get("vat_total", 0) or 0)
                grand_total = float(extracted_fields["grand_total"])
                calculated_total = round(work_total + parts_total, 2)
                calculated_total_with_vat = round(calculated_total + vat_total, 2)
                if not amounts_match(calculated_total, grand_total) and not amounts_match(
                    calculated_total_with_vat,
                    grand_total,
                ):
                    checks.append(
                        {
                            "check_type": "ocr_total_mismatch",
                            "severity": CheckSeverity.SUSPICIOUS,
                            "title": "Сумма строк не совпадает с итоговой суммой",
                            "details": "Нужна ручная проверка итогов заказ-наряда",
                            "payload": {
                                "work_total": work_total,
                                "parts_total": parts_total,
                                "vat_total": vat_total,
                                "calculated_total": calculated_total,
                                "calculated_total_with_vat": calculated_total_with_vat,
                                "grand_total": grand_total,
                            },
                        }
                    )

            for reason in manual_review_reasons:
                checks.append(build_manual_review_check(reason, extracted_fields=extracted_fields))

            replace_ocr_checks(db, repair.id, checks)

            recognized_fields_count = len(confidence_map)
            repair.is_preliminary = True
            repair.is_partially_recognized = recognized_fields_count < 4
            has_blocking_checks = any(item["severity"] in {CheckSeverity.SUSPICIOUS, CheckSeverity.ERROR} for item in checks)
            if not recognized_fields_count:
                repair.status = RepairStatus.OCR_ERROR
            elif has_blocking_checks:
                repair.status = RepairStatus.SUSPICIOUS
            else:
                repair.status = RepairStatus.IN_REVIEW

            if recognized_fields_count >= 4:
                document.status = DocumentStatus.RECOGNIZED
                message = "Document processed automatically"
            elif recognized_fields_count > 0:
                document.status = DocumentStatus.PARTIALLY_RECOGNIZED
                message = "Document processed partially and sent for review"
            elif document.source_type == "pdf":
                document.status = DocumentStatus.OCR_ERROR
                message = "Document processing did not extract text"
            else:
                document.status = DocumentStatus.NEEDS_REVIEW
                message = "Image uploaded; manual review is required"

            document.ocr_confidence = average_confidence(confidence_map)
            document.review_queue_priority = 100 if document.status != DocumentStatus.RECOGNIZED else 20

            version_number = max([version.version_number for version in document.versions], default=0) + 1
            text_excerpt = normalize_text(text.replace("\n", " "))[:500] if text else None
            parsed_payload = {
                "processor": "hybrid_document_ocr_v2",
                "ocr_profile_scope": profile_selection.profile_scope,
                "ocr_profile_source": profile_selection.source,
                "ocr_profile_reason": profile_selection.reason,
                "document_kind": document.kind.value,
                "extracted_from": extracted_from,
                "text_length": len(text),
                "text_excerpt": text_excerpt,
                "extracted_fields": extracted_fields,
                "extracted_items": extracted_items,
                "manual_review_reasons": manual_review_reasons,
                "normalization_notes": normalization_notes,
                "service_not_done": parsed.get("service_not_done", []),
                "labor_norm_applicability": {
                    "eligible": labor_norm_applicability.eligible,
                    "scope": labor_norm_applicability.scope,
                    "reason_code": labor_norm_applicability.reason_code,
                    "reason": labor_norm_applicability.reason,
                    "brand_family": labor_norm_applicability.brand_family,
                    "catalog_name": labor_norm_applicability.catalog_name,
                    "matched_count": labor_norm_summary.matched_count,
                    "unmatched_count": labor_norm_summary.unmatched_count,
                },
            }
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=version_number,
                    storage_key=document.storage_key,
                    parsed_payload=parsed_payload,
                    field_confidence_map=confidence_map,
                    change_summary=message,
                )
            )

            final_job_status = (
                ImportStatus.COMPLETED if document.status == DocumentStatus.RECOGNIZED else ImportStatus.COMPLETED_WITH_CONFLICTS
            )
            mark_job_completed(
                db,
                job,
                status=final_job_status,
                summary={
                    "document_id": document.id,
                    "document_status": document.status.value,
                    "recognized_fields_count": recognized_fields_count,
                    "works_count": len(extracted_items["works"]),
                    "parts_count": len(extracted_items["parts"]),
                    "manual_review_reasons": manual_review_reasons,
                    "normalization_notes": normalization_notes,
                    "confidence": document.ocr_confidence,
                },
            )
            logger.info(
                "document_processing_completed",
                extra={"document_id": document.id, "job_id": job.id, "document_status": document.status.value},
            )
            db.commit()
    except Exception as exc:
        logger.exception("document_processing_failed", extra={"document_id": document_id, "job_id": job_id})
        db.rollback()
        document = load_document_for_processing(db, document_id)
        job = db.get(ImportJob, job_id)
        if document is None or document.repair is None or job is None:
            raise
        document.status = DocumentStatus.OCR_ERROR
        document.review_queue_priority = 100
        document.ocr_confidence = 0
        document.repair.status = RepairStatus.OCR_ERROR
        replace_ocr_checks(
            db,
            document.repair.id,
            [
                {
                    "check_type": "ocr_processing_failed",
                    "severity": CheckSeverity.ERROR,
                    "title": "Ошибка автоматической обработки документа",
                    "details": str(exc),
                    "payload": {"error": str(exc)},
                }
            ],
        )
        mark_job_failed(
            db,
            job,
            error_message=str(exc),
            summary={"document_id": document.id, "document_status": document.status.value},
        )
        db.commit()
        message = "Document processing failed"

    refreshed_document = load_document_for_processing(db, document.id)
    refreshed_job = db.get(ImportJob, job_id)
    if refreshed_document is None or refreshed_job is None:
        raise ValueError("Processed document could not be reloaded")
    return ProcessingResult(document=refreshed_document, job=refreshed_job, message=message)
