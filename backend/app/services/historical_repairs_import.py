from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import BinaryIO, Optional

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.audit import AuditLog
from app.models.enums import CatalogStatus, CheckSeverity, ImportStatus, RepairStatus, ServiceStatus
from app.models.imports import ImportConflict, ImportJob
from app.models.repair import Repair, RepairCheck, RepairPart, RepairWork
from app.models.service import Service
from app.models.user import User
from app.models.vehicle import Vehicle
from app.services.document_processing import normalize_compare_token, normalize_identifier_token
from app.services.service_catalog import ensure_service_catalog_synced, resolve_service_by_name


EXPECTED_HEADERS = (
    "ТС.Государственный номер",
    "ТС.Вид ТС",
    "Период",
    "Пробег",
    "Поставщик",
    "Колонна",
    "Регистратор",
    "Номенклатура.Автогрупп (Номенклатура)",
    "Номенклатура.Артикул",
    "Статья затрат",
    "ТС.Модель",
    "Номенклатура",
    "Количество",
    "Сумма",
)
PLACEHOLDER_EXTERNAL_ID = "__batch_import_placeholder__"
ORDER_NUMBER_PATTERN = re.compile(r"(?:Заказ-наряд|Заказ наряд)\s*(?P<value>[A-Za-zА-Яа-я0-9/_-]+)", re.IGNORECASE)
IMPORT_REASON_PREFIX = "historical_import:"
WORK_EXPENSE_KEYWORDS = (
    "то и ремонт",
    "дополнительные расходы",
    "буксировка",
    "эвакуация",
    "комплектация",
    "предпродажная подготовка",
    "расходы",
)
PART_EXPENSE_KEYWORDS = ("запчаст", "з/ч", "зч ")
WORK_GROUP_KEYWORDS = ("услуг", "работ", "ремонт", "обслужив", "шиномонтаж")
WORK_NAME_KEYWORDS = (
    "ремонт",
    "замен",
    "диагност",
    "мойк",
    "шиномонтаж",
    "балансиров",
    "свар",
    "буксиров",
    "эваку",
    "регулиров",
    "установк",
    "демонтаж",
    "монтаж",
)


@dataclass
class HistoricalRepairLine:
    source_row: int
    nomenclature: str
    article: str | None
    quantity: float
    amount: float
    expense_item: str | None
    auto_group: str | None


@dataclass
class HistoricalRepairGroup:
    source_key: str
    raw_plate: str | None
    normalized_plate: str | None
    repair_date: date
    raw_service_name: str | None
    registrator: str | None
    order_number: str | None
    column_name: str | None
    vehicle_type_label: str | None
    vehicle_model: str | None
    mileage: int = 0
    lines: list[HistoricalRepairLine] = field(default_factory=list)


@dataclass
class HistoricalRepairImportResult:
    job_id: int
    status: ImportStatus
    source_filename: str
    rows_total: int = 0
    grouped_repairs: int = 0
    created_repairs: int = 0
    duplicate_repairs: int = 0
    conflicts_created: int = 0
    created_services: int = 0
    created_works: int = 0
    created_parts: int = 0
    repair_limit_applied: int | None = None
    first_repair_id: int | None = None
    recent_repair_ids: list[int] = field(default_factory=list)
    sample_conflicts: list[str] = field(default_factory=list)

    @property
    def message(self) -> str:
        if self.created_repairs == 0 and self.conflicts_created > 0:
            return "Импорт завершён без создания ремонтов, обнаружены конфликты"
        if self.duplicate_repairs > 0 or self.conflicts_created > 0:
            return "Импорт истории завершён с конфликтами"
        return "Исторические ремонты успешно импортированы"

    def build_summary(self) -> dict[str, object]:
        return {
            "rows_total": self.rows_total,
            "grouped_repairs": self.grouped_repairs,
            "created_repairs": self.created_repairs,
            "duplicate_repairs": self.duplicate_repairs,
            "conflicts_created": self.conflicts_created,
            "created_services": self.created_services,
            "created_works": self.created_works,
            "created_parts": self.created_parts,
            "repair_limit_applied": self.repair_limit_applied,
            "first_repair_id": self.first_repair_id,
            "recent_repair_ids": self.recent_repair_ids,
            "sample_conflicts": self.sample_conflicts,
        }


def normalize_optional_text(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def normalize_amount(value: object) -> float:
    if value in (None, ""):
        return 0.0
    return round(float(value), 2)


def normalize_quantity(value: object) -> float:
    if value in (None, ""):
        return 1.0
    quantity = float(value)
    return quantity if quantity > 0 else 1.0


def normalize_repair_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    try:
        text = str(value).strip()
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def normalize_mileage(value: object) -> int:
    if value in (None, ""):
        return 0
    return max(0, int(round(float(value))))


def extract_order_number(value: str | None) -> str | None:
    if not value:
        return None
    match = ORDER_NUMBER_PATTERN.search(value)
    if match is None:
        return None
    normalized = normalize_identifier_token(match.group("value"))
    return normalized or None


def normalize_plate(value: str | None) -> str | None:
    normalized = normalize_compare_token(value)
    return normalized or None


def build_source_key(
    *,
    raw_plate: str | None,
    raw_service_name: str | None,
    registrator: str | None,
    repair_date: date,
) -> str:
    parts = [
        normalize_identifier_token(raw_plate) or "-",
        normalize_identifier_token(raw_service_name) or "-",
        normalize_identifier_token(registrator) or "-",
        repair_date.isoformat(),
    ]
    return "|".join(parts)


def build_import_comment(group: HistoricalRepairGroup) -> str:
    parts = ["Исторический импорт"]
    if group.registrator:
        parts.append(group.registrator)
    if group.column_name:
        parts.append(f"Колонна: {group.column_name}")
    if group.vehicle_model:
        parts.append(f"Модель: {group.vehicle_model}")
    return " · ".join(parts)


def classify_line_kind(line: HistoricalRepairLine) -> str:
    expense = (line.expense_item or "").strip().lower()
    auto_group = (line.auto_group or "").strip().lower()
    nomenclature = line.nomenclature.strip().lower()

    if any(keyword in expense for keyword in PART_EXPENSE_KEYWORDS):
        return "part"
    if any(keyword in auto_group for keyword in WORK_GROUP_KEYWORDS):
        return "work"
    if any(keyword in nomenclature for keyword in WORK_NAME_KEYWORDS):
        return "work"
    if any(keyword in expense for keyword in WORK_EXPENSE_KEYWORDS):
        return "work"
    return "part"


def validate_headers(headers: tuple[object, ...]) -> None:
    normalized_headers = tuple(str(item).strip() if item is not None else "" for item in headers)
    if normalized_headers != EXPECTED_HEADERS:
        raise ValueError("Неожиданный формат файла истории ремонтов. Ожидалась выгрузка `2025 для ИИ.xlsx`.")


def append_conflict(
    db: Session,
    result: HistoricalRepairImportResult,
    *,
    job_id: int,
    entity_type: str,
    conflict_key: str,
    message: str,
    incoming_payload: dict[str, object] | None = None,
    existing_payload: dict[str, object] | None = None,
) -> None:
    db.add(
        ImportConflict(
            import_job_id=job_id,
            entity_type=entity_type,
            conflict_key=conflict_key,
            incoming_payload=incoming_payload,
            existing_payload=existing_payload,
            status="pending",
        )
    )
    result.conflicts_created += 1
    if len(result.sample_conflicts) < 8:
        result.sample_conflicts.append(message)


def build_vehicle_lookup(db: Session) -> tuple[dict[str, Vehicle | None], list[tuple[str, Vehicle]]]:
    exact_plate_map: dict[str, Vehicle | None] = {}
    plate_candidates: list[tuple[str, Vehicle]] = []
    for vehicle in db.scalars(select(Vehicle)).all():
        if vehicle.external_id == PLACEHOLDER_EXTERNAL_ID:
            continue
        normalized_plate = normalize_plate(vehicle.plate_number)
        if not normalized_plate:
            continue
        if normalized_plate in exact_plate_map and exact_plate_map[normalized_plate] is not None:
            exact_plate_map[normalized_plate] = None
        else:
            exact_plate_map[normalized_plate] = vehicle
        plate_candidates.append((normalized_plate, vehicle))
    return exact_plate_map, plate_candidates


def find_vehicle(
    normalized_plate: str | None,
    exact_plate_map: dict[str, Vehicle | None],
    plate_candidates: list[tuple[str, Vehicle]],
) -> Vehicle | None:
    if not normalized_plate:
        return None
    direct_match = exact_plate_map.get(normalized_plate)
    if direct_match is not None:
        return direct_match

    partial_matches: dict[int, Vehicle] = {}
    for candidate_plate, vehicle in plate_candidates:
        if len(normalized_plate) < 6:
            continue
        if candidate_plate.startswith(normalized_plate) or normalized_plate.startswith(candidate_plate):
            partial_matches[vehicle.id] = vehicle
    if len(partial_matches) == 1:
        return next(iter(partial_matches.values()))
    return None


def load_existing_historical_keys(db: Session) -> set[str]:
    keys: set[str] = set()
    for reason in db.scalars(select(Repair.reason).where(Repair.reason.is_not(None))).all():
        if not reason or not reason.startswith(IMPORT_REASON_PREFIX):
            continue
        keys.add(reason[len(IMPORT_REASON_PREFIX):].strip())
    return keys


def resolve_or_create_service(
    db: Session,
    current_admin: User,
    cache: dict[str, Service | None],
    result: HistoricalRepairImportResult,
    raw_service_name: str | None,
) -> Service | None:
    normalized_name = normalize_optional_text(raw_service_name)
    if not normalized_name:
        return None

    cache_key = normalized_name.casefold()
    if cache_key in cache:
        return cache[cache_key]

    service = resolve_service_by_name(db, normalized_name)
    if service is None:
        service = Service(
            name=normalized_name[:255],
            status=ServiceStatus.PRELIMINARY,
            created_by_user_id=current_admin.id,
            comment="Создано при импорте исторических ремонтов",
        )
        db.add(service)
        db.flush()
        result.created_services += 1
    cache[cache_key] = service
    return service


def parse_groups(file_obj: BinaryIO) -> tuple[dict[str, HistoricalRepairGroup], int]:
    file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    iterator = worksheet.iter_rows(min_row=1, values_only=True)
    header_row = next(iterator, None)
    if header_row is None:
        raise ValueError("Файл истории ремонтов пуст")
    validate_headers(header_row)

    groups: dict[str, HistoricalRepairGroup] = {}
    rows_total = 0

    for row_index, row in enumerate(iterator, start=2):
        rows_total += 1
        (
            raw_plate,
            vehicle_type_label,
            period_value,
            mileage_value,
            raw_service_name,
            column_name,
            registrator,
            auto_group,
            article,
            expense_item,
            vehicle_model,
            nomenclature,
            quantity,
            amount,
        ) = row

        repair_date = normalize_repair_date(period_value)
        if repair_date is None:
            continue

        raw_registrator = normalize_optional_text(registrator)
        source_key = build_source_key(
            raw_plate=normalize_optional_text(raw_plate),
            raw_service_name=normalize_optional_text(raw_service_name),
            registrator=raw_registrator,
            repair_date=repair_date,
        )
        group = groups.get(source_key)
        if group is None:
            group = HistoricalRepairGroup(
                source_key=source_key,
                raw_plate=normalize_optional_text(raw_plate),
                normalized_plate=normalize_plate(normalize_optional_text(raw_plate)),
                repair_date=repair_date,
                raw_service_name=normalize_optional_text(raw_service_name),
                registrator=raw_registrator,
                order_number=extract_order_number(raw_registrator),
                column_name=normalize_optional_text(column_name),
                vehicle_type_label=normalize_optional_text(vehicle_type_label),
                vehicle_model=normalize_optional_text(vehicle_model),
            )
            groups[source_key] = group

        group.mileage = max(group.mileage, normalize_mileage(mileage_value))
        group.lines.append(
            HistoricalRepairLine(
                source_row=row_index,
                nomenclature=normalize_optional_text(nomenclature) or "Без названия",
                article=normalize_optional_text(article),
                quantity=normalize_quantity(quantity),
                amount=normalize_amount(amount),
                expense_item=normalize_optional_text(expense_item),
                auto_group=normalize_optional_text(auto_group),
            )
        )

    workbook.close()
    return groups, rows_total


def import_historical_repairs(
    db: Session,
    *,
    file_obj: BinaryIO,
    filename: str,
    current_admin: User,
    repair_limit: int | None = None,
) -> HistoricalRepairImportResult:
    if repair_limit is not None and repair_limit <= 0:
        raise ValueError("Лимит ремонтов должен быть положительным числом")

    job = ImportJob(
        import_type="historical_repairs",
        source_filename=filename,
        status=ImportStatus.PROCESSING,
        summary={},
        error_message=None,
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    result = HistoricalRepairImportResult(
        job_id=job.id,
        status=ImportStatus.PROCESSING,
        source_filename=filename,
        repair_limit_applied=repair_limit,
    )

    try:
        ensure_service_catalog_synced(db, commit=False)
        groups, rows_total = parse_groups(file_obj)
        result.rows_total = rows_total
        result.grouped_repairs = len(groups)

        exact_plate_map, plate_candidates = build_vehicle_lookup(db)
        existing_historical_keys = load_existing_historical_keys(db)
        service_cache: dict[str, Service | None] = {}

        for group in groups.values():
            if repair_limit is not None and result.created_repairs >= repair_limit:
                break

            if group.source_key in existing_historical_keys:
                result.duplicate_repairs += 1
                append_conflict(
                    db,
                    result,
                    job_id=job.id,
                    entity_type="repair",
                    conflict_key=group.source_key,
                    message=(
                        f"Дубликат исторического ремонта: {group.order_number or group.registrator or group.source_key}"
                    ),
                    incoming_payload={
                        "repair_date": group.repair_date.isoformat(),
                        "plate_number": group.raw_plate,
                        "service_name": group.raw_service_name,
                        "order_number": group.order_number,
                    },
                    existing_payload={"reason": f"{IMPORT_REASON_PREFIX}{group.source_key}"},
                )
                continue

            vehicle = find_vehicle(group.normalized_plate, exact_plate_map, plate_candidates)
            if vehicle is None:
                append_conflict(
                    db,
                    result,
                    job_id=job.id,
                    entity_type="repair",
                    conflict_key=group.source_key,
                    message=(
                        f"Не найдена техника для импорта: {group.raw_plate or 'без номера'} · "
                        f"{group.order_number or group.registrator or group.repair_date.isoformat()}"
                    ),
                    incoming_payload={
                        "repair_date": group.repair_date.isoformat(),
                        "plate_number": group.raw_plate,
                        "service_name": group.raw_service_name,
                        "order_number": group.order_number,
                        "registrator": group.registrator,
                    },
                )
                continue

            service = resolve_or_create_service(db, current_admin, service_cache, result, group.raw_service_name)
            repair = Repair(
                order_number=group.order_number,
                repair_date=group.repair_date,
                vehicle_id=vehicle.id,
                service_id=service.id if service is not None else None,
                created_by_user_id=current_admin.id,
                mileage=group.mileage,
                reason=f"{IMPORT_REASON_PREFIX}{group.source_key}",
                employee_comment=build_import_comment(group),
                work_total=0,
                parts_total=0,
                vat_total=0,
                grand_total=0,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
                is_partially_recognized=False,
                is_manually_completed=False,
            )
            db.add(repair)
            db.flush()

            work_total = 0.0
            parts_total = 0.0
            for line in group.lines:
                line_total = round(line.amount, 2)
                price = round(line_total / line.quantity, 2) if line.quantity else line_total
                if classify_line_kind(line) == "work":
                    db.add(
                        RepairWork(
                            repair_id=repair.id,
                            work_code=line.article,
                            work_name=line.nomenclature,
                            quantity=line.quantity,
                            price=price,
                            line_total=line_total,
                            status=CatalogStatus.PRELIMINARY,
                            reference_payload={
                                "source": "historical_import",
                                "source_row": line.source_row,
                                "expense_item": line.expense_item,
                                "auto_group": line.auto_group,
                            },
                        )
                    )
                    work_total += line_total
                    result.created_works += 1
                else:
                    db.add(
                        RepairPart(
                            repair_id=repair.id,
                            article=line.article,
                            part_name=line.nomenclature,
                            quantity=line.quantity,
                            price=price,
                            line_total=line_total,
                            status=CatalogStatus.PRELIMINARY,
                        )
                    )
                    parts_total += line_total
                    result.created_parts += 1

            repair.work_total = round(work_total, 2)
            repair.parts_total = round(parts_total, 2)
            repair.grand_total = round(work_total + parts_total, 2)
            if group.mileage <= 0:
                db.add(
                    RepairCheck(
                        repair_id=repair.id,
                        check_type="historical_mileage_missing",
                        severity=CheckSeverity.WARNING,
                        title="В исторической записи отсутствует пробег",
                        details="Для этой строки истории в исходном Excel не было пробега, сохранено значение 0.",
                        calculation_payload={"source": "historical_import"},
                        is_resolved=False,
                    )
                )

            db.add(
                AuditLog(
                    user_id=current_admin.id,
                    entity_type="repair",
                    entity_id=str(repair.id),
                    action_type="historical_import_created",
                    old_value=None,
                    new_value={
                        "repair_id": repair.id,
                        "job_id": job.id,
                        "source_filename": filename,
                        "source_key": group.source_key,
                        "order_number": repair.order_number,
                        "repair_date": repair.repair_date.isoformat(),
                        "vehicle_id": repair.vehicle_id,
                        "service_id": repair.service_id,
                        "grand_total": float(repair.grand_total),
                    },
                )
            )

            existing_historical_keys.add(group.source_key)
            result.created_repairs += 1
            if result.first_repair_id is None:
                result.first_repair_id = repair.id
            if len(result.recent_repair_ids) < 10:
                result.recent_repair_ids.append(repair.id)

        result.status = (
            ImportStatus.COMPLETED_WITH_CONFLICTS if result.conflicts_created > 0 or result.duplicate_repairs > 0 else ImportStatus.COMPLETED
        )
        job.status = result.status
        job.summary = result.build_summary()
        job.error_message = None
        db.add(job)
        db.commit()
        return result
    except Exception as exc:
        db.rollback()
        failed_job = db.get(ImportJob, job.id)
        if failed_job is not None:
            failed_job.status = ImportStatus.FAILED
            failed_job.summary = result.build_summary()
            failed_job.error_message = str(exc)
            db.add(failed_job)
            db.commit()
        raise
