from __future__ import annotations

import tempfile
import unittest
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image
from sqlalchemy import text
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.exc import IntegrityError

from app.api.deps import get_db
from app.api import documents as documents_api
from app.api import jobs as jobs_api
from app.api import upload_validation as upload_validation_api
from app.api.repairs import build_export_warning_rows, build_report_status_summary
from app.core.paths import set_storage_root
from app.core.security import get_password_hash
from app.db.base import Base
from app.main import app
from app.models.audit import AuditLog
from app.models.document import Document
from app.models.document import DocumentVersion
from app.models.enums import CatalogStatus, CheckSeverity, DocumentKind, DocumentStatus, ImportStatus, RepairStatus, ServiceStatus, UserRole, VehicleStatus, VehicleType
from app.models.imports import ImportJob
from app.models.repair import Repair, RepairCheck, RepairWork
from app.models.service import Service
from app.models.user import User
from app.models.vehicle import Vehicle
from tests.sqlite_test_utils import create_sqlite_test_engine, reset_database


class DocumentJobsApiTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.storage_root = Path(cls.temp_dir.name) / "storage"
        cls.storage_root.mkdir(parents=True, exist_ok=True)
        cls.engine = create_sqlite_test_engine(enforce_foreign_keys=True)
        cls.SessionLocal = sessionmaker(bind=cls.engine, autoflush=False, autocommit=False, future=True)
        Base.metadata.create_all(bind=cls.engine)

        def override_get_db():
            db = cls.SessionLocal()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[get_db] = override_get_db
        set_storage_root(cls.storage_root)
        cls.client = TestClient(app)

    @classmethod
    def tearDownClass(cls) -> None:
        app.dependency_overrides.clear()
        cls.engine.dispose()
        cls.temp_dir.cleanup()

    def setUp(self) -> None:
        reset_database(self.engine, Base.metadata)
        with self.SessionLocal() as db:
            db.add_all(
                [
                    User(
                        full_name="Admin User",
                        login="admin",
                        email="admin@example.com",
                        password_hash=get_password_hash("secret123"),
                        role=UserRole.ADMIN,
                        is_active=True,
                    ),
                    User(
                        full_name="Employee User",
                        login="employee",
                        email="employee@example.com",
                        password_hash=get_password_hash("secret123"),
                        role=UserRole.EMPLOYEE,
                        is_active=True,
                    ),
                ]
            )
            db.commit()

    def _get_auth_headers(self, username: str = "admin", password: str = "secret123") -> dict[str, str]:
        response = self.client.post(
            "/api/auth/login",
            data={"username": username, "password": password},
        )
        self.assertEqual(response.status_code, 200, response.text)
        token = response.json()["access_token"]
        return {"Authorization": f"Bearer {token}"}

    def _upload_order_document(self, headers: dict[str, str]) -> dict:
        response = self.client.post(
            "/api/documents/upload",
            headers=headers,
            files={"file": ("job-test.pdf", b"%PDF-1.4\n%test\n", "application/pdf")},
            data={"kind": "order"},
        )
        self.assertEqual(response.status_code, 200, response.text)
        return response.json()

    def _make_png_bytes(self, color: tuple[int, int, int]) -> bytes:
        image = Image.new("RGB", (64, 48), color=color)
        buffer = BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()

    def test_upload_creates_queued_job_and_retry_reuses_failed_job(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]
        self.assertIsNotNone(job_id)
        self.assertEqual(payload["import_status"], "queued")
        self.assertEqual(payload["document"]["latest_import_job"]["status"], "queued")

        job_response = self.client.get(f"/api/jobs/{job_id}", headers=headers)
        self.assertEqual(job_response.status_code, 200, job_response.text)
        self.assertEqual(job_response.json()["status"], "queued")

        list_response = self.client.get("/api/documents?limit=8", headers=headers)
        self.assertEqual(list_response.status_code, 200, list_response.text)
        listed_document = next(item for item in list_response.json()["items"] if item["id"] == document_id)
        self.assertEqual(listed_document["latest_import_job"]["id"], job_id)
        self.assertEqual(listed_document["latest_import_job"]["status"], "queued")

        with self.SessionLocal() as db:
            self._mark_job_failed(db, job_id, document_id)

        retry_response = self.client.post(f"/api/jobs/{job_id}/retry", headers=headers)
        self.assertEqual(retry_response.status_code, 200, retry_response.text)
        retry_payload = retry_response.json()
        self.assertEqual(retry_payload["job"]["id"], job_id)
        self.assertEqual(retry_payload["job"]["status"], "retry")

        retried_job_response = self.client.get(f"/api/jobs/{job_id}", headers=headers)
        self.assertEqual(retried_job_response.status_code, 200, retried_job_response.text)
        self.assertEqual(retried_job_response.json()["status"], "retry")

    def test_get_job_tolerates_non_object_summary(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        job_id = payload["job_id"]
        self.assertIsNotNone(job_id)
        assert job_id is not None

        with self.SessionLocal() as db:
            job = db.get(ImportJob, job_id)
            self.assertIsNotNone(job)
            assert job is not None
            job.summary = ["legacy-broken-summary"]
            db.commit()

        response = self.client.get(f"/api/jobs/{job_id}", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["summary"], {})

    def test_retry_endpoint_reuses_existing_active_job_for_same_document(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        failed_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            self._mark_job_failed(db, failed_job_id, document_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            active_job = ImportJob(
                document_id=document_id,
                import_type="document_ocr",
                source_filename=document.original_filename,
                status=ImportStatus.QUEUED,
                summary={"document_id": document_id, "stage": "queued_existing_active"},
                error_message=None,
                attempts=0,
                started_at=None,
                finished_at=None,
            )
            db.add(active_job)
            db.commit()
            db.refresh(active_job)
            active_job_id = active_job.id

        retry_response = self.client.post(f"/api/jobs/{failed_job_id}/retry", headers=headers)
        self.assertEqual(retry_response.status_code, 200, retry_response.text)
        retry_payload = retry_response.json()
        self.assertEqual(retry_payload["job"]["id"], active_job_id)
        self.assertEqual(retry_payload["job"]["status"], "queued")

        with self.SessionLocal() as db:
            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            self.assertEqual(len(jobs), 2)
            self.assertEqual(sum(1 for job in jobs if job.status in {ImportStatus.QUEUED, ImportStatus.RETRY, ImportStatus.PROCESSING}), 1)
            failed_job = db.get(ImportJob, failed_job_id)
            self.assertIsNotNone(failed_job)
            assert failed_job is not None
            self.assertEqual(failed_job.status, ImportStatus.FAILED)

    def test_retry_endpoint_replaces_stale_ocr_payload_with_queued_reprocessing_version(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        retry_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            job = db.get(ImportJob, retry_job_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(job)
            self.assertIsNotNone(document)
            assert job is not None
            assert document is not None

            job.status = ImportStatus.COMPLETED
            job.summary = {"document_id": document_id, "stage": "completed"}
            document.status = DocumentStatus.RECOGNIZED
            document.ocr_confidence = 0.88
            next_version_number = max((item.version_number for item in document.versions), default=0) + 1
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=next_version_number,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "processor": "legacy-ocr",
                        "ocr_status": "completed",
                        "extracted_fields": {"order_number": "ZN-RETRY-STALE-001"},
                        "manual_review_reasons": ["service_name_missing"],
                    },
                    field_confidence_map={"order_number": 0.88},
                    change_summary="Stale OCR payload before retry",
                )
            )
            db.commit()

        retry_response = self.client.post(f"/api/jobs/{retry_job_id}/retry", headers=headers)

        self.assertEqual(retry_response.status_code, 200, retry_response.text)

        documents_response = self.client.get("/api/documents?limit=20", headers=headers)
        self.assertEqual(documents_response.status_code, 200, documents_response.text)
        document_payload = next(item for item in documents_response.json()["items"] if item["id"] == document_id)
        self.assertEqual(document_payload["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(document_payload["ocr_confidence"], None)
        self.assertEqual(document_payload["parsed_payload"]["ocr_status"], "queued")
        self.assertEqual(document_payload["parsed_payload"]["pipeline"], "reprocessing")
        self.assertNotIn("extracted_fields", document_payload["parsed_payload"])

    def test_retry_endpoint_rolls_back_when_enqueue_fails(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        failed_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            self._mark_job_failed(db, failed_job_id, document_id)

        with patch.object(jobs_api, "queue_document_processing", side_effect=RuntimeError("retry queue unavailable")):
            with self.assertRaises(RuntimeError):
                self.client.post(f"/api/jobs/{failed_job_id}/retry", headers=headers)

        with self.SessionLocal() as db:
            failed_job = db.get(ImportJob, failed_job_id)
            self.assertIsNotNone(failed_job)
            assert failed_job is not None
            self.assertEqual(failed_job.status, ImportStatus.FAILED)
            self.assertEqual(db.query(ImportJob).filter(ImportJob.document_id == document_id).count(), 1)

    def test_retry_endpoint_rejects_archived_document(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        failed_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            self._mark_job_failed(db, failed_job_id, document_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.ARCHIVED
            document.review_queue_priority = 0
            db.commit()

        retry_response = self.client.post(f"/api/jobs/{failed_job_id}/retry", headers=headers)
        self.assertEqual(retry_response.status_code, 409, retry_response.text)
        self.assertEqual(retry_response.json()["detail"], "Archived documents cannot be modified")

        with self.SessionLocal() as db:
            failed_job = db.get(ImportJob, failed_job_id)
            self.assertIsNotNone(failed_job)
            assert failed_job is not None
            self.assertEqual(failed_job.status, ImportStatus.FAILED)

    def test_employee_cannot_view_job_for_archived_vehicle(self) -> None:
        admin_headers = self._get_auth_headers("admin")
        payload = self._upload_order_document(admin_headers)
        job_id = payload["job_id"]
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            vehicle = db.get(Vehicle, repair.vehicle_id)
            employee = db.query(User).filter(User.login == "employee").one()
            self.assertIsNotNone(vehicle)
            assert vehicle is not None

            repair.created_by_user_id = employee.id
            vehicle.external_id = "truck-job-employee-visible"
            vehicle.plate_number = "JOB-EMP-116"
            vehicle.status = VehicleStatus.ACTIVE
            db.commit()
            vehicle_id = vehicle.id

        employee_headers = self._get_auth_headers("employee")
        visible_response = self.client.get(f"/api/jobs/{job_id}", headers=employee_headers)
        self.assertEqual(visible_response.status_code, 200, visible_response.text)

        with self.SessionLocal() as db:
            vehicle = db.get(Vehicle, vehicle_id)
            self.assertIsNotNone(vehicle)
            assert vehicle is not None
            vehicle.status = VehicleStatus.ARCHIVED
            db.commit()

        archived_response = self.client.get(f"/api/jobs/{job_id}", headers=employee_headers)
        self.assertEqual(archived_response.status_code, 404, archived_response.text)

    def test_process_endpoint_reuses_existing_active_job(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        initial_job_id = payload["job_id"]

        first_response = self.client.post(f"/api/documents/{document_id}/process", headers=headers)
        self.assertEqual(first_response.status_code, 200, first_response.text)
        self.assertEqual(first_response.json()["job_id"], initial_job_id)

        second_response = self.client.post(f"/api/documents/{document_id}/process", headers=headers)
        self.assertEqual(second_response.status_code, 200, second_response.text)
        self.assertEqual(second_response.json()["job_id"], initial_job_id)

        with self.SessionLocal() as db:
            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            document = db.get(Document, document_id)
            self.assertEqual(len(jobs), 1)
            self.assertEqual(jobs[0].id, initial_job_id)
            self.assertEqual(jobs[0].status, ImportStatus.QUEUED)
            self.assertIsNotNone(document)
            assert document is not None
            self.assertEqual(len(document.versions), 1)
            latest_version = max(document.versions, key=lambda item: item.version_number)
            self.assertEqual(latest_version.change_summary, "Initial upload")

    def test_process_endpoint_replaces_stale_ocr_payload_with_queued_reprocessing_version(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        with self.SessionLocal() as db:
            self._mark_job_completed(db, job_id, document_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.RECOGNIZED
            document.ocr_confidence = 0.91
            next_version_number = max((item.version_number for item in document.versions), default=0) + 1
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=next_version_number,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "processor": "legacy-ocr",
                        "ocr_status": "completed",
                        "extracted_fields": {"order_number": "ZN-STALE-001"},
                        "manual_review_reasons": ["vehicle_not_found"],
                    },
                    field_confidence_map={"order_number": 0.91},
                    change_summary="Stale OCR payload",
                )
            )
            db.commit()

        response = self.client.post(f"/api/documents/{document_id}/process", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        response_payload = response.json()
        self.assertEqual(response_payload["document"]["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(response_payload["document"]["parsed_payload"]["ocr_status"], "queued")
        self.assertEqual(response_payload["document"]["parsed_payload"]["pipeline"], "reprocessing")
        self.assertNotIn("extracted_fields", response_payload["document"]["parsed_payload"])
        self.assertEqual(response_payload["document"]["ocr_confidence"], None)

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            latest_version = max(document.versions, key=lambda item: item.version_number)
            self.assertEqual(latest_version.change_summary, "Queued for reprocessing")
            self.assertEqual(latest_version.parsed_payload["ocr_status"], "queued")
            self.assertEqual(latest_version.parsed_payload["pipeline"], "reprocessing")

    def test_database_rejects_second_active_ocr_job_for_same_document(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None

            db.add(
                ImportJob(
                    document_id=document_id,
                    import_type="document_ocr",
                    source_filename=document.original_filename,
                    status=ImportStatus.PROCESSING,
                    summary={"document_id": document_id, "stage": "duplicate_processing"},
                    error_message=None,
                    attempts=1,
                    started_at=datetime(2025, 1, 3, 10, 0, 0),
                    finished_at=None,
                )
            )
            with self.assertRaises(IntegrityError):
                db.commit()
            db.rollback()

            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            self.assertEqual(len(jobs), 1)
            self.assertEqual(jobs[0].status, ImportStatus.QUEUED)

    def test_process_endpoint_reuses_active_processing_job_even_if_latest_job_is_completed(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        initial_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            initial_job = db.get(ImportJob, initial_job_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(initial_job)
            self.assertIsNotNone(document)
            assert initial_job is not None
            assert document is not None

            initial_job.status = ImportStatus.PROCESSING
            initial_job.started_at = datetime(2025, 1, 1, 9, 0, 0)
            db.add(
                ImportJob(
                    document_id=document_id,
                    import_type="document_ocr",
                    source_filename=document.original_filename,
                    status=ImportStatus.COMPLETED,
                    summary={"document_id": document_id, "stage": "completed"},
                    error_message=None,
                    attempts=1,
                    started_at=None,
                    finished_at=None,
                )
            )
            db.commit()

        response = self.client.post(f"/api/documents/{document_id}/process", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["job_id"], initial_job_id)
        self.assertEqual(response.json()["import_status"], "processing")

        with self.SessionLocal() as db:
            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            self.assertEqual(len(jobs), 2)
            self.assertEqual(sum(1 for job in jobs if job.status == ImportStatus.PROCESSING), 1)

    def test_process_endpoint_reuses_active_job_and_replaces_stale_ocr_payload(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        initial_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            initial_job = db.get(ImportJob, initial_job_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(initial_job)
            self.assertIsNotNone(document)
            assert initial_job is not None
            assert document is not None

            initial_job.status = ImportStatus.PROCESSING
            initial_job.started_at = datetime(2025, 1, 1, 9, 0, 0)
            document.status = DocumentStatus.RECOGNIZED
            document.ocr_confidence = 0.93
            next_version_number = max((item.version_number for item in document.versions), default=0) + 1
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=next_version_number,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "processor": "legacy-ocr",
                        "ocr_status": "completed",
                        "extracted_fields": {"order_number": "ZN-ACTIVE-STALE-001"},
                    },
                    field_confidence_map={"order_number": 0.93},
                    change_summary="Stale OCR payload during active processing",
                )
            )
            db.commit()

        response = self.client.post(f"/api/documents/{document_id}/process", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        response_payload = response.json()
        self.assertEqual(response_payload["job_id"], initial_job_id)
        self.assertEqual(response_payload["import_status"], ImportStatus.PROCESSING.value)
        self.assertEqual(response_payload["document"]["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(response_payload["document"]["ocr_confidence"], None)
        self.assertEqual(response_payload["document"]["parsed_payload"]["ocr_status"], "queued")
        self.assertEqual(response_payload["document"]["parsed_payload"]["pipeline"], "reprocessing")
        self.assertNotIn("extracted_fields", response_payload["document"]["parsed_payload"])

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            self.assertIsNotNone(document)
            assert document is not None
            self.assertEqual(len(jobs), 1)
            latest_version = max(document.versions, key=lambda item: item.version_number)
            self.assertEqual(latest_version.change_summary, "Queued for reprocessing")
            self.assertEqual(latest_version.parsed_payload["ocr_status"], "queued")
            self.assertEqual(latest_version.parsed_payload["pipeline"], "reprocessing")

    def test_document_list_uses_canonical_active_import_job_when_newer_completed_job_exists(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        initial_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            initial_job = db.get(ImportJob, initial_job_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(initial_job)
            self.assertIsNotNone(document)
            assert initial_job is not None
            assert document is not None

            initial_job.status = ImportStatus.PROCESSING
            initial_job.started_at = datetime(2025, 1, 1, 9, 0, 0)
            db.add(
                ImportJob(
                    document_id=document_id,
                    import_type="document_ocr",
                    source_filename=document.original_filename,
                    status=ImportStatus.COMPLETED,
                    summary={"document_id": document_id, "stage": "completed_after_active"},
                    error_message=None,
                    attempts=1,
                    started_at=None,
                    finished_at=None,
                )
            )
            db.commit()

        response = self.client.get("/api/documents?limit=20", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        listed_document = next(item for item in response.json()["items"] if item["id"] == document_id)
        self.assertEqual(listed_document["latest_import_job"]["id"], initial_job_id)
        self.assertEqual(listed_document["latest_import_job"]["status"], ImportStatus.PROCESSING.value)

    def test_process_endpoint_rejects_document_from_archived_service(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        initial_job_id = payload["job_id"]
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None

            service = Service(
                name="Archived Single Process Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
            )
            db.add(service)
            db.flush()

            repair.service_id = service.id
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.NEEDS_REVIEW
            document.review_queue_priority = 100
            db.commit()

        response = self.client.post(f"/api/documents/{document_id}/process", headers=headers)

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Repairs for archived services cannot be modified")

        with self.SessionLocal() as db:
            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            self.assertEqual([job.id for job in jobs], [initial_job_id])

    def test_worker_claim_fails_job_for_document_from_archived_service_without_ocr_mutation(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            service = Service(
                name="Archived Worker Process Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
            )
            db.add(service)
            db.flush()

            repair.service_id = service.id
            db.commit()

        with patch("app.services.document_processing.extract_document_text") as extract_document_text:
            from app.services.import_jobs import claim_next_document_processing_job

            with self.SessionLocal() as db:
                next_job = claim_next_document_processing_job(db)
                self.assertIsNone(next_job)

            extract_document_text.assert_not_called()

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            repair = db.get(Repair, repair_id)
            job = db.get(ImportJob, job_id)
            self.assertIsNotNone(document)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(job)
            assert document is not None
            assert repair is not None
            assert job is not None

            self.assertEqual(document.status, DocumentStatus.UPLOADED)
            self.assertEqual(repair.status, RepairStatus.DRAFT)
            self.assertEqual(job.status, ImportStatus.FAILED)
            self.assertEqual(job.error_message, "Repairs for archived services cannot be modified")
            self.assertTrue((job.summary or {}).get("failed_during_claim"))

    def test_worker_claim_skips_archived_service_job_and_returns_next_operational_job(self) -> None:
        headers = self._get_auth_headers()
        archived_payload = self._upload_order_document(headers)
        archived_document_id = archived_payload["document"]["id"]
        archived_job_id = archived_payload["job_id"]
        archived_repair_id = archived_payload["document"]["repair"]["id"]

        active_payload = self._upload_order_document(headers)
        active_job_id = active_payload["job_id"]

        with self.SessionLocal() as db:
            archived_repair = db.get(Repair, archived_repair_id)
            self.assertIsNotNone(archived_repair)
            assert archived_repair is not None

            service = Service(
                name="Archived Claim Skip Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
            )
            db.add(service)
            db.flush()

            archived_repair.service_id = service.id
            db.commit()

        with self.SessionLocal() as db:
            from app.services.import_jobs import claim_next_document_processing_job

            claimed_job = claim_next_document_processing_job(db)
            self.assertIsNotNone(claimed_job)
            assert claimed_job is not None
            self.assertEqual(claimed_job.id, active_job_id)

        with self.SessionLocal() as db:
            archived_document = db.get(Document, archived_document_id)
            archived_job = db.get(ImportJob, archived_job_id)
            active_job = db.get(ImportJob, active_job_id)
            self.assertIsNotNone(archived_document)
            self.assertIsNotNone(archived_job)
            self.assertIsNotNone(active_job)
            assert archived_document is not None
            assert archived_job is not None
            assert active_job is not None

            self.assertEqual(archived_document.status, DocumentStatus.UPLOADED)
            self.assertEqual(archived_job.status, ImportStatus.FAILED)
            self.assertEqual(archived_job.error_message, "Repairs for archived services cannot be modified")
            self.assertTrue((archived_job.summary or {}).get("failed_during_claim"))
            self.assertEqual(active_job.status, ImportStatus.PROCESSING)

    def test_worker_claim_skips_job_with_missing_vehicle_relation_and_returns_next_operational_job(self) -> None:
        headers = self._get_auth_headers()
        broken_payload = self._upload_order_document(headers)
        broken_document_id = broken_payload["document"]["id"]
        broken_job_id = broken_payload["job_id"]
        broken_repair_id = broken_payload["document"]["repair"]["id"]

        active_payload = self._upload_order_document(headers)
        active_job_id = active_payload["job_id"]

        with self.SessionLocal() as db:
            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(text("UPDATE repairs SET vehicle_id = 999999 WHERE id = :repair_id"), {"repair_id": broken_repair_id})
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        with self.SessionLocal() as db:
            from app.services.import_jobs import claim_next_document_processing_job

            claimed_job = claim_next_document_processing_job(db)
            self.assertIsNotNone(claimed_job)
            assert claimed_job is not None
            self.assertEqual(claimed_job.id, active_job_id)

        with self.SessionLocal() as db:
            broken_document = db.get(Document, broken_document_id)
            broken_job = db.get(ImportJob, broken_job_id)
            active_job = db.get(ImportJob, active_job_id)
            self.assertIsNotNone(broken_document)
            self.assertIsNotNone(broken_job)
            self.assertIsNotNone(active_job)
            assert broken_document is not None
            assert broken_job is not None
            assert active_job is not None

            self.assertEqual(broken_document.status, DocumentStatus.UPLOADED)
            self.assertEqual(broken_job.status, ImportStatus.FAILED)
            self.assertEqual(broken_job.error_message, "Document vehicle relation is incomplete")
            self.assertTrue((broken_job.summary or {}).get("failed_during_claim"))
            self.assertEqual(active_job.status, ImportStatus.PROCESSING)

    def test_reprocess_existing_only_primary_uses_canonical_source_document_when_primary_flag_drifted(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        canonical_document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        with self.SessionLocal() as db:
            self._mark_job_completed(db, job_id, canonical_document_id)
            repair = db.get(Repair, repair_id)
            canonical_document = db.get(Document, canonical_document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(canonical_document)
            assert repair is not None
            assert canonical_document is not None

            canonical_document.is_primary = False
            db.commit()

        response = self.client.post("/api/documents/reprocess-existing?only_primary=true&limit=10", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["processed_count"], 1)
        self.assertEqual(payload["document_ids"], [canonical_document_id])

    def test_process_pending_replaces_stale_ocr_payload_with_queued_reprocessing_version(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        with self.SessionLocal() as db:
            self._mark_job_completed(db, job_id, document_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.NEEDS_REVIEW
            document.ocr_confidence = 0.77
            next_version_number = max((item.version_number for item in document.versions), default=0) + 1
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=next_version_number,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "processor": "legacy-ocr",
                        "ocr_status": "completed",
                        "extracted_fields": {"order_number": "ZN-PENDING-STALE-001"},
                    },
                    field_confidence_map={"order_number": 0.77},
                    change_summary="Stale OCR payload before process-pending",
                )
            )
            db.commit()

        response = self.client.post("/api/documents/process-pending?limit=10", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        batch_payload = response.json()
        self.assertEqual(batch_payload["processed_count"], 1)
        self.assertEqual(batch_payload["document_ids"], [document_id])

        documents_response = self.client.get("/api/documents?limit=20", headers=headers)
        self.assertEqual(documents_response.status_code, 200, documents_response.text)
        document_payload = next(item for item in documents_response.json()["items"] if item["id"] == document_id)
        self.assertEqual(document_payload["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(document_payload["ocr_confidence"], None)
        self.assertEqual(document_payload["parsed_payload"]["ocr_status"], "queued")
        self.assertEqual(document_payload["parsed_payload"]["pipeline"], "reprocessing")
        self.assertNotIn("extracted_fields", document_payload["parsed_payload"])

    def test_process_pending_does_not_count_document_when_active_job_and_queued_state_already_exist(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        response = self.client.post("/api/documents/process-pending?limit=10", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        batch_payload = response.json()
        self.assertEqual(batch_payload["processed_count"], 0)
        self.assertEqual(batch_payload["document_ids"], [])
        self.assertEqual(batch_payload["job_ids"], [])

        with self.SessionLocal() as db:
            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            self.assertEqual([job.id for job in jobs], [job_id])
            self.assertEqual(jobs[0].status, ImportStatus.QUEUED)

    def test_document_list_returns_canonical_primary_flag_when_primary_marker_drifted(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            document.is_primary = False
            db.commit()

        response = self.client.get("/api/documents?limit=20", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        listed_document = next(item for item in response.json()["items"] if item["id"] == document_id)
        self.assertTrue(listed_document["is_primary"])

    def test_process_pending_rolls_back_batch_when_later_queueing_fails(self) -> None:
        headers = self._get_auth_headers()
        first_payload = self._upload_order_document(headers)
        second_payload = self._upload_order_document(headers)

        with self.SessionLocal() as db:
            first_job = db.get(ImportJob, first_payload["job_id"])
            second_job = db.get(ImportJob, second_payload["job_id"])
            first_document = db.get(Document, first_payload["document"]["id"])
            second_document = db.get(Document, second_payload["document"]["id"])
            self.assertIsNotNone(first_job)
            self.assertIsNotNone(second_job)
            self.assertIsNotNone(first_document)
            self.assertIsNotNone(second_document)
            assert first_job is not None
            assert second_job is not None
            assert first_document is not None
            assert second_document is not None
            first_job.status = ImportStatus.COMPLETED
            second_job.status = ImportStatus.COMPLETED
            first_document.status = DocumentStatus.UPLOADED
            second_document.status = DocumentStatus.UPLOADED
            db.commit()

        original_queue_document_processing = documents_api.queue_document_processing_result
        call_count = 0

        def queue_side_effect(db: Session, document_id: int, *, retry_failed: bool = False, recheck: bool = False):
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise RuntimeError("batch queue failed")
            return original_queue_document_processing(db, document_id, retry_failed=retry_failed, recheck=recheck)

        with patch.object(documents_api, "queue_document_processing_result", side_effect=queue_side_effect):
            with self.assertRaises(RuntimeError):
                self.client.post("/api/documents/process-pending?limit=10", headers=headers)

        with self.SessionLocal() as db:
            jobs = db.query(ImportJob).order_by(ImportJob.id.asc()).all()
            self.assertEqual(len(jobs), 2)
            self.assertTrue(all(job.status == ImportStatus.COMPLETED for job in jobs))

    def test_process_pending_skips_documents_from_archived_repairs(self) -> None:
        headers = self._get_auth_headers()
        archived_payload = self._upload_order_document(headers)
        active_payload = self._upload_order_document(headers)

        with self.SessionLocal() as db:
            self._mark_job_completed(db, archived_payload["job_id"], archived_payload["document"]["id"])
            self._mark_job_completed(db, active_payload["job_id"], active_payload["document"]["id"])

            archived_repair = db.get(Repair, archived_payload["document"]["repair"]["id"])
            archived_document = db.get(Document, archived_payload["document"]["id"])
            active_document = db.get(Document, active_payload["document"]["id"])
            self.assertIsNotNone(archived_repair)
            self.assertIsNotNone(archived_document)
            self.assertIsNotNone(active_document)
            assert archived_repair is not None
            assert archived_document is not None
            assert active_document is not None

            archived_repair.status = RepairStatus.ARCHIVED
            archived_document.status = DocumentStatus.NEEDS_REVIEW
            archived_document.review_queue_priority = 100
            active_document.status = DocumentStatus.NEEDS_REVIEW
            active_document.review_queue_priority = 100
            db.commit()

        response = self.client.post("/api/documents/process-pending?limit=10", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["processed_count"], 1)
        self.assertEqual(payload["document_ids"], [active_payload["document"]["id"]])

    def test_process_pending_skips_documents_from_archived_services(self) -> None:
        headers = self._get_auth_headers()
        archived_payload = self._upload_order_document(headers)
        active_payload = self._upload_order_document(headers)

        with self.SessionLocal() as db:
            self._mark_job_completed(db, archived_payload["job_id"], archived_payload["document"]["id"])
            self._mark_job_completed(db, active_payload["job_id"], active_payload["document"]["id"])

            archived_repair = db.get(Repair, archived_payload["document"]["repair"]["id"])
            archived_document = db.get(Document, archived_payload["document"]["id"])
            active_document = db.get(Document, active_payload["document"]["id"])
            self.assertIsNotNone(archived_repair)
            self.assertIsNotNone(archived_document)
            self.assertIsNotNone(active_document)
            assert archived_repair is not None
            assert archived_document is not None
            assert active_document is not None
            archived_service = Service(
                name="Archived Pending Batch Service",
                city="Kazan",
                status=ServiceStatus.CONFIRMED,
                created_by_user_id=1,
                confirmed_by_user_id=1,
            )
            db.add(archived_service)
            db.flush()
            archived_repair.service_id = archived_service.id

            archived_service.status = ServiceStatus.ARCHIVED
            archived_document.status = DocumentStatus.NEEDS_REVIEW
            archived_document.review_queue_priority = 100
            active_document.status = DocumentStatus.NEEDS_REVIEW
            active_document.review_queue_priority = 100
            db.commit()

        response = self.client.post("/api/documents/process-pending?limit=10", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["processed_count"], 1)
        self.assertEqual(payload["document_ids"], [active_payload["document"]["id"]])

    def test_reprocess_existing_skips_documents_from_archived_repairs(self) -> None:
        headers = self._get_auth_headers()
        archived_payload = self._upload_order_document(headers)
        active_payload = self._upload_order_document(headers)

        with self.SessionLocal() as db:
            self._mark_job_completed(db, archived_payload["job_id"], archived_payload["document"]["id"])
            self._mark_job_completed(db, active_payload["job_id"], active_payload["document"]["id"])

            archived_repair = db.get(Repair, archived_payload["document"]["repair"]["id"])
            archived_document = db.get(Document, archived_payload["document"]["id"])
            active_document = db.get(Document, active_payload["document"]["id"])
            self.assertIsNotNone(archived_repair)
            self.assertIsNotNone(archived_document)
            self.assertIsNotNone(active_document)
            assert archived_repair is not None
            assert archived_document is not None
            assert active_document is not None

            archived_repair.status = RepairStatus.ARCHIVED
            archived_document.status = DocumentStatus.CONFIRMED
            archived_document.review_queue_priority = 100
            active_document.status = DocumentStatus.CONFIRMED
            active_document.review_queue_priority = 100
            db.commit()

        response = self.client.post("/api/documents/reprocess-existing?limit=10", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["processed_count"], 1)
        self.assertEqual(payload["document_ids"], [active_payload["document"]["id"]])

    def test_reprocess_existing_skips_documents_from_archived_services(self) -> None:
        headers = self._get_auth_headers()
        archived_payload = self._upload_order_document(headers)
        active_payload = self._upload_order_document(headers)

        with self.SessionLocal() as db:
            self._mark_job_completed(db, archived_payload["job_id"], archived_payload["document"]["id"])
            self._mark_job_completed(db, active_payload["job_id"], active_payload["document"]["id"])

            archived_repair = db.get(Repair, archived_payload["document"]["repair"]["id"])
            archived_document = db.get(Document, archived_payload["document"]["id"])
            active_document = db.get(Document, active_payload["document"]["id"])
            self.assertIsNotNone(archived_repair)
            self.assertIsNotNone(archived_document)
            self.assertIsNotNone(active_document)
            assert archived_repair is not None
            assert archived_document is not None
            assert active_document is not None
            archived_service = Service(
                name="Archived Reprocess Batch Service",
                city="Kazan",
                status=ServiceStatus.CONFIRMED,
                created_by_user_id=1,
                confirmed_by_user_id=1,
            )
            db.add(archived_service)
            db.flush()
            archived_repair.service_id = archived_service.id

            archived_service.status = ServiceStatus.ARCHIVED
            archived_document.status = DocumentStatus.CONFIRMED
            archived_document.review_queue_priority = 100
            active_document.status = DocumentStatus.CONFIRMED
            active_document.review_queue_priority = 100
            db.commit()

        response = self.client.post("/api/documents/reprocess-existing?limit=10", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["processed_count"], 1)
        self.assertEqual(payload["document_ids"], [active_payload["document"]["id"]])

    def test_reprocess_existing_replaces_stale_ocr_payload_with_queued_reprocessing_version(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        with self.SessionLocal() as db:
            self._mark_job_completed(db, job_id, document_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.CONFIRMED
            document.ocr_confidence = 0.93
            next_version_number = max((item.version_number for item in document.versions), default=0) + 1
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=next_version_number,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "processor": "legacy-ocr",
                        "ocr_status": "completed",
                        "extracted_fields": {"order_number": "ZN-REPROCESS-STALE-001"},
                    },
                    field_confidence_map={"order_number": 0.93},
                    change_summary="Stale OCR payload before reprocess-existing",
                )
            )
            db.commit()

        response = self.client.post("/api/documents/reprocess-existing?limit=10", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        batch_payload = response.json()
        self.assertEqual(batch_payload["processed_count"], 1)
        self.assertEqual(batch_payload["document_ids"], [document_id])

        documents_response = self.client.get("/api/documents?limit=20", headers=headers)
        self.assertEqual(documents_response.status_code, 200, documents_response.text)
        document_payload = next(item for item in documents_response.json()["items"] if item["id"] == document_id)
        self.assertEqual(document_payload["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(document_payload["ocr_confidence"], None)
        self.assertEqual(document_payload["parsed_payload"]["ocr_status"], "queued")
        self.assertEqual(document_payload["parsed_payload"]["pipeline"], "reprocessing")
        self.assertNotIn("extracted_fields", document_payload["parsed_payload"])

    def test_reprocess_existing_does_not_count_document_when_active_job_and_queued_state_already_exist(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        response = self.client.post("/api/documents/reprocess-existing?limit=10", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        batch_payload = response.json()
        self.assertEqual(batch_payload["processed_count"], 0)
        self.assertEqual(batch_payload["document_ids"], [])
        self.assertEqual(batch_payload["job_ids"], [])
        self.assertEqual(batch_payload["status_counts"], {})

        with self.SessionLocal() as db:
            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            self.assertEqual([job.id for job in jobs], [job_id])
            self.assertEqual(jobs[0].status, ImportStatus.QUEUED)

    def test_reprocess_existing_rolls_back_batch_when_later_queueing_fails(self) -> None:
        headers = self._get_auth_headers()
        first_payload = self._upload_order_document(headers)
        second_payload = self._upload_order_document(headers)

        with self.SessionLocal() as db:
            self._mark_job_completed(db, first_payload["job_id"], first_payload["document"]["id"])
            self._mark_job_completed(db, second_payload["job_id"], second_payload["document"]["id"])

        original_queue_document_processing = documents_api.queue_document_processing_result
        call_count = 0

        def queue_side_effect(db: Session, document_id: int, *, retry_failed: bool = False, recheck: bool = False):
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise RuntimeError("batch reprocess failed")
            return original_queue_document_processing(db, document_id, retry_failed=retry_failed, recheck=recheck)

        with patch.object(documents_api, "queue_document_processing_result", side_effect=queue_side_effect):
            with self.assertRaises(RuntimeError):
                self.client.post("/api/documents/reprocess-existing?limit=10", headers=headers)

        with self.SessionLocal() as db:
            jobs = db.query(ImportJob).order_by(ImportJob.id.asc()).all()
            self.assertEqual(len(jobs), 2)
            self.assertTrue(all(job.status == ImportStatus.COMPLETED for job in jobs))

    def test_repair_detail_returns_executive_report_and_document_job_status(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        repair_payload = response.json()

        self.assertEqual(repair_payload["id"], repair_id)
        self.assertIn("executive_report", repair_payload)
        self.assertEqual(repair_payload["source_document_id"], document_id)
        self.assertIsInstance(repair_payload["executive_report"]["headline"], str)
        self.assertIsInstance(repair_payload["executive_report"]["summary"], str)
        self.assertIsInstance(repair_payload["executive_report"]["status"], str)
        self.assertIsInstance(repair_payload["executive_report"]["full_report_sections"], list)
        self.assertGreaterEqual(len(repair_payload["documents"]), 1)

        document_payload = next(item for item in repair_payload["documents"] if item["id"] == document_id)
        self.assertEqual(document_payload["latest_import_job"]["id"], job_id)
        self.assertEqual(document_payload["latest_import_job"]["status"], "queued")

    def test_repair_detail_uses_canonical_active_import_job_when_newer_completed_job_exists(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        document_id = payload["document"]["id"]
        initial_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            initial_job = db.get(ImportJob, initial_job_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(initial_job)
            self.assertIsNotNone(document)
            assert initial_job is not None
            assert document is not None

            initial_job.status = ImportStatus.PROCESSING
            initial_job.started_at = datetime(2025, 1, 2, 10, 0, 0)
            db.add(
                ImportJob(
                    document_id=document_id,
                    import_type="document_ocr",
                    source_filename=document.original_filename,
                    status=ImportStatus.COMPLETED,
                    summary={"document_id": document_id, "stage": "completed_after_active"},
                    error_message=None,
                    attempts=1,
                    started_at=None,
                    finished_at=None,
                )
            )
            db.commit()

        response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        repair_payload = response.json()
        document_payload = next(item for item in repair_payload["documents"] if item["id"] == document_id)
        self.assertEqual(document_payload["latest_import_job"]["id"], initial_job_id)
        self.assertEqual(document_payload["latest_import_job"]["status"], ImportStatus.PROCESSING.value)

    def test_repair_detail_does_not_expose_foreign_source_document_id_from_legacy_relation_drift(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        local_document_id = payload["document"]["id"]

        with self.SessionLocal() as db:
            local_repair = db.get(Repair, repair_id)
            self.assertIsNotNone(local_repair)
            assert local_repair is not None

            foreign_vehicle = Vehicle(
                external_id="truck-foreign-source",
                vehicle_type=VehicleType.TRUCK,
                plate_number="X123YZ116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(foreign_vehicle)
            db.flush()

            foreign_repair = Repair(
                order_number="ZN-FOREIGN-001",
                repair_date=date(2025, 1, 20),
                vehicle_id=foreign_vehicle.id,
                created_by_user_id=1,
                mileage=45000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            db.add(foreign_repair)
            db.flush()

            foreign_document = Document(
                repair_id=foreign_repair.id,
                uploaded_by_user_id=1,
                original_filename="foreign-source.pdf",
                storage_key="documents/test/foreign-source.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(foreign_document)
            db.flush()

            local_document = db.get(Document, local_document_id)
            self.assertIsNotNone(local_document)
            assert local_document is not None
            local_document.is_primary = False
            local_repair.source_document_id = foreign_document.id
            db.commit()

        response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        repair_payload = response.json()

        self.assertEqual(repair_payload["source_document_id"], local_document_id)
        self.assertEqual([item["id"] for item in repair_payload["documents"]], [local_document_id])
        self.assertTrue(repair_payload["documents"][0]["is_primary"])

    def test_repair_detail_returns_not_found_for_legacy_missing_vehicle_relation(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(text("UPDATE repairs SET vehicle_id = 999999 WHERE id = :repair_id"), {"repair_id": repair_id})
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Repair not found")

    def test_document_and_repair_detail_tolerate_non_object_metadata_payloads(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        document_id = payload["document"]["id"]

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(document)
            self.assertIsNotNone(repair)
            assert document is not None
            assert repair is not None
            next_version_number = max((item.version_number for item in document.versions), default=0) + 1

            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=next_version_number,
                    storage_key=document.storage_key,
                    parsed_payload=["legacy-broken-parsed-payload"],
                    field_confidence_map=["legacy-broken-confidence-map"],
                    change_summary="legacy broken metadata",
                )
            )
            db.add(
                RepairWork(
                    repair_id=repair.id,
                    work_code="A-1",
                    work_name="Legacy work",
                    quantity=1,
                    standard_hours=None,
                    actual_hours=None,
                    price=100,
                    line_total=100,
                    status=CatalogStatus.PRELIMINARY,
                    reference_payload=["legacy-broken-reference-payload"],
                )
            )
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="legacy_metadata_check",
                    severity=CheckSeverity.WARNING,
                    title="Legacy payload",
                    details=None,
                    calculation_payload=["legacy-broken-calculation-payload"],
                    is_resolved=False,
                )
            )
            db.add(
                AuditLog(
                    user_id=1,
                    entity_type="repair",
                    entity_id=str(repair.id),
                    action_type="legacy_metadata",
                    old_value=["legacy-broken-old-value"],
                    new_value=["legacy-broken-new-value"],
                )
            )
            db.commit()

        documents_response = self.client.get("/api/documents?limit=20", headers=headers)
        self.assertEqual(documents_response.status_code, 200, documents_response.text)
        document_payload = next(item for item in documents_response.json()["items"] if item["id"] == document_id)
        self.assertEqual(document_payload["parsed_payload"], {})

        repair_response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)
        self.assertEqual(repair_response.status_code, 200, repair_response.text)
        repair_payload = repair_response.json()

        repair_document_payload = next(item for item in repair_payload["documents"] if item["id"] == document_id)
        self.assertEqual(repair_document_payload["versions"][0]["parsed_payload"], {})
        self.assertEqual(repair_document_payload["versions"][0]["field_confidence_map"], {})
        self.assertEqual(repair_payload["works"][0]["reference_payload"], {})
        self.assertEqual(repair_payload["checks"][0]["calculation_payload"], {})
        self.assertEqual(repair_payload["history"][0]["old_value"], {})
        self.assertEqual(repair_payload["history"][0]["new_value"], {})

    def test_audit_log_tolerates_non_object_payloads(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            db.add(
                AuditLog(
                    user_id=1,
                    entity_type="repair",
                    entity_id="1",
                    action_type="legacy_audit_metadata",
                    old_value=["legacy-broken-old-value"],
                    new_value=["legacy-broken-new-value"],
                )
            )
            db.commit()

        response = self.client.get("/api/audit?limit=50", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        audit_item = next(item for item in payload["items"] if item["action_type"] == "legacy_audit_metadata")
        self.assertEqual(audit_item["old_value"], {})
        self.assertEqual(audit_item["new_value"], {})

    def test_noop_document_patch_returns_not_found_for_legacy_missing_vehicle_relation(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(text("UPDATE repairs SET vehicle_id = 999999 WHERE id = :repair_id"), {"repair_id": repair_id})
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        response = self.client.patch(f"/api/documents/{document_id}", headers=headers, json={})

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Document not found")

    def test_status_document_patch_returns_not_found_without_mutation_for_legacy_missing_vehicle_relation(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(text("UPDATE repairs SET vehicle_id = 999999 WHERE id = :repair_id"), {"repair_id": repair_id})
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        response = self.client.patch(
            f"/api/documents/{document_id}",
            headers=headers,
            json={"status": DocumentStatus.CONFIRMED.value},
        )

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Document not found")

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            self.assertEqual(document.status, DocumentStatus.UPLOADED)

    def test_process_document_returns_not_found_for_legacy_missing_vehicle_relation(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(text("UPDATE repairs SET vehicle_id = 999999 WHERE id = :repair_id"), {"repair_id": repair_id})
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        response = self.client.post(f"/api/documents/{document_id}/process", headers=headers)

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Document not found")

    def test_restore_document_returns_not_found_without_mutation_for_legacy_missing_vehicle_relation(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.ARCHIVED
            db.commit()

            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(text("UPDATE repairs SET vehicle_id = 999999 WHERE id = :repair_id"), {"repair_id": repair_id})
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        response = self.client.post(f"/api/documents/{document_id}/restore", headers=headers)

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Document not found")

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            self.assertEqual(document.status, DocumentStatus.ARCHIVED)

    def test_archived_document_cannot_become_primary(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            archived_document = Document(
                repair_id=repair_id,
                uploaded_by_user_id=1,
                original_filename="archived-repeat.pdf",
                storage_key="documents/test/archived-repeat.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.ARCHIVED,
                is_primary=False,
                review_queue_priority=0,
            )
            db.add(archived_document)
            db.commit()
            archived_document_id = archived_document.id

        response = self.client.post(f"/api/documents/{archived_document_id}/set-primary", headers=headers)

        self.assertEqual(response.status_code, 400, response.text)
        self.assertEqual(response.json()["detail"], "Archived documents cannot be primary")

    def test_set_primary_returns_not_found_for_legacy_missing_vehicle_relation(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        repair_id = payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(text("UPDATE repairs SET vehicle_id = 999999 WHERE id = :repair_id"), {"repair_id": repair_id})
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        response = self.client.post(f"/api/documents/{document_id}/set-primary", headers=headers)

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Document not found")

    def test_comparison_review_returns_not_found_without_mutation_for_legacy_missing_vehicle_relation(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        primary_document_id = payload["document"]["id"]

        with self.SessionLocal() as db:
            candidate_document = Document(
                repair_id=repair_id,
                uploaded_by_user_id=1,
                original_filename="candidate-repeat.pdf",
                storage_key="documents/test/candidate-repeat.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=20,
            )
            db.add(candidate_document)
            db.commit()
            candidate_document_id = candidate_document.id

            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(text("UPDATE repairs SET vehicle_id = 999999 WHERE id = :repair_id"), {"repair_id": repair_id})
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        response = self.client.post(
            f"/api/documents/{candidate_document_id}/compare/review",
            headers=headers,
            json={
                "with_document_id": primary_document_id,
                "action": "keep_current_primary",
                "comment": "Should fail before mutation",
            },
        )

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Document not found")

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            primary_document = db.get(Document, primary_document_id)
            candidate_document = db.get(Document, candidate_document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(primary_document)
            self.assertIsNotNone(candidate_document)
            assert repair is not None
            assert primary_document is not None
            assert candidate_document is not None
            self.assertEqual(repair.source_document_id, primary_document_id)
            self.assertTrue(primary_document.is_primary)
            self.assertFalse(candidate_document.is_primary)
            self.assertIsNone(candidate_document.notes)

    def test_document_cannot_be_compared_with_itself(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]

        response = self.client.get(
            f"/api/documents/{document_id}/compare",
            headers=headers,
            params={"with_document_id": document_id},
        )

        self.assertEqual(response.status_code, 400, response.text)
        self.assertEqual(response.json()["detail"], "Cannot compare a document with itself")

    def test_upload_multiple_images_merges_into_single_pdf_and_queues_job(self) -> None:
        headers = self._get_auth_headers()

        response = self.client.post(
            "/api/documents/upload",
            headers=headers,
            files=[
                ("files", ("page-1.png", self._make_png_bytes((220, 20, 60)), "image/png")),
                ("files", ("page-2.png", self._make_png_bytes((65, 105, 225)), "image/png")),
            ],
            data={"kind": "order"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document"]["mime_type"], "application/pdf")
        self.assertEqual(payload["document"]["source_type"], "pdf")
        self.assertEqual(payload["document"]["parsed_payload"]["upload_mode"], "merged_images")
        self.assertEqual(len(payload["document"]["parsed_payload"]["uploaded_files"]), 2)
        self.assertTrue(payload["document"]["original_filename"].endswith(".pdf"))

        with self.SessionLocal() as db:
            document = db.get(Document, payload["document"]["id"])
            self.assertIsNotNone(document)
            assert document is not None
            stored_path = self.storage_root / document.storage_key
            self.assertTrue(stored_path.exists())
            self.assertTrue(stored_path.read_bytes().startswith(b"%PDF"))

    def test_upload_multiple_images_returns_400_when_image_merge_reader_fails(self) -> None:
        headers = self._get_auth_headers()

        with patch("app.services.pdf_tools.Image.open", side_effect=OSError("broken image stream")):
            response = self.client.post(
                "/api/documents/upload",
                headers=headers,
                files=[
                    ("files", ("page-1.png", b"\x89PNG\r\n\x1a\nbroken-1", "image/png")),
                    ("files", ("page-2.png", b"\x89PNG\r\n\x1a\nbroken-2", "image/png")),
                ],
                data={"kind": "order"},
            )

        self.assertEqual(response.status_code, 400, response.text)
        self.assertEqual(response.json()["detail"], "Один из выбранных файлов не удалось прочитать как изображение")

    def test_single_file_upload_validates_size_before_reading_content(self) -> None:
        headers = self._get_auth_headers()

        with (
            patch.object(upload_validation_api.settings, "max_document_upload_size_bytes", 1),
            patch.object(
                documents_api,
                "read_uploaded_file_bytes",
                side_effect=AssertionError("upload content must not be read before validation"),
            ),
        ):
            response = self.client.post(
                "/api/documents/upload",
                headers=headers,
                files={"file": ("too-large.pdf", b"%PDF-1.4\n%too-large\n", "application/pdf")},
                data={"kind": "order"},
            )

        self.assertEqual(response.status_code, 413, response.text)

    def test_worker_processing_uses_overridden_storage_root(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        sample_text = """
Общество с ограниченной ответственностью "ЛидерТрак"
НАРЯД-ЗАКАЗ № ЛТ250012276 от 25.12.2025
Автомобиль:
FH13A42T, гос. номер: 879КВА716, шасси: YV2RT40A7LA856012, пробег: 172274
Причина
обращения:
ТО основное с заменой масла+салонный фильтр
На холостом ходу при малых оборотах вибрации по кабине.
Выполненные сервисные услуги и использованные материалы
1 ZZ000253133903 Масло моторное синтетическое DONGFENG Diesel Ultra CS, EURO-6 10W40, бочка 205 л. 37 литр 447,16 992,68 15 717,50 3 143,50 18 861,00
Рекомендации:
неисправности по вибрации по кабине на момент осмотра не обнаружено, подушки двс в норме
"""

        with patch("app.services.document_processing.extract_document_text", return_value=(sample_text, "pdf_text", None)):
            from app.services.import_jobs import claim_next_document_processing_job, run_document_processing_job

            with self.SessionLocal() as db:
                next_job = claim_next_document_processing_job(db)
                self.assertIsNotNone(next_job)
                assert next_job is not None
                self.assertEqual(next_job.id, job_id)

            with self.SessionLocal() as db:
                attached_job = db.get(ImportJob, job_id)
                self.assertIsNotNone(attached_job)
                assert attached_job is not None
                run_document_processing_job(db, attached_job)

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            self.assertNotEqual(document.status, DocumentStatus.OCR_ERROR)
            latest_job = db.get(ImportJob, job_id)
            self.assertIsNotNone(latest_job)
            assert latest_job is not None
            self.assertIn(latest_job.status, {ImportStatus.COMPLETED, ImportStatus.COMPLETED_WITH_CONFLICTS})
            self.assertTrue((self.storage_root / document.storage_key).exists())

    def test_database_rejects_queued_duplicate_while_document_is_processing(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        active_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            active_job = db.get(ImportJob, active_job_id)
            self.assertIsNotNone(active_job)
            assert active_job is not None

            active_job.status = ImportStatus.PROCESSING
            active_job.started_at = datetime(2025, 1, 2, 10, 0, 0)
            db.commit()

        with self.SessionLocal() as db:
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None

            db.add(
                ImportJob(
                    document_id=document_id,
                    import_type="document_ocr",
                    source_filename=document.original_filename,
                    status=ImportStatus.QUEUED,
                    summary={"document_id": document_id, "stage": "queued_duplicate"},
                    error_message=None,
                    attempts=0,
                    started_at=None,
                    finished_at=None,
                )
            )
            with self.assertRaises(IntegrityError):
                db.commit()
            db.rollback()

            jobs = db.query(ImportJob).filter(ImportJob.document_id == document_id).order_by(ImportJob.id.asc()).all()
            self.assertEqual(len(jobs), 1)
            self.assertEqual(jobs[0].id, active_job_id)
            self.assertEqual(jobs[0].status, ImportStatus.PROCESSING)

    def test_worker_claim_marks_legacy_queued_duplicate_as_failed(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        canonical_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            db.execute(text("drop index if exists uq_import_jobs_active_document_ocr"))
            document = db.get(Document, document_id)
            self.assertIsNotNone(document)
            assert document is not None
            duplicate_job = ImportJob(
                document_id=document_id,
                import_type="document_ocr",
                source_filename=document.original_filename,
                status=ImportStatus.RETRY,
                summary={"document_id": document_id, "stage": "legacy_duplicate_retry"},
                error_message=None,
                attempts=1,
                started_at=None,
                finished_at=None,
            )
            db.add(duplicate_job)
            db.commit()
            db.refresh(duplicate_job)
            duplicate_job_id = duplicate_job.id

        with self.SessionLocal() as db:
            from app.services.import_jobs import claim_next_document_processing_job

            claimed_job = claim_next_document_processing_job(db)
            self.assertIsNotNone(claimed_job)
            assert claimed_job is not None
            self.assertEqual(claimed_job.id, canonical_job_id)

        with self.SessionLocal() as db:
            canonical_job = db.get(ImportJob, canonical_job_id)
            duplicate_job = db.get(ImportJob, duplicate_job_id)
            self.assertIsNotNone(canonical_job)
            self.assertIsNotNone(duplicate_job)
            assert canonical_job is not None
            assert duplicate_job is not None
            self.assertEqual(canonical_job.status, ImportStatus.PROCESSING)
            self.assertEqual(duplicate_job.status, ImportStatus.FAILED)
            self.assertEqual(duplicate_job.error_message, "Superseded by canonical OCR job during worker claim")
            self.assertEqual((duplicate_job.summary or {}).get("superseded_by_job_id"), canonical_job_id)

    def test_upload_rolls_back_document_and_file_when_queueing_fails(self) -> None:
        headers = self._get_auth_headers()
        files_before = sorted(
            str(path.relative_to(self.storage_root))
            for path in self.storage_root.rglob("*")
            if path.is_file()
        )

        with patch.object(documents_api, "queue_document_processing", side_effect=RuntimeError("queue unavailable")):
            with self.assertRaises(RuntimeError):
                self.client.post(
                    "/api/documents/upload",
                    headers=headers,
                    files={"file": ("job-test.pdf", b"%PDF-1.4\n%test\n", "application/pdf")},
                    data={"kind": "order"},
                )

        with self.SessionLocal() as db:
            self.assertEqual(db.query(Document).count(), 0)
            self.assertEqual(db.query(Repair).count(), 0)
            self.assertEqual(db.query(ImportJob).count(), 0)

        files_after = sorted(
            str(path.relative_to(self.storage_root))
            for path in self.storage_root.rglob("*")
            if path.is_file()
        )
        self.assertEqual(files_after, files_before)

    def test_upload_rejects_archived_vehicle(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            archived_vehicle = Vehicle(
                external_id="truck-archived-upload",
                vehicle_type=VehicleType.TRUCK,
                plate_number="A777AA116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ARCHIVED,
            )
            db.add(archived_vehicle)
            db.commit()
            archived_vehicle_id = archived_vehicle.id

        response = self.client.post(
            "/api/documents/upload",
            headers=headers,
            files={"file": ("job-test.pdf", b"%PDF-1.4\n%test\n", "application/pdf")},
            data={"kind": "order", "vehicle_id": str(archived_vehicle_id)},
        )

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Archived vehicles cannot be used in operational actions")

        with self.SessionLocal() as db:
            self.assertEqual(db.query(Document).count(), 0)
            self.assertEqual(db.query(Repair).count(), 0)
            self.assertEqual(db.query(ImportJob).count(), 0)

    def test_upload_rejects_archived_placeholder_vehicle(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            db.add(
                Vehicle(
                    external_id="__batch_import_placeholder__",
                    vehicle_type=VehicleType.TRUCK,
                    plate_number="PLACEHOLDER-ARCHIVED-UPLOAD",
                    brand="Placeholder",
                    model="Upload",
                    status=VehicleStatus.ARCHIVED,
                )
            )
            db.commit()

        response = self.client.post(
            "/api/documents/upload",
            headers=headers,
            files={"file": ("job-test.pdf", b"%PDF-1.4\n%test\n", "application/pdf")},
            data={"kind": "order"},
        )

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Archived vehicles cannot be used in operational actions")

        with self.SessionLocal() as db:
            self.assertEqual(db.query(Document).count(), 0)
            self.assertEqual(db.query(Repair).count(), 0)
            self.assertEqual(db.query(ImportJob).count(), 0)

    def test_upload_to_repair_rejects_archived_service(self) -> None:
        headers = self._get_auth_headers()
        initial_payload = self._upload_order_document(headers)
        repair_id = initial_payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            archived_service = Service(
                name="Archived Upload Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
                confirmed_by_user_id=1,
            )
            db.add(archived_service)
            db.flush()

            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            repair.service_id = archived_service.id
            db.commit()

        response = self.client.post(
            "/api/documents/upload-to-repair",
            headers=headers,
            files={"file": ("repair-attach.pdf", b"%PDF-1.4\n%repair-attach\n", "application/pdf")},
            data={"repair_id": str(repair_id), "kind": "repeat_scan"},
        )

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Repairs for archived services cannot be modified")

    def test_link_vehicle_rolls_back_when_queueing_fails(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-ROLLBACK",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            target_vehicle = Vehicle(
                external_id="truck-target-rollback",
                vehicle_type=VehicleType.TRUCK,
                plate_number="M111MM116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, target_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-LINK-ROLLBACK-001",
                repair_date=date(2025, 1, 21),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=150000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="rollback-order.pdf",
                storage_key="documents/test/rollback-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id
            placeholder_vehicle_id = placeholder_vehicle.id
            target_vehicle_id = target_vehicle.id

        with patch.object(documents_api, "queue_document_processing", side_effect=RuntimeError("queue unavailable")):
            with self.assertRaises(RuntimeError):
                self.client.post(
                    f"/api/documents/{document_id}/link-vehicle",
                    headers=headers,
                    json={"vehicle_id": target_vehicle_id},
                )

        with self.SessionLocal() as db:
            refreshed_document = db.get(Document, document_id)
            refreshed_repair = db.get(Repair, repair_id)
            self.assertIsNotNone(refreshed_document)
            self.assertIsNotNone(refreshed_repair)
            assert refreshed_document is not None
            assert refreshed_repair is not None

            self.assertEqual(refreshed_document.status, DocumentStatus.NEEDS_REVIEW)
            self.assertEqual(refreshed_repair.vehicle_id, placeholder_vehicle_id)
            self.assertEqual(db.query(ImportJob).filter(ImportJob.document_id == document_id).count(), 0)

    def test_link_vehicle_response_includes_latest_import_job(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-LINK-SUCCESS",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            target_vehicle = Vehicle(
                external_id="truck-target-success",
                vehicle_type=VehicleType.TRUCK,
                plate_number="T700TT116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, target_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-LINK-SUCCESS-001",
                repair_date=date(2025, 1, 25),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=154000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="link-success-order.pdf",
                storage_key="documents/test/link-success-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id
            target_vehicle_id = target_vehicle.id

        response = self.client.post(
            f"/api/documents/{document_id}/link-vehicle",
            headers=headers,
            json={"vehicle_id": target_vehicle_id},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["repair_id"], repair_id)
        self.assertEqual(payload["job_id"], payload["document"]["latest_import_job"]["id"])
        self.assertEqual(payload["import_status"], payload["document"]["latest_import_job"]["status"])
        self.assertEqual(payload["document"]["vehicle"]["id"], target_vehicle_id)

    def test_link_vehicle_writes_processing_audit_when_reprocessing_reopens_repair(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-LINK-AUDIT",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            target_vehicle = Vehicle(
                external_id="truck-target-link-audit",
                vehicle_type=VehicleType.TRUCK,
                plate_number="А700АА116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, target_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-LINK-AUDIT-001",
                repair_date=date(2025, 2, 1),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=160000,
                status=RepairStatus.EMPLOYEE_CONFIRMED,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="link-audit-order.pdf",
                storage_key="documents/test/link-audit-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.CONFIRMED,
                is_primary=True,
                review_queue_priority=20,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id
            target_vehicle_id = target_vehicle.id

        response = self.client.post(
            f"/api/documents/{document_id}/link-vehicle",
            headers=headers,
            json={"vehicle_id": target_vehicle_id},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document"]["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(payload["document"]["repair"]["status"], RepairStatus.IN_REVIEW.value)

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertEqual(document.status, DocumentStatus.UPLOADED)

            audit_entry = db.query(AuditLog).filter(
                AuditLog.entity_type == "document",
                AuditLog.entity_id == str(document_id),
                AuditLog.action_type == "document_processing_queued",
            ).order_by(AuditLog.id.desc()).first()
            self.assertIsNotNone(audit_entry)
            assert audit_entry is not None
            self.assertEqual(audit_entry.old_value["status"], DocumentStatus.CONFIRMED.value)
            self.assertEqual(audit_entry.new_value["status"], DocumentStatus.UPLOADED.value)
            self.assertEqual(audit_entry.old_value["repair_status"], RepairStatus.EMPLOYEE_CONFIRMED.value)
            self.assertEqual(audit_entry.new_value["repair_status"], RepairStatus.IN_REVIEW.value)

    def test_link_vehicle_clears_vehicle_manual_review_state_before_reprocessing(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-LINK-WARNING",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            target_vehicle = Vehicle(
                external_id="truck-target-warning",
                vehicle_type=VehicleType.TRUCK,
                plate_number="Р700РР116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, target_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-LINK-WARNING-001",
                repair_date=date(2025, 1, 28),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=157000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="link-warning-order.pdf",
                storage_key="documents/test/link-warning-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_fields": {"plate_number": "Р700РР116"},
                        "manual_review_reasons": ["vehicle_not_found"],
                    },
                    field_confidence_map={},
                    change_summary="Initial OCR payload",
                )
            )
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_vehicle_not_found",
                    severity=CheckSeverity.WARNING,
                    title="Техника не найдена в базе",
                    details="Нужна ручная привязка техники",
                    calculation_payload={"reason": "vehicle_not_found"},
                    is_resolved=False,
                )
            )
            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id
            target_vehicle_id = target_vehicle.id

        response = self.client.post(
            f"/api/documents/{document_id}/link-vehicle",
            headers=headers,
            json={"vehicle_id": target_vehicle_id},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document"]["vehicle"]["id"], target_vehicle_id)
        self.assertEqual(payload["document"]["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(payload["document"]["parsed_payload"]["ocr_status"], "queued")
        self.assertEqual(payload["document"]["parsed_payload"]["pipeline"], "reprocessing")
        self.assertNotIn("manual_review_reasons", payload["document"]["parsed_payload"])

        repair_response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)
        self.assertEqual(repair_response.status_code, 200, repair_response.text)
        repair_payload = repair_response.json()
        vehicle_checks = [
            item
            for item in repair_payload["checks"]
            if item["check_type"] in {"ocr_vehicle_not_found", "ocr_vehicle_missing"}
        ]
        self.assertEqual(len(vehicle_checks), 1)
        self.assertTrue(vehicle_checks[0]["is_resolved"])

    def test_upload_to_repair_rolls_back_document_and_file_when_queueing_fails(self) -> None:
        headers = self._get_auth_headers()
        initial_payload = self._upload_order_document(headers)
        repair_id = initial_payload["document"]["repair"]["id"]

        with self.SessionLocal() as db:
            initial_document_ids = {
                row[0]
                for row in db.query(Document.id).filter(Document.repair_id == repair_id).all()
            }

        files_before = sorted(
            str(path.relative_to(self.storage_root))
            for path in self.storage_root.rglob("*")
            if path.is_file()
        )

        with patch.object(documents_api, "queue_document_processing", side_effect=RuntimeError("queue unavailable")):
            with self.assertRaises(RuntimeError):
                self.client.post(
                    "/api/documents/upload-to-repair",
                    headers=headers,
                    files={"file": ("repair-attach.pdf", b"%PDF-1.4\n%repair-attach\n", "application/pdf")},
                    data={"repair_id": str(repair_id), "kind": "repeat_scan"},
                )

        with self.SessionLocal() as db:
            document_ids_after = {
                row[0]
                for row in db.query(Document.id).filter(Document.repair_id == repair_id).all()
            }
            self.assertEqual(document_ids_after, initial_document_ids)
            self.assertEqual(db.query(ImportJob).count(), 1)

        files_after = sorted(
            str(path.relative_to(self.storage_root))
            for path in self.storage_root.rglob("*")
            if path.is_file()
        )
        self.assertEqual(files_after, files_before)

    def test_create_vehicle_rolls_back_when_queueing_fails(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-CREATE-ROLLBACK",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            db.add(placeholder_vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-CREATE-ROLLBACK-001",
                repair_date=date(2025, 1, 22),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=151000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="create-rollback-order.pdf",
                storage_key="documents/test/create-rollback-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id
            placeholder_vehicle_id = placeholder_vehicle.id

        with patch.object(documents_api, "queue_document_processing", side_effect=RuntimeError("queue unavailable")):
            with self.assertRaises(RuntimeError):
                self.client.post(
                    f"/api/documents/{document_id}/create-vehicle",
                    headers=headers,
                    json={
                        "vehicle_type": "truck",
                        "plate_number": "К123КК116",
                        "vin": "YV2RT40A7LA856012",
                        "brand": "Volvo",
                        "model": "FH",
                        "year": 2020,
                        "comment": "rollback create vehicle",
                    },
                )

        with self.SessionLocal() as db:
            refreshed_document = db.get(Document, document_id)
            refreshed_repair = db.get(Repair, repair_id)
            self.assertIsNotNone(refreshed_document)
            self.assertIsNotNone(refreshed_repair)
            assert refreshed_document is not None
            assert refreshed_repair is not None

            self.assertEqual(refreshed_document.status, DocumentStatus.NEEDS_REVIEW)
            self.assertEqual(refreshed_repair.vehicle_id, placeholder_vehicle_id)
            self.assertEqual(
                db.query(Vehicle).filter(Vehicle.plate_number == "К123КК116").count(),
                0,
            )
            self.assertEqual(db.query(ImportJob).filter(ImportJob.document_id == document_id).count(), 0)

    def test_create_vehicle_ignores_archived_matching_vehicle_and_creates_new_vehicle(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-CREATE-ARCHIVED",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            archived_vehicle = Vehicle(
                external_id="truck-archived-match",
                vehicle_type=VehicleType.TRUCK,
                plate_number="К123КК116",
                vin="YV2RT40A7LA856012",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ARCHIVED,
            )
            db.add_all([placeholder_vehicle, archived_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-CREATE-ARCHIVED-001",
                repair_date=date(2025, 1, 23),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=152000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="create-archived-order.pdf",
                storage_key="documents/test/create-archived-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id
            placeholder_vehicle_id = placeholder_vehicle.id

        response = self.client.post(
            f"/api/documents/{document_id}/create-vehicle",
            headers=headers,
            json={
                "vehicle_type": "truck",
                "plate_number": "К123КК116",
                "vin": "YV2RT40A7LA856012",
                "brand": "Volvo",
                "model": "FH",
                "year": 2020,
                "comment": "archived match must be rejected",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertTrue(payload["created_new_vehicle"])
        self.assertEqual(payload["repair_id"], repair_id)
        self.assertNotEqual(payload["document"]["vehicle"]["id"], placeholder_vehicle_id)
        self.assertEqual(payload["document"]["vehicle"]["plate_number"], "К123КК116")

        with self.SessionLocal() as db:
            refreshed_repair = db.get(Repair, repair_id)
            self.assertIsNotNone(refreshed_repair)
            assert refreshed_repair is not None
            self.assertNotEqual(refreshed_repair.vehicle_id, placeholder_vehicle_id)
            self.assertEqual(
                db.query(Vehicle)
                .filter(
                    Vehicle.plate_number == "К123КК116",
                    Vehicle.status == VehicleStatus.ACTIVE,
                )
                .count(),
                1,
            )
            self.assertEqual(db.query(ImportJob).filter(ImportJob.document_id == document_id).count(), 1)

    def test_create_vehicle_response_includes_latest_import_job(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-CREATE-SUCCESS",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            db.add(placeholder_vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-CREATE-SUCCESS-001",
                repair_date=date(2025, 1, 26),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=155000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="create-success-order.pdf",
                storage_key="documents/test/create-success-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id

        response = self.client.post(
            f"/api/documents/{document_id}/create-vehicle",
            headers=headers,
            json={
                "vehicle_type": "truck",
                "plate_number": "С700СС116",
                "vin": "YV2RT40A7LA856099",
                "brand": "Volvo",
                "model": "FH",
                "year": 2020,
                "comment": "create success",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertTrue(payload["created_new_vehicle"])
        self.assertEqual(payload["repair_id"], repair_id)
        self.assertEqual(payload["job_id"], payload["document"]["latest_import_job"]["id"])
        self.assertEqual(payload["import_status"], payload["document"]["latest_import_job"]["status"])
        self.assertEqual(payload["document"]["vehicle"]["plate_number"], "С700СС116")

    def test_create_vehicle_clears_vehicle_manual_review_state_before_reprocessing(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-CREATE-WARNING",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            db.add(placeholder_vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-CREATE-WARNING-001",
                repair_date=date(2025, 1, 29),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=158000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="create-warning-order.pdf",
                storage_key="documents/test/create-warning-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_fields": {"plate_number": "С800СС116"},
                        "manual_review_reasons": ["vehicle_missing"],
                    },
                    field_confidence_map={},
                    change_summary="Initial OCR payload",
                )
            )
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_vehicle_missing",
                    severity=CheckSeverity.WARNING,
                    title="Не удалось определить технику",
                    details="Нужна ручная привязка техники",
                    calculation_payload={"reason": "vehicle_missing"},
                    is_resolved=False,
                )
            )
            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id

        response = self.client.post(
            f"/api/documents/{document_id}/create-vehicle",
            headers=headers,
            json={
                "vehicle_type": "truck",
                "plate_number": "С800СС116",
                "vin": "YV2RT40A7LA856100",
                "brand": "Volvo",
                "model": "FH",
                "year": 2020,
                "comment": "create warning clear",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document"]["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(payload["document"]["parsed_payload"]["ocr_status"], "queued")
        self.assertEqual(payload["document"]["parsed_payload"]["pipeline"], "reprocessing")
        self.assertNotIn("manual_review_reasons", payload["document"]["parsed_payload"])

        repair_response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)
        self.assertEqual(repair_response.status_code, 200, repair_response.text)
        repair_payload = repair_response.json()
        vehicle_checks = [
            item
            for item in repair_payload["checks"]
            if item["check_type"] in {"ocr_vehicle_not_found", "ocr_vehicle_missing"}
        ]
        self.assertEqual(len(vehicle_checks), 1)
        self.assertTrue(vehicle_checks[0]["is_resolved"])

    def test_create_vehicle_prefers_active_match_over_archived_duplicate(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-CREATE-ACTIVE-MATCH",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            archived_vehicle = Vehicle(
                external_id="truck-archived-duplicate",
                vehicle_type=VehicleType.TRUCK,
                plate_number="К456КК116",
                vin="YV2RT40A7LA856013",
                brand="Volvo",
                model="FH archived",
                status=VehicleStatus.ARCHIVED,
            )
            active_vehicle = Vehicle(
                external_id="truck-active-duplicate",
                vehicle_type=VehicleType.TRUCK,
                plate_number="К456КК116",
                vin="YV2RT40A7LA856013",
                brand="Volvo",
                model="FH active",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, archived_vehicle, active_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-CREATE-ACTIVE-MATCH-001",
                repair_date=date(2025, 1, 24),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=153000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="create-active-match-order.pdf",
                storage_key="documents/test/create-active-match-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id
            active_vehicle_id = active_vehicle.id

        response = self.client.post(
            f"/api/documents/{document_id}/create-vehicle",
            headers=headers,
            json={
                "vehicle_type": "truck",
                "plate_number": "К456КК116",
                "vin": "YV2RT40A7LA856013",
                "brand": "Volvo",
                "model": "FH",
                "year": 2020,
                "comment": "active match should win",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertFalse(payload["created_new_vehicle"])
        self.assertEqual(payload["document"]["vehicle"]["id"], active_vehicle_id)
        self.assertEqual(payload["repair_id"], repair_id)
        self.assertEqual(payload["document"]["vehicle"]["plate_number"], "К456КК116")

        with self.SessionLocal() as db:
            refreshed_repair = db.get(Repair, repair_id)
            self.assertIsNotNone(refreshed_repair)
            assert refreshed_repair is not None
            self.assertEqual(refreshed_repair.vehicle_id, active_vehicle_id)
            self.assertEqual(db.query(ImportJob).filter(ImportJob.document_id == document_id).count(), 1)

    def test_create_vehicle_reuses_active_vehicle_for_shifted_ocr_plate_format(self) -> None:
        headers = self._get_auth_headers()

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-CREATE-SHIFTED",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            active_vehicle = Vehicle(
                external_id="truck-shifted-match",
                vehicle_type=VehicleType.TRUCK,
                plate_number="К879ВА/716",
                vin=None,
                brand="Volvo",
                model="FH active",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, active_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-CREATE-SHIFTED-MATCH-001",
                repair_date=date(2025, 1, 27),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=1,
                mileage=156000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="create-shifted-match-order.pdf",
                storage_key="documents/test/create-shifted-match-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()

            document_id = document.id
            repair_id = repair.id
            active_vehicle_id = active_vehicle.id

        response = self.client.post(
            f"/api/documents/{document_id}/create-vehicle",
            headers=headers,
            json={
                "vehicle_type": "truck",
                "plate_number": "879КВА716",
                "vin": None,
                "brand": "Volvo",
                "model": "FH",
                "year": 2020,
                "comment": "shifted plate should match existing vehicle",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertFalse(payload["created_new_vehicle"])
        self.assertEqual(payload["document"]["vehicle"]["id"], active_vehicle_id)
        self.assertEqual(payload["repair_id"], repair_id)

        with self.SessionLocal() as db:
            refreshed_repair = db.get(Repair, repair_id)
            self.assertIsNotNone(refreshed_repair)
            assert refreshed_repair is not None
            self.assertEqual(refreshed_repair.vehicle_id, active_vehicle_id)
            self.assertEqual(db.query(Vehicle).filter(Vehicle.status == VehicleStatus.ACTIVE).count(), 2)
            self.assertEqual(db.query(ImportJob).filter(ImportJob.document_id == document_id).count(), 1)

    def test_export_status_and_warnings_follow_executive_report_findings(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        job_id = payload["job_id"]

        sample_text = """
Общество с ограниченной ответственностью "ЛидерТрак"
НАРЯД-ЗАКАЗ № ЛТ250012276 от 25.12.2025
Автомобиль:
FH13A42T, гос. номер: 879КВА716, шасси: YV2RT40A7LA856012, пробег: 172274
Причина
обращения:
ТО основное с заменой масла+салонный фильтр
На холостом ходу при малых оборотах вибрации по кабине.
Выполненные сервисные услуги и использованные материалы
1 ZZ000253133903 Масло моторное синтетическое DONGFENG Diesel Ultra CS, EURO-6 10W40, бочка 205 л. 37 литр 447,16 992,68 15 717,50 3 143,50 18 861,00
Рекомендации:
неисправности по вибрации по кабине на момент осмотра не обнаружено, подушки двс в норме
"""

        with patch("app.services.document_processing.extract_document_text", return_value=(sample_text, "pdf_text", None)):
            from app.services.import_jobs import claim_next_document_processing_job, run_document_processing_job

            with self.SessionLocal() as db:
                next_job = claim_next_document_processing_job(db)
                self.assertIsNotNone(next_job)
                assert next_job is not None
                self.assertEqual(next_job.id, job_id)

            with self.SessionLocal() as db:
                attached_job = db.get(ImportJob, job_id)
                self.assertIsNotNone(attached_job)
                assert attached_job is not None
                run_document_processing_job(db, attached_job)

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            report_status, report_status_comment = build_report_status_summary(repair)
            warning_rows = build_export_warning_rows(repair)

        self.assertEqual(report_status, "Есть критичные несоответствия")
        self.assertIn("ручная проверка", report_status_comment)
        self.assertTrue(any(row[2] == "По заявленной вибрации проблема не подтверждена" for row in warning_rows))
        self.assertTrue(any(row[2] == "Моторное масло требует проверки на соответствие Volvo" for row in warning_rows))

    def test_archived_export_warning_rows_use_archived_source_document_ocr_confidence(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        document_id = payload["document"]["id"]

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            repair.status = RepairStatus.ARCHIVED
            repair.source_document_id = document.id
            document.status = DocumentStatus.ARCHIVED
            document.ocr_confidence = 0.61
            db.commit()
            db.refresh(repair)

            warning_rows = build_export_warning_rows(repair, include_archived_fallback=True)

        self.assertTrue(
            any(row[2] == "Низкая уверенность OCR по основному документу" for row in warning_rows),
            warning_rows,
        )

    def test_report_status_stays_in_ocr_queue_when_source_document_job_is_processing(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        document_id = payload["document"]["id"]
        job_id = payload["job_id"]

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            job = db.get(ImportJob, job_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            self.assertIsNotNone(job)
            assert repair is not None
            assert document is not None
            assert job is not None

            repair.source_document_id = document.id
            document.status = DocumentStatus.RECOGNIZED
            job.status = ImportStatus.PROCESSING
            db.add(repair)
            db.add(document)
            db.add(job)
            db.commit()
            db.refresh(repair)

            report_status, report_status_comment = build_report_status_summary(repair)

        self.assertEqual(report_status, "В очереди OCR")
        self.assertIn("обработке", report_status_comment)

    def test_report_status_uses_canonical_active_import_job_when_newer_completed_job_exists(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        document_id = payload["document"]["id"]
        initial_job_id = payload["job_id"]

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            initial_job = db.get(ImportJob, initial_job_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            self.assertIsNotNone(initial_job)
            assert repair is not None
            assert document is not None
            assert initial_job is not None

            repair.source_document_id = document.id
            document.status = DocumentStatus.RECOGNIZED
            initial_job.status = ImportStatus.PROCESSING
            initial_job.started_at = datetime(2025, 1, 2, 10, 0, 0)
            db.add(
                ImportJob(
                    document_id=document_id,
                    import_type="document_ocr",
                    source_filename=document.original_filename,
                    status=ImportStatus.COMPLETED,
                    summary={"document_id": document_id, "stage": "completed_after_active"},
                    error_message=None,
                    attempts=1,
                    started_at=None,
                    finished_at=None,
                )
            )
            db.commit()
            db.refresh(repair)

            report_status, report_status_comment = build_report_status_summary(repair)

        self.assertEqual(report_status, "В очереди OCR")
        self.assertIn("обработке", report_status_comment)

    def test_attach_multiple_images_to_existing_repair_merges_into_single_pdf(self) -> None:
        headers = self._get_auth_headers()
        initial_payload = self._upload_order_document(headers)
        repair_id = initial_payload["document"]["repair"]["id"]

        response = self.client.post(
            "/api/documents/upload-to-repair",
            headers=headers,
            files=[
                ("files", ("scan-1.png", self._make_png_bytes((34, 139, 34)), "image/png")),
                ("files", ("scan-2.png", self._make_png_bytes((255, 140, 0)), "image/png")),
            ],
            data={"repair_id": str(repair_id), "kind": "repeat_scan"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document"]["repair"]["id"], repair_id)
        self.assertEqual(payload["document"]["mime_type"], "application/pdf")
        self.assertEqual(payload["document"]["source_type"], "pdf")
        self.assertEqual(payload["document"]["parsed_payload"]["upload_mode"], "merged_images")
        self.assertFalse(payload["document"]["is_primary"])

    def test_repair_and_vehicle_pdf_exports_return_pdf_files(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        repair_id = payload["document"]["repair"]["id"]
        vehicle_id = payload["document"]["vehicle"]["id"]

        repair_response = self.client.get(f"/api/repairs/{repair_id}/export.pdf", headers=headers)
        self.assertEqual(repair_response.status_code, 200, repair_response.text)
        self.assertIn("application/pdf", repair_response.headers["content-type"])
        self.assertIn('.pdf"', repair_response.headers["content-disposition"])
        self.assertTrue(repair_response.content.startswith(b"%PDF"))

        vehicle_response = self.client.get(f"/api/vehicles/{vehicle_id}/export.pdf", headers=headers)
        self.assertEqual(vehicle_response.status_code, 200, vehicle_response.text)
        self.assertIn("application/pdf", vehicle_response.headers["content-type"])
        self.assertIn('.pdf"', vehicle_response.headers["content-disposition"])
        self.assertTrue(vehicle_response.content.startswith(b"%PDF"))

    def test_document_vehicle_and_repair_download_endpoints_support_head_requests(self) -> None:
        headers = self._get_auth_headers()
        payload = self._upload_order_document(headers)
        document_id = payload["document"]["id"]
        repair_id = payload["document"]["repair"]["id"]
        vehicle_id = payload["document"]["vehicle"]["id"]

        document_download_response = self.client.head(f"/api/documents/{document_id}/download", headers=headers)
        self.assertEqual(document_download_response.status_code, 200, document_download_response.text)
        self.assertIn("application/pdf", document_download_response.headers["content-type"])
        self.assertIn('.pdf"', document_download_response.headers["content-disposition"])

        document_xlsx_response = self.client.head(f"/api/documents/{document_id}/export", headers=headers)
        self.assertEqual(document_xlsx_response.status_code, 200, document_xlsx_response.text)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            document_xlsx_response.headers["content-type"],
        )
        self.assertIn('.xlsx"', document_xlsx_response.headers["content-disposition"])

        document_pdf_response = self.client.head(f"/api/documents/{document_id}/export.pdf", headers=headers)
        self.assertEqual(document_pdf_response.status_code, 200, document_pdf_response.text)
        self.assertIn("application/pdf", document_pdf_response.headers["content-type"])
        self.assertIn('.pdf"', document_pdf_response.headers["content-disposition"])

        repair_xlsx_response = self.client.head(f"/api/repairs/{repair_id}/export", headers=headers)
        self.assertEqual(repair_xlsx_response.status_code, 200, repair_xlsx_response.text)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            repair_xlsx_response.headers["content-type"],
        )
        self.assertIn('.xlsx"', repair_xlsx_response.headers["content-disposition"])

        vehicle_xlsx_response = self.client.head(f"/api/vehicles/{vehicle_id}/export", headers=headers)
        self.assertEqual(vehicle_xlsx_response.status_code, 200, vehicle_xlsx_response.text)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            vehicle_xlsx_response.headers["content-type"],
        )
        self.assertIn('.xlsx"', vehicle_xlsx_response.headers["content-disposition"])

        vehicle_pdf_response = self.client.head(f"/api/vehicles/{vehicle_id}/export.pdf", headers=headers)
        self.assertEqual(vehicle_pdf_response.status_code, 200, vehicle_pdf_response.text)
        self.assertIn("application/pdf", vehicle_pdf_response.headers["content-type"])
        self.assertIn('.pdf"', vehicle_pdf_response.headers["content-disposition"])

    def test_document_report_and_exports_use_selected_document_context(self) -> None:
        headers = self._get_auth_headers()
        initial_payload = self._upload_order_document(headers)
        repair_id = initial_payload["document"]["repair"]["id"]
        primary_document_id = initial_payload["document"]["id"]
        primary_filename = initial_payload["document"]["original_filename"]

        attach_response = self.client.post(
            "/api/documents/upload-to-repair",
            headers=headers,
            files={"file": ("repeat-report-scan.pdf", b"%PDF-1.4\n%repeat-report\n", "application/pdf")},
            data={"repair_id": str(repair_id), "kind": "repeat_scan"},
        )
        self.assertEqual(attach_response.status_code, 200, attach_response.text)
        attached_payload = attach_response.json()
        report_document_id = attached_payload["document"]["id"]
        report_filename = attached_payload["document"]["original_filename"]

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            primary_document = db.get(Document, primary_document_id)
            report_document = db.get(Document, report_document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(primary_document)
            self.assertIsNotNone(report_document)
            assert repair is not None
            assert primary_document is not None
            assert report_document is not None

            repair.source_document_id = primary_document.id
            primary_document.status = DocumentStatus.RECOGNIZED
            primary_document.ocr_confidence = 0.97
            report_document.status = DocumentStatus.RECOGNIZED
            report_document.ocr_confidence = 0.51
            next_version_number = max((item.version_number for item in report_document.versions), default=0) + 1
            db.add(
                DocumentVersion(
                    document_id=report_document.id,
                    version_number=next_version_number,
                    storage_key=report_document.storage_key,
                    parsed_payload={
                        "processor": "selected-document-report-test",
                        "ocr_status": "completed",
                        "manual_review_reasons": ["service_name_missing"],
                        "extracted_fields": {"order_number": "ZN-DOC-REPORT-001"},
                    },
                    field_confidence_map={"order_number": 0.91},
                    change_summary="Selected document report payload",
                )
            )
            db.commit()

        report_response = self.client.get(f"/api/documents/{report_document_id}/report", headers=headers)
        self.assertEqual(report_response.status_code, 200, report_response.text)
        report_payload = report_response.json()
        self.assertEqual(report_payload["document_id"], report_document_id)
        self.assertEqual(report_payload["source_document_id"], primary_document_id)
        self.assertEqual(report_payload["report_document_id"], report_document_id)
        self.assertEqual(report_payload["source_document_filename"], primary_filename)
        self.assertEqual(report_payload["report_document_filename"], report_filename)
        self.assertFalse(report_payload["is_primary_document"])
        finding_titles = [item["title"] for item in report_payload["executive_report"]["findings"]]
        self.assertIn("Низкая уверенность OCR по основному документу", finding_titles)

        export_response = self.client.get(f"/api/documents/{report_document_id}/export", headers=headers)
        self.assertEqual(export_response.status_code, 200, export_response.text)
        workbook = load_workbook(filename=BytesIO(export_response.content))
        summary_rows = {
            str(row[0]): row[1]
            for row in workbook["Отчет"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        self.assertEqual(summary_rows["Документ отчета"], report_filename)
        self.assertEqual(summary_rows["Основной документ"], primary_filename)

        document_rows = [
            row
            for row in workbook["Документы"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        primary_row = next(row for row in document_rows if row[0] == primary_document_id)
        report_row = next(row for row in document_rows if row[0] == report_document_id)
        self.assertEqual(primary_row[4], "Да")
        self.assertEqual(report_row[5], "Да")

        pdf_response = self.client.get(f"/api/documents/{report_document_id}/export.pdf", headers=headers)
        self.assertEqual(pdf_response.status_code, 200, pdf_response.text)
        self.assertIn("application/pdf", pdf_response.headers["content-type"])
        self.assertIn('.pdf"', pdf_response.headers["content-disposition"])
        self.assertTrue(pdf_response.content.startswith(b"%PDF"))

    def _mark_job_failed(self, db: Session, job_id: int, document_id: int) -> None:
        job = db.get(ImportJob, job_id)
        document = db.get(Document, document_id)
        self.assertIsNotNone(job)
        self.assertIsNotNone(document)
        assert job is not None
        assert document is not None
        job.status = ImportStatus.FAILED
        job.error_message = "ocr failed"
        document.status = DocumentStatus.OCR_ERROR
        document.review_queue_priority = 100
        if document.repair is not None:
            document.repair.status = RepairStatus.OCR_ERROR
        db.add(job)
        db.add(document)
        db.commit()

    def _mark_job_completed(self, db: Session, job_id: int, document_id: int) -> None:
        job = db.get(ImportJob, job_id)
        document = db.get(Document, document_id)
        self.assertIsNotNone(job)
        self.assertIsNotNone(document)
        assert job is not None
        assert document is not None
        job.status = ImportStatus.COMPLETED
        job.error_message = None
        document.status = DocumentStatus.RECOGNIZED
        document.review_queue_priority = 0
        if document.repair is not None:
            document.repair.status = RepairStatus.IN_REVIEW
        db.add(job)
        db.add(document)
        db.commit()


if __name__ == "__main__":
    unittest.main()
