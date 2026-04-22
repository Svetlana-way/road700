from __future__ import annotations

import tempfile
import unittest
from datetime import date
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.orm import Session, sessionmaker

from app.api import services as services_api
from app.api import documents as documents_api
from app.api.repairs import build_repair_export_workbook, build_repair_pdf_sections
from app.api.deps import get_db
from app.core.paths import get_storage_root, set_storage_root
from app.core.security import get_password_hash
from app.db.base import Base
from app.main import app
from app.models.audit import AuditLog
from app.models.document import Document, DocumentVersion
from app.models.enums import CatalogStatus, CheckSeverity, DocumentKind, DocumentStatus, ImportStatus, RepairStatus, ServiceStatus, UserRole, VehicleStatus, VehicleType
from app.models.imports import ImportConflict, ImportJob
from app.models.repair import Repair, RepairCheck, RepairPart, RepairWork
from app.models.service import Service
from app.models.user import User
from app.models.vehicle import Vehicle, VehicleAssignmentHistory
from app.services.historical_repairs_import import IMPORT_REASON_PREFIX
from app.services.service_catalog import (
    ServiceCatalogEntry,
    ensure_service_catalog_synced,
    find_service_catalog_entry,
    find_service_name_in_text,
    resolve_service_by_name,
)
from tests.sqlite_test_utils import create_sqlite_test_engine, reset_database


class ReviewAndServicesApiTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.storage_root = Path(cls.temp_dir.name) / "storage"
        cls.storage_root.mkdir(parents=True, exist_ok=True)
        cls.original_storage_root = get_storage_root()
        set_storage_root(cls.storage_root)
        cls.engine = create_sqlite_test_engine(enforce_foreign_keys=True)
        cls.SessionLocal = sessionmaker(bind=cls.engine, autoflush=False, autocommit=False, future=True)
        Base.metadata.create_all(bind=cls.engine)

        def override_get_db():
            db = cls.SessionLocal()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[get_db] = override_get_db
        cls.client = TestClient(app)

    @classmethod
    def tearDownClass(cls) -> None:
        app.dependency_overrides.clear()
        set_storage_root(cls.original_storage_root)
        cls.engine.dispose()
        cls.temp_dir.cleanup()

    def setUp(self) -> None:
        reset_database(self.engine, Base.metadata)

        with self.SessionLocal() as db:
            admin = User(
                full_name="Admin User",
                login="admin",
                email="admin@example.com",
                password_hash=get_password_hash("secret123"),
                role=UserRole.ADMIN,
                is_active=True,
            )
            employee = User(
                full_name="Employee User",
                login="employee",
                email="employee@example.com",
                password_hash=get_password_hash("secret123"),
                role=UserRole.EMPLOYEE,
                is_active=True,
            )
            db.add_all([admin, employee])
            db.flush()

            vehicle = Vehicle(
                external_id="truck-1",
                vehicle_type=VehicleType.TRUCK,
                plate_number="A123BC116",
                brand="Dong Feng",
                model="KL",
                status=VehicleStatus.ACTIVE,
            )
            service = Service(
                name="Service Alpha",
                city="Kazan",
                status=ServiceStatus.CONFIRMED,
                created_by_user_id=admin.id,
                confirmed_by_user_id=admin.id,
            )
            db.add_all([vehicle, service])
            db.flush()

            db.add(
                VehicleAssignmentHistory(
                    vehicle_id=vehicle.id,
                    user_id=employee.id,
                    starts_at=date(2025, 1, 1),
                    ends_at=None,
                    assigned_by_user_id=admin.id,
                    comment="Primary assignment",
                )
            )

            repair = Repair(
                order_number="ZN-001",
                repair_date=date(2025, 1, 15),
                vehicle_id=vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=120000,
                grand_total=15000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            db.add(repair)
            db.flush()

            db.add(
                Document(
                    repair_id=repair.id,
                    uploaded_by_user_id=employee.id,
                    original_filename="repair-order.pdf",
                    storage_key="documents/test/repair-order.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.NEEDS_REVIEW,
                    is_primary=True,
                    review_queue_priority=100,
                )
            )
            db.commit()

    def _get_auth_headers(self, username: str, password: str = "secret123") -> dict[str, str]:
        response = self.client.post(
            "/api/auth/login",
            data={"username": username, "password": password},
        )
        self.assertEqual(response.status_code, 200, response.text)
        token = response.json()["access_token"]
        return {"Authorization": f"Bearer {token}"}

    def test_review_queue_counts_and_filtered_total_match_category_distribution(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            suspicious_vehicle = Vehicle(
                external_id="truck-2",
                vehicle_type=VehicleType.TRUCK,
                plate_number="B234CD116",
                brand="Dong Feng",
                model="GX",
                status=VehicleStatus.ACTIVE,
            )
            ocr_vehicle = Vehicle(
                external_id="truck-3",
                vehicle_type=VehicleType.TRUCK,
                plate_number="C345DE116",
                brand="FAW",
                model="J7",
                status=VehicleStatus.ACTIVE,
            )
            employee_confirmation_vehicle = Vehicle(
                external_id="truck-4",
                vehicle_type=VehicleType.TRUCK,
                plate_number="D456EF116",
                brand="Sitrak",
                model="C7H",
                status=VehicleStatus.ACTIVE,
            )
            partial_vehicle = Vehicle(
                external_id="truck-5",
                vehicle_type=VehicleType.TRUCK,
                plate_number="E567FG116",
                brand="Howo",
                model="TX",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([suspicious_vehicle, ocr_vehicle, employee_confirmation_vehicle, partial_vehicle])
            db.flush()

            suspicious_repair = Repair(
                order_number="ZN-002",
                repair_date=date(2025, 1, 16),
                vehicle_id=suspicious_vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=130000,
                grand_total=10000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            ocr_repair = Repair(
                order_number="ZN-003",
                repair_date=date(2025, 1, 17),
                vehicle_id=ocr_vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=140000,
                grand_total=11000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            employee_confirmation_repair = Repair(
                order_number="ZN-004",
                repair_date=date(2025, 1, 18),
                vehicle_id=employee_confirmation_vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=150000,
                grand_total=12000,
                status=RepairStatus.EMPLOYEE_CONFIRMED,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            partial_repair = Repair(
                order_number="ZN-005",
                repair_date=date(2025, 1, 19),
                vehicle_id=partial_vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=160000,
                grand_total=13000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
                is_partially_recognized=True,
            )
            db.add_all([suspicious_repair, ocr_repair, employee_confirmation_repair, partial_repair])
            db.flush()

            suspicious_document = Document(
                repair_id=suspicious_repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="suspicious-order.pdf",
                storage_key="documents/test/suspicious-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.RECOGNIZED,
                is_primary=True,
                review_queue_priority=80,
            )
            ocr_document = Document(
                repair_id=ocr_repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="ocr-error-order.pdf",
                storage_key="documents/test/ocr-error-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.OCR_ERROR,
                is_primary=True,
                review_queue_priority=90,
            )
            employee_confirmation_document = Document(
                repair_id=employee_confirmation_repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="employee-confirmation-order.pdf",
                storage_key="documents/test/employee-confirmation-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.CONFIRMED,
                is_primary=True,
                review_queue_priority=60,
            )
            partial_document = Document(
                repair_id=partial_repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="partial-order.pdf",
                storage_key="documents/test/partial-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.RECOGNIZED,
                is_primary=True,
                review_queue_priority=70,
            )
            db.add_all([suspicious_document, ocr_document, employee_confirmation_document, partial_document])
            db.flush()

            db.add(
                RepairCheck(
                    repair_id=suspicious_repair.id,
                    check_type="ocr_total_mismatch",
                    severity=CheckSeverity.SUSPICIOUS,
                    title="Сумма строк не совпадает",
                    details="Требуется проверка",
                    is_resolved=False,
                )
            )
            db.commit()

        all_response = self.client.get("/api/review/queue?limit=10&category=all", headers=headers)
        self.assertEqual(all_response.status_code, 200, all_response.text)
        all_payload = all_response.json()
        self.assertEqual(all_payload["total"], 5)
        self.assertEqual(all_payload["counts"]["all"], 5)
        self.assertEqual(all_payload["counts"]["manual_review"], 1)
        self.assertEqual(all_payload["counts"]["suspicious"], 1)
        self.assertEqual(all_payload["counts"]["ocr_error"], 1)
        self.assertEqual(all_payload["counts"]["employee_confirmation"], 1)
        self.assertEqual(all_payload["counts"]["partial_recognition"], 1)

        suspicious_response = self.client.get("/api/review/queue?limit=10&category=suspicious", headers=headers)
        self.assertEqual(suspicious_response.status_code, 200, suspicious_response.text)
        suspicious_payload = suspicious_response.json()
        self.assertEqual(suspicious_payload["total"], 1)
        self.assertEqual(len(suspicious_payload["items"]), 1)
        self.assertEqual(suspicious_payload["items"][0]["category"], "suspicious")
        self.assertEqual(suspicious_payload["items"][0]["document"]["original_filename"], "suspicious-order.pdf")
        self.assertEqual(suspicious_payload["counts"], all_payload["counts"])

    def test_review_queue_excludes_documents_with_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            service = db.get(Service, 1)
            self.assertIsNotNone(service)
            assert service is not None
            service.status = ServiceStatus.ARCHIVED
            db.commit()

        response = self.client.get("/api/review/queue?limit=10&category=all", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["total"], 0)
        self.assertEqual(payload["counts"]["all"], 0)

    def test_review_queue_prioritizes_ocr_error_over_blocking_checks(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            vehicle = Vehicle(
                external_id="truck-review-ocr-priority",
                vehicle_type=VehicleType.TRUCK,
                plate_number="Q336QQ116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-REVIEW-OCR-PRIORITY-001",
                repair_date=date(2025, 1, 22),
                vehicle_id=vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=152000,
                grand_total=12200,
                status=RepairStatus.OCR_ERROR,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="ocr-error-with-blocking-check.pdf",
                storage_key="documents/test/ocr-error-with-blocking-check.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.OCR_ERROR,
                is_primary=True,
                review_queue_priority=95,
                ocr_confidence=0.12,
            )
            db.add(document)
            db.flush()

            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_processing_failed",
                    severity=CheckSeverity.ERROR,
                    title="OCR не обработал документ",
                    details="Файл поврежден",
                    is_resolved=False,
                )
            )
            db.commit()

        ocr_response = self.client.get("/api/review/queue?limit=10&category=ocr_error", headers=headers)
        self.assertEqual(ocr_response.status_code, 200, ocr_response.text)
        ocr_payload = ocr_response.json()

        suspicious_response = self.client.get("/api/review/queue?limit=10&category=suspicious", headers=headers)
        self.assertEqual(suspicious_response.status_code, 200, suspicious_response.text)
        suspicious_payload = suspicious_response.json()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()

        self.assertEqual(ocr_payload["counts"]["ocr_error"], 1)
        self.assertEqual(ocr_payload["total"], 1)
        self.assertEqual(ocr_payload["items"][0]["category"], "ocr_error")
        self.assertEqual(ocr_payload["items"][0]["document"]["original_filename"], "ocr-error-with-blocking-check.pdf")
        self.assertEqual(suspicious_payload["total"], 0)
        self.assertEqual(quality_payload["documents_ocr_error"], 1)

    def test_review_queue_localizes_manual_review_reasons(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            document = db.get(Document, 1)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.RECOGNIZED
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_fields": {},
                        "manual_review_reasons": ["vehicle_not_found", "service_name_missing"],
                    },
                    field_confidence_map={},
                    change_summary="Manual review labels smoke",
                )
            )
            db.commit()

        response = self.client.get("/api/review/queue?limit=10&category=manual_review", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()

        self.assertEqual(payload["total"], 1)
        self.assertEqual(
            payload["items"][0]["manual_review_reasons"],
            ["Техника не найдена в базе", "Не удалось определить сервис"],
        )
        self.assertIn("Техника не найдена в базе", payload["items"][0]["issue_titles"])
        self.assertIn("Не удалось определить сервис", payload["items"][0]["issue_titles"])

    def test_review_queue_excludes_uploaded_documents_awaiting_ocr(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            document = db.get(Document, 1)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.UPLOADED
            document.ocr_confidence = None
            db.add(
                ImportJob(
                    document_id=document.id,
                    import_type="document_ocr",
                    source_filename=document.original_filename,
                    status=ImportStatus.QUEUED,
                    summary={"stage": "queued"},
                    error_message=None,
                    attempts=0,
                )
            )
            db.commit()

        response = self.client.get("/api/review/queue?limit=10&category=all", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["total"], 0)
        self.assertEqual(payload["counts"]["all"], 0)
        self.assertEqual(payload["items"], [])

    def test_review_action_rejects_uploaded_document_awaiting_ocr(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            document = db.get(Document, 1)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.UPLOADED
            document.ocr_confidence = None
            db.commit()

        response = self.client.post(
            "/api/review/queue/1/action",
            headers=headers,
            json={"action": "send_to_review", "comment": "Need another pass"},
        )

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Document is not in review queue")

    def test_review_action_rejects_archived_service(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            service = db.get(Service, 1)
            self.assertIsNotNone(service)
            assert service is not None
            service.status = ServiceStatus.ARCHIVED
            db.commit()

        response = self.client.post(
            "/api/review/queue/1/action",
            headers=headers,
            json={"action": "send_to_review", "comment": "Need another pass"},
        )

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Document not found")

    def test_employee_can_execute_review_action_for_assigned_vehicle(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            document = db.get(Document, 1)
            self.assertIsNotNone(document)
            document_id = document.id

        response = self.client.post(
            f"/api/review/queue/{document_id}/action",
            headers=headers,
            json={"action": "send_to_review", "comment": "Need another pass"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document_status"], "needs_review")
        self.assertEqual(payload["repair_status"], "in_review")

    def test_employee_loses_review_access_when_assigned_vehicle_is_archived(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            vehicle = db.get(Vehicle, 1)
            document = db.get(Document, 1)
            repair = db.get(Repair, 1)
            self.assertIsNotNone(vehicle)
            self.assertIsNotNone(document)
            self.assertIsNotNone(repair)
            assert vehicle is not None
            assert document is not None
            assert repair is not None
            repair.created_by_user_id = 1
            vehicle.status = VehicleStatus.ARCHIVED
            document_id = document.id
            db.commit()

        response = self.client.post(
            f"/api/review/queue/{document_id}/action",
            headers=headers,
            json={"action": "send_to_review", "comment": "Need another pass"},
        )

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["detail"], "Document not found")

    def test_employee_cannot_confirm_repair_with_unresolved_findings(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            self.assertIsNotNone(repair)
            assert repair is not None
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_total_mismatch",
                    severity=CheckSeverity.SUSPICIOUS,
                    title="Сумма строк не совпадает с итоговой суммой",
                    details="Требуется ручная проверка",
                    is_resolved=False,
                )
            )
            db.commit()

        response = self.client.post(
            "/api/review/queue/1/action",
            headers=headers,
            json={"action": "employee_confirm", "comment": "Проверил"},
        )

        self.assertEqual(response.status_code, 400, response.text)
        self.assertIn("Сначала разберите все предупреждения", response.json()["detail"])

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertEqual(document.status, DocumentStatus.NEEDS_REVIEW)

    def test_employee_resolving_last_check_does_not_bypass_review_confirmation(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            document.status = DocumentStatus.RECOGNIZED
            check = RepairCheck(
                repair_id=repair.id,
                check_type="ocr_service_missing",
                severity=CheckSeverity.WARNING,
                title="Не удалось определить сервис",
                details="Сервис у ремонта не назначен. Нужна ручная проверка.",
                is_resolved=False,
            )
            db.add(check)
            db.commit()
            check_id = check.id

        response = self.client.patch(
            f"/api/repairs/1/checks/{check_id}",
            headers=headers,
            json={"is_resolved": True, "comment": "Проверил предупреждение"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], RepairStatus.IN_REVIEW.value)
        patched_check = next(item for item in payload["checks"] if item["id"] == check_id)
        self.assertTrue(patched_check["is_resolved"])

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertEqual(document.status, DocumentStatus.RECOGNIZED)

    def test_employee_workflow_runs_end_to_end_between_employee_and_admin(self) -> None:
        employee_headers = self._get_auth_headers("employee")
        admin_headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            warning_check = RepairCheck(
                repair_id=repair.id,
                check_type="ocr_service_missing",
                severity=CheckSeverity.WARNING,
                title="Не удалось определить сервис",
                details="Нужна ручная проверка сервиса",
                is_resolved=False,
            )
            db.add(warning_check)
            db.commit()
            check_id = warning_check.id

        blocked_response = self.client.post(
            "/api/review/queue/1/action",
            headers=employee_headers,
            json={"action": "employee_confirm", "comment": "Проверил и подтверждаю"},
        )
        self.assertEqual(blocked_response.status_code, 400, blocked_response.text)
        self.assertIn("Сначала разберите все предупреждения", blocked_response.json()["detail"])

        resolve_response = self.client.patch(
            f"/api/repairs/1/checks/{check_id}",
            headers=employee_headers,
            json={"is_resolved": True, "comment": "Сервис проверен вручную"},
        )
        self.assertEqual(resolve_response.status_code, 200, resolve_response.text)
        self.assertEqual(resolve_response.json()["status"], RepairStatus.IN_REVIEW.value)

        employee_confirm_response = self.client.post(
            "/api/review/queue/1/action",
            headers=employee_headers,
            json={"action": "employee_confirm", "comment": "Подтверждаю после проверки"},
        )
        self.assertEqual(employee_confirm_response.status_code, 200, employee_confirm_response.text)
        employee_confirm_payload = employee_confirm_response.json()
        self.assertEqual(employee_confirm_payload["document_status"], DocumentStatus.CONFIRMED.value)
        self.assertEqual(employee_confirm_payload["repair_status"], RepairStatus.EMPLOYEE_CONFIRMED.value)
        self.assertIsNotNone(employee_confirm_payload["queue_item"])
        self.assertEqual(employee_confirm_payload["queue_item"]["category"], "employee_confirmation")

        employee_confirmation_queue = self.client.get(
            "/api/review/queue?limit=10&category=employee_confirmation",
            headers=admin_headers,
        )
        self.assertEqual(employee_confirmation_queue.status_code, 200, employee_confirmation_queue.text)
        employee_confirmation_payload = employee_confirmation_queue.json()
        self.assertEqual(employee_confirmation_payload["total"], 1)
        self.assertEqual(employee_confirmation_payload["items"][0]["document"]["id"], 1)
        self.assertEqual(employee_confirmation_payload["items"][0]["repair"]["status"], RepairStatus.EMPLOYEE_CONFIRMED.value)

        send_back_response = self.client.post(
            "/api/review/queue/1/action",
            headers=admin_headers,
            json={"action": "send_to_review", "comment": "Нужна повторная проверка администратором"},
        )
        self.assertEqual(send_back_response.status_code, 200, send_back_response.text)
        send_back_payload = send_back_response.json()
        self.assertEqual(send_back_payload["document_status"], DocumentStatus.NEEDS_REVIEW.value)
        self.assertEqual(send_back_payload["repair_status"], RepairStatus.IN_REVIEW.value)
        self.assertIsNotNone(send_back_payload["queue_item"])
        self.assertEqual(send_back_payload["queue_item"]["category"], "manual_review")

        second_employee_confirm_response = self.client.post(
            "/api/review/queue/1/action",
            headers=employee_headers,
            json={"action": "employee_confirm", "comment": "Подтверждаю повторно после возврата"},
        )
        self.assertEqual(second_employee_confirm_response.status_code, 200, second_employee_confirm_response.text)
        self.assertEqual(second_employee_confirm_response.json()["repair_status"], RepairStatus.EMPLOYEE_CONFIRMED.value)

        admin_confirm_response = self.client.post(
            "/api/review/queue/1/action",
            headers=admin_headers,
            json={"action": "confirm", "comment": "Финально подтверждено администратором"},
        )
        self.assertEqual(admin_confirm_response.status_code, 200, admin_confirm_response.text)
        admin_confirm_payload = admin_confirm_response.json()
        self.assertEqual(admin_confirm_payload["document_status"], DocumentStatus.CONFIRMED.value)
        self.assertEqual(admin_confirm_payload["repair_status"], RepairStatus.CONFIRMED.value)
        self.assertIsNone(admin_confirm_payload["queue_item"])

        final_queue_response = self.client.get(
            "/api/review/queue?limit=10&category=employee_confirmation",
            headers=admin_headers,
        )
        self.assertEqual(final_queue_response.status_code, 200, final_queue_response.text)
        self.assertEqual(final_queue_response.json()["total"], 0)

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.CONFIRMED)
            self.assertFalse(repair.is_preliminary)
            self.assertEqual(document.status, DocumentStatus.CONFIRMED)
            self.assertIn("Подтверждено сотрудником", document.notes or "")
            self.assertIn("Возвращено в ручную проверку", document.notes or "")
            self.assertIn("Подтверждено администратором", document.notes or "")

    def test_check_resolution_reopens_admin_confirmed_repair_as_preliminary(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            repair.status = RepairStatus.CONFIRMED
            repair.is_preliminary = False
            document.status = DocumentStatus.CONFIRMED
            check = RepairCheck(
                repair_id=repair.id,
                check_type="ocr_service_missing_confirmed",
                severity=CheckSeverity.WARNING,
                title="Не удалось определить сервис после подтверждения",
                details="Нужна ручная проверка",
                is_resolved=False,
            )
            db.add(check)
            db.commit()
            check_id = check.id

        response = self.client.patch(
            f"/api/repairs/1/checks/{check_id}",
            headers=headers,
            json={"is_resolved": True, "comment": "Проверил подтвержденный ремонт"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], RepairStatus.IN_REVIEW.value)
        self.assertTrue(payload["is_preliminary"])

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            self.assertIsNotNone(repair)
            assert repair is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertTrue(repair.is_preliminary)

    def test_employee_cannot_confirm_repair_twice_after_employee_confirmation(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            repair.status = RepairStatus.EMPLOYEE_CONFIRMED
            repair.is_preliminary = True
            document.status = DocumentStatus.CONFIRMED
            db.commit()

        response = self.client.post(
            "/api/review/queue/1/action",
            headers=headers,
            json={"action": "employee_confirm", "comment": "Повторное подтверждение"},
        )

        self.assertEqual(response.status_code, 409, response.text)
        self.assertIn("already confirmed by employee", response.json()["detail"])

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.EMPLOYEE_CONFIRMED)
            self.assertEqual(document.status, DocumentStatus.CONFIRMED)
            self.assertEqual(document.notes, None)

    def test_upload_to_repair_reopens_employee_confirmed_repair_for_review(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            repair.status = RepairStatus.EMPLOYEE_CONFIRMED
            repair.is_preliminary = True
            document.status = DocumentStatus.CONFIRMED
            db.commit()

        with patch.object(documents_api, "queue_document_processing", autospec=True) as queue_mock:
            queue_mock.return_value = type(
                "QueuedJobStub",
                (),
                {"id": 601, "status": type("JobStatusStub", (), {"value": "queued"})()},
            )()

            response = self.client.post(
                "/api/documents/upload-to-repair",
                headers=headers,
                data={"repair_id": "1", "kind": "repeat_scan"},
                files={"file": ("repeat-scan.pdf", b"%PDF-1.4\n%repeat\n", "application/pdf")},
            )

        self.assertEqual(response.status_code, 200, response.text)

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            self.assertIsNotNone(repair)
            assert repair is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertTrue(repair.is_preliminary)

            audit_entry = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "document",
                    AuditLog.action_type == "document_attached",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(audit_entry)
            assert audit_entry is not None
            self.assertEqual(audit_entry.new_value["repair_status"], RepairStatus.IN_REVIEW.value)
            self.assertTrue(audit_entry.new_value["is_preliminary"])

    def test_reprocess_reopens_employee_confirmed_repair_for_review(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            repair.status = RepairStatus.EMPLOYEE_CONFIRMED
            repair.is_preliminary = True
            document.status = DocumentStatus.CONFIRMED
            document.review_queue_priority = 20
            db.commit()

        response = self.client.post("/api/documents/1/process", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document"]["status"], DocumentStatus.UPLOADED.value)
        self.assertEqual(payload["import_status"], "queued")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertTrue(repair.is_preliminary)
            self.assertEqual(document.status, DocumentStatus.UPLOADED)
            self.assertEqual(document.review_queue_priority, 100)

            audit_entry = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "document",
                    AuditLog.entity_id == str(document.id),
                    AuditLog.action_type == "document_processing_queued",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(audit_entry)
            assert audit_entry is not None
            self.assertEqual(audit_entry.old_value["status"], DocumentStatus.CONFIRMED.value)
            self.assertEqual(audit_entry.new_value["status"], DocumentStatus.UPLOADED.value)
            self.assertEqual(audit_entry.old_value["repair_status"], RepairStatus.EMPLOYEE_CONFIRMED.value)
            self.assertEqual(audit_entry.new_value["repair_status"], RepairStatus.IN_REVIEW.value)

    def test_review_field_update_resets_employee_confirmed_repair_back_to_in_review(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            repair.status = RepairStatus.EMPLOYEE_CONFIRMED
            repair.is_preliminary = True
            document.status = DocumentStatus.CONFIRMED
            db.commit()

        response = self.client.patch(
            "/api/repairs/1/review-fields",
            headers=headers,
            json={"mileage": 120500},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], RepairStatus.IN_REVIEW.value)
        self.assertEqual(payload["mileage"], 120500)

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertEqual(document.status, DocumentStatus.CONFIRMED)
            self.assertTrue(repair.is_preliminary)

    def test_review_field_update_reopens_admin_confirmed_repair_as_preliminary(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            repair.status = RepairStatus.CONFIRMED
            repair.is_preliminary = False
            document.status = DocumentStatus.CONFIRMED
            db.commit()

        response = self.client.patch(
            "/api/repairs/1/review-fields",
            headers=headers,
            json={"mileage": 120700},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], RepairStatus.IN_REVIEW.value)
        self.assertTrue(payload["is_preliminary"])

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            self.assertIsNotNone(repair)
            assert repair is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertTrue(repair.is_preliminary)

    def test_archived_service_rejects_review_field_updates(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            service = db.get(Service, 1)
            self.assertIsNotNone(service)
            assert service is not None
            service.status = ServiceStatus.ARCHIVED
            db.commit()

        response = self.client.patch(
            "/api/repairs/1/review-fields",
            headers=headers,
            json={"employee_comment": "This must be rejected for archived service"},
        )

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Repairs for archived services cannot be modified")

    def test_review_field_update_syncs_manual_review_payload(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            document = db.get(Document, 1)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.RECOGNIZED
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_fields": {},
                        "manual_review_reasons": [
                            "order_number_missing",
                            "mileage_missing",
                            "repair_date_invalid",
                        ],
                    },
                    field_confidence_map={},
                    change_summary="Initial OCR payload",
                )
            )
            db.commit()

        response = self.client.patch(
            "/api/repairs/1/review-fields",
            headers=headers,
            json={
                "order_number": "ZN-REVIEW-SYNC-001",
                "mileage": 120500,
                "repair_date": "2025-01-16",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        latest_payload = payload["documents"][0]["versions"][0]["parsed_payload"]
        self.assertEqual(latest_payload["manual_review_reasons"], [])
        self.assertEqual(latest_payload["extracted_fields"]["order_number"], "ZN-REVIEW-SYNC-001")
        self.assertEqual(latest_payload["extracted_fields"]["mileage"], 120500)
        self.assertEqual(latest_payload["extracted_fields"]["repair_date"], "2025-01-16")
        finding_titles = [item["title"] for item in payload["executive_report"]["findings"]]
        self.assertNotIn("Документ содержит признаки неполного или спорного распознавания", finding_titles)

    def test_employee_can_access_own_preliminary_repair_after_server_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign",
                vehicle_type=VehicleType.TRUCK,
                plate_number="B456CD116",
                brand="Sitrak",
                model="C7H",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-001",
                repair_date=date(2025, 1, 16),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=1000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-order.pdf",
                storage_key="documents/test/uploaded-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.UPLOADED,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            repair_id = repair.id
            document_id = document.id

        repair_response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)
        self.assertEqual(repair_response.status_code, 200, repair_response.text)
        self.assertEqual(repair_response.json()["id"], repair_id)

        documents_response = self.client.get("/api/documents?limit=8", headers=headers)
        self.assertEqual(documents_response.status_code, 200, documents_response.text)
        visible_document_ids = [item["id"] for item in documents_response.json()["items"]]
        self.assertIn(document_id, visible_document_ids)

    def test_employee_repair_detail_hides_archived_documents_and_document_history(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            repair = db.get(Repair, 1)
            self.assertIsNotNone(admin)
            self.assertIsNotNone(repair)
            assert admin is not None
            assert repair is not None

            archived_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="employee-hidden-archived.pdf",
                storage_key="documents/test/employee-hidden-archived.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.ARCHIVED,
                is_primary=False,
                review_queue_priority=0,
            )
            db.add(archived_document)
            db.flush()

            db.add(
                AuditLog(
                    user_id=admin.id,
                    entity_type="document",
                    entity_id=str(archived_document.id),
                    action_type="employee_hidden_archived_document_note",
                    old_value=None,
                    new_value={
                        "document_id": archived_document.id,
                        "original_filename": archived_document.original_filename,
                    },
                )
            )
            db.commit()
            archived_document_id = archived_document.id

        response = self.client.get("/api/repairs/1", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        visible_document_ids = [item["id"] for item in payload["documents"]]
        self.assertIn(1, visible_document_ids)
        self.assertNotIn(archived_document_id, visible_document_ids)
        document_history_actions = [item["action_type"] for item in payload["document_history"]]
        self.assertNotIn("employee_hidden_archived_document_note", document_history_actions)

    def test_employee_can_execute_review_action_for_own_preliminary_repair_after_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-2",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-2",
                vehicle_type=VehicleType.TRUCK,
                plate_number="C789EF116",
                brand="Howo",
                model="T5G",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-REVIEW-001",
                repair_date=date(2025, 1, 17),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=2000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-review-order.pdf",
                storage_key="documents/test/uploaded-review-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            document_id = document.id

        response = self.client.post(
            f"/api/review/queue/{document_id}/action",
            headers=headers,
            json={"action": "send_to_review", "comment": "Повторная ручная проверка"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document_status"], DocumentStatus.NEEDS_REVIEW.value)
        self.assertEqual(payload["repair_status"], RepairStatus.IN_REVIEW.value)

    def test_employee_loses_review_access_for_own_preliminary_repair_after_vehicle_archived(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)
            assert employee is not None

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-2-ARCH",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-2-archived",
                vehicle_type=VehicleType.TRUCK,
                plate_number="C790EF116",
                brand="Howo",
                model="T5G",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-REVIEW-ARCH-001",
                repair_date=date(2025, 1, 17),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=2001,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-review-archived-order.pdf",
                storage_key="documents/test/uploaded-review-archived-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            document_id = document.id
            vehicle_id = foreign_vehicle.id

        visible_response = self.client.post(
            f"/api/review/queue/{document_id}/action",
            headers=headers,
            json={"action": "send_to_review", "comment": "Повторная ручная проверка"},
        )
        self.assertEqual(visible_response.status_code, 200, visible_response.text)

        with self.SessionLocal() as db:
            vehicle = db.get(Vehicle, vehicle_id)
            self.assertIsNotNone(vehicle)
            assert vehicle is not None
            vehicle.status = VehicleStatus.ARCHIVED
            db.commit()

        archived_response = self.client.post(
            f"/api/review/queue/{document_id}/action",
            headers=headers,
            json={"action": "send_to_review", "comment": "Повторная ручная проверка"},
        )

        self.assertEqual(archived_response.status_code, 404, archived_response.text)
        self.assertEqual(archived_response.json()["detail"], "Document not found")

    def test_employee_can_reprocess_own_preliminary_document_after_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-3",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-3",
                vehicle_type=VehicleType.TRUCK,
                plate_number="D012GH116",
                brand="Foton",
                model="Auman",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-PROCESS-001",
                repair_date=date(2025, 1, 18),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=3000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-process-order.pdf",
                storage_key="documents/test/uploaded-process-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            document_id = document.id

        with patch.object(documents_api, "queue_document_processing", autospec=True) as queue_mock:
            queue_mock.return_value = type(
                "QueuedJobStub",
                (),
                {"id": 501, "status": type("JobStatusStub", (), {"value": "queued"})()},
            )()

            response = self.client.post(
                f"/api/documents/{document_id}/process",
                headers=headers,
            )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["document"]["id"], document_id)
        self.assertEqual(payload["job_id"], 501)
        self.assertEqual(payload["import_status"], "queued")
        queue_mock.assert_called_once()

    def test_dashboard_summary_includes_own_preliminary_repair_after_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-4",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-4",
                vehicle_type=VehicleType.TRUCK,
                plate_number="E345IJ116",
                brand="Shaanxi",
                model="X3000",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-DASH-001",
                repair_date=date(2025, 1, 19),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=4000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-dashboard-order.pdf",
                storage_key="documents/test/uploaded-dashboard-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()

        response = self.client.get("/api/dashboard/summary", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["vehicles_total"], 1)
        self.assertEqual(payload["repairs_total"], 2)
        self.assertEqual(payload["documents_total"], 2)
        self.assertEqual(payload["documents_review_queue"], 2)

    def test_dashboard_details_include_own_preliminary_document_after_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-5",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-5",
                vehicle_type=VehicleType.TRUCK,
                plate_number="F678KL116",
                brand="JAC",
                model="K7",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-DQ-001",
                repair_date=date(2025, 1, 20),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=5000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-dashboard-details-order.pdf",
                storage_key="documents/test/uploaded-dashboard-details-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=130,
                ocr_confidence=0.42,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            document_id = document.id

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()
        self.assertEqual(quality_payload["documents_needs_review"], 2)
        self.assertEqual(quality_payload["documents_low_confidence"], 1)

        details_response = self.client.get("/api/dashboard/data-quality/details?limit=8", headers=headers)
        self.assertEqual(details_response.status_code, 200, details_response.text)
        details_payload = details_response.json()
        self.assertEqual(details_payload["counts"]["documents"], 2)
        visible_document_ids = [item["document_id"] for item in details_payload["documents"]]
        self.assertIn(document_id, visible_document_ids)

    def test_dashboard_summary_excludes_repairs_and_documents_with_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            service = db.get(Service, 1)
            self.assertIsNotNone(service)
            assert service is not None
            service.status = ServiceStatus.ARCHIVED
            db.commit()

        response = self.client.get("/api/dashboard/summary", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["vehicles_total"], 1)
        self.assertEqual(payload["repairs_total"], 0)
        self.assertEqual(payload["documents_total"], 0)
        self.assertEqual(payload["documents_review_queue"], 0)

    def test_dashboard_summary_review_queue_matches_documents_blocked_by_repair_status(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            suspicious_vehicle = Vehicle(
                external_id="truck-dashboard-summary-suspicious",
                vehicle_type=VehicleType.TRUCK,
                plate_number="S111SS116",
                brand="Dong Feng",
                model="GX",
                status=VehicleStatus.ACTIVE,
            )
            employee_confirmation_vehicle = Vehicle(
                external_id="truck-dashboard-summary-employee-confirmed",
                vehicle_type=VehicleType.TRUCK,
                plate_number="S222SS116",
                brand="Sitrak",
                model="C7H",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([suspicious_vehicle, employee_confirmation_vehicle])
            db.flush()

            suspicious_repair = Repair(
                order_number="ZN-DASH-QUEUE-002",
                repair_date=date(2025, 1, 16),
                vehicle_id=suspicious_vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=130000,
                grand_total=10000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            employee_confirmation_repair = Repair(
                order_number="ZN-DASH-QUEUE-003",
                repair_date=date(2025, 1, 17),
                vehicle_id=employee_confirmation_vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=140000,
                grand_total=11000,
                status=RepairStatus.EMPLOYEE_CONFIRMED,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            db.add_all([suspicious_repair, employee_confirmation_repair])
            db.flush()

            db.add_all(
                [
                    Document(
                        repair_id=suspicious_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="dashboard-summary-suspicious.pdf",
                        storage_key="documents/test/dashboard-summary-suspicious.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.RECOGNIZED,
                        is_primary=True,
                        review_queue_priority=80,
                    ),
                    Document(
                        repair_id=employee_confirmation_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="dashboard-summary-employee-confirmed.pdf",
                        storage_key="documents/test/dashboard-summary-employee-confirmed.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.CONFIRMED,
                        is_primary=True,
                        review_queue_priority=60,
                    ),
                ]
            )
            db.flush()
            db.add(
                RepairCheck(
                    repair_id=suspicious_repair.id,
                    check_type="ocr_total_mismatch",
                    severity=CheckSeverity.SUSPICIOUS,
                    title="Сумма строк не совпадает",
                    details="Требуется проверка",
                    is_resolved=False,
                )
            )
            db.commit()

        review_response = self.client.get("/api/review/queue?limit=20&category=all", headers=headers)
        self.assertEqual(review_response.status_code, 200, review_response.text)
        review_payload = review_response.json()

        summary_response = self.client.get("/api/dashboard/summary", headers=headers)
        self.assertEqual(summary_response.status_code, 200, summary_response.text)
        summary_payload = summary_response.json()

        self.assertEqual(review_payload["counts"]["all"], 3)
        self.assertEqual(summary_payload["documents_review_queue"], review_payload["counts"]["all"])

    def test_dashboard_data_quality_matches_review_queue_for_documents_blocked_by_repair_status(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            suspicious_vehicle = Vehicle(
                external_id="truck-dashboard-quality-suspicious",
                vehicle_type=VehicleType.TRUCK,
                plate_number="Q111QQ116",
                brand="Dong Feng",
                model="GX",
                status=VehicleStatus.ACTIVE,
            )
            employee_confirmation_vehicle = Vehicle(
                external_id="truck-dashboard-quality-employee-confirmed",
                vehicle_type=VehicleType.TRUCK,
                plate_number="Q222QQ116",
                brand="Sitrak",
                model="C7H",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([suspicious_vehicle, employee_confirmation_vehicle])
            db.flush()

            suspicious_repair = Repair(
                order_number="ZN-DASH-DQ-002",
                repair_date=date(2025, 1, 18),
                vehicle_id=suspicious_vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=130000,
                grand_total=10000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            employee_confirmation_repair = Repair(
                order_number="ZN-DASH-DQ-003",
                repair_date=date(2025, 1, 19),
                vehicle_id=employee_confirmation_vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=140000,
                grand_total=11000,
                status=RepairStatus.EMPLOYEE_CONFIRMED,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            db.add_all([suspicious_repair, employee_confirmation_repair])
            db.flush()

            db.add_all(
                [
                    Document(
                        repair_id=suspicious_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="dashboard-quality-suspicious.pdf",
                        storage_key="documents/test/dashboard-quality-suspicious.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.RECOGNIZED,
                        is_primary=True,
                        review_queue_priority=80,
                        ocr_confidence=0.95,
                    ),
                    Document(
                        repair_id=employee_confirmation_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="dashboard-quality-employee-confirmed.pdf",
                        storage_key="documents/test/dashboard-quality-employee-confirmed.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.CONFIRMED,
                        is_primary=True,
                        review_queue_priority=60,
                        ocr_confidence=0.95,
                    ),
                ]
            )
            db.flush()
            db.add(
                RepairCheck(
                    repair_id=suspicious_repair.id,
                    check_type="ocr_total_mismatch",
                    severity=CheckSeverity.SUSPICIOUS,
                    title="Сумма строк не совпадает",
                    details="Требуется проверка",
                    is_resolved=False,
                )
            )
            db.commit()

        review_response = self.client.get("/api/review/queue?limit=20&category=all", headers=headers)
        self.assertEqual(review_response.status_code, 200, review_response.text)
        review_payload = review_response.json()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()

        details_response = self.client.get("/api/dashboard/data-quality/details?limit=20", headers=headers)
        self.assertEqual(details_response.status_code, 200, details_response.text)
        details_payload = details_response.json()

        self.assertEqual(review_payload["counts"]["all"], 3)
        self.assertEqual(quality_payload["documents_needs_review"], review_payload["counts"]["all"])
        self.assertEqual(details_payload["counts"]["documents"], review_payload["counts"]["all"])

    def test_dashboard_data_quality_counts_repair_level_ocr_errors(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            vehicle = Vehicle(
                external_id="truck-dashboard-quality-ocr-repair-status",
                vehicle_type=VehicleType.TRUCK,
                plate_number="Q333QQ116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-DASH-DQ-OCR-001",
                repair_date=date(2025, 1, 20),
                vehicle_id=vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=150000,
                grand_total=12000,
                status=RepairStatus.OCR_ERROR,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            db.add(repair)
            db.flush()

            db.add(
                Document(
                    repair_id=repair.id,
                    uploaded_by_user_id=employee.id,
                    original_filename="dashboard-quality-repair-ocr-error.pdf",
                    storage_key="documents/test/dashboard-quality-repair-ocr-error.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.RECOGNIZED,
                    is_primary=True,
                    review_queue_priority=70,
                    ocr_confidence=0.97,
                )
            )
            db.commit()

        review_response = self.client.get("/api/review/queue?limit=20&category=ocr_error", headers=headers)
        self.assertEqual(review_response.status_code, 200, review_response.text)
        review_payload = review_response.json()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()

        self.assertEqual(review_payload["counts"]["ocr_error"], 1)
        self.assertEqual(quality_payload["documents_ocr_error"], review_payload["counts"]["ocr_error"])

    def test_dashboard_suspicious_repair_counts_include_open_suspicious_checks(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            vehicle = Vehicle(
                external_id="truck-dashboard-suspicious-checks",
                vehicle_type=VehicleType.TRUCK,
                plate_number="Q444QQ116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-DASH-SUSP-001",
                repair_date=date(2025, 1, 21),
                vehicle_id=vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=155000,
                grand_total=12500,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            db.add(repair)
            db.flush()

            db.add(
                Document(
                    repair_id=repair.id,
                    uploaded_by_user_id=employee.id,
                    original_filename="dashboard-suspicious-checks.pdf",
                    storage_key="documents/test/dashboard-suspicious-checks.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.RECOGNIZED,
                    is_primary=True,
                    review_queue_priority=75,
                    ocr_confidence=0.97,
                )
            )
            db.flush()
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_total_mismatch",
                    severity=CheckSeverity.SUSPICIOUS,
                    title="Сумма строк не совпадает",
                    details="Требуется проверка",
                    is_resolved=False,
                )
            )
            db.commit()

        review_response = self.client.get("/api/review/queue?limit=20&category=suspicious", headers=headers)
        self.assertEqual(review_response.status_code, 200, review_response.text)
        review_payload = review_response.json()

        summary_response = self.client.get("/api/dashboard/summary", headers=headers)
        self.assertEqual(summary_response.status_code, 200, summary_response.text)
        summary_payload = summary_response.json()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()

        self.assertEqual(review_payload["counts"]["suspicious"], 1)
        self.assertEqual(summary_payload["repairs_suspicious"], review_payload["counts"]["suspicious"])
        self.assertEqual(quality_payload["repairs_suspicious"], review_payload["counts"]["suspicious"])

    def test_dashboard_suspicious_repair_counts_exclude_ocr_error_repairs_with_blocking_checks(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            vehicle = Vehicle(
                external_id="truck-dashboard-ocr-excluded-from-suspicious",
                vehicle_type=VehicleType.TRUCK,
                plate_number="Q445QQ116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-DASH-OCR-SUSP-001",
                repair_date=date(2025, 1, 22),
                vehicle_id=vehicle.id,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=156000,
                grand_total=12600,
                status=RepairStatus.OCR_ERROR,
                is_preliminary=True,
                is_partially_recognized=False,
            )
            db.add(repair)
            db.flush()

            db.add(
                Document(
                    repair_id=repair.id,
                    uploaded_by_user_id=employee.id,
                    original_filename="dashboard-ocr-error-not-suspicious.pdf",
                    storage_key="documents/test/dashboard-ocr-error-not-suspicious.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.OCR_ERROR,
                    is_primary=True,
                    review_queue_priority=75,
                    ocr_confidence=0.11,
                )
            )
            db.flush()
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_processing_failed",
                    severity=CheckSeverity.ERROR,
                    title="OCR не обработал документ",
                    details="Файл поврежден",
                    is_resolved=False,
                )
            )
            db.commit()

        suspicious_review_response = self.client.get("/api/review/queue?limit=20&category=suspicious", headers=headers)
        self.assertEqual(suspicious_review_response.status_code, 200, suspicious_review_response.text)
        suspicious_review_payload = suspicious_review_response.json()

        ocr_review_response = self.client.get("/api/review/queue?limit=20&category=ocr_error", headers=headers)
        self.assertEqual(ocr_review_response.status_code, 200, ocr_review_response.text)
        ocr_review_payload = ocr_review_response.json()

        summary_response = self.client.get("/api/dashboard/summary", headers=headers)
        self.assertEqual(summary_response.status_code, 200, summary_response.text)
        summary_payload = summary_response.json()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()

        self.assertEqual(suspicious_review_payload["counts"]["suspicious"], 0)
        self.assertEqual(suspicious_review_payload["total"], 0)
        self.assertEqual(ocr_review_payload["counts"]["ocr_error"], 1)
        self.assertEqual(summary_payload["repairs_suspicious"], 0)
        self.assertEqual(quality_payload["repairs_suspicious"], 0)
        self.assertEqual(quality_payload["documents_ocr_error"], 1)

    def test_dashboard_excludes_uploaded_documents_awaiting_ocr_from_review_metrics(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            document = db.get(Document, 1)
            self.assertIsNotNone(document)
            assert document is not None
            document.status = DocumentStatus.UPLOADED
            document.ocr_confidence = None
            db.add(
                ImportJob(
                    document_id=document.id,
                    import_type="document_ocr",
                    source_filename=document.original_filename,
                    status=ImportStatus.QUEUED,
                    summary={"stage": "queued"},
                    error_message=None,
                    attempts=0,
                )
            )
            db.commit()

        summary_response = self.client.get("/api/dashboard/summary", headers=headers)
        self.assertEqual(summary_response.status_code, 200, summary_response.text)
        summary_payload = summary_response.json()
        self.assertEqual(summary_payload["documents_total"], 1)
        self.assertEqual(summary_payload["documents_review_queue"], 0)

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()
        self.assertEqual(quality_payload["documents_needs_review"], 0)
        self.assertEqual(quality_payload["documents_ocr_error"], 0)
        self.assertEqual(quality_payload["documents_low_confidence"], 0)

        details_response = self.client.get("/api/dashboard/data-quality/details?limit=8", headers=headers)
        self.assertEqual(details_response.status_code, 200, details_response.text)
        details_payload = details_response.json()
        self.assertEqual(details_payload["counts"]["documents"], 0)
        self.assertEqual(details_payload["documents"], [])

    def test_dashboard_data_quality_excludes_documents_with_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            service = db.get(Service, 1)
            self.assertIsNotNone(service)
            assert service is not None
            service.status = ServiceStatus.ARCHIVED
            db.commit()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()
        self.assertEqual(quality_payload["documents_needs_review"], 0)
        self.assertEqual(quality_payload["documents_ocr_error"], 0)

        details_response = self.client.get("/api/dashboard/data-quality/details?limit=8", headers=headers)
        self.assertEqual(details_response.status_code, 200, details_response.text)
        details_payload = details_response.json()
        self.assertEqual(details_payload["counts"]["documents"], 0)
        self.assertEqual(details_payload["documents"], [])

    def test_dashboard_data_quality_excludes_import_conflicts_for_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            document = db.scalar(select(Document).where(Document.original_filename == "repair-order.pdf"))
            self.assertIsNotNone(document)
            assert document is not None
            repair = db.get(Repair, document.repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            service = db.get(Service, repair.service_id)
            self.assertIsNotNone(service)
            assert service is not None

            job = ImportJob(
                document_id=document.id,
                import_type="historical_repairs",
                source_filename="archived-service-history.xlsx",
                status=ImportStatus.COMPLETED_WITH_CONFLICTS,
                summary={"stage": "completed"},
                error_message=None,
                attempts=1,
            )
            db.add(job)
            db.flush()
            db.add(
                ImportConflict(
                    import_job_id=job.id,
                    entity_type="repair",
                    conflict_key="archived-service-conflict",
                    incoming_payload={"repair_id": repair.id},
                    existing_payload=None,
                    resolution_payload=None,
                    status="pending",
                )
            )
            service.status = ServiceStatus.ARCHIVED
            db.commit()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()
        self.assertEqual(quality_payload["import_conflicts_pending"], 0)

        details_response = self.client.get("/api/dashboard/data-quality/details?limit=8", headers=headers)
        self.assertEqual(details_response.status_code, 200, details_response.text)
        details_payload = details_response.json()
        self.assertEqual(details_payload["counts"]["conflicts"], 0)
        self.assertEqual(details_payload["conflicts"], [])

    def test_dashboard_data_quality_excludes_preliminary_work_and_part_for_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.scalar(select(Repair).where(Repair.order_number == "ZN-001"))
            self.assertIsNotNone(repair)
            assert repair is not None
            service = db.get(Service, repair.service_id)
            self.assertIsNotNone(service)
            assert service is not None

            db.add_all(
                [
                    RepairWork(
                        repair_id=repair.id,
                        work_code="PRE-DASH-001",
                        work_name="Предварительная работа в архивном сервисе",
                        quantity=1,
                        price=1800,
                        line_total=1800,
                        status=CatalogStatus.PRELIMINARY,
                    ),
                    RepairPart(
                        repair_id=repair.id,
                        article="PRE-DASH-PART-001",
                        part_name="Предварительная деталь в архивном сервисе",
                        quantity=1,
                        price=2400,
                        line_total=2400,
                        status=CatalogStatus.PRELIMINARY,
                    ),
                ]
            )
            service.status = ServiceStatus.ARCHIVED
            db.commit()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()
        self.assertEqual(quality_payload["works_preliminary"], 0)
        self.assertEqual(quality_payload["parts_preliminary"], 0)

        details_response = self.client.get("/api/dashboard/data-quality/details?limit=8", headers=headers)
        self.assertEqual(details_response.status_code, 200, details_response.text)
        details_payload = details_response.json()
        self.assertEqual(details_payload["counts"]["works"], 0)
        self.assertEqual(details_payload["counts"]["parts"], 0)
        self.assertEqual(details_payload["works"], [])
        self.assertEqual(details_payload["parts"], [])

    def test_dashboard_data_quality_service_items_ignore_archived_repairs(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            preliminary_service = Service(
                name="Dashboard Preliminary Archived Tail",
                city="Kazan",
                status=ServiceStatus.PRELIMINARY,
                created_by_user_id=1,
            )
            db.add(preliminary_service)
            db.flush()

            archived_repair = Repair(
                order_number="ZN-DASH-SVC-ARCH-001",
                repair_date=date(2025, 1, 30),
                vehicle_id=1,
                service_id=preliminary_service.id,
                created_by_user_id=1,
                mileage=133000,
                reason="Archived repair should not count in service dashboard item",
                grand_total=15000,
                status=RepairStatus.ARCHIVED,
                is_preliminary=False,
            )
            db.add(archived_repair)
            db.commit()

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()
        self.assertEqual(quality_payload["services_preliminary"], 0)

        details_response = self.client.get("/api/dashboard/data-quality/details?limit=20", headers=headers)
        self.assertEqual(details_response.status_code, 200, details_response.text)
        details_payload = details_response.json()
        self.assertEqual(details_payload["counts"]["services"], 0)
        self.assertFalse(
            any(item["name"] == "Dashboard Preliminary Archived Tail" for item in details_payload["services"])
        )

    def test_dashboard_data_quality_includes_visible_preliminary_service_confirmed_by_admin(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            service_item = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(service_item)
            assert service_item is not None

            service_item.name = "Dashboard Renamed Preliminary Service"
            service_item.status = ServiceStatus.PRELIMINARY
            service_item.created_by_user_id = None
            service_item.confirmed_by_user_id = 1
            db.add(service_item)
            db.commit()

        services_response = self.client.get("/api/services", headers=headers)
        self.assertEqual(services_response.status_code, 200, services_response.text)
        self.assertTrue(
            any(item["name"] == "Dashboard Renamed Preliminary Service" for item in services_response.json()["items"])
        )

        quality_response = self.client.get("/api/dashboard/data-quality", headers=headers)
        self.assertEqual(quality_response.status_code, 200, quality_response.text)
        quality_payload = quality_response.json()
        self.assertEqual(quality_payload["services_preliminary"], 1)

        details_response = self.client.get("/api/dashboard/data-quality/details?limit=8", headers=headers)
        self.assertEqual(details_response.status_code, 200, details_response.text)
        details_payload = details_response.json()
        self.assertEqual(details_payload["counts"]["services"], 1)
        self.assertTrue(
            any(item["name"] == "Dashboard Renamed Preliminary Service" for item in details_payload["services"])
        )

    def test_dashboard_data_quality_details_without_limit_return_full_problem_document_list(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(service)
            assert service is not None

            created_document_ids: list[int] = []
            for index in range(9):
                vehicle = Vehicle(
                    external_id=f"truck-dashboard-full-{index}",
                    vehicle_type=VehicleType.TRUCK,
                    plate_number=f"K10{index}KK116",
                    brand="Volvo",
                    model="FH",
                    status=VehicleStatus.ACTIVE,
                )
                db.add(vehicle)
                db.flush()

                repair = Repair(
                    order_number=f"ZN-DASH-FULL-{index:03d}",
                    repair_date=date(2025, 2, index + 1),
                    vehicle_id=vehicle.id,
                    service_id=service.id,
                    created_by_user_id=admin.id,
                    mileage=150000 + index,
                    grand_total=10000,
                    status=RepairStatus.IN_REVIEW,
                    is_preliminary=True,
                    is_partially_recognized=False,
                )
                db.add(repair)
                db.flush()

                document = Document(
                    repair_id=repair.id,
                    uploaded_by_user_id=admin.id,
                    original_filename=f"dashboard-full-quality-{index}.pdf",
                    storage_key=f"documents/test/dashboard-full-quality-{index}.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.NEEDS_REVIEW,
                    is_primary=True,
                    review_queue_priority=100 - index,
                    ocr_confidence=0.5,
                )
                db.add(document)
                db.flush()
                created_document_ids.append(document.id)

            db.commit()

        response = self.client.get("/api/dashboard/data-quality/details", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()

        expected_total = 10
        self.assertEqual(payload["counts"]["documents"], expected_total)
        self.assertEqual(len(payload["documents"]), expected_total)
        returned_document_ids = {item["document_id"] for item in payload["documents"]}
        self.assertTrue(set(created_document_ids).issubset(returned_document_ids))

    def test_dashboard_work_and_part_items_ignore_legacy_foreign_source_document_id(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            local_vehicle = Vehicle(
                external_id="truck-dashboard-local",
                vehicle_type=VehicleType.TRUCK,
                plate_number="E555EE116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-dashboard-foreign",
                vehicle_type=VehicleType.TRUCK,
                plate_number="F666FF116",
                brand="Scania",
                model="R",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([local_vehicle, foreign_vehicle])
            db.flush()

            local_repair = Repair(
                order_number="ZN-DASH-LOCAL-001",
                repair_date=date(2025, 1, 24),
                vehicle_id=local_vehicle.id,
                created_by_user_id=admin.id,
                mileage=21000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            foreign_repair = Repair(
                order_number="ZN-DASH-FOREIGN-001",
                repair_date=date(2025, 1, 25),
                vehicle_id=foreign_vehicle.id,
                created_by_user_id=admin.id,
                mileage=22000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add_all([local_repair, foreign_repair])
            db.flush()

            local_document = Document(
                repair_id=local_repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="dashboard-local-primary.pdf",
                storage_key="documents/test/dashboard-local-primary.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            foreign_document = Document(
                repair_id=foreign_repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="dashboard-foreign-primary.pdf",
                storage_key="documents/test/dashboard-foreign-primary.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add_all([local_document, foreign_document])
            db.flush()

            local_repair.source_document_id = foreign_document.id
            db.add_all(
                [
                    RepairWork(
                        repair_id=local_repair.id,
                        work_code="PRE-001",
                        work_name="Предварительная диагностика",
                        quantity=1,
                        price=1500,
                        line_total=1500,
                        status=CatalogStatus.PRELIMINARY,
                    ),
                    RepairPart(
                        repair_id=local_repair.id,
                        article="P-001",
                        part_name="Тестовая деталь",
                        quantity=1,
                        price=2500,
                        line_total=2500,
                        status=CatalogStatus.PRELIMINARY,
                    ),
                ]
            )
            db.commit()
            local_document_id = local_document.id

        response = self.client.get("/api/dashboard/data-quality/details?limit=8", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()

        work_item = next(item for item in payload["works"] if item["work_name"] == "Предварительная диагностика")
        part_item = next(item for item in payload["parts"] if item["part_name"] == "Тестовая деталь")
        self.assertEqual(work_item["document_id"], local_document_id)
        self.assertEqual(part_item["document_id"], local_document_id)

    def test_dashboard_work_and_part_items_keep_null_document_id_when_repair_has_only_attachments(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            vehicle = Vehicle(
                external_id="truck-dashboard-attachment-only",
                vehicle_type=VehicleType.TRUCK,
                plate_number="T101TT116",
                brand="DAF",
                model="XF",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-DASH-ATT-001",
                repair_date=date(2025, 1, 29),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=23000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            attachment_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="dashboard-attachment.pdf",
                storage_key="documents/test/dashboard-attachment.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ATTACHMENT,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=40,
            )
            db.add(attachment_document)
            db.flush()

            db.add_all(
                [
                    RepairWork(
                        repair_id=repair.id,
                        work_code="ATT-001",
                        work_name="Работа без source order",
                        quantity=1,
                        price=900,
                        line_total=900,
                        status=CatalogStatus.PRELIMINARY,
                    ),
                    RepairPart(
                        repair_id=repair.id,
                        article="ATT-PART-001",
                        part_name="Деталь без source order",
                        quantity=1,
                        price=1200,
                        line_total=1200,
                        status=CatalogStatus.PRELIMINARY,
                    ),
                ]
            )
            db.commit()
            repair_id = repair.id

        response = self.client.get("/api/dashboard/data-quality/details?limit=25", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()

        work_item = next(item for item in payload["works"] if item["repair_id"] == repair_id)
        part_item = next(item for item in payload["parts"] if item["repair_id"] == repair_id)
        self.assertIsNone(work_item["document_id"])
        self.assertIsNone(part_item["document_id"])

    def test_admin_repair_detail_keeps_source_document_empty_when_repair_has_only_attachments(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            vehicle = Vehicle(
                external_id="truck-detail-attachment-only",
                vehicle_type=VehicleType.TRUCK,
                plate_number="M202MM116",
                brand="MAN",
                model="TGX",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-DETAIL-ATT-001",
                repair_date=date(2025, 1, 30),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=24000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            attachment_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="detail-attachment.pdf",
                storage_key="documents/test/detail-attachment.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ATTACHMENT,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=35,
            )
            db.add(attachment_document)
            db.commit()
            repair_id = repair.id
            attachment_document_id = attachment_document.id

        response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()

        self.assertIsNone(payload["source_document_id"])
        attachment_payload = next(item for item in payload["documents"] if item["id"] == attachment_document_id)
        self.assertFalse(attachment_payload["is_primary"])

    def test_repair_detail_exposes_labor_norm_reference_fields_for_work_items(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            self.assertIsNotNone(repair)
            assert repair is not None

            db.add_all(
                [
                    RepairWork(
                        repair_id=repair.id,
                        work_code=None,
                        work_name="Датчик ABS, замена",
                        quantity=1,
                        price=1377.5,
                        line_total=1377.5,
                        status=CatalogStatus.PRELIMINARY,
                        reference_payload={
                            "labor_norm_reference_status": "catalog_gap",
                            "labor_norm_item_reason_code": "catalog_gap",
                            "labor_norm_item_reason": "работа выглядит осмысленной, но для датчика ABS не указана ось или сторона, поэтому выбрать норму из каталога нельзя",
                            "labor_norm_next_step": "Чтобы подобрать норму, в заказ-наряде нужно указать ось или сторону датчика ABS.",
                        },
                    ),
                    RepairWork(
                        repair_id=repair.id,
                        work_code="11090111",
                        work_name="Воздушный фильтр - замена",
                        quantity=1,
                        standard_hours=0.5,
                        price=1710,
                        line_total=1710,
                        status=CatalogStatus.CONFIRMED,
                        reference_payload={
                            "labor_norm_reference_status": "matched",
                            "labor_norm_code": "11090111",
                            "labor_norm_name": "Диагностика неисправностей и замена элемента воздушного фильтра",
                            "labor_norm_matched_by": "code",
                            "labor_norm_match_score": 1.0,
                        },
                    ),
                ]
            )
            db.commit()

        response = self.client.get("/api/repairs/1", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        gap_work = next(item for item in payload["works"] if item["work_name"] == "Датчик ABS, замена")
        matched_work = next(item for item in payload["works"] if item["work_name"] == "Воздушный фильтр - замена")

        self.assertEqual(gap_work["reference_status"], "catalog_gap")
        self.assertEqual(gap_work["reference_reason_code"], "catalog_gap_missing_abs_position")
        self.assertIn("ось", gap_work["reference_reason"])
        self.assertIn("ось", gap_work["reference_next_step"])
        self.assertIn("Датчик ABS", gap_work["reference_rewrite_draft"])
        self.assertIsNone(gap_work["reference_match_code"])

        self.assertEqual(matched_work["reference_status"], "matched")
        self.assertEqual(matched_work["reference_match_code"], "11090111")
        self.assertEqual(
            matched_work["reference_match_name"],
            "Диагностика неисправностей и замена элемента воздушного фильтра",
        )
        self.assertEqual(matched_work["reference_matched_by"], "code")
        self.assertEqual(matched_work["reference_match_score"], 1.0)

    def test_document_list_exposes_labor_norm_reference_fields_in_parsed_payload_works(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            document = db.get(Document, 1)
            self.assertIsNotNone(document)
            assert document is not None

            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_items": {
                            "works": [
                                {
                                    "work_code": None,
                                    "work_name": "Датчик ABS, замена",
                                    "reference_payload": {
                                        "labor_norm_reference_status": "catalog_gap",
                                        "labor_norm_item_reason_code": "catalog_gap",
                                        "labor_norm_item_reason": "работа выглядит осмысленной, но для датчика ABS не указана ось или сторона, поэтому выбрать норму из каталога нельзя",
                                        "labor_norm_next_step": "Чтобы подобрать норму, в заказ-наряде нужно указать ось или сторону датчика ABS.",
                                    },
                                },
                                {
                                    "work_code": "11090111",
                                    "work_name": "Воздушный фильтр - замена",
                                    "reference_payload": {
                                        "labor_norm_reference_status": "matched",
                                        "labor_norm_code": "11090111",
                                        "labor_norm_name": "Диагностика неисправностей и замена элемента воздушного фильтра",
                                        "labor_norm_matched_by": "code",
                                        "labor_norm_match_score": 1.0,
                                    },
                                },
                            ]
                        }
                    },
                    field_confidence_map={},
                    change_summary="seed parsed payload",
                )
            )
            db.commit()

        response = self.client.get("/api/documents?limit=8", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        document_payload = next(item for item in payload["items"] if item["id"] == 1)
        works = document_payload["parsed_payload"]["extracted_items"]["works"]
        gap_work = next(item for item in works if item["work_name"] == "Датчик ABS, замена")
        matched_work = next(item for item in works if item["work_name"] == "Воздушный фильтр - замена")

        self.assertEqual(gap_work["reference_status"], "catalog_gap")
        self.assertEqual(gap_work["reference_reason_code"], "catalog_gap_missing_abs_position")
        self.assertIn("ось", gap_work["reference_reason"])
        self.assertIn("ось", gap_work["reference_next_step"])
        self.assertIn("Датчик ABS", gap_work["reference_rewrite_draft"])
        self.assertIsNone(gap_work["reference_match_code"])

        self.assertEqual(matched_work["reference_status"], "matched")
        self.assertEqual(matched_work["reference_match_code"], "11090111")
        self.assertEqual(
            matched_work["reference_match_name"],
            "Диагностика неисправностей и замена элемента воздушного фильтра",
        )
        self.assertEqual(matched_work["reference_matched_by"], "code")
        self.assertEqual(matched_work["reference_match_score"], 1.0)

    def test_repair_detail_exposes_labor_norm_reference_fields_in_document_versions(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            document = db.get(Document, 1)
            self.assertIsNotNone(document)
            assert document is not None

            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_items": {
                            "works": [
                                {
                                    "work_code": None,
                                    "work_name": "Датчик ABS, замена",
                                    "reference_payload": {
                                        "labor_norm_reference_status": "catalog_gap",
                                        "labor_norm_item_reason_code": "catalog_gap",
                                        "labor_norm_item_reason": "работа выглядит осмысленной, но для датчика ABS не указана ось или сторона, поэтому выбрать норму из каталога нельзя",
                                        "labor_norm_next_step": "Чтобы подобрать норму, в заказ-наряде нужно указать ось или сторону датчика ABS.",
                                    },
                                },
                                {
                                    "work_code": "11090111",
                                    "work_name": "Воздушный фильтр - замена",
                                    "reference_payload": {
                                        "labor_norm_reference_status": "matched",
                                        "labor_norm_code": "11090111",
                                        "labor_norm_name": "Диагностика неисправностей и замена элемента воздушного фильтра",
                                        "labor_norm_matched_by": "code",
                                        "labor_norm_match_score": 1.0,
                                    },
                                },
                            ]
                        }
                    },
                    field_confidence_map={},
                    change_summary="seed repair version payload",
                )
            )
            db.commit()

        response = self.client.get("/api/repairs/1", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        version_payload = payload["documents"][0]["versions"][0]["parsed_payload"]
        works = version_payload["extracted_items"]["works"]
        gap_work = next(item for item in works if item["work_name"] == "Датчик ABS, замена")
        matched_work = next(item for item in works if item["work_name"] == "Воздушный фильтр - замена")

        self.assertEqual(gap_work["reference_status"], "catalog_gap")
        self.assertEqual(gap_work["reference_reason_code"], "catalog_gap_missing_abs_position")
        self.assertIn("ось", gap_work["reference_reason"])
        self.assertIn("ось", gap_work["reference_next_step"])
        self.assertIn("Датчик ABS", gap_work["reference_rewrite_draft"])
        self.assertIsNone(gap_work["reference_match_code"])

        self.assertEqual(matched_work["reference_status"], "matched")
        self.assertEqual(matched_work["reference_match_code"], "11090111")
        self.assertEqual(
            matched_work["reference_match_name"],
            "Диагностика неисправностей и замена элемента воздушного фильтра",
        )
        self.assertEqual(matched_work["reference_matched_by"], "code")
        self.assertEqual(matched_work["reference_match_score"], 1.0)

    def test_export_surfaces_customer_summary_in_pdf_and_xlsx_summary(self) -> None:
        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            self.assertIsNotNone(repair)
            assert repair is not None

            db.add_all(
                [
                    RepairWork(
                        repair_id=repair.id,
                        work_code=None,
                        work_name="Датчик ABS, замена",
                        quantity=1,
                        price=1377.5,
                        line_total=1377.5,
                        status=CatalogStatus.PRELIMINARY,
                        reference_payload={
                            "labor_norm_reference_status": "catalog_gap",
                            "labor_norm_next_step": "Чтобы подобрать норму, в заказ-наряде нужно указать ось или сторону датчика ABS.",
                        },
                    ),
                    RepairWork(
                        repair_id=repair.id,
                        work_code="11090111",
                        work_name="Воздушный фильтр - замена",
                        quantity=1,
                        price=1710.0,
                        line_total=1710.0,
                        status=CatalogStatus.CONFIRMED,
                        reference_payload={"labor_norm_reference_status": "matched"},
                    ),
                    RepairWork(
                        repair_id=repair.id,
                        work_code="0101",
                        work_name="Мойка технологическая седельного тягача",
                        quantity=1,
                        price=1000.0,
                        line_total=1000.0,
                        status=CatalogStatus.PRELIMINARY,
                        reference_payload={"labor_norm_reference_status": "outside_catalog_service"},
                    ),
                ]
            )
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_total_mismatch",
                    severity=CheckSeverity.SUSPICIOUS,
                    title="Сумма строк не совпадает с итоговой суммой",
                    details="Нужна ручная проверка итогов заказ-наряда",
                    calculation_payload={},
                    is_resolved=False,
                )
            )
            db.commit()
            db.refresh(repair)

            pdf_sections = build_repair_pdf_sections(repair)
            workbook = build_repair_export_workbook(repair)

        summary_section = next(items for title, items in pdf_sections if title == "Сводка")
        joined_summary = " ".join(summary_section)
        self.assertIn("Сводка для заказчика:", joined_summary)
        self.assertIn("Подтверждено по каталогу: 1", joined_summary)
        self.assertIn("Вне каталога: 1", joined_summary)
        self.assertIn("Нужна ручная проверка: 1", joined_summary)
        self.assertIn("OCR-риск: 1", joined_summary)

        summary_rows = {
            str(row[0]): row[1]
            for row in workbook["Отчет"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        self.assertIn("1 строк.", str(summary_rows["Подтверждено по каталогу"]))
        self.assertIn("1 строк.", str(summary_rows["Вне каталога"]))
        self.assertIn("1 строк.", str(summary_rows["Нужна ручная проверка"]))
        self.assertIn("1 сигналов.", str(summary_rows["OCR-риск"]))

    def test_employee_cannot_read_audit_log_even_for_own_preliminary_repair_after_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(admin)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-6",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-6",
                vehicle_type=VehicleType.TRUCK,
                plate_number="G901MN116",
                brand="FAW",
                model="J7",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-AUDIT-001",
                repair_date=date(2025, 1, 21),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=6000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-audit-order.pdf",
                storage_key="documents/test/uploaded-audit-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.flush()

            db.add_all(
                [
                    AuditLog(
                        user_id=admin.id,
                        entity_type="repair",
                        entity_id=str(repair.id),
                        action_type="admin_repair_note",
                        old_value=None,
                        new_value={"repair_id": repair.id},
                    ),
                    AuditLog(
                        user_id=admin.id,
                        entity_type="document",
                        entity_id=str(document.id),
                        action_type="admin_document_note",
                        old_value=None,
                        new_value={"document_id": document.id},
                    ),
                ]
            )
            db.commit()

        response = self.client.get("/api/audit?limit=50", headers=headers)

        self.assertEqual(response.status_code, 403, response.text)
        self.assertEqual(response.json()["detail"], "Admin access required")

    def test_global_search_supports_offset_per_section_pagination(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(admin)
            self.assertIsNotNone(service)
            assert admin is not None
            assert service is not None

            for index in range(9):
                vehicle = Vehicle(
                    external_id=f"bulk-search-{index:03d}",
                    vehicle_type=VehicleType.TRUCK,
                    plate_number=f"Q10{index}QQ116",
                    brand="Bulk Search",
                    model=f"Model {index}",
                    status=VehicleStatus.ACTIVE,
                )
                db.add(vehicle)
                db.flush()

                repair = Repair(
                    order_number=f"BULK-SEARCH-{index:03d}",
                    repair_date=date(2025, 3, index + 1),
                    vehicle_id=vehicle.id,
                    service_id=service.id,
                    created_by_user_id=admin.id,
                    mileage=200000 + index,
                    grand_total=10000 + index,
                    status=RepairStatus.IN_REVIEW,
                    is_preliminary=True,
                    is_partially_recognized=False,
                )
                db.add(repair)
                db.flush()

                db.add(
                    Document(
                        repair_id=repair.id,
                        uploaded_by_user_id=admin.id,
                        original_filename=f"bulk-search-{index:03d}.pdf",
                        storage_key=f"documents/test/bulk-search-{index:03d}.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.NEEDS_REVIEW,
                        is_primary=True,
                        review_queue_priority=100,
                    )
                )

            db.commit()

        first_page_response = self.client.get(
            "/api/search/global?q=bulk-search&limit_per_section=8&offset_per_section=0",
            headers=headers,
        )
        second_page_response = self.client.get(
            "/api/search/global?q=bulk-search&limit_per_section=8&offset_per_section=8",
            headers=headers,
        )

        self.assertEqual(first_page_response.status_code, 200, first_page_response.text)
        self.assertEqual(second_page_response.status_code, 200, second_page_response.text)

        first_page_payload = first_page_response.json()
        second_page_payload = second_page_response.json()

        self.assertEqual(first_page_payload["documents_total"], 9)
        self.assertEqual(first_page_payload["repairs_total"], 9)
        self.assertEqual(first_page_payload["vehicles_total"], 9)
        self.assertEqual(len(first_page_payload["documents"]), 8)
        self.assertEqual(len(first_page_payload["repairs"]), 8)
        self.assertEqual(len(first_page_payload["vehicles"]), 8)
        self.assertEqual(len(second_page_payload["documents"]), 1)
        self.assertEqual(len(second_page_payload["repairs"]), 1)
        self.assertEqual(len(second_page_payload["vehicles"]), 1)

        self.assertEqual(
            {item["document_id"] for item in first_page_payload["documents"]} &
            {item["document_id"] for item in second_page_payload["documents"]},
            set(),
        )
        self.assertEqual(
            {item["repair_id"] for item in first_page_payload["repairs"]} &
            {item["repair_id"] for item in second_page_payload["repairs"]},
            set(),
        )
        self.assertEqual(
            {item["vehicle_id"] for item in first_page_payload["vehicles"]} &
            {item["vehicle_id"] for item in second_page_payload["vehicles"]},
            set(),
        )

    def test_employee_can_download_own_preliminary_document_after_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-7",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-7",
                vehicle_type=VehicleType.TRUCK,
                plate_number="H234OP116",
                brand="Sitrak",
                model="C7H Max",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-DOWNLOAD-001",
                repair_date=date(2025, 1, 22),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=7000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-download-order.pdf",
                storage_key="documents/test/uploaded-download-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            document_id = document.id
            storage_key = document.storage_key

        file_path = self.storage_root / storage_key
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_bytes(b"%PDF-1.4\n%relink-download-test\n")

        response = self.client.get(
            f"/api/documents/{document_id}/download",
            headers=headers,
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertIn("application/pdf", response.headers["content-type"])
        self.assertTrue(response.content.startswith(b"%PDF-1.4"))

    def test_employee_cannot_download_own_preliminary_document_after_vehicle_archived(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)
            assert employee is not None

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-7-ARCH",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-7-archived",
                vehicle_type=VehicleType.TRUCK,
                plate_number="H235OP116",
                brand="Sitrak",
                model="C7H Max",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-DOWNLOAD-ARCH-001",
                repair_date=date(2025, 1, 22),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=7100,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-download-archived-order.pdf",
                storage_key="documents/test/uploaded-download-archived-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            document_id = document.id
            vehicle_id = foreign_vehicle.id
            storage_key = document.storage_key

        file_path = self.storage_root / storage_key
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_bytes(b"%PDF-1.4\n%relink-download-archived-test\n")

        visible_response = self.client.get(
            f"/api/documents/{document_id}/download",
            headers=headers,
        )
        self.assertEqual(visible_response.status_code, 200, visible_response.text)

        with self.SessionLocal() as db:
            vehicle = db.get(Vehicle, vehicle_id)
            self.assertIsNotNone(vehicle)
            assert vehicle is not None
            vehicle.status = VehicleStatus.ARCHIVED
            db.commit()

        archived_response = self.client.get(
            f"/api/documents/{document_id}/download",
            headers=headers,
        )
        self.assertEqual(archived_response.status_code, 404, archived_response.text)

    def test_document_download_blocks_storage_path_traversal_from_persisted_storage_key(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            vehicle = db.scalar(select(Vehicle).where(Vehicle.external_id == "truck-1"))
            self.assertIsNotNone(admin)
            self.assertIsNotNone(vehicle)
            assert admin is not None
            assert vehicle is not None

            repair = Repair(
                order_number="ZN-UPL-DOWNLOAD-TRAVERSAL-001",
                repair_date=date(2025, 1, 23),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=7200,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=False,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="outside-secret.txt",
                storage_key="../outside-secret.txt",
                mime_type="text/plain",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()
            document_id = document.id

        outside_file = self.storage_root.parent / "outside-secret.txt"
        outside_file.write_text("should-not-leak", encoding="utf-8")
        try:
            response = self.client.get(
                f"/api/documents/{document_id}/download",
                headers=headers,
            )
            self.assertEqual(response.status_code, 404, response.text)
            self.assertEqual(response.json()["detail"], "Document file not found")
        finally:
            outside_file.unlink(missing_ok=True)

    def test_employee_can_export_own_preliminary_repair_after_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-8",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-8",
                vehicle_type=VehicleType.TRUCK,
                plate_number="J567RS116",
                brand="Mercedes-Benz",
                model="Actros",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-EXPORT-001",
                repair_date=date(2025, 1, 23),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=8000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-export-order.pdf",
                storage_key="documents/test/uploaded-export-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            repair_id = repair.id

        xlsx_response = self.client.get(
            f"/api/repairs/{repair_id}/export",
            headers=headers,
        )
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.text)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            xlsx_response.headers["content-type"],
        )
        self.assertTrue(xlsx_response.content.startswith(b"PK"))

        pdf_response = self.client.get(
            f"/api/repairs/{repair_id}/export.pdf",
            headers=headers,
        )
        self.assertEqual(pdf_response.status_code, 200, pdf_response.text)
        self.assertIn("application/pdf", pdf_response.headers["content-type"])
        self.assertTrue(pdf_response.content.startswith(b"%PDF"))

    def test_repair_export_endpoints_support_head_requests(self) -> None:
        headers = self._get_auth_headers("employee")

        xlsx_response = self.client.head("/api/repairs/1/export", headers=headers)
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.text)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            xlsx_response.headers["content-type"],
        )
        self.assertIn('.xlsx"', xlsx_response.headers["content-disposition"])

        pdf_response = self.client.head("/api/repairs/1/export.pdf", headers=headers)
        self.assertEqual(pdf_response.status_code, 200, pdf_response.text)
        self.assertIn("application/pdf", pdf_response.headers["content-type"])
        self.assertIn('.pdf"', pdf_response.headers["content-disposition"])

    def test_employee_export_hides_archived_documents_in_xlsx_and_pdf_sections(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            repair = db.get(Repair, 1)
            self.assertIsNotNone(admin)
            self.assertIsNotNone(repair)
            assert admin is not None
            assert repair is not None

            archived_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="employee-hidden-export-archived.pdf",
                storage_key="documents/test/employee-hidden-export-archived.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.ARCHIVED,
                is_primary=False,
                review_queue_priority=0,
            )
            db.add(archived_document)
            db.commit()
            repair_id = repair.id

        xlsx_response = self.client.get(f"/api/repairs/{repair_id}/export", headers=headers)
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.text)

        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        document_rows = [
            row
            for row in workbook["Документы"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        exported_filenames = [row[1] for row in document_rows]
        self.assertIn("repair-order.pdf", exported_filenames)
        self.assertNotIn("employee-hidden-export-archived.pdf", exported_filenames)

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            pdf_sections = build_repair_pdf_sections(repair, include_archived_documents=False)

        documents_section = next(items for title, items in pdf_sections if title == "Документы")
        self.assertFalse(any("employee-hidden-export-archived.pdf" in item for item in documents_section))

    def test_employee_export_does_not_use_archived_document_as_report_source(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            document.status = DocumentStatus.ARCHIVED
            document.is_primary = False
            document.review_queue_priority = 0
            repair.source_document_id = document.id
            db.commit()
            repair_id = repair.id

        xlsx_response = self.client.get(f"/api/repairs/{repair_id}/export", headers=headers)
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.text)

        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        summary_rows = {
            str(row[0]): row[1]
            for row in workbook["Отчет"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        self.assertIn(summary_rows["Основной документ"], ("", None))
        self.assertIn(summary_rows["Статус основного документа"], ("", None))

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            pdf_sections = build_repair_pdf_sections(repair, include_archived_documents=False)

        summary_section = next(items for title, items in pdf_sections if title == "Сводка")
        self.assertIn("Основной документ: Не выбран", summary_section)
        self.assertIn("Статус основного документа: Не определён", summary_section)

    def test_export_uses_canonical_primary_marker_in_xlsx_and_pdf_documents_section(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            document.is_primary = False
            repair.source_document_id = document.id
            db.commit()
            repair_id = repair.id
            document_filename = document.original_filename

        xlsx_response = self.client.get(f"/api/repairs/{repair_id}/export", headers=headers)
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.text)

        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        document_rows = [
            row
            for row in workbook["Документы"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        exported_row = next(row for row in document_rows if row[1] == document_filename)
        self.assertEqual(exported_row[4], "Да")

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            pdf_sections = build_repair_pdf_sections(repair, include_archived_documents=False)

        documents_section = next(items for title, items in pdf_sections if title == "Документы")
        self.assertTrue(any(f"{document_filename} | order | needs_review | основной" in item for item in documents_section))

    def test_admin_export_uses_archived_source_document_for_archived_repair(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            document.status = DocumentStatus.ARCHIVED
            document.is_primary = False
            document.review_queue_priority = 0
            repair.status = RepairStatus.ARCHIVED
            repair.is_preliminary = False
            repair.source_document_id = document.id
            db.commit()
            repair_id = repair.id
            document_filename = document.original_filename

        xlsx_response = self.client.get(f"/api/repairs/{repair_id}/export", headers=headers)
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.text)

        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        summary_rows = {
            str(row[0]): row[1]
            for row in workbook["Отчет"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        self.assertEqual(summary_rows["Основной документ"], document_filename)
        self.assertEqual(summary_rows["Статус основного документа"], DocumentStatus.ARCHIVED.value)

        document_rows = [
            row
            for row in workbook["Документы"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        exported_row = next(row for row in document_rows if row[1] == document_filename)
        self.assertEqual(exported_row[4], "Да")

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            pdf_sections = build_repair_pdf_sections(repair, include_archived_documents=True)

        summary_section = next(items for title, items in pdf_sections if title == "Сводка")
        self.assertIn(f"Основной документ: {document_filename}", summary_section)
        self.assertIn(f"Статус основного документа: {DocumentStatus.ARCHIVED.value}", summary_section)
        documents_section = next(items for title, items in pdf_sections if title == "Документы")
        self.assertTrue(any(f"{document_filename} | order | archived | основной" in item for item in documents_section))

    def test_admin_repair_detail_uses_archived_source_document_for_archived_repair(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            document.status = DocumentStatus.ARCHIVED
            document.is_primary = False
            document.review_queue_priority = 0
            document.ocr_confidence = 0.61
            repair.status = RepairStatus.ARCHIVED
            repair.is_preliminary = False
            repair.source_document_id = document.id
            db.commit()
            repair_id = repair.id
            document_id = document.id

        response = self.client.get(f"/api/repairs/{repair_id}", headers=headers)
        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()

        self.assertEqual(payload["source_document_id"], document_id)
        finding_titles = [item["title"] for item in payload["executive_report"]["findings"]]
        self.assertIn("Низкая уверенность OCR по основному документу", finding_titles)
        archived_document = next(item for item in payload["documents"] if item["id"] == document_id)
        self.assertTrue(archived_document["is_primary"])

    def test_employee_cannot_export_own_preliminary_repair_after_vehicle_archived(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)
            assert employee is not None

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-8-ARCH",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-8-archived",
                vehicle_type=VehicleType.TRUCK,
                plate_number="J568RS116",
                brand="Mercedes-Benz",
                model="Actros",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-EXPORT-ARCH-001",
                repair_date=date(2025, 1, 23),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=8100,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="uploaded-export-archived-order.pdf",
                storage_key="documents/test/uploaded-export-archived-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            repair_id = repair.id
            vehicle_id = foreign_vehicle.id

        visible_xlsx_response = self.client.get(
            f"/api/repairs/{repair_id}/export",
            headers=headers,
        )
        self.assertEqual(visible_xlsx_response.status_code, 200, visible_xlsx_response.text)

        visible_pdf_response = self.client.get(
            f"/api/repairs/{repair_id}/export.pdf",
            headers=headers,
        )
        self.assertEqual(visible_pdf_response.status_code, 200, visible_pdf_response.text)

        with self.SessionLocal() as db:
            vehicle = db.get(Vehicle, vehicle_id)
            self.assertIsNotNone(vehicle)
            assert vehicle is not None
            vehicle.status = VehicleStatus.ARCHIVED
            db.commit()

        archived_xlsx_response = self.client.get(
            f"/api/repairs/{repair_id}/export",
            headers=headers,
        )
        self.assertEqual(archived_xlsx_response.status_code, 404, archived_xlsx_response.text)

        archived_pdf_response = self.client.get(
            f"/api/repairs/{repair_id}/export.pdf",
            headers=headers,
        )
        self.assertEqual(archived_pdf_response.status_code, 404, archived_pdf_response.text)

    def test_repair_export_includes_workflow_stage_and_final_report_sections(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            vehicle = Vehicle(
                external_id="truck-export-report-1",
                vehicle_type=VehicleType.TRUCK,
                plate_number="M456OP116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-REPORT-001",
                repair_date=date(2025, 1, 25),
                vehicle_id=vehicle.id,
                created_by_user_id=employee.id,
                mileage=154000,
                reason="Вибрация на холостом ходу",
                employee_comment="Сотрудник проверил документ и передал на финальное решение",
                work_total=12000,
                parts_total=18000,
                vat_total=6000,
                grand_total=36000,
                expected_total=34000,
                status=RepairStatus.EMPLOYEE_CONFIRMED,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            db.add_all(
                [
                    RepairWork(
                        repair_id=repair.id,
                        work_code="DIAG-01",
                        work_name="Диагностика подвески",
                        quantity=1,
                        standard_hours=1.0,
                        actual_hours=1.2,
                        price=12000,
                        line_total=12000,
                        status=CatalogStatus.CONFIRMED,
                    ),
                    RepairPart(
                        repair_id=repair.id,
                        article="FILTER-01",
                        part_name="Фильтр салона",
                        quantity=1,
                        unit_name="шт",
                        price=18000,
                        line_total=18000,
                        status=CatalogStatus.CONFIRMED,
                    ),
                ]
            )

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="final-report-order.pdf",
                storage_key="documents/test/final-report-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()
            repair_id = repair.id

        xlsx_response = self.client.get(f"/api/repairs/{repair_id}/export", headers=headers)
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.text)

        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        self.assertIn("Итоговый отчет", workbook.sheetnames)

        summary_rows = {
            str(row[0]): row[1]
            for row in workbook["Отчет"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        self.assertEqual(summary_rows["Этап workflow"], "Ожидает финального подтверждения администратора")
        self.assertIn("финальный review администратора", str(summary_rows["Комментарий к workflow"]))

        executive_rows = [
            row
            for row in workbook["Итоговый отчет"].iter_rows(min_row=2, values_only=True)
            if row and row[0] and row[1]
        ]
        self.assertTrue(any(row[0] == "1. Финансы" and "Итого по заказ-наряду" in str(row[1]) for row in executive_rows))
        self.assertTrue(any(row[0] == "2. Что фактически сделано" for row in executive_rows))
        self.assertTrue(any(row[0] == "8. Рекомендация" for row in executive_rows))

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            pdf_sections = build_repair_pdf_sections(repair)

        pdf_section_titles = [title for title, _items in pdf_sections]
        self.assertIn("Короткий отчёт для руководителя", pdf_section_titles)
        self.assertIn("1. Финансы", pdf_section_titles)
        self.assertIn("7. Итог", pdf_section_titles)
        summary_section = next(items for title, items in pdf_sections if title == "Сводка")
        self.assertIn("Этап workflow: Ожидает финального подтверждения администратора", summary_section)

    def test_repair_export_marks_blocking_checks_as_suspicious_workflow(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)
            assert employee is not None

            vehicle = Vehicle(
                external_id="truck-export-suspicious-1",
                vehicle_type=VehicleType.TRUCK,
                plate_number="H321BC116",
                brand="Scania",
                model="R",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-REPORT-SUSP-001",
                repair_date=date(2025, 1, 26),
                vehicle_id=vehicle.id,
                created_by_user_id=employee.id,
                mileage=167000,
                reason="Проверка на дребезг",
                work_total=14000,
                parts_total=6000,
                vat_total=4000,
                grand_total=24000,
                expected_total=24000,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="blocking-check-order.pdf",
                storage_key="documents/test/blocking-check-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.RECOGNIZED,
                is_primary=True,
                review_queue_priority=80,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="export_workflow_blocking_check",
                    severity=CheckSeverity.ERROR,
                    title="Есть блокирующее замечание",
                    details="Тест проверяет workflow summary при открытом blocking check",
                    is_resolved=False,
                )
            )
            db.commit()
            repair_id = repair.id

        xlsx_response = self.client.get(f"/api/repairs/{repair_id}/export", headers=headers)
        self.assertEqual(xlsx_response.status_code, 200, xlsx_response.text)

        workbook = load_workbook(filename=BytesIO(xlsx_response.content))
        summary_rows = {
            str(row[0]): row[1]
            for row in workbook["Отчет"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        self.assertEqual(summary_rows["Этап workflow"], "Подозрительный ремонт")
        self.assertIn("управленческого решения", str(summary_rows["Комментарий к workflow"]))

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            self.assertIsNotNone(repair)
            assert repair is not None
            pdf_sections = build_repair_pdf_sections(repair)

        summary_section = next(items for title, items in pdf_sections if title == "Сводка")
        self.assertIn("Этап workflow: Подозрительный ремонт", summary_section)

    def test_employee_can_compare_own_preliminary_documents_after_vehicle_relink(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-9",
                brand="Placeholder",
                model="Upload",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-foreign-9",
                vehicle_type=VehicleType.TRUCK,
                plate_number="K890TU116",
                brand="MAN",
                model="TGX",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([placeholder_vehicle, foreign_vehicle])
            db.flush()

            repair = Repair(
                order_number="ZN-UPL-COMPARE-001",
                repair_date=date(2025, 1, 24),
                vehicle_id=placeholder_vehicle.id,
                created_by_user_id=employee.id,
                mileage=9000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            primary_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="compare-primary.pdf",
                storage_key="documents/test/compare-primary.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            repeat_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=employee.id,
                original_filename="compare-repeat.pdf",
                storage_key="documents/test/compare-repeat.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=90,
            )
            db.add_all([primary_document, repeat_document])
            db.flush()

            db.add_all(
                [
                    DocumentVersion(
                        document_id=primary_document.id,
                        version_number=1,
                        storage_key=primary_document.storage_key,
                        parsed_payload={
                            "extracted_fields": {
                                "order_number": "ZN-UPL-COMPARE-001",
                                "grand_total": 10000,
                            },
                            "extracted_items": {
                                "works": [{"name": "ТО"}],
                                "parts": [{"name": "Фильтр"}],
                            },
                        },
                        field_confidence_map={},
                        change_summary="Primary OCR",
                    ),
                    DocumentVersion(
                        document_id=repeat_document.id,
                        version_number=1,
                        storage_key=repeat_document.storage_key,
                        parsed_payload={
                            "extracted_fields": {
                                "order_number": "ZN-UPL-COMPARE-001-R2",
                                "grand_total": 12500,
                            },
                            "extracted_items": {
                                "works": [{"name": "ТО"}, {"name": "Диагностика"}],
                                "parts": [{"name": "Фильтр"}, {"name": "Масло"}],
                            },
                        },
                        field_confidence_map={},
                        change_summary="Repeat OCR",
                    ),
                ]
            )

            repair.source_document_id = primary_document.id
            repair.vehicle_id = foreign_vehicle.id
            db.commit()
            primary_document_id = primary_document.id
            repeat_document_id = repeat_document.id

        response = self.client.get(
            f"/api/documents/{repeat_document_id}/compare",
            headers=headers,
            params={"with_document_id": primary_document_id},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["left_document"]["id"], repeat_document_id)
        self.assertEqual(payload["right_document"]["id"], primary_document_id)
        self.assertEqual(payload["works_count_left"], 2)
        self.assertEqual(payload["works_count_right"], 1)
        self.assertEqual(payload["parts_count_left"], 2)
        self.assertEqual(payload["parts_count_right"], 1)
        differing_fields = {item["field_name"] for item in payload["compared_fields"] if item["is_different"]}
        self.assertIn("order_number", differing_fields)
        self.assertIn("grand_total", differing_fields)

    def test_admin_can_review_comparison_and_promote_document_to_primary(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)

            comparison_vehicle = Vehicle(
                external_id="truck-compare-admin",
                vehicle_type=VehicleType.TRUCK,
                plate_number="L123UV116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(comparison_vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-ADMIN-COMPARE-001",
                repair_date=date(2025, 1, 25),
                vehicle_id=comparison_vehicle.id,
                created_by_user_id=admin.id,
                mileage=10000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            primary_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="admin-primary.pdf",
                storage_key="documents/test/admin-primary.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            candidate_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="admin-repeat.pdf",
                storage_key="documents/test/admin-repeat.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=90,
            )
            db.add_all([primary_document, candidate_document])
            db.flush()

            repair.source_document_id = primary_document.id
            db.commit()
            repair_id = repair.id
            primary_document_id = primary_document.id
            candidate_document_id = candidate_document.id
            repair_id = repair.id

        response = self.client.post(
            f"/api/documents/{candidate_document_id}/compare/review",
            headers=headers,
            json={
                "with_document_id": primary_document_id,
                "action": "make_document_primary",
                "comment": "Повторный скан качественнее",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["action"], "make_document_primary")
        self.assertEqual(payload["document_id"], candidate_document_id)
        self.assertEqual(payload["repair_id"], repair_id)
        self.assertEqual(payload["source_document_id"], candidate_document_id)

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            new_primary = db.get(Document, candidate_document_id)
            old_primary = db.get(Document, primary_document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(new_primary)
            self.assertIsNotNone(old_primary)
            assert repair is not None
            assert new_primary is not None
            assert old_primary is not None
            self.assertEqual(repair.source_document_id, candidate_document_id)
            self.assertTrue(new_primary.is_primary)
            self.assertFalse(old_primary.is_primary)
            self.assertIn("Повторный скан качественнее", new_primary.notes or "")

    def test_set_primary_does_not_reopen_workflow_or_write_audit_when_canonical_source_is_unchanged(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            vehicle = Vehicle(
                external_id="truck-set-primary-noop",
                vehicle_type=VehicleType.TRUCK,
                plate_number="S555SP116",
                brand="Volvo",
                model="FM",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-SET-PRIMARY-NOOP",
                repair_date=date(2025, 1, 29),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=22000,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="set-primary-noop.pdf",
                storage_key="documents/test/set-primary-noop.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.CONFIRMED,
                is_primary=False,
                review_queue_priority=100,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()
            repair_id = repair.id
            document_id = document.id

        response = self.client.post(
            f"/api/documents/{document_id}/set-primary",
            headers=headers,
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertTrue(payload["is_primary"])

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.source_document_id, document_id)
            self.assertEqual(repair.status, RepairStatus.CONFIRMED)
            self.assertTrue(document.is_primary)

            self.assertIsNone(
                db.scalar(
                    select(AuditLog)
                    .where(
                        AuditLog.entity_type == "repair",
                        AuditLog.entity_id == str(repair_id),
                        AuditLog.action_type == "primary_document_changed",
                    )
                    .order_by(AuditLog.id.desc())
                )
            )
            self.assertIsNone(
                db.scalar(
                    select(AuditLog)
                    .where(
                        AuditLog.entity_type == "document",
                        AuditLog.entity_id == str(document_id),
                        AuditLog.action_type == "set_primary",
                    )
                    .order_by(AuditLog.id.desc())
                )
            )

    def test_restore_document_writes_primary_document_audit_when_it_becomes_source(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            vehicle = Vehicle(
                external_id="truck-restore-primary-audit",
                vehicle_type=VehicleType.TRUCK,
                plate_number="R777RT116",
                brand="MAN",
                model="TGS",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-RESTORE-PRIMARY-AUDIT",
                repair_date=date(2025, 1, 30),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=26000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="restore-primary-audit.pdf",
                storage_key="documents/test/restore-primary-audit.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.ARCHIVED,
                is_primary=False,
                review_queue_priority=0,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = None
            db.commit()
            repair_id = repair.id
            document_id = document.id

        response = self.client.post(
            f"/api/documents/{document_id}/restore",
            headers=headers,
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], DocumentStatus.NEEDS_REVIEW.value)
        self.assertTrue(payload["is_primary"])

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            audit_entry = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "repair",
                    AuditLog.entity_id == str(repair_id),
                    AuditLog.action_type == "primary_document_changed",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            self.assertIsNotNone(audit_entry)
            assert repair is not None
            assert document is not None
            assert audit_entry is not None
            self.assertEqual(repair.source_document_id, document_id)
            self.assertTrue(document.is_primary)
            self.assertIsNone(audit_entry.old_value["source_document_id"])
            self.assertEqual(audit_entry.new_value["source_document_id"], document_id)

    def test_restore_document_reopens_confirmed_repair_for_review(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            vehicle = Vehicle(
                external_id="truck-restore-reopen-confirmed",
                vehicle_type=VehicleType.TRUCK,
                plate_number="V888VV116",
                brand="Scania",
                model="G",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-RESTORE-REOPEN-001",
                repair_date=date(2025, 1, 31),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=31000,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="restore-reopen-confirmed.pdf",
                storage_key="documents/test/restore-reopen-confirmed.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.ARCHIVED,
                is_primary=True,
                review_queue_priority=0,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()
            repair_id = repair.id
            document_id = document.id

        response = self.client.post(
            f"/api/documents/{document_id}/restore",
            headers=headers,
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], DocumentStatus.NEEDS_REVIEW.value)
        self.assertTrue(payload["is_primary"])

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertTrue(repair.is_preliminary)
            self.assertEqual(document.status, DocumentStatus.NEEDS_REVIEW)
            self.assertEqual(document.review_queue_priority, 20)

    def test_archive_document_reopens_confirmed_repair_for_review(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            vehicle = Vehicle(
                external_id="truck-archive-reopen-confirmed",
                vehicle_type=VehicleType.TRUCK,
                plate_number="W999WW116",
                brand="Mercedes",
                model="Arocs",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-ARCHIVE-REOPEN-001",
                repair_date=date(2025, 2, 1),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=33000,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="archive-reopen-confirmed.pdf",
                storage_key="documents/test/archive-reopen-confirmed.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.CONFIRMED,
                is_primary=True,
                review_queue_priority=20,
            )
            db.add(document)
            db.flush()

            repair.source_document_id = document.id
            db.commit()
            repair_id = repair.id
            document_id = document.id

        response = self.client.post(
            f"/api/documents/{document_id}/archive",
            headers=headers,
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], DocumentStatus.ARCHIVED.value)

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            document = db.get(Document, document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertTrue(repair.is_preliminary)
            self.assertEqual(document.status, DocumentStatus.ARCHIVED)
            self.assertEqual(document.review_queue_priority, 0)

    def test_comparison_review_accepts_canonical_source_document_even_if_primary_flag_drifted(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            vehicle = Vehicle(
                external_id="truck-compare-primary-drift",
                vehicle_type=VehicleType.TRUCK,
                plate_number="P777PP116",
                brand="Mercedes",
                model="Actros",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-COMPARE-PRIMARY-DRIFT",
                repair_date=date(2025, 1, 28),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=18000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            primary_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="canonical-source.pdf",
                storage_key="documents/test/canonical-source.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=100,
            )
            candidate_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="candidate-repeat.pdf",
                storage_key="documents/test/candidate-repeat-drift.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=90,
            )
            db.add_all([primary_document, candidate_document])
            db.flush()

            repair.source_document_id = primary_document.id
            db.commit()
            repair_id = repair.id
            primary_document_id = primary_document.id
            candidate_document_id = candidate_document.id

        response = self.client.post(
            f"/api/documents/{candidate_document_id}/compare/review",
            headers=headers,
            json={
                "with_document_id": primary_document_id,
                "action": "keep_current_primary",
                "comment": "canonical source should stay accepted",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["action"], "keep_current_primary")
        self.assertEqual(payload["source_document_id"], primary_document_id)

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            primary_document = db.get(Document, primary_document_id)
            candidate_document = db.get(Document, candidate_document_id)
            audit_entry = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "repair",
                    AuditLog.entity_id == str(repair_id),
                    AuditLog.action_type == "document_comparison_reviewed",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(repair)
            self.assertIsNotNone(primary_document)
            self.assertIsNotNone(candidate_document)
            self.assertIsNotNone(audit_entry)
            assert repair is not None
            assert primary_document is not None
            assert candidate_document is not None
            assert audit_entry is not None
            self.assertEqual(repair.source_document_id, primary_document_id)
            self.assertTrue(primary_document.is_primary)
            self.assertFalse(candidate_document.is_primary)
            old_documents = {item["id"]: item for item in audit_entry.old_value["documents"]}
            new_documents = {item["id"]: item for item in audit_entry.new_value["documents"]}
            self.assertTrue(old_documents[primary_document_id]["is_primary"])
            self.assertFalse(old_documents[candidate_document_id]["is_primary"])
            self.assertTrue(new_documents[primary_document_id]["is_primary"])
            self.assertFalse(new_documents[candidate_document_id]["is_primary"])

    def test_comparison_review_ignores_legacy_foreign_source_document_id_in_response_and_audit(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            local_vehicle = Vehicle(
                external_id="truck-compare-drift-local",
                vehicle_type=VehicleType.TRUCK,
                plate_number="M321OP116",
                brand="MAN",
                model="TGX",
                status=VehicleStatus.ACTIVE,
            )
            foreign_vehicle = Vehicle(
                external_id="truck-compare-drift-foreign",
                vehicle_type=VehicleType.TRUCK,
                plate_number="N654QR116",
                brand="Scania",
                model="R",
                status=VehicleStatus.ACTIVE,
            )
            db.add_all([local_vehicle, foreign_vehicle])
            db.flush()

            local_repair = Repair(
                order_number="ZN-COMPARE-DRIFT-001",
                repair_date=date(2025, 1, 26),
                vehicle_id=local_vehicle.id,
                created_by_user_id=admin.id,
                mileage=12000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            foreign_repair = Repair(
                order_number="ZN-COMPARE-DRIFT-FOREIGN",
                repair_date=date(2025, 1, 27),
                vehicle_id=foreign_vehicle.id,
                created_by_user_id=admin.id,
                mileage=15000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add_all([local_repair, foreign_repair])
            db.flush()

            primary_document = Document(
                repair_id=local_repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="drift-primary.pdf",
                storage_key="documents/test/drift-primary.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            candidate_document = Document(
                repair_id=local_repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="drift-repeat.pdf",
                storage_key="documents/test/drift-repeat.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=90,
            )
            foreign_document = Document(
                repair_id=foreign_repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="foreign-source.pdf",
                storage_key="documents/test/foreign-source-drift.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            db.add_all([primary_document, candidate_document, foreign_document])
            db.flush()

            local_repair.source_document_id = foreign_document.id
            db.commit()
            primary_document_id = primary_document.id
            candidate_document_id = candidate_document.id
            local_repair_id = local_repair.id

        response = self.client.post(
            f"/api/documents/{candidate_document_id}/compare/review",
            headers=headers,
            json={
                "with_document_id": primary_document_id,
                "action": "keep_current_primary",
                "comment": "Оставляем текущий основной документ",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["action"], "keep_current_primary")
        self.assertEqual(payload["source_document_id"], primary_document_id)

        with self.SessionLocal() as db:
            repair = db.get(Repair, local_repair_id)
            primary_document = db.get(Document, primary_document_id)
            candidate_document = db.get(Document, candidate_document_id)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(primary_document)
            self.assertIsNotNone(candidate_document)
            assert repair is not None
            assert primary_document is not None
            assert candidate_document is not None
            self.assertEqual(repair.source_document_id, primary_document_id)
            self.assertTrue(primary_document.is_primary)
            self.assertFalse(candidate_document.is_primary)

            audit_entry = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "repair",
                    AuditLog.entity_id == str(local_repair_id),
                    AuditLog.action_type == "document_comparison_reviewed",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(audit_entry)
            assert audit_entry is not None
            self.assertEqual(audit_entry.old_value["source_document_id"], primary_document_id)
            self.assertEqual(audit_entry.new_value["source_document_id"], primary_document_id)

    def test_comparison_review_ignores_legacy_archived_source_document_id_in_response_and_audit(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            self.assertIsNotNone(admin)
            assert admin is not None

            vehicle = Vehicle(
                external_id="truck-compare-drift-archived",
                vehicle_type=VehicleType.TRUCK,
                plate_number="T321TT116",
                brand="DAF",
                model="XG",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ZN-COMPARE-DRIFT-ARCHIVED",
                repair_date=date(2025, 1, 27),
                vehicle_id=vehicle.id,
                created_by_user_id=admin.id,
                mileage=14000,
                status=RepairStatus.IN_REVIEW,
                is_preliminary=True,
            )
            db.add(repair)
            db.flush()

            primary_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="archived-drift-primary.pdf",
                storage_key="documents/test/archived-drift-primary.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=True,
                review_queue_priority=100,
            )
            candidate_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="archived-drift-repeat.pdf",
                storage_key="documents/test/archived-drift-repeat.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.REPEAT_SCAN,
                status=DocumentStatus.NEEDS_REVIEW,
                is_primary=False,
                review_queue_priority=90,
            )
            archived_document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=admin.id,
                original_filename="archived-drift-archived.pdf",
                storage_key="documents/test/archived-drift-archived.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.ARCHIVED,
                is_primary=False,
                review_queue_priority=0,
            )
            db.add_all([primary_document, candidate_document, archived_document])
            db.flush()

            repair.source_document_id = archived_document.id
            db.commit()
            primary_document_id = primary_document.id
            candidate_document_id = candidate_document.id
            repair_id = repair.id

        response = self.client.post(
            f"/api/documents/{candidate_document_id}/compare/review",
            headers=headers,
            json={
                "with_document_id": primary_document_id,
                "action": "keep_current_primary",
                "comment": "Архивный source_document_id должен игнорироваться",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["action"], "keep_current_primary")
        self.assertEqual(payload["source_document_id"], primary_document_id)

        with self.SessionLocal() as db:
            repair = db.get(Repair, repair_id)
            audit_entry = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "repair",
                    AuditLog.entity_id == str(repair_id),
                    AuditLog.action_type == "document_comparison_reviewed",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(repair)
            self.assertIsNotNone(audit_entry)
            assert repair is not None
            assert audit_entry is not None
            self.assertEqual(repair.source_document_id, primary_document_id)
            self.assertEqual(audit_entry.old_value["source_document_id"], primary_document_id)
            self.assertEqual(audit_entry.new_value["source_document_id"], primary_document_id)

    def test_manual_service_assignment_updates_single_warning_without_duplicates(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            repair.service_id = None
            document.status = DocumentStatus.RECOGNIZED
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_fields": {},
                        "manual_review_reasons": ["service_name_missing"],
                    },
                    field_confidence_map={},
                    change_summary="Initial OCR payload",
                )
            )
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_service_missing",
                    severity=CheckSeverity.WARNING,
                    title="Не удалось определить сервис",
                    details="Сервис у ремонта не назначен. Нужна ручная проверка.",
                    calculation_payload={"reason": "service_name_missing"},
                    is_resolved=False,
                )
            )
            db.commit()

        assign_response = self.client.patch(
            "/api/repairs/1/service",
            headers=headers,
            json={"service_name": "Service Alpha"},
        )
        self.assertEqual(assign_response.status_code, 200, assign_response.text)

        repeat_assign_response = self.client.patch(
            "/api/repairs/1/service",
            headers=headers,
            json={"service_name": "Service Alpha"},
        )
        self.assertEqual(repeat_assign_response.status_code, 200, repeat_assign_response.text)

        clear_response = self.client.patch(
            "/api/repairs/1/service",
            headers=headers,
            json={"service_name": ""},
        )
        self.assertEqual(clear_response.status_code, 200, clear_response.text)

        restore_response = self.client.patch(
            "/api/repairs/1/service",
            headers=headers,
            json={"service_name": "Service Alpha"},
        )
        self.assertEqual(restore_response.status_code, 200, restore_response.text)

        payload = restore_response.json()
        service_checks = [item for item in payload["checks"] if item["check_type"] == "ocr_service_missing"]
        self.assertEqual(len(service_checks), 1)
        self.assertTrue(service_checks[0]["is_resolved"])
        self.assertEqual(payload["service"]["name"], "Service Alpha")
        latest_payload = payload["documents"][0]["versions"][0]["parsed_payload"]
        self.assertEqual(latest_payload["manual_review_reasons"], [])
        self.assertEqual(latest_payload["extracted_fields"]["service_name"], "Service Alpha")

    def test_manual_service_assignment_reopens_admin_confirmed_repair_as_preliminary(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            service = Service(
                name="Confirmed Repair Reopen Service",
                city="Kazan",
                status=ServiceStatus.CONFIRMED,
                created_by_user_id=1,
                confirmed_by_user_id=1,
            )
            db.add(service)
            db.flush()

            repair.status = RepairStatus.CONFIRMED
            repair.is_preliminary = False
            repair.service_id = None
            document.status = DocumentStatus.CONFIRMED
            db.commit()

        response = self.client.patch(
            "/api/repairs/1/service",
            headers=headers,
            json={"service_name": "Confirmed Repair Reopen Service"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], RepairStatus.IN_REVIEW.value)
        self.assertTrue(payload["is_preliminary"])
        self.assertEqual(payload["service"]["name"], "Confirmed Repair Reopen Service")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            self.assertIsNotNone(repair)
            assert repair is not None
            self.assertEqual(repair.status, RepairStatus.IN_REVIEW)
            self.assertTrue(repair.is_preliminary)

    def test_manual_service_assignment_tolerates_non_object_check_payload(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            repair.service_id = None
            document.status = DocumentStatus.RECOGNIZED
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_fields": {},
                        "manual_review_reasons": ["service_name_missing"],
                    },
                    field_confidence_map={},
                    change_summary="Initial OCR payload",
                )
            )
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_service_missing",
                    severity=CheckSeverity.WARNING,
                    title="Не удалось определить сервис",
                    details="Сервис у ремонта не назначен. Нужна ручная проверка.",
                    calculation_payload=["legacy-broken-payload"],
                    is_resolved=False,
                )
            )
            db.commit()

        response = self.client.patch(
            "/api/repairs/1/service",
            headers=headers,
            json={"service_name": "Service Alpha"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        service_checks = [item for item in payload["checks"] if item["check_type"] == "ocr_service_missing"]
        self.assertEqual(len(service_checks), 1)
        self.assertTrue(service_checks[0]["is_resolved"])
        self.assertEqual(payload["service"]["name"], "Service Alpha")

    def test_manual_service_assignment_stores_canonical_service_name_for_alias(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            repair.service_id = None
            document.status = DocumentStatus.RECOGNIZED
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_fields": {},
                        "manual_review_reasons": ["service_name_missing"],
                    },
                    field_confidence_map={},
                    change_summary="Initial OCR payload",
                )
            )
            db.commit()

        response = self.client.patch(
            "/api/repairs/1/service",
            headers=headers,
            json={"service_name": "AXB"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["service"]["name"], "ООО «АХВ Трак Сервис»")
        latest_payload = payload["documents"][0]["versions"][0]["parsed_payload"]
        self.assertEqual(latest_payload["manual_review_reasons"], [])
        self.assertEqual(latest_payload["extracted_fields"]["service_name"], "ООО «АХВ Трак Сервис»")

    def test_admin_manual_update_syncs_service_and_review_payload_state(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None

            repair.service_id = None
            document.status = DocumentStatus.RECOGNIZED
            db.add(
                DocumentVersion(
                    document_id=document.id,
                    version_number=1,
                    storage_key=document.storage_key,
                    parsed_payload={
                        "extracted_fields": {},
                        "manual_review_reasons": ["service_name_missing", "order_number_missing"],
                    },
                    field_confidence_map={},
                    change_summary="Initial OCR payload",
                )
            )
            db.add(
                RepairCheck(
                    repair_id=repair.id,
                    check_type="ocr_service_missing",
                    severity=CheckSeverity.WARNING,
                    title="Не удалось определить сервис",
                    details="Сервис у ремонта не назначен. Нужна ручная проверка.",
                    calculation_payload={"reason": "service_name_missing"},
                    is_resolved=False,
                )
            )
            db.commit()

        response = self.client.patch(
            "/api/repairs/1",
            headers=headers,
            json={
                "service_name": "Service Alpha",
                "order_number": "ZN-MANUAL-SYNC-001",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["service"]["name"], "Service Alpha")
        service_checks = [item for item in payload["checks"] if item["check_type"] == "ocr_service_missing"]
        self.assertEqual(len(service_checks), 1)
        self.assertTrue(service_checks[0]["is_resolved"])
        latest_payload = payload["documents"][0]["versions"][0]["parsed_payload"]
        self.assertEqual(latest_payload["manual_review_reasons"], [])
        self.assertEqual(latest_payload["extracted_fields"]["service_name"], "Service Alpha")
        self.assertEqual(latest_payload["extracted_fields"]["order_number"], "ZN-MANUAL-SYNC-001")

    def test_admin_created_service_keeps_confirmed_status(self) -> None:
        headers = self._get_auth_headers("admin")

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock, patch.object(
            services_api,
            "find_service_catalog_entry",
            autospec=True,
            return_value=None,
        ):
            sync_mock.return_value = ()
            response = self.client.post(
                "/api/services",
                headers=headers,
                json={
                    "name": "Custom Confirmed Service",
                    "city": "Moscow",
                    "status": "confirmed",
                },
            )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["status"], "confirmed")
        self.assertIsNotNone(payload["confirmed_by_user_id"])

        with self.SessionLocal() as db:
            audit_entry = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "service",
                    AuditLog.action_type == "service_created",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(audit_entry)
            assert audit_entry is not None
            self.assertEqual(audit_entry.user_id, 1)
            self.assertEqual(audit_entry.new_value["name"], "Custom Confirmed Service")
            self.assertEqual(audit_entry.new_value["status"], "confirmed")
            self.assertEqual(audit_entry.new_value["confirmed_by_user_id"], payload["confirmed_by_user_id"])

    def test_service_update_writes_audit_log_with_old_and_new_values(self) -> None:
        headers = self._get_auth_headers("admin")

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock, patch.object(
            services_api,
            "find_service_catalog_entry",
            autospec=True,
            return_value=None,
        ):
            sync_mock.return_value = ()
            response = self.client.patch(
                "/api/services/1",
                headers=headers,
                json={
                    "name": "Service Alpha Updated",
                    "city": "Moscow",
                    "status": "preliminary",
                    "comment": "Archived for operational use",
                },
            )

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["name"], "Service Alpha Updated")
        self.assertEqual(payload["status"], "preliminary")

        with self.SessionLocal() as db:
            audit_entry = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "service",
                    AuditLog.entity_id == "1",
                    AuditLog.action_type == "service_updated",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(audit_entry)
            assert audit_entry is not None
            self.assertEqual(audit_entry.user_id, 1)
            self.assertEqual(audit_entry.old_value["name"], "Service Alpha")
            self.assertEqual(audit_entry.old_value["status"], "confirmed")
            self.assertEqual(audit_entry.new_value["name"], "Service Alpha Updated")
            self.assertEqual(audit_entry.new_value["city"], "Moscow")
            self.assertEqual(audit_entry.new_value["status"], "preliminary")
            self.assertEqual(audit_entry.new_value["comment"], "Archived for operational use")

    def test_service_archive_and_restore_use_explicit_endpoints(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            service_item = Service(
                name="Archivable Service",
                city="Kazan",
                status=ServiceStatus.CONFIRMED,
                created_by_user_id=1,
                confirmed_by_user_id=1,
            )
            db.add(service_item)
            db.commit()
            service_id = service_item.id

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock:
            sync_mock.return_value = ()

            patch_response = self.client.patch(
                f"/api/services/{service_id}",
                headers=headers,
                json={"status": "archived"},
            )
            self.assertEqual(patch_response.status_code, 409, patch_response.text)
            self.assertIn("explicit archive/restore endpoints", patch_response.json()["detail"])

            archive_response = self.client.post(f"/api/services/{service_id}/archive", headers=headers)
            self.assertEqual(archive_response.status_code, 200, archive_response.text)
            self.assertEqual(archive_response.json()["status"], "archived")

            restore_response = self.client.post(f"/api/services/{service_id}/restore", headers=headers)
            self.assertEqual(restore_response.status_code, 200, restore_response.text)
            self.assertEqual(restore_response.json()["status"], "confirmed")

        with self.SessionLocal() as db:
            archive_audit = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "service",
                    AuditLog.entity_id == str(service_id),
                    AuditLog.action_type == "service_archived",
                )
                .order_by(AuditLog.id.desc())
            )
            restore_audit = db.scalar(
                select(AuditLog)
                .where(
                    AuditLog.entity_type == "service",
                    AuditLog.entity_id == str(service_id),
                    AuditLog.action_type == "service_restored",
                )
                .order_by(AuditLog.id.desc())
            )
            self.assertIsNotNone(archive_audit)
            self.assertIsNotNone(restore_audit)
            assert archive_audit is not None
            assert restore_audit is not None
            self.assertEqual(archive_audit.old_value["status"], "confirmed")
            self.assertEqual(archive_audit.new_value["status"], "archived")
            self.assertEqual(restore_audit.old_value["status"], "archived")
            self.assertEqual(restore_audit.new_value["status"], "confirmed")

    def test_service_restore_rejects_non_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock:
            sync_mock.return_value = ()
            response = self.client.post("/api/services/1/restore", headers=headers)

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Service is not archived")

    def test_service_archive_rejects_already_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            service_item = Service(
                name="Already Archived Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
                confirmed_by_user_id=1,
            )
            db.add(service_item)
            db.commit()
            service_id = service_item.id

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock:
            sync_mock.return_value = ()
            response = self.client.post(f"/api/services/{service_id}/archive", headers=headers)

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Service is already archived")

    def test_preliminary_service_update_does_not_mark_service_as_confirmed(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            service_item = Service(
                name="Preliminary Service",
                city="Kazan",
                status=ServiceStatus.PRELIMINARY,
                created_by_user_id=1,
                confirmed_by_user_id=None,
            )
            db.add(service_item)
            db.commit()
            service_id = service_item.id

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock, patch.object(
            services_api,
            "find_service_catalog_entry",
            autospec=True,
            return_value=None,
        ):
            sync_mock.return_value = ()
            update_response = self.client.patch(
                f"/api/services/{service_id}",
                headers=headers,
                json={"city": "Moscow", "comment": "Touched by admin"},
            )
            self.assertEqual(update_response.status_code, 200, update_response.text)
            self.assertEqual(update_response.json()["status"], "preliminary")
            self.assertIsNone(update_response.json()["confirmed_by_user_id"])

            archive_response = self.client.post(f"/api/services/{service_id}/archive", headers=headers)
            self.assertEqual(archive_response.status_code, 200, archive_response.text)
            self.assertEqual(archive_response.json()["status"], "archived")

            restore_response = self.client.post(f"/api/services/{service_id}/restore", headers=headers)
            self.assertEqual(restore_response.status_code, 200, restore_response.text)
            self.assertEqual(restore_response.json()["status"], "preliminary")
            self.assertIsNone(restore_response.json()["confirmed_by_user_id"])

    def test_service_with_active_repairs_cannot_be_archived(self) -> None:
        headers = self._get_auth_headers("admin")

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock:
            sync_mock.return_value = ()
            response = self.client.post("/api/services/1/archive", headers=headers)

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Нельзя архивировать сервис с активными ремонтами")

        with self.SessionLocal() as db:
            service_item = db.get(Service, 1)
            self.assertIsNotNone(service_item)
            assert service_item is not None
            self.assertEqual(service_item.status, ServiceStatus.CONFIRMED)

    def test_service_with_only_archived_repairs_can_be_archived_despite_stale_active_document(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            vehicle = Vehicle(
                external_id="truck-service-archive-drift",
                vehicle_type=VehicleType.TRUCK,
                plate_number="SVC-DRIFT-116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            service_item = Service(
                name="Archivable Drift Service",
                city="Kazan",
                status=ServiceStatus.CONFIRMED,
                created_by_user_id=1,
                confirmed_by_user_id=1,
            )
            db.add_all([vehicle, service_item])
            db.flush()

            archived_repair = Repair(
                order_number="ARCHIVE-SVC-DRIFT-001",
                repair_date=date(2025, 2, 2),
                vehicle_id=vehicle.id,
                service_id=service_item.id,
                created_by_user_id=1,
                mileage=6000,
                status=RepairStatus.ARCHIVED,
                is_preliminary=False,
            )
            db.add(archived_repair)
            db.flush()

            db.add(
                Document(
                    repair_id=archived_repair.id,
                    uploaded_by_user_id=1,
                    original_filename="archived-service-drift-order.pdf",
                    storage_key="documents/test/archived-service-drift-order.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.NEEDS_REVIEW,
                    is_primary=False,
                    review_queue_priority=20,
                )
            )
            db.commit()
            service_id = service_item.id

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock:
            sync_mock.return_value = ()
            response = self.client.post(f"/api/services/{service_id}/archive", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["status"], "archived")

    def test_list_services_does_not_persist_sync_side_effects(self) -> None:
        headers = self._get_auth_headers("admin")

        def mutate_without_commit(db: Session, *, commit: bool = False):
            service_item = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(service_item)
            service_item.city = "Transient City"
            db.add(service_item)
            db.flush()
            return (service_item,)

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True, side_effect=mutate_without_commit):
            response = self.client.get("/api/services", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        service_payload = next(item for item in payload["items"] if item["name"] == "Service Alpha")
        self.assertEqual(service_payload["city"], "Transient City")

        with self.SessionLocal() as db:
            persisted = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(persisted)
            self.assertEqual(persisted.city, "Kazan")

    def test_employee_service_list_hides_archived_services_even_with_archived_filter(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            archived_service = Service(
                name="Archived Employee Hidden Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
                confirmed_by_user_id=1,
            )
            db.add(archived_service)
            db.commit()

        default_response = self.client.get("/api/services", headers=headers)

        self.assertEqual(default_response.status_code, 200, default_response.text)
        default_payload = default_response.json()
        self.assertFalse(any(item["status"] == "archived" for item in default_payload["items"]))

        archived_response = self.client.get("/api/services?status=archived", headers=headers)

        self.assertEqual(archived_response.status_code, 200, archived_response.text)
        archived_payload = archived_response.json()
        self.assertEqual(archived_payload["total"], 0)
        self.assertEqual(archived_payload["items"], [])
        self.assertEqual(archived_payload["cities"], [])

    def test_service_list_cities_follow_search_filter(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            db.add_all(
                [
                    Service(
                        name="Alpha Fleet Branch",
                        city="Kazan",
                        status=ServiceStatus.CONFIRMED,
                        created_by_user_id=1,
                        confirmed_by_user_id=1,
                    ),
                    Service(
                        name="Beta Truck Branch",
                        city="Moscow",
                        status=ServiceStatus.CONFIRMED,
                        created_by_user_id=1,
                        confirmed_by_user_id=1,
                    ),
                ]
            )
            db.commit()

        response = self.client.get("/api/services?q=alpha", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual({item["name"] for item in payload["items"]}, {"Alpha Fleet Branch", "Service Alpha"})
        self.assertEqual(payload["cities"], ["Kazan"])

    def test_confirmed_service_stays_visible_after_rename_outside_catalog(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            service_item = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(service_item)
            service_item.name = "Renamed Catalog Service"
            service_item.created_by_user_id = None
            service_item.confirmed_by_user_id = 1
            db.add(service_item)
            db.commit()

        with patch.object(services_api, "ensure_service_catalog_synced", autospec=True) as sync_mock, patch.object(
            services_api,
            "get_service_catalog_names",
            autospec=True,
            return_value=(),
        ):
            sync_mock.return_value = ()
            response = self.client.get("/api/services", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertTrue(any(item["name"] == "Renamed Catalog Service" for item in payload["items"]))

    def test_catalog_sync_preserves_admin_confirmed_overrides(self) -> None:
        with self.SessionLocal() as db:
            admin = db.scalar(select(User).where(User.login == "admin"))
            service_item = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(admin)
            self.assertIsNotNone(service_item)
            service_item.city = "Manual City"
            service_item.contact = "manual@example.com"
            service_item.comment = "Manual comment"
            service_item.status = ServiceStatus.CONFIRMED
            service_item.confirmed_by_user_id = admin.id
            db.add(service_item)
            db.commit()

            with patch(
                "app.services.service_catalog.get_service_catalog_entries",
                return_value=(
                    ServiceCatalogEntry(
                        name="Service Alpha",
                        city="Catalog City",
                        contact="catalog@example.com",
                        comment="Catalog comment",
                        aliases=("Service Alpha",),
                    ),
                ),
            ):
                ensure_service_catalog_synced(db, commit=False)

            db.refresh(service_item)
            self.assertEqual(service_item.city, "Manual City")
            self.assertEqual(service_item.contact, "manual@example.com")
            self.assertEqual(service_item.comment, "Manual comment")

    def test_builtin_fallback_catalog_syncs_axb_logistics_and_klever_services(self) -> None:
        with self.SessionLocal() as db:
            ensure_service_catalog_synced(db, commit=False)

            axb_service = db.scalar(select(Service).where(Service.name == "ООО «АХВ Трак Сервис»"))
            logistics_service = db.scalar(select(Service).where(Service.name == "ООО «ЛОГИСТИКА»"))
            klever_service = db.scalar(select(Service).where(Service.name == "ООО «КЛЕВЕР ТРАК»"))

            self.assertIsNotNone(axb_service)
            self.assertIsNotNone(logistics_service)
            self.assertIsNotNone(klever_service)
            self.assertEqual(axb_service.status, ServiceStatus.CONFIRMED)
            self.assertEqual(logistics_service.status, ServiceStatus.CONFIRMED)
            self.assertEqual(klever_service.status, ServiceStatus.CONFIRMED)

    def test_builtin_fallback_catalog_resolves_axb_logistics_and_klever_aliases(self) -> None:
        with self.SessionLocal() as db:
            axb_entry = find_service_catalog_entry("AXB")
            logistics_entry = find_service_catalog_entry('Общество с ограниченной ответственностью "ЛОГИСТИКА"')
            klever_entry = find_service_catalog_entry("Клевер Трак")

            self.assertIsNotNone(axb_entry)
            self.assertIsNotNone(logistics_entry)
            self.assertIsNotNone(klever_entry)
            self.assertEqual(axb_entry.name, "ООО «АХВ Трак Сервис»")
            self.assertEqual(logistics_entry.name, "ООО «ЛОГИСТИКА»")
            self.assertEqual(klever_entry.name, "ООО «КЛЕВЕР ТРАК»")

            axb_service = resolve_service_by_name(db, "AXB")
            logistics_service = resolve_service_by_name(
                db,
                'Общество с ограниченной ответственностью "ЛОГИСТИКА"',
            )
            klever_service = resolve_service_by_name(db, "Клевер Трак")

            self.assertIsNotNone(axb_service)
            self.assertIsNotNone(logistics_service)
            self.assertIsNotNone(klever_service)
            self.assertEqual(axb_service.name, "ООО «АХВ Трак Сервис»")
            self.assertEqual(logistics_service.name, "ООО «ЛОГИСТИКА»")

    def test_resolve_service_by_name_ignores_archived_catalog_service(self) -> None:
        with self.SessionLocal() as db:
            ensure_service_catalog_synced(db, commit=False)
            axb_service = db.scalar(select(Service).where(Service.name == "ООО «АХВ Трак Сервис»"))
            self.assertIsNotNone(axb_service)
            assert axb_service is not None

            axb_service.status = ServiceStatus.ARCHIVED
            db.commit()

            self.assertIsNone(resolve_service_by_name(db, "AXB"))

    def test_find_service_name_in_text_ignores_archived_catalog_service_when_db_provided(self) -> None:
        with self.SessionLocal() as db:
            ensure_service_catalog_synced(db, commit=False)
            axb_service = db.scalar(select(Service).where(Service.name == "ООО «АХВ Трак Сервис»"))
            self.assertIsNotNone(axb_service)
            assert axb_service is not None

            axb_service.status = ServiceStatus.ARCHIVED
            db.commit()

            self.assertIsNone(find_service_name_in_text("AXB заказ-наряд", db=db))

    def test_service_catalog_sync_does_not_mutate_archived_catalog_service_fields(self) -> None:
        with self.SessionLocal() as db:
            ensure_service_catalog_synced(db, commit=False)
            axb_service = db.scalar(select(Service).where(Service.name == "ООО «АХВ Трак Сервис»"))
            self.assertIsNotNone(axb_service)
            assert axb_service is not None

            axb_service.status = ServiceStatus.ARCHIVED
            axb_service.city = "Архивный город"
            axb_service.contact = "archived@example.com"
            axb_service.comment = "Archived comment"
            db.commit()

            ensure_service_catalog_synced(db, commit=False)
            db.refresh(axb_service)

            self.assertEqual(axb_service.status, ServiceStatus.ARCHIVED)
            self.assertEqual(axb_service.city, "Архивный город")
            self.assertEqual(axb_service.contact, "archived@example.com")
            self.assertEqual(axb_service.comment, "Archived comment")

    def test_cannot_assign_employee_to_archived_vehicle(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            archived_vehicle = Vehicle(
                external_id="truck-archived",
                vehicle_type=VehicleType.TRUCK,
                plate_number="X999XX116",
                brand="Archived",
                model="Truck",
                status=VehicleStatus.ARCHIVED,
            )
            db.add(archived_vehicle)
            db.commit()
            archived_vehicle_id = archived_vehicle.id

        response = self.client.post(
            "/api/users/2/vehicle-assignments",
            headers=headers,
            json={
                "vehicle_id": archived_vehicle_id,
                "starts_at": "2025-02-01",
            },
        )

        self.assertEqual(response.status_code, 400, response.text)
        self.assertEqual(response.json()["detail"], "Нельзя назначить сотрудника на архивную технику")

    def test_vehicle_list_excludes_archived_by_default_and_allows_explicit_archive_filter(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            archived_vehicle = Vehicle(
                external_id="truck-archived-list",
                vehicle_type=VehicleType.TRUCK,
                plate_number="Y999YY116",
                brand="Archived",
                model="List",
                status=VehicleStatus.ARCHIVED,
            )
            db.add(archived_vehicle)
            db.commit()
            archived_vehicle_id = archived_vehicle.id

        default_response = self.client.get("/api/vehicles?limit=50", headers=headers)

        self.assertEqual(default_response.status_code, 200, default_response.text)
        default_payload = default_response.json()
        self.assertEqual(default_payload["total"], 1)
        self.assertEqual([item["id"] for item in default_payload["items"]], [1])

        archived_response = self.client.get("/api/vehicles?limit=50&status=archived", headers=headers)

        self.assertEqual(archived_response.status_code, 200, archived_response.text)
        archived_payload = archived_response.json()
        self.assertEqual(archived_payload["total"], 1)
        self.assertEqual([item["id"] for item in archived_payload["items"]], [archived_vehicle_id])

    def test_employee_visibility_excludes_placeholder_vehicle_even_with_legacy_assignment(self) -> None:
        headers = self._get_auth_headers("employee")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            self.assertIsNotNone(employee)
            assert employee is not None

            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-ACCESS",
                brand="Placeholder",
                model="Legacy",
                status=VehicleStatus.ACTIVE,
            )
            db.add(placeholder_vehicle)
            db.flush()

            db.add(
                VehicleAssignmentHistory(
                    vehicle_id=placeholder_vehicle.id,
                    user_id=employee.id,
                    starts_at=date(2025, 1, 1),
                    ends_at=None,
                    assigned_by_user_id=1,
                    comment="legacy placeholder assignment",
                )
            )
            db.commit()

        vehicles_response = self.client.get("/api/vehicles?limit=50", headers=headers)

        self.assertEqual(vehicles_response.status_code, 200, vehicles_response.text)
        vehicles_payload = vehicles_response.json()
        self.assertEqual(vehicles_payload["total"], 1)
        self.assertEqual([item["external_id"] for item in vehicles_payload["items"]], ["truck-1"])

        dashboard_response = self.client.get("/api/dashboard/summary", headers=headers)

        self.assertEqual(dashboard_response.status_code, 200, dashboard_response.text)
        dashboard_payload = dashboard_response.json()
        self.assertEqual(dashboard_payload["vehicles_total"], 1)

    def test_admin_vehicle_views_exclude_placeholder_vehicle(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            db.add(
                Vehicle(
                    external_id="__batch_import_placeholder__",
                    vehicle_type=VehicleType.TRUCK,
                    plate_number="PLACEHOLDER-ADMIN",
                    brand="Placeholder",
                    model="System",
                    status=VehicleStatus.ACTIVE,
                )
            )
            db.commit()

        vehicles_response = self.client.get("/api/vehicles?limit=50", headers=headers)

        self.assertEqual(vehicles_response.status_code, 200, vehicles_response.text)
        vehicles_payload = vehicles_response.json()
        self.assertEqual(vehicles_payload["total"], 1)
        self.assertEqual([item["external_id"] for item in vehicles_payload["items"]], ["truck-1"])

        dashboard_response = self.client.get("/api/dashboard/summary", headers=headers)

        self.assertEqual(dashboard_response.status_code, 200, dashboard_response.text)
        dashboard_payload = dashboard_response.json()
        self.assertEqual(dashboard_payload["vehicles_total"], 1)

        search_response = self.client.get("/api/search/global?q=PLACEHOLDER-ADMIN", headers=headers)

        self.assertEqual(search_response.status_code, 200, search_response.text)
        search_payload = search_response.json()
        self.assertEqual(search_payload["vehicles_total"], 0)
        self.assertEqual(search_payload["vehicles"], [])

    def test_placeholder_vehicle_cannot_be_archived_via_patch(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            placeholder_vehicle = Vehicle(
                external_id="__batch_import_placeholder__",
                vehicle_type=VehicleType.TRUCK,
                plate_number="PLACEHOLDER-PATCH",
                brand="Placeholder",
                model="Vehicle",
                status=VehicleStatus.ACTIVE,
            )
            db.add(placeholder_vehicle)
            db.commit()
            placeholder_vehicle_id = placeholder_vehicle.id

        response = self.client.patch(
            f"/api/vehicles/{placeholder_vehicle_id}",
            headers=headers,
            json={"status": "archived"},
        )

        self.assertEqual(response.status_code, 400, response.text)
        self.assertEqual(response.json()["detail"], "Нельзя архивировать системную placeholder-технику")

        with self.SessionLocal() as db:
            refreshed_vehicle = db.get(Vehicle, placeholder_vehicle_id)
            self.assertIsNotNone(refreshed_vehicle)
            assert refreshed_vehicle is not None
            self.assertEqual(refreshed_vehicle.status, VehicleStatus.ACTIVE)

    def test_vehicle_with_active_repair_cannot_be_archived_via_patch(self) -> None:
        headers = self._get_auth_headers("admin")

        response = self.client.patch(
            "/api/vehicles/1",
            headers=headers,
            json={"status": "archived"},
        )

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["detail"], "Нельзя архивировать технику с активными ремонтами")

        with self.SessionLocal() as db:
            vehicle = db.get(Vehicle, 1)
            self.assertIsNotNone(vehicle)
            assert vehicle is not None
            self.assertEqual(vehicle.status, VehicleStatus.ACTIVE)

    def test_vehicle_with_only_archived_repairs_can_be_archived_via_patch(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            vehicle = Vehicle(
                external_id="truck-archivable",
                vehicle_type=VehicleType.TRUCK,
                plate_number="ARCHIVABLE-116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ARCHIVE-VEH-001",
                repair_date=date(2025, 2, 1),
                vehicle_id=vehicle.id,
                created_by_user_id=1,
                mileage=5000,
                status=RepairStatus.ARCHIVED,
                is_preliminary=False,
            )
            db.add(repair)
            db.flush()

            db.add(
                Document(
                    repair_id=repair.id,
                    uploaded_by_user_id=1,
                    original_filename="archived-vehicle-order.pdf",
                    storage_key="documents/test/archived-vehicle-order.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.ARCHIVED,
                    is_primary=False,
                    review_queue_priority=0,
                )
            )
            db.commit()
            vehicle_id = vehicle.id

        response = self.client.patch(
            f"/api/vehicles/{vehicle_id}",
            headers=headers,
            json={"status": "archived"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["status"], "archived")

        with self.SessionLocal() as db:
            refreshed_vehicle = db.get(Vehicle, vehicle_id)
            self.assertIsNotNone(refreshed_vehicle)
            assert refreshed_vehicle is not None
            self.assertEqual(refreshed_vehicle.status, VehicleStatus.ARCHIVED)

    def test_vehicle_with_only_archived_repairs_can_be_archived_despite_stale_active_document(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            vehicle = Vehicle(
                external_id="truck-archivable-drift",
                vehicle_type=VehicleType.TRUCK,
                plate_number="ARCHDRIFT-116",
                brand="Volvo",
                model="FH",
                status=VehicleStatus.ACTIVE,
            )
            db.add(vehicle)
            db.flush()

            repair = Repair(
                order_number="ARCHIVE-VEH-DRIFT-001",
                repair_date=date(2025, 2, 3),
                vehicle_id=vehicle.id,
                created_by_user_id=1,
                mileage=7000,
                status=RepairStatus.ARCHIVED,
                is_preliminary=False,
            )
            db.add(repair)
            db.flush()

            db.add(
                Document(
                    repair_id=repair.id,
                    uploaded_by_user_id=1,
                    original_filename="archived-vehicle-drift-order.pdf",
                    storage_key="documents/test/archived-vehicle-drift-order.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.NEEDS_REVIEW,
                    is_primary=False,
                    review_queue_priority=20,
                )
            )
            db.commit()
            vehicle_id = vehicle.id

        response = self.client.patch(
            f"/api/vehicles/{vehicle_id}",
            headers=headers,
            json={"status": "archived"},
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["status"], "archived")

    def test_vehicle_detail_excludes_archived_repairs_and_archived_service_history(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            archived_service = Service(
                name="Archived Vehicle Detail Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
            )
            db.add(archived_service)
            db.flush()

            active_historical_repair = Repair(
                order_number="ZN-HIST-VEH-001",
                repair_date=date(2025, 1, 20),
                vehicle_id=1,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=121000,
                grand_total=25000,
                status=RepairStatus.CONFIRMED,
                reason=f"{IMPORT_REASON_PREFIX} historical active vehicle detail",
                is_preliminary=False,
            )
            archived_repair = Repair(
                order_number="ZN-ARCH-VEH-001",
                repair_date=date(2025, 1, 25),
                vehicle_id=1,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=122000,
                grand_total=30000,
                status=RepairStatus.ARCHIVED,
                is_preliminary=False,
            )
            archived_service_historical_repair = Repair(
                order_number="ZN-HIST-VEH-ARCH-SERVICE",
                repair_date=date(2025, 1, 24),
                vehicle_id=1,
                service_id=archived_service.id,
                created_by_user_id=employee.id,
                mileage=123000,
                grand_total=35000,
                status=RepairStatus.CONFIRMED,
                reason=f"{IMPORT_REASON_PREFIX} archived service vehicle detail",
                is_preliminary=False,
            )
            archived_historical_repair = Repair(
                order_number="ZN-HIST-VEH-ARCH-REPAIR",
                repair_date=date(2025, 1, 23),
                vehicle_id=1,
                service_id=service.id,
                created_by_user_id=employee.id,
                mileage=124000,
                grand_total=45000,
                status=RepairStatus.ARCHIVED,
                reason=f"{IMPORT_REASON_PREFIX} archived repair vehicle detail",
                is_preliminary=False,
            )
            db.add_all(
                [
                    active_historical_repair,
                    archived_repair,
                    archived_service_historical_repair,
                    archived_historical_repair,
                ]
            )
            db.flush()
            db.add(
                RepairCheck(
                    repair_id=active_historical_repair.id,
                    check_type="historical_vehicle_summary_suspicious",
                    severity=CheckSeverity.SUSPICIOUS,
                    title="Исторический ремонт требует проверки",
                    details="Скрытый suspicious check для проверки summary техники",
                    is_resolved=False,
                )
            )

            db.add_all(
                [
                    Document(
                        repair_id=active_historical_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="vehicle-active-history.pdf",
                        storage_key="documents/test/vehicle-active-history.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.NEEDS_REVIEW,
                        is_primary=False,
                        review_queue_priority=10,
                    ),
                    Document(
                        repair_id=active_historical_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="vehicle-active-history-archived.pdf",
                        storage_key="documents/test/vehicle-active-history-archived.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.REPEAT_SCAN,
                        status=DocumentStatus.ARCHIVED,
                        is_primary=False,
                        review_queue_priority=0,
                    ),
                    Document(
                        repair_id=archived_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="vehicle-archived-history.pdf",
                        storage_key="documents/test/vehicle-archived-history.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.NEEDS_REVIEW,
                        is_primary=False,
                        review_queue_priority=5,
                    ),
                ]
            )
            db.commit()

        response = self.client.get("/api/vehicles/1", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["historical_repairs_total"], 1)
        self.assertEqual(payload["historical_last_repair_date"], "2025-01-20")
        self.assertEqual(payload["history_summary"]["repairs_total"], 2)
        self.assertEqual(payload["history_summary"]["documents_total"], 2)
        self.assertEqual(payload["history_summary"]["confirmed_repairs"], 0)
        self.assertEqual(payload["history_summary"]["suspicious_repairs"], 1)
        self.assertEqual(payload["history_summary"]["last_repair_date"], "2025-01-20")
        self.assertEqual(payload["historical_history_summary"]["repairs_total"], 1)
        self.assertEqual(payload["historical_history_summary"]["services_total"], 1)
        self.assertEqual(payload["historical_history_summary"]["total_spend"], 25000.0)
        self.assertEqual(payload["historical_history_summary"]["first_repair_date"], "2025-01-20")
        self.assertEqual(payload["historical_history_summary"]["last_repair_date"], "2025-01-20")
        self.assertEqual(
            [repair["order_number"] for repair in payload["repair_history"]],
            ["ZN-HIST-VEH-001", "ZN-001"],
        )
        active_history_item = next(item for item in payload["repair_history"] if item["order_number"] == "ZN-HIST-VEH-001")
        self.assertEqual(active_history_item["documents_total"], 1)
        self.assertEqual(
            [repair["order_number"] for repair in payload["historical_repair_history"]],
            ["ZN-HIST-VEH-001"],
        )

    def test_vehicle_detail_and_export_skip_legacy_broken_assignment_users(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            broken_assignment = VehicleAssignmentHistory(
                vehicle_id=1,
                user_id=2,
                starts_at=date(2025, 2, 1),
                ends_at=None,
                assigned_by_user_id=1,
                comment="broken vehicle detail assignment",
            )
            db.add(broken_assignment)
            db.flush()
            broken_assignment_id = broken_assignment.id
            db.commit()

        with self.SessionLocal() as db:
            db.execute(text("PRAGMA foreign_keys = OFF"))
            db.execute(
                text("UPDATE vehicle_assignment_history SET user_id = 999999 WHERE id = :assignment_id"),
                {"assignment_id": broken_assignment_id},
            )
            db.commit()
            db.execute(text("PRAGMA foreign_keys = ON"))

        detail_response = self.client.get("/api/vehicles/1", headers=headers)

        self.assertEqual(detail_response.status_code, 200, detail_response.text)
        detail_payload = detail_response.json()
        self.assertEqual(len(detail_payload["active_assignments"]), 1)
        self.assertEqual(detail_payload["active_assignments"][0]["comment"], "Primary assignment")

        export_response = self.client.get("/api/vehicles/1/export", headers=headers)

        self.assertEqual(export_response.status_code, 200, export_response.text)
        workbook = load_workbook(filename=BytesIO(export_response.content))
        assignment_rows = [
            row
            for row in workbook["Закрепления"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        self.assertEqual(len(assignment_rows), 1)
        self.assertEqual(assignment_rows[0][0], "Employee User")

    def test_vehicle_export_excludes_archived_repairs_and_archived_service_history(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            employee = db.scalar(select(User).where(User.login == "employee"))
            service = db.scalar(select(Service).where(Service.name == "Service Alpha"))
            self.assertIsNotNone(employee)
            self.assertIsNotNone(service)
            assert employee is not None
            assert service is not None

            archived_service = Service(
                name="Archived Vehicle Export Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
            )
            db.add(archived_service)
            db.flush()

            db.add_all(
                [
                    Repair(
                        order_number="ZN-HIST-EXPORT-001",
                        repair_date=date(2025, 1, 21),
                        vehicle_id=1,
                        service_id=service.id,
                        created_by_user_id=employee.id,
                        mileage=125000,
                        grand_total=26000,
                        status=RepairStatus.CONFIRMED,
                        reason=f"{IMPORT_REASON_PREFIX} historical export vehicle",
                        is_preliminary=False,
                    ),
                    Repair(
                        order_number="ZN-HIST-EXPORT-ARCH-SERVICE",
                        repair_date=date(2025, 1, 22),
                        vehicle_id=1,
                        service_id=archived_service.id,
                        created_by_user_id=employee.id,
                        mileage=126000,
                        grand_total=36000,
                        status=RepairStatus.CONFIRMED,
                        reason=f"{IMPORT_REASON_PREFIX} archived service export vehicle",
                        is_preliminary=False,
                    ),
                    Repair(
                        order_number="ZN-EXPORT-ARCH-001",
                        repair_date=date(2025, 1, 23),
                        vehicle_id=1,
                        service_id=service.id,
                        created_by_user_id=employee.id,
                        mileage=127000,
                        grand_total=28000,
                        status=RepairStatus.ARCHIVED,
                        is_preliminary=False,
                    ),
                ]
            )
            db.flush()

            historical_export_repair = db.scalar(select(Repair).where(Repair.order_number == "ZN-HIST-EXPORT-001"))
            self.assertIsNotNone(historical_export_repair)
            assert historical_export_repair is not None

            db.add_all(
                [
                    Document(
                        repair_id=historical_export_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="vehicle-export-active-history.pdf",
                        storage_key="documents/test/vehicle-export-active-history.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.NEEDS_REVIEW,
                        is_primary=False,
                        review_queue_priority=10,
                    ),
                    Document(
                        repair_id=historical_export_repair.id,
                        uploaded_by_user_id=employee.id,
                        original_filename="vehicle-export-active-history-archived.pdf",
                        storage_key="documents/test/vehicle-export-active-history-archived.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.REPEAT_SCAN,
                        status=DocumentStatus.ARCHIVED,
                        is_primary=False,
                        review_queue_priority=0,
                    ),
                ]
            )
            db.commit()

        response = self.client.get("/api/vehicles/1/export", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers["content-type"],
        )

        workbook = load_workbook(filename=BytesIO(response.content))
        summary_rows = {
            str(row[0]): row[1]
            for row in workbook["Техника"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        }
        self.assertEqual(summary_rows["Ремонтов"], 2)
        self.assertEqual(summary_rows["Исторических ремонтов 2025"], 1)

        repairs_rows = [
            row
            for row in workbook["Ремонты"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        self.assertEqual([row[1] for row in repairs_rows], ["ZN-HIST-EXPORT-001", "ZN-001"])
        export_history_row = next(row for row in repairs_rows if row[1] == "ZN-HIST-EXPORT-001")
        self.assertEqual(export_history_row[7], 1)

        historical_rows = [
            row
            for row in workbook["История 2025"].iter_rows(min_row=2, values_only=True)
            if row and row[0]
        ]
        self.assertEqual([row[1] for row in historical_rows], ["ZN-HIST-EXPORT-001"])

    def test_global_search_excludes_archived_vehicle_matches(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            db.add(
                Vehicle(
                    external_id="truck-archived-search",
                    vehicle_type=VehicleType.TRUCK,
                    plate_number="ARCH-SEARCH-116",
                    brand="Archived",
                    model="Search",
                    status=VehicleStatus.ARCHIVED,
                )
            )
            db.commit()

        response = self.client.get("/api/search/global?q=ARCH-SEARCH-116", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["vehicles_total"], 0)
        self.assertEqual(payload["vehicles"], [])
        self.assertEqual(payload["repairs_total"], 0)
        self.assertEqual(payload["repairs"], [])
        self.assertEqual(payload["documents_total"], 0)
        self.assertEqual(payload["documents"], [])

    def test_global_search_excludes_repairs_and_documents_with_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            archived_service = Service(
                name="Archived Search Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
            )
            db.add(archived_service)
            db.flush()

            repair = Repair(
                order_number="ZN-ARCH-SERVICE-001",
                repair_date=date(2025, 1, 25),
                vehicle_id=1,
                service_id=archived_service.id,
                created_by_user_id=1,
                mileage=128000,
                reason="Archived service search regression",
                grand_total=18000,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
            )
            db.add(repair)
            db.flush()

            document = Document(
                repair_id=repair.id,
                uploaded_by_user_id=1,
                original_filename="archived-service-order.pdf",
                storage_key="documents/test/archived-service-order.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.CONFIRMED,
                is_primary=True,
                review_queue_priority=0,
            )
            db.add(document)
            db.flush()
            repair.source_document_id = document.id
            db.commit()

        response = self.client.get("/api/search/global?q=Archived%20Search%20Service", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["repairs_total"], 0)
        self.assertEqual(payload["repairs"], [])
        self.assertEqual(payload["documents_total"], 0)
        self.assertEqual(payload["documents"], [])

    def test_document_list_excludes_documents_from_archived_vehicle_and_archived_service(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            archived_vehicle = Vehicle(
                external_id="truck-doc-archived",
                vehicle_type=VehicleType.TRUCK,
                plate_number="DOC-ARCH-116",
                brand="Archived",
                model="Vehicle",
                status=VehicleStatus.ARCHIVED,
            )
            archived_service = Service(
                name="Archived Documents Service",
                city="Kazan",
                status=ServiceStatus.ARCHIVED,
                created_by_user_id=1,
            )
            db.add_all([archived_vehicle, archived_service])
            db.flush()

            archived_vehicle_repair = Repair(
                order_number="ZN-DOC-ARCH-VEH-001",
                repair_date=date(2025, 1, 26),
                vehicle_id=archived_vehicle.id,
                created_by_user_id=1,
                mileage=129000,
                reason="Archived vehicle document list regression",
                grand_total=11000,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
            )
            archived_service_repair = Repair(
                order_number="ZN-DOC-ARCH-SVC-001",
                repair_date=date(2025, 1, 27),
                vehicle_id=1,
                service_id=archived_service.id,
                created_by_user_id=1,
                mileage=130000,
                reason="Archived service document list regression",
                grand_total=12000,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
            )
            db.add_all([archived_vehicle_repair, archived_service_repair])
            db.flush()

            db.add_all(
                [
                    Document(
                        repair_id=archived_vehicle_repair.id,
                        uploaded_by_user_id=1,
                        original_filename="archived-vehicle-doc.pdf",
                        storage_key="documents/test/archived-vehicle-doc.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.CONFIRMED,
                        is_primary=True,
                        review_queue_priority=0,
                    ),
                    Document(
                        repair_id=archived_service_repair.id,
                        uploaded_by_user_id=1,
                        original_filename="archived-service-doc.pdf",
                        storage_key="documents/test/archived-service-doc.pdf",
                        mime_type="application/pdf",
                        source_type="pdf",
                        kind=DocumentKind.ORDER,
                        status=DocumentStatus.CONFIRMED,
                        is_primary=True,
                        review_queue_priority=0,
                    ),
                ]
            )
            db.commit()

        response = self.client.get("/api/documents?limit=20", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        visible_filenames = [item["original_filename"] for item in payload["items"]]
        self.assertIn("repair-order.pdf", visible_filenames)
        self.assertNotIn("archived-vehicle-doc.pdf", visible_filenames)
        self.assertNotIn("archived-service-doc.pdf", visible_filenames)

    def test_document_list_excludes_archived_documents_inside_active_repairs(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            archived_document = Document(
                repair_id=1,
                uploaded_by_user_id=1,
                original_filename="archived-active-repair-doc.pdf",
                storage_key="documents/test/archived-active-repair-doc.pdf",
                mime_type="application/pdf",
                source_type="pdf",
                kind=DocumentKind.ORDER,
                status=DocumentStatus.ARCHIVED,
                is_primary=False,
                review_queue_priority=0,
            )
            db.add(archived_document)
            db.commit()

        response = self.client.get("/api/documents?limit=20", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        visible_filenames = [item["original_filename"] for item in payload["items"]]
        self.assertIn("repair-order.pdf", visible_filenames)
        self.assertNotIn("archived-active-repair-doc.pdf", visible_filenames)

    def test_global_search_excludes_archived_repair_and_archived_document_matches(self) -> None:
        headers = self._get_auth_headers("admin")

        with self.SessionLocal() as db:
            archived_repair = Repair(
                order_number="ZN-ARCH-SEARCH-REPAIR-001",
                repair_date=date(2025, 1, 28),
                vehicle_id=1,
                service_id=1,
                created_by_user_id=1,
                mileage=131000,
                reason="Archived repair search regression",
                grand_total=13000,
                status=RepairStatus.ARCHIVED,
                is_preliminary=False,
            )
            active_repair_with_archived_doc = Repair(
                order_number="ZN-ACTIVE-ARCH-DOC-001",
                repair_date=date(2025, 1, 29),
                vehicle_id=1,
                service_id=1,
                created_by_user_id=1,
                mileage=132000,
                reason="Archived document search regression",
                grand_total=14000,
                status=RepairStatus.CONFIRMED,
                is_preliminary=False,
            )
            db.add_all([archived_repair, active_repair_with_archived_doc])
            db.flush()

            db.add(
                Document(
                    repair_id=active_repair_with_archived_doc.id,
                    uploaded_by_user_id=1,
                    original_filename="archived-search-document.pdf",
                    storage_key="documents/test/archived-search-document.pdf",
                    mime_type="application/pdf",
                    source_type="pdf",
                    kind=DocumentKind.ORDER,
                    status=DocumentStatus.ARCHIVED,
                    is_primary=False,
                    review_queue_priority=0,
                )
            )
            db.commit()

        repair_response = self.client.get("/api/search/global?q=ZN-ARCH-SEARCH-REPAIR-001", headers=headers)
        self.assertEqual(repair_response.status_code, 200, repair_response.text)
        repair_payload = repair_response.json()
        self.assertEqual(repair_payload["repairs_total"], 0)
        self.assertEqual(repair_payload["repairs"], [])
        self.assertEqual(repair_payload["documents_total"], 0)
        self.assertEqual(repair_payload["documents"], [])

        document_response = self.client.get("/api/search/global?q=archived-search-document.pdf", headers=headers)
        self.assertEqual(document_response.status_code, 200, document_response.text)
        document_payload = document_response.json()
        self.assertEqual(document_payload["documents_total"], 0)
        self.assertEqual(document_payload["documents"], [])

    def test_delete_endpoint_archives_repair_instead_of_removing_it(self) -> None:
        headers = self._get_auth_headers("admin")

        response = self.client.delete("/api/repairs/1", headers=headers)

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["message"], "Заказ-наряд и связанные документы отправлены в архив")

        with self.SessionLocal() as db:
            repair = db.get(Repair, 1)
            document = db.get(Document, 1)
            self.assertIsNotNone(repair)
            self.assertIsNotNone(document)
            assert repair is not None
            assert document is not None
            self.assertEqual(repair.status, RepairStatus.ARCHIVED)
            self.assertEqual(document.status, DocumentStatus.ARCHIVED)
            self.assertIsNone(repair.source_document_id)
            self.assertFalse(document.is_primary)


if __name__ == "__main__":
    unittest.main()
